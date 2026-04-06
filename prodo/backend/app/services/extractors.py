# mypy: ignore-errors
"""
PDF Table Extraction using Multiple Tools.

Supports:
- Tabula (tabula-py): Best for well-structured tables
- Camelot: Best for complex table layouts with borders
- PyMuPDF (fitz): Fast, general purpose
- pdfplumber: Detailed layout analysis
- Marker: PDF to markdown conversion

Features:
- Multiple extraction methods with automatic fallback
- OCR support for scanned PDFs (via Tesseract/EasyOCR)
- Intelligent table detection and confidence scoring
- Layout-aware text extraction
- Parallel processing for multi-page documents
- Smart header detection

Each extractor has different strengths - use compare_extractors() to find the best one.
"""
from __future__ import annotations

import concurrent.futures
import copy
import hashlib
import json
import logging
import os
import re
import tempfile
import threading
import time
import warnings
from abc import ABC, abstractmethod
from dataclasses import asdict, dataclass, field, replace
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

from backend.app.services.config import get_settings

logger = logging.getLogger("neura.extraction.pdf")

@dataclass
class ExtractionConfig:
    """Configuration for PDF extraction."""
    max_pages: int = 100
    max_auto_pages: int = 25
    max_compare_pages: int = 10
    max_tables_per_page: int = 10
    min_rows_for_table: int = 2
    min_cols_for_table: int = 2
    ocr_enabled: bool = True
    ocr_language: str = "eng"
    parallel_pages: bool = True
    max_workers: int = 4
    confidence_threshold: float = 0.5
    detect_headers: bool = True
    preserve_layout: bool = True

DEFAULT_CONFIG = ExtractionConfig()

# Cache/dedupe configuration (set to 0 to disable)
_CACHE_TTL_SECONDS = int(os.getenv("NEURA_PDF_EXTRACT_CACHE_TTL_SECONDS", "300"))
_CACHE_MAX_ITEMS = int(os.getenv("NEURA_PDF_EXTRACT_CACHE_MAX_ITEMS", "128"))
_CACHE_DEDUPE_ENABLED = os.getenv("NEURA_PDF_EXTRACT_DEDUPE", "true").strip().lower() in {"1", "true", "yes"}
_CACHE_WAIT_SECONDS = float(os.getenv("NEURA_PDF_EXTRACT_DEDUPE_WAIT_SECONDS", "30"))

@dataclass
class ExtractedTable:
    """A table extracted from a PDF."""
    id: str
    page: int
    headers: List[str]
    rows: List[List[str]]
    confidence: float = 1.0
    method: str = "unknown"
    bbox: Optional[Tuple[float, float, float, float]] = None  # x0, y0, x1, y1
    metadata: Dict[str, Any] = field(default_factory=dict)

    def __post_init__(self):
        """Validate and clean up table data."""
        # Ensure headers are strings
        self.headers = [str(h).strip() if h else "" for h in self.headers]
        # Ensure rows are list of string lists
        self.rows = [
            [str(cell).strip() if cell else "" for cell in row]
            for row in self.rows
        ]

    @property
    def row_count(self) -> int:
        """Number of data rows."""
        return len(self.rows)

    @property
    def col_count(self) -> int:
        """Number of columns."""
        return len(self.headers) if self.headers else (len(self.rows[0]) if self.rows else 0)

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary."""
        return {
            "id": self.id,
            "page": self.page,
            "headers": self.headers,
            "rows": self.rows,
            "confidence": self.confidence,
            "method": self.method,
            "bbox": self.bbox,
            "metadata": self.metadata,
            "row_count": self.row_count,
            "col_count": self.col_count,
        }

    def get_column(self, col_name: str) -> List[str]:
        """Get all values from a column by header name."""
        try:
            idx = self.headers.index(col_name)
            return [row[idx] if idx < len(row) else "" for row in self.rows]
        except ValueError:
            return []

@dataclass
class ExtractionResult:
    """Result of PDF extraction."""
    tables: List[ExtractedTable]
    text: str
    page_count: int
    method: str
    errors: List[str] = field(default_factory=list)
    metadata: Dict[str, Any] = field(default_factory=dict)
    extraction_time_ms: float = 0.0
    ocr_used: bool = False

    def __post_init__(self) -> None:
        self.metadata = dict(self.metadata or {})
        self.metadata.setdefault("error_count", len(self.errors))
        self.metadata.setdefault(
            "partial_success",
            bool(self.errors) and (bool(self.tables) or bool(self.text)),
        )

    @property
    def has_tables(self) -> bool:
        """Check if any tables were extracted."""
        return len(self.tables) > 0

    @property
    def total_rows(self) -> int:
        """Total number of rows across all tables."""
        return sum(t.row_count for t in self.tables)

    def get_table_by_page(self, page: int) -> List[ExtractedTable]:
        """Get all tables from a specific page."""
        return [t for t in self.tables if t.page == page]

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary."""
        return {
            "tables": [t.to_dict() for t in self.tables],
            "text": self.text,
            "page_count": self.page_count,
            "method": self.method,
            "errors": self.errors,
            "metadata": self.metadata,
            "extraction_time_ms": self.extraction_time_ms,
            "ocr_used": self.ocr_used,
            "has_tables": self.has_tables,
            "total_rows": self.total_rows,
        }

@dataclass
class _ExtractionCacheEntry:
    result: ExtractionResult
    created_at: float
    last_access: float
    hits: int = 0

_EXTRACTION_CACHE: dict[str, _ExtractionCacheEntry] = {}
_EXTRACTION_INFLIGHT: dict[str, threading.Event] = {}
_EXTRACTION_CACHE_LOCK = threading.Lock()

def _cache_enabled() -> bool:
    return _CACHE_DEDUPE_ENABLED and _CACHE_TTL_SECONDS > 0 and _CACHE_MAX_ITEMS > 0

def _build_cache_key(
    pdf_path: Path,
    method: str,
    pages: Optional[List[int]],
    config: ExtractionConfig,
) -> str:
    try:
        stat = pdf_path.stat()
        mtime_ns = stat.st_mtime_ns
        size = stat.st_size
    except OSError:
        mtime_ns = 0
        size = 0
    payload = {
        "path": str(pdf_path),
        "mtime_ns": mtime_ns,
        "size": size,
        "method": method,
        "pages": pages,
        "config": asdict(config),
    }
    raw = json.dumps(payload, sort_keys=True, default=str)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()

def _clone_result(result: ExtractionResult) -> ExtractionResult:
    return copy.deepcopy(result)

def _mark_cache_hit(entry: _ExtractionCacheEntry) -> ExtractionResult:
    cached = _clone_result(entry.result)
    cached.metadata = dict(cached.metadata or {})
    cached.metadata["cache_hit"] = True
    cached.metadata["cache_age_s"] = int(time.time() - entry.created_at)
    return cached

def _prune_cache_locked(now: float) -> None:
    expired = [key for key, entry in _EXTRACTION_CACHE.items() if now - entry.created_at > _CACHE_TTL_SECONDS]
    for key in expired:
        _EXTRACTION_CACHE.pop(key, None)
    if len(_EXTRACTION_CACHE) <= _CACHE_MAX_ITEMS:
        return
    ordered = sorted(_EXTRACTION_CACHE.items(), key=lambda item: item[1].last_access)
    for key, _entry in ordered[: max(0, len(_EXTRACTION_CACHE) - _CACHE_MAX_ITEMS)]:
        _EXTRACTION_CACHE.pop(key, None)

def _get_cached_result(cache_key: str) -> Optional[ExtractionResult]:
    if not _cache_enabled():
        return None
    now = time.time()
    with _EXTRACTION_CACHE_LOCK:
        entry = _EXTRACTION_CACHE.get(cache_key)
        if not entry:
            return None
        if now - entry.created_at > _CACHE_TTL_SECONDS:
            _EXTRACTION_CACHE.pop(cache_key, None)
            return None
        entry.hits += 1
        entry.last_access = now
        return _mark_cache_hit(entry)

def _store_cache_result(cache_key: str, result: ExtractionResult) -> None:
    if not _cache_enabled():
        return
    now = time.time()
    with _EXTRACTION_CACHE_LOCK:
        _EXTRACTION_CACHE[cache_key] = _ExtractionCacheEntry(
            result=_clone_result(result),
            created_at=now,
            last_access=now,
            hits=0,
        )
        _prune_cache_locked(now)

def _acquire_inflight(cache_key: str) -> tuple[threading.Event, bool]:
    with _EXTRACTION_CACHE_LOCK:
        existing = _EXTRACTION_INFLIGHT.get(cache_key)
        if existing:
            return existing, False
        event = threading.Event()
        _EXTRACTION_INFLIGHT[cache_key] = event
        return event, True

def _release_inflight(cache_key: str, event: threading.Event) -> None:
    with _EXTRACTION_CACHE_LOCK:
        if _EXTRACTION_INFLIGHT.get(cache_key) is event:
            _EXTRACTION_INFLIGHT.pop(cache_key, None)
    event.set()

def _get_ocr_max_pixels() -> int:
    raw = os.getenv("NEURA_OCR_MAX_PIXELS", "20000000")
    try:
        value = int(raw)
    except (TypeError, ValueError):
        value = 20000000
    return max(value, 1000000)

def _load_image_for_ocr(image_bytes: bytes):
    import io
    from PIL import Image, UnidentifiedImageError

    max_pixels = _get_ocr_max_pixels()
    if max_pixels:
        Image.MAX_IMAGE_PIXELS = max_pixels

    with warnings.catch_warnings():
        warnings.simplefilter("error", Image.DecompressionBombWarning)
        try:
            image = Image.open(io.BytesIO(image_bytes))
            image.load()
            return image
        except (Image.DecompressionBombError, Image.DecompressionBombWarning) as exc:
            raise ValueError("Image exceeds pixel limit") from exc
        except UnidentifiedImageError as exc:
            raise ValueError("Unsupported image format") from exc

class OCREngine:
    """OCR engine abstraction supporting multiple backends.

    Priority: GLM-OCR (best) → EasyOCR → Tesseract.
    """

    def __init__(self, language: str = "eng"):
        self.language = language
        self._engine: Optional[str] = None
        self._lock = threading.Lock()
        self._easyocr_reader = None
        self._easyocr_lock = threading.Lock()

    def _detect_engine(self) -> str:
        """Detect available OCR engine."""
        if self._engine:
            return self._engine

        with self._lock:
            # Try GLM-OCR first (best accuracy for industrial documents)
            try:
                from backend.app.services.llm import get_llm_config
                config = get_llm_config()
                if config.vision_enabled and config.vision_model:
                    self._engine = "glm_ocr"
                    return self._engine
            except Exception:
                pass

            # Try EasyOCR
            try:
                import easyocr  # noqa: F401
                self._engine = "easyocr"
                return self._engine
            except ImportError:
                pass

            # Try Tesseract
            try:
                import pytesseract
                pytesseract.get_tesseract_version()
                self._engine = "tesseract"
                return self._engine
            except ImportError:
                logger.debug("pytesseract not installed")
            except Exception as e:
                logger.debug(f"Tesseract not available: {e}")

            self._engine = "none"
            return self._engine

    def extract_text(self, image_bytes: bytes) -> str:
        """Extract text from image bytes using OCR."""
        engine = self._detect_engine()

        if engine == "none":
            raise RuntimeError("OCR engine not available")

        if engine == "glm_ocr":
            try:
                return self._ocr_glm(image_bytes)
            except Exception as exc:
                logger.warning(f"OCR extraction failed (glm_ocr): {exc}")
                # Fall through to easyocr/tesseract
                for fallback in (self._ocr_easyocr_safe, self._ocr_tesseract_safe):
                    text = fallback(image_bytes)
                    if text:
                        return text
                raise RuntimeError("OCR extraction failed") from exc

        if engine == "easyocr":
            try:
                return self._ocr_easyocr(image_bytes)
            except Exception as exc:
                logger.warning(f"OCR extraction failed (easyocr): {exc}")
                text = self._ocr_tesseract_safe(image_bytes)
                if text:
                    return text
                raise RuntimeError("OCR extraction failed") from exc

        if engine == "tesseract":
            return self._ocr_tesseract(image_bytes)

        raise RuntimeError("OCR engine not available")

    def _ocr_glm(self, image_bytes: bytes) -> str:
        """Extract text using GLM-OCR via Ollama."""
        from backend.app.services.infra_services import ocr_extract
        text = ocr_extract(image_bytes)
        if not text:
            raise RuntimeError("GLM-OCR returned no text")
        return text

    def _ocr_easyocr(self, image_bytes: bytes) -> str:
        """Extract text using EasyOCR."""
        import easyocr
        import numpy as np

        image_array = np.array(_load_image_for_ocr(image_bytes))

        reader = self._easyocr_reader
        if reader is None:
            with self._easyocr_lock:
                reader = self._easyocr_reader
                if reader is None:
                    reader = easyocr.Reader([self.language[:2]], gpu=False)
                    self._easyocr_reader = reader
        results = reader.readtext(image_array)
        return " ".join([text for _, text, _ in results])

    def _ocr_easyocr_safe(self, image_bytes: bytes) -> Optional[str]:
        """EasyOCR fallback that returns None instead of raising."""
        try:
            return self._ocr_easyocr(image_bytes)
        except Exception:
            return None

    def _ocr_tesseract(self, image_bytes: bytes) -> str:
        """Extract text using Tesseract."""
        import pytesseract
        image = _load_image_for_ocr(image_bytes)
        return pytesseract.image_to_string(image, lang=self.language)

    def _ocr_tesseract_safe(self, image_bytes: bytes) -> Optional[str]:
        """Tesseract fallback that returns None instead of raising."""
        try:
            import pytesseract  # noqa: F401
            return self._ocr_tesseract(image_bytes)
        except Exception:
            return None

    def is_available(self) -> bool:
        """Check if OCR is available."""
        return self._detect_engine() != "none"

# Global OCR engines per language
_ocr_engines: Dict[str, OCREngine] = {}

def get_ocr_engine(language: str = "eng") -> OCREngine:
    """Get or create OCR engine for a specific language."""
    lang = (language or "eng").strip() or "eng"
    engine = _ocr_engines.get(lang)
    if engine is None:
        engine = OCREngine(lang)
        _ocr_engines[lang] = engine
    return engine

def _resolve_config(config: Optional[ExtractionConfig]) -> ExtractionConfig:
    return config or DEFAULT_CONFIG

def _safe_error_message(prefix: str, exc: Exception | None = None) -> str:
    if exc is None:
        return prefix
    return f"{prefix} ({exc.__class__.__name__})"

def _limit_config_pages(config: ExtractionConfig, limit: Optional[int]) -> ExtractionConfig:
    if limit is None:
        return config
    try:
        limit_value = int(limit)
    except (TypeError, ValueError):
        return config
    if limit_value <= 0:
        return config
    if config.max_pages > limit_value:
        return replace(config, max_pages=limit_value)
    return config

def _is_relative_to(path: Path, base: Path) -> bool:
    try:
        path.relative_to(base)
        return True
    except ValueError:
        return False

def _is_unc_path(path: Path) -> bool:
    raw = str(path)
    return raw.startswith("\\\\") or raw.startswith("//")

def _resolve_pdf_path(pdf_path: Union[str, Path]) -> tuple[Optional[Path], Optional[str]]:
    if pdf_path is None:
        return None, "pdf_path is required"
    try:
        candidate = Path(pdf_path)
    except TypeError:
        return None, "pdf_path must be a valid filesystem path"

    if candidate.suffix.lower() != ".pdf":
        return None, "pdf_path must point to a .pdf file"

    try:
        resolved = candidate.resolve()
    except Exception:
        resolved = candidate

    if not resolved.exists():
        return None, "pdf_path does not exist"
    if not resolved.is_file():
        return None, "pdf_path must be a file"

    settings = get_settings()
    if not settings.allow_unsafe_pdf_paths:
        if _is_unc_path(resolved):
            return None, "UNC paths are not allowed"
        allowed_roots = [
            Path(settings.uploads_root).resolve(),
            Path(settings.excel_uploads_root).resolve(),
            Path(tempfile.gettempdir()).resolve(),
        ]
        if not any(_is_relative_to(resolved, root) for root in allowed_roots):
            return None, "pdf_path must be within uploads or temp directories"

    return resolved, None

def _normalize_pages_input(
    pages: Optional[List[int]],
) -> tuple[Optional[List[int]], Optional[str]]:
    if pages is None:
        return None, None
    if not isinstance(pages, (list, tuple)):
        return None, "pages must be a list of non-negative integers"
    if not pages:
        return None, "pages must contain at least one page index"
    normalized: list[int] = []
    for page in pages:
        if isinstance(page, bool) or not isinstance(page, int):
            return None, "pages must be a list of non-negative integers"
        if page < 0:
            return None, "pages must be non-negative"
        normalized.append(page)
    deduped: list[int] = []
    seen: set[int] = set()
    for page in normalized:
        if page in seen:
            continue
        seen.add(page)
        deduped.append(page)
    return deduped, None

def _resolve_pages_to_process(
    pdf_path: Union[str, Path],
    pages: Optional[List[int]],
    config: ExtractionConfig,
    *,
    page_count: Optional[int] = None,
    max_pages: Optional[int] = None,
) -> tuple[List[int], Optional[str]]:
    normalized, error = _normalize_pages_input(pages)
    if error:
        return [], error

    page_count = page_count if page_count is not None else _get_pdf_page_count(pdf_path)

    if normalized is None:
        if page_count is None:
            normalized = list(range(config.max_pages))
        else:
            normalized = list(range(page_count))
    else:
        if page_count is not None:
            out_of_range = [p for p in normalized if p >= page_count]
            if out_of_range:
                max_index = max(page_count - 1, 0)
                return [], f"pages out of range (max index {max_index})"

    limit = max_pages if max_pages is not None else config.max_pages
    if limit and len(normalized) > limit:
        normalized = normalized[:limit]

    return normalized, None

def _get_pdf_page_count(pdf_path: Union[str, Path]) -> Optional[int]:
    try:
        import fitz
    except ImportError:
        return None
    try:
        doc = fitz.open(pdf_path)
        count = doc.page_count
        doc.close()
        return count
    except Exception:
        return None

def _resolve_page_numbers(
    pdf_path: Union[str, Path],
    pages: Optional[List[int]],
    config: ExtractionConfig,
    *,
    page_count: Optional[int] = None,
    max_pages: Optional[int] = None,
) -> tuple[List[int], Optional[str]]:
    return _resolve_pages_to_process(
        pdf_path,
        pages,
        config,
        page_count=page_count,
        max_pages=max_pages,
    )

def _normalize_table_data(
    data: List[List[Any]],
    config: ExtractionConfig,
) -> Tuple[List[str], List[List[str]]]:
    if not data:
        return [], []

    cleaned_rows = [
        [_clean_cell_value(cell) for cell in row]
        for row in data
    ]

    if not cleaned_rows:
        return [], []

    if config.detect_headers:
        header_candidate = cleaned_rows[0]
        if _is_header_row(header_candidate, cleaned_rows):
            headers = [h if h else f"Column_{i+1}" for i, h in enumerate(header_candidate)]
            rows = cleaned_rows[1:]
        else:
            headers = [f"Column_{i+1}" for i in range(len(cleaned_rows[0]))]
            rows = cleaned_rows
    else:
        header_candidate = cleaned_rows[0]
        headers = [h if h else f"Column_{i+1}" for i, h in enumerate(header_candidate)]
        rows = cleaned_rows[1:]

    normalized_rows: List[List[str]] = []
    for row in rows:
        normalized_row = [row[i] if i < len(row) else "" for i in range(len(headers))]
        normalized_rows.append(normalized_row)

    return headers, normalized_rows

def _apply_table_confidence(
    table: ExtractedTable,
    config: ExtractionConfig,
    base_confidence: float,
) -> float:
    confidence = _calculate_table_confidence(table, config)
    return min(1.0, max(0.0, base_confidence * confidence))

def _table_meets_requirements(table: ExtractedTable, config: ExtractionConfig) -> bool:
    if table.row_count < config.min_rows_for_table:
        return False
    if table.col_count < config.min_cols_for_table:
        return False
    if table.confidence < config.confidence_threshold:
        return False
    return True

def _is_header_row(row: List[str], all_rows: List[List[str]]) -> bool:
    """
    Detect if a row is likely a header row.

    Uses heuristics:
    - Headers are often shorter than data
    - Headers contain fewer numbers
    - Headers have distinct patterns (all caps, title case)
    """
    if not row or not all_rows:
        return False

    # Count numeric values in the row
    num_numeric = sum(1 for cell in row if _is_numeric(cell))
    num_total = len([c for c in row if c.strip()])

    if num_total == 0:
        return False

    numeric_ratio = num_numeric / num_total

    # Headers usually have fewer numeric values
    if numeric_ratio < 0.3:
        # Check if other rows have more numeric values
        if len(all_rows) > 1:
            other_numeric_ratios = []
            for other_row in all_rows[1:min(5, len(all_rows))]:
                other_num = sum(1 for cell in other_row if _is_numeric(cell))
                other_total = len([c for c in other_row if c.strip()])
                if other_total > 0:
                    other_numeric_ratios.append(other_num / other_total)

            if other_numeric_ratios:
                avg_other_ratio = sum(other_numeric_ratios) / len(other_numeric_ratios)
                if avg_other_ratio > numeric_ratio + 0.2:
                    return True

    # Check for common header patterns
    header_patterns = [
        lambda c: c.isupper(),  # ALL CAPS
        lambda c: c.istitle(),  # Title Case
        lambda c: c.lower() in ("id", "name", "date", "amount", "total", "qty", "price", "description"),
    ]

    pattern_matches = sum(
        1 for cell in row
        if cell.strip() and any(p(cell.strip()) for p in header_patterns)
    )

    return pattern_matches >= len(row) // 2

def _is_numeric(value: str) -> bool:
    """Check if a string represents a numeric value."""
    if not value or not value.strip():
        return False

    cleaned = value.strip().replace(",", "").replace("$", "").replace("%", "")
    cleaned = cleaned.lstrip("-+")

    try:
        float(cleaned)
        return True
    except ValueError:
        return False

def _clean_cell_value(value: Any) -> str:
    """Clean and normalize a cell value."""
    if value is None:
        return ""

    text = str(value).strip()

    # Remove excessive whitespace
    text = re.sub(r'\s+', ' ', text)

    # Remove null characters
    text = text.replace('\x00', '')

    return text

def _calculate_table_confidence(
    table: ExtractedTable,
    config: ExtractionConfig,
) -> float:
    """Calculate confidence score for extracted table."""
    confidence = 1.0

    # Penalize tables with few rows
    if table.row_count < config.min_rows_for_table:
        confidence *= 0.5

    # Penalize tables with few columns
    if table.col_count < config.min_cols_for_table:
        confidence *= 0.5

    # Penalize tables with many empty cells
    empty_cells = sum(
        1 for row in table.rows
        for cell in row
        if not cell.strip()
    )
    total_cells = table.row_count * table.col_count
    if total_cells > 0:
        empty_ratio = empty_cells / total_cells
        if empty_ratio > 0.5:
            confidence *= (1 - empty_ratio)

    # Penalize tables with inconsistent row lengths
    expected_cols = table.col_count
    inconsistent_rows = sum(
        1 for row in table.rows
        if len(row) != expected_cols
    )
    if table.row_count > 0:
        inconsistent_ratio = inconsistent_rows / table.row_count
        confidence *= (1 - inconsistent_ratio * 0.5)

    # Boost confidence if headers look valid
    if table.headers and all(h.strip() for h in table.headers):
        confidence *= 1.1

    return min(1.0, max(0.0, confidence))

class PDFExtractor(ABC):
    """Abstract base class for PDF extractors."""

    name: str = "base"

    @abstractmethod
    def is_available(self) -> bool:
        """Check if this extractor is available."""
        pass

    @abstractmethod
    def extract_tables(
        self,
        pdf_path: Union[str, Path],
        pages: Optional[List[int]] = None,
        config: Optional[ExtractionConfig] = None,
    ) -> ExtractionResult:
        """Extract tables from a PDF."""
        pass

    def extract_text(
        self,
        pdf_path: Union[str, Path],
        pages: Optional[List[int]] = None,
        config: Optional[ExtractionConfig] = None,
    ) -> str:
        """Extract text from a PDF."""
        # Default implementation - subclasses can override
        result = self.extract_tables(pdf_path, pages, config=config)
        return result.text

class TabulaExtractor(PDFExtractor):
    """
    Tabula-based PDF table extraction.

    Best for:
    - Well-structured tables with clear borders
    - Tables spanning multiple pages
    - Consistent column layouts

    Requires: tabula-py (pip install tabula-py) + Java Runtime
    """

    name = "tabula"

    def is_available(self) -> bool:
        try:
            import tabula
            # Try to check if Java is available
            return True
        except ImportError:
            return False

    def extract_tables(
        self,
        pdf_path: Union[str, Path],
        pages: Optional[List[int]] = None,
        config: Optional[ExtractionConfig] = None,
    ) -> ExtractionResult:
        try:
            import tabula
        except ImportError:
            return ExtractionResult(
                tables=[],
                text="",
                page_count=0,
                method=self.name,
                errors=["tabula-py not installed. Run: pip install tabula-py"],
            )

        config = _resolve_config(config)
        pdf_path, path_error = _resolve_pdf_path(pdf_path)
        if path_error:
            return ExtractionResult(
                tables=[],
                text="",
                page_count=0,
                method=self.name,
                errors=[path_error],
            )
        tables: List[ExtractedTable] = []
        errors: List[str] = []
        base_confidence = 0.85

        try:
            pages_to_process, page_error = _resolve_page_numbers(pdf_path, pages, config)
            if page_error:
                return ExtractionResult(
                    tables=[],
                    text="",
                    page_count=0,
                    method=self.name,
                    errors=[page_error],
                )

            for page_num in pages_to_process:
                page_table_count = 0
                try:
                    dfs = tabula.read_pdf(
                        str(pdf_path),
                        pages=page_num + 1,  # tabula uses 1-based
                        multiple_tables=True,
                        pandas_options={"header": None},
                    )
                except Exception as exc:
                    errors.append(_safe_error_message(f"Tabula failed on page {page_num + 1}", exc))
                    continue

                for i, df in enumerate(dfs):
                    if page_table_count >= config.max_tables_per_page:
                        break
                    if df is None or df.empty:
                        continue

                    data = df.fillna("").astype(str).values.tolist()
                    headers, rows = _normalize_table_data(data, config)
                    if not headers or not rows:
                        continue

                    table = ExtractedTable(
                        id=f"tabula_p{page_num+1}_t{i+1}",
                        page=page_num + 1,
                        headers=headers,
                        rows=rows,
                        confidence=base_confidence,
                        method=self.name,
                    )
                    table.confidence = _apply_table_confidence(table, config, base_confidence)
                    if not _table_meets_requirements(table, config):
                        continue
                    tables.append(table)
                    page_table_count += 1

            # Get page count
            page_count = _get_pdf_page_count(pdf_path)
            if page_count is None:
                page_count = len(pages_to_process)

            return ExtractionResult(
                tables=tables,
                text="",  # tabula focuses on tables
                page_count=page_count,
                method=self.name,
                errors=errors,
            )

        except Exception as e:
            logger.error(f"Tabula extraction failed: {e}")
            return ExtractionResult(
                tables=[],
                text="",
                page_count=0,
                method=self.name,
                errors=[_safe_error_message("Tabula extraction failed", e)],
            )

class CamelotExtractor(PDFExtractor):
    """
    Camelot-based PDF table extraction.

    Best for:
    - Complex table layouts
    - Tables with merged cells
    - Tables with visible borders (lattice mode)
    - Tables without borders (stream mode)

    Requires: camelot-py (pip install camelot-py[cv])
    """

    name = "camelot"

    def is_available(self) -> bool:
        try:
            import camelot
            return True
        except ImportError:
            return False

    def extract_tables(
        self,
        pdf_path: Union[str, Path],
        pages: Optional[List[int]] = None,
        config: Optional[ExtractionConfig] = None,
        flavor: str = "lattice",  # or "stream"
    ) -> ExtractionResult:
        try:
            import camelot
        except ImportError:
            return ExtractionResult(
                tables=[],
                text="",
                page_count=0,
                method=self.name,
                errors=["camelot-py not installed. Run: pip install camelot-py[cv]"],
            )

        config = _resolve_config(config)
        pdf_path, path_error = _resolve_pdf_path(pdf_path)
        if path_error:
            return ExtractionResult(
                tables=[],
                text="",
                page_count=0,
                method=self.name,
                errors=[path_error],
            )
        tables: List[ExtractedTable] = []
        errors: List[str] = []

        try:
            pages_to_process, page_error = _resolve_page_numbers(pdf_path, pages, config)
            if page_error:
                return ExtractionResult(
                    tables=[],
                    text="",
                    page_count=0,
                    method=self.name,
                    errors=[page_error],
                )
            page_spec = ",".join(str(p + 1) for p in pages_to_process) if pages_to_process else "1"

            # Try lattice mode first (for tables with borders)
            try:
                camelot_tables = camelot.read_pdf(
                    str(pdf_path),
                    pages=page_spec,
                    flavor=flavor,
                )
            except Exception as e:
                logger.warning(f"Camelot {flavor} mode failed, trying alternative: {e}")
                alt_flavor = "stream" if flavor == "lattice" else "lattice"
                camelot_tables = camelot.read_pdf(
                    str(pdf_path),
                    pages=page_spec,
                    flavor=alt_flavor,
                )

            page_table_counts: Dict[int, int] = {}

            for i, ct in enumerate(camelot_tables):
                page_number = ct.page if hasattr(ct, 'page') else 1
                page_table_counts.setdefault(page_number, 0)
                if page_table_counts[page_number] >= config.max_tables_per_page:
                    continue
                df = ct.df
                if df is None or df.empty:
                    continue

                # Convert DataFrame to table
                data = df.fillna("").astype(str).values.tolist()

                if len(data) < 1:
                    continue

                headers, rows = _normalize_table_data(data, config)
                if not headers or not rows:
                    continue

                # Get bounding box
                bbox = None
                if hasattr(ct, '_bbox'):
                    bbox = ct._bbox

                base_confidence = ct.accuracy / 100.0 if hasattr(ct, 'accuracy') else 0.8
                table = ExtractedTable(
                    id=f"camelot_table_{i+1}",
                    page=page_number,
                    headers=headers,
                    rows=rows,
                    confidence=base_confidence,
                    method=self.name,
                    bbox=bbox,
                    metadata={"flavor": flavor},
                )
                table.confidence = _apply_table_confidence(table, config, base_confidence)
                if not _table_meets_requirements(table, config):
                    continue
                tables.append(table)
                page_table_counts[page_number] += 1

            # Get page count
            try:
                import fitz
                doc = fitz.open(pdf_path)
                page_count = doc.page_count
                doc.close()
            except ImportError:
                logger.debug("PyMuPDF not available for page count in Camelot extractor")
                page_count = 0
            except Exception as e:
                logger.debug(f"Could not get page count in Camelot extractor: {e}")
                page_count = 0

            return ExtractionResult(
                tables=tables,
                text="",
                page_count=page_count,
                method=self.name,
                errors=errors,
            )

        except Exception as e:
            logger.error(f"Camelot extraction failed: {e}")
            return ExtractionResult(
                tables=[],
                text="",
                page_count=0,
                method=self.name,
                errors=[_safe_error_message("Camelot extraction failed", e)],
            )

class PyMuPDFExtractor(PDFExtractor):
    """
    PyMuPDF (fitz) based PDF extraction.

    Best for:
    - Fast extraction
    - General purpose
    - Text extraction with layout

    Requires: pymupdf (pip install pymupdf)
    """

    name = "pymupdf"

    def is_available(self) -> bool:
        try:
            import fitz
            return True
        except ImportError:
            return False

    def extract_tables(
        self,
        pdf_path: Union[str, Path],
        pages: Optional[List[int]] = None,
        config: Optional[ExtractionConfig] = None,
    ) -> ExtractionResult:
        try:
            import fitz
        except ImportError:
            return ExtractionResult(
                tables=[],
                text="",
                page_count=0,
                method=self.name,
                errors=["pymupdf not installed. Run: pip install pymupdf"],
            )

        config = _resolve_config(config)
        pdf_path, path_error = _resolve_pdf_path(pdf_path)
        if path_error:
            return ExtractionResult(
                tables=[],
                text="",
                page_count=0,
                method=self.name,
                errors=[path_error],
            )
        tables: List[ExtractedTable] = []
        text_parts: List[str] = []
        errors: List[str] = []
        ocr_used = False
        base_confidence = 0.9

        def _extract_page(page_num: int) -> tuple[int, str, List[ExtractedTable], List[str], bool]:
            local_tables: List[ExtractedTable] = []
            local_errors: List[str] = []
            local_ocr = False
            try:
                doc_local = fitz.open(pdf_path)
                if page_num >= doc_local.page_count:
                    doc_local.close()
                    return page_num, "", [], [], False
                page = doc_local[page_num]
                page_text = page.get_text("text") or ""
                if not page_text.strip() and config.ocr_enabled:
                    try:
                        pix = page.get_pixmap(dpi=150)
                        ocr_text = get_ocr_engine(config.ocr_language).extract_text(pix.tobytes("png"))
                        if ocr_text.strip():
                            page_text = ocr_text
                            local_ocr = True
                    except Exception as exc:
                        local_errors.append(_safe_error_message(f"OCR failed on page {page_num + 1}", exc))

                page_text = page_text or ""
                page_text_block = f"--- Page {page_num + 1} ---\n{page_text}"

                try:
                    page_tables = page.find_tables()
                    tables_iter = page_tables.tables if hasattr(page_tables, "tables") else page_tables
                    page_table_count = 0
                    for i, table in enumerate(tables_iter):
                        if page_table_count >= config.max_tables_per_page:
                            break
                        if getattr(table, "row_count", 0) == 0:
                            continue

                        data = table.extract()
                        if not data or len(data) < 1:
                            continue

                        headers, rows = _normalize_table_data(data, config)
                        if not headers or not rows:
                            continue

                        extracted = ExtractedTable(
                            id=f"pymupdf_p{page_num+1}_t{i+1}",
                            page=page_num + 1,
                            headers=headers,
                            rows=rows,
                            confidence=base_confidence,
                            method=self.name,
                        )
                        extracted.confidence = _apply_table_confidence(extracted, config, base_confidence)
                        if not _table_meets_requirements(extracted, config):
                            continue
                        local_tables.append(extracted)
                        page_table_count += 1
                except Exception as exc:
                    local_errors.append(_safe_error_message(f"Table extraction failed on page {page_num + 1}", exc))

                doc_local.close()
                return page_num, page_text_block, local_tables, local_errors, local_ocr
            except Exception as exc:
                return page_num, "", [], [_safe_error_message(f"Page {page_num + 1} failed", exc)], False

        try:
            doc = fitz.open(pdf_path)
            page_count = doc.page_count
            doc.close()

            pages_to_process, page_error = _resolve_page_numbers(
                pdf_path,
                pages,
                config,
                page_count=page_count,
            )
            if page_error:
                return ExtractionResult(
                    tables=[],
                    text="",
                    page_count=page_count,
                    method=self.name,
                    errors=[page_error],
                )

            if config.parallel_pages and len(pages_to_process) > 1:
                with concurrent.futures.ThreadPoolExecutor(max_workers=config.max_workers) as executor:
                    futures = [executor.submit(_extract_page, page_num) for page_num in pages_to_process]
                    results = [f.result() for f in concurrent.futures.as_completed(futures)]
                for page_num, page_text_block, local_tables, local_errors, local_ocr in sorted(results, key=lambda r: r[0]):
                    if page_text_block:
                        text_parts.append(page_text_block)
                    tables.extend(local_tables)
                    errors.extend(local_errors)
                    ocr_used = ocr_used or local_ocr
            else:
                for page_num in pages_to_process:
                    _, page_text_block, local_tables, local_errors, local_ocr = _extract_page(page_num)
                    if page_text_block:
                        text_parts.append(page_text_block)
                    tables.extend(local_tables)
                    errors.extend(local_errors)
                    ocr_used = ocr_used or local_ocr

            return ExtractionResult(
                tables=tables,
                text="\n\n".join(text_parts),
                page_count=page_count,
                method=self.name,
                errors=errors,
                ocr_used=ocr_used,
            )

        except Exception as e:
            logger.error(f"PyMuPDF extraction failed: {e}")
            return ExtractionResult(
                tables=[],
                text="",
                page_count=0,
                method=self.name,
                errors=[_safe_error_message("PyMuPDF extraction failed", e)],
            )

class PDFPlumberExtractor(PDFExtractor):
    """
    pdfplumber based PDF extraction.

    Best for:
    - Detailed layout analysis
    - Character-level extraction
    - Complex document structures

    Requires: pdfplumber (pip install pdfplumber)
    """

    name = "pdfplumber"

    def is_available(self) -> bool:
        try:
            import pdfplumber
            return True
        except ImportError:
            return False

    def extract_tables(
        self,
        pdf_path: Union[str, Path],
        pages: Optional[List[int]] = None,
        config: Optional[ExtractionConfig] = None,
    ) -> ExtractionResult:
        try:
            import pdfplumber
        except ImportError:
            return ExtractionResult(
                tables=[],
                text="",
                page_count=0,
                method=self.name,
                errors=["pdfplumber not installed. Run: pip install pdfplumber"],
            )

        config = _resolve_config(config)
        pdf_path, path_error = _resolve_pdf_path(pdf_path)
        if path_error:
            return ExtractionResult(
                tables=[],
                text="",
                page_count=0,
                method=self.name,
                errors=[path_error],
            )
        tables: List[ExtractedTable] = []
        text_parts: List[str] = []
        errors: List[str] = []
        ocr_used = False
        base_confidence = 0.85

        def _extract_page(page_num: int) -> tuple[int, str, List[ExtractedTable], List[str], bool]:
            local_tables: List[ExtractedTable] = []
            local_errors: List[str] = []
            local_ocr = False
            try:
                with pdfplumber.open(pdf_path) as pdf_local:
                    if page_num >= len(pdf_local.pages):
                        return page_num, "", [], [], False
                    page = pdf_local.pages[page_num]
                    page_text = page.extract_text() or ""
                    if not page_text.strip() and config.ocr_enabled:
                        try:
                            image = page.to_image(resolution=150).original
                            from io import BytesIO
                            buffer = BytesIO()
                            image.save(buffer, format="PNG")
                            ocr_text = get_ocr_engine(config.ocr_language).extract_text(buffer.getvalue())
                            if ocr_text.strip():
                                page_text = ocr_text
                                local_ocr = True
                        except Exception as exc:
                            local_errors.append(_safe_error_message(f"OCR failed on page {page_num + 1}", exc))

                    page_text_block = f"--- Page {page_num + 1} ---\n{page_text}"

                    try:
                        page_tables = page.extract_tables()
                        page_table_count = 0
                        for i, table_data in enumerate(page_tables):
                            if page_table_count >= config.max_tables_per_page:
                                break
                            if not table_data or len(table_data) < 1:
                                continue

                            headers, rows = _normalize_table_data(table_data, config)
                            if not headers or not rows:
                                continue

                            extracted = ExtractedTable(
                                id=f"pdfplumber_p{page_num+1}_t{i+1}",
                                page=page_num + 1,
                                headers=headers,
                                rows=rows,
                                confidence=base_confidence,
                                method=self.name,
                            )
                            extracted.confidence = _apply_table_confidence(extracted, config, base_confidence)
                            if not _table_meets_requirements(extracted, config):
                                continue
                            local_tables.append(extracted)
                            page_table_count += 1
                    except Exception as exc:
                        local_errors.append(_safe_error_message(f"Table extraction failed on page {page_num + 1}", exc))

                return page_num, page_text_block, local_tables, local_errors, local_ocr
            except Exception as exc:
                return page_num, "", [], [_safe_error_message(f"Page {page_num + 1} failed", exc)], False

        try:
            with pdfplumber.open(pdf_path) as pdf:
                page_count = len(pdf.pages)

            pages_to_process, page_error = _resolve_page_numbers(
                pdf_path,
                pages,
                config,
                page_count=page_count,
            )
            if page_error:
                return ExtractionResult(
                    tables=[],
                    text="",
                    page_count=page_count,
                    method=self.name,
                    errors=[page_error],
                )

            if config.parallel_pages and len(pages_to_process) > 1:
                with concurrent.futures.ThreadPoolExecutor(max_workers=config.max_workers) as executor:
                    futures = [executor.submit(_extract_page, page_num) for page_num in pages_to_process]
                    results = [f.result() for f in concurrent.futures.as_completed(futures)]
                for page_num, page_text_block, local_tables, local_errors, local_ocr in sorted(results, key=lambda r: r[0]):
                    if page_text_block:
                        text_parts.append(page_text_block)
                    tables.extend(local_tables)
                    errors.extend(local_errors)
                    ocr_used = ocr_used or local_ocr
            else:
                for page_num in pages_to_process:
                    _, page_text_block, local_tables, local_errors, local_ocr = _extract_page(page_num)
                    if page_text_block:
                        text_parts.append(page_text_block)
                    tables.extend(local_tables)
                    errors.extend(local_errors)
                    ocr_used = ocr_used or local_ocr

            return ExtractionResult(
                tables=tables,
                text="\n\n".join(text_parts),
                page_count=page_count,
                method=self.name,
                errors=errors,
                ocr_used=ocr_used,
            )

        except Exception as e:
            logger.error(f"pdfplumber extraction failed: {e}")
            return ExtractionResult(
                tables=[],
                text="",
                page_count=0,
                method=self.name,
                errors=[_safe_error_message("pdfplumber extraction failed", e)],
            )

# Registry of available extractors
EXTRACTORS: Dict[str, type] = {
    "tabula": TabulaExtractor,
    "camelot": CamelotExtractor,
    "pymupdf": PyMuPDFExtractor,
    "pdfplumber": PDFPlumberExtractor,
}

def get_available_extractors() -> List[str]:
    """Get list of available extractors."""
    available = []
    for name, cls in EXTRACTORS.items():
        try:
            extractor = cls()
            if extractor.is_available():
                available.append(name)
        except Exception as e:
            logger.debug(f"Extractor '{name}' not available: {e}")
    return available

def extract_pdf_tables(
    pdf_path: Union[str, Path],
    method: str = "auto",
    pages: Optional[List[int]] = None,
    config: Optional[ExtractionConfig] = None,
) -> ExtractionResult:
    """Extract tables from a PDF using the specified method."""
    config = _resolve_config(config)
    resolved_path, path_error = _resolve_pdf_path(pdf_path)
    if path_error:
        return ExtractionResult(
            tables=[],
            text="",
            page_count=0,
            method=method,
            errors=[path_error],
        )
    normalized_pages, _ = _normalize_pages_input(pages)
    effective_pages = normalized_pages if normalized_pages is not None else pages
    if method == "auto":
        config = _limit_config_pages(config, config.max_auto_pages)

    cache_key = None
    if _cache_enabled():
        cache_key = _build_cache_key(resolved_path, method, effective_pages, config)
        cached = _get_cached_result(cache_key)
        if cached is not None:
            return cached

        event, is_owner = _acquire_inflight(cache_key)
        if not is_owner:
            event.wait(_CACHE_WAIT_SECONDS)
            cached = _get_cached_result(cache_key)
            if cached is not None:
                return cached
    else:
        event = None
        is_owner = False
    if method == "auto":
        try:
            result = extract_with_best_method(resolved_path, pages, config=config)
            if cache_key:
                _store_cache_result(cache_key, result)
            return result
        finally:
            if cache_key and is_owner and event is not None:
                _release_inflight(cache_key, event)

    if method not in EXTRACTORS:
        if cache_key and is_owner and event is not None:
            _release_inflight(cache_key, event)
        return ExtractionResult(
            tables=[],
            text="",
            page_count=0,
            method=method,
            errors=[f"Unknown extraction method: {method}. Available: {list(EXTRACTORS.keys())}"],
        )

    extractor = EXTRACTORS[method]()
    if not extractor.is_available():
        if cache_key and is_owner and event is not None:
            _release_inflight(cache_key, event)
        return ExtractionResult(
            tables=[],
            text="",
            page_count=0,
            method=method,
            errors=[f"Extractor '{method}' is not available. Check dependencies."],
        )

    try:
        result = extractor.extract_tables(resolved_path, pages, config=config)
        if cache_key:
            _store_cache_result(cache_key, result)
        return result
    finally:
        if cache_key and is_owner and event is not None:
            _release_inflight(cache_key, event)

def extract_with_best_method(
    pdf_path: Union[str, Path],
    pages: Optional[List[int]] = None,
    config: Optional[ExtractionConfig] = None,
) -> ExtractionResult:
    """
    Try multiple extractors and return the best result.

    Priority order: pymupdf > pdfplumber > camelot > tabula
    """
    config = _resolve_config(config)
    config = _limit_config_pages(config, config.max_auto_pages)
    pdf_path, path_error = _resolve_pdf_path(pdf_path)
    if path_error:
        return ExtractionResult(
            tables=[],
            text="",
            page_count=0,
            method="auto",
            errors=[path_error],
        )
    best_result: Optional[ExtractionResult] = None
    method_errors: list[str] = []

    # Priority order - faster methods first
    priority = ["pymupdf", "pdfplumber", "camelot", "tabula"]

    for method in priority:
        if method not in EXTRACTORS:
            continue

        extractor = EXTRACTORS[method]()
        if not extractor.is_available():
            continue

        try:
            result = extractor.extract_tables(pdf_path, pages, config=config)

            if result.errors:
                method_errors.extend([f"{method}: {err}" for err in result.errors])
                if not result.tables:
                    continue

            if best_result is None:
                best_result = result
            elif len(result.tables) > len(best_result.tables):
                best_result = result
            elif (len(result.tables) == len(best_result.tables) and
                  result.tables and best_result.tables and
                  result.tables[0].confidence > best_result.tables[0].confidence):
                best_result = result

            # If we got good results, no need to try more extractors
            if result.tables and not result.errors:
                break

        except Exception as e:
            logger.warning(f"Extractor {method} failed: {e}")
            method_errors.append(_safe_error_message(f"{method} extractor failed", e))
            continue

    if best_result is None:
        return ExtractionResult(
            tables=[],
            text="",
            page_count=0,
            method="auto",
            errors=method_errors or ["No extractors were able to process this PDF"],
        )

    return best_result

def compare_extractors(
    pdf_path: Union[str, Path],
    pages: Optional[List[int]] = None,
    config: Optional[ExtractionConfig] = None,
) -> Dict[str, ExtractionResult]:
    """Compare results from all available extractors."""
    results = {}

    config = _resolve_config(config)
    config = _limit_config_pages(config, config.max_compare_pages)
    for name, cls in EXTRACTORS.items():
        try:
            extractor = cls()
            if extractor.is_available():
                results[name] = extractor.extract_tables(pdf_path, pages, config=config)
        except Exception as e:
            results[name] = ExtractionResult(
                tables=[],
                text="",
                page_count=0,
                method=name,
                errors=[_safe_error_message("Extractor failed", e)],
            )

    return results

    name: str
    data_type: DataType
    non_empty_count: int
    empty_count: int
    unique_count: int
    min_value: Optional[Any] = None
    max_value: Optional[Any] = None
    sample_values: List[Any] = field(default_factory=list)

    @property
    def fill_rate(self) -> float:
        """Percentage of non-empty values."""
        total = self.non_empty_count + self.empty_count
        return self.non_empty_count / total if total > 0 else 0.0

@dataclass
class ExtractionConfig:
    """Configuration for Excel extraction."""
    max_rows: int = 50000
    max_sheets: int = 50
    max_columns: int = 500
    detect_headers: bool = True
    infer_types: bool = True
    compute_stats: bool = True
    chunk_size: int = 1000
    encodings_to_try: Tuple[str, ...] = ("utf-8", "latin-1", "cp1252", "iso-8859-1")
    date_formats: Tuple[str, ...] = (
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%Y/%m/%d",
        "%d-%m-%Y",
        "%m-%d-%Y",
    )

DEFAULT_CONFIG = ExtractionConfig()

@dataclass
class ExcelSheet:
    """Data from a single Excel sheet."""
    name: str
    headers: List[str]
    rows: List[List[Any]]
    row_count: int
    column_count: int
    metadata: Dict[str, Any] = field(default_factory=dict)
    column_stats: List[ColumnStats] = field(default_factory=list)
    header_confidence: float = 1.0

    def __post_init__(self):
        """Validate sheet data."""
        # Ensure headers are strings
        self.headers = [str(h) if h is not None else "" for h in self.headers]

    def get_column(self, col_name: str) -> List[Any]:
        """Get all values from a column by header name."""
        try:
            idx = self.headers.index(col_name)
            return [row[idx] if idx < len(row) else None for row in self.rows]
        except ValueError:
            return []

    def get_column_by_index(self, idx: int) -> List[Any]:
        """Get all values from a column by index."""
        if idx < 0 or idx >= self.column_count:
            return []
        return [row[idx] if idx < len(row) else None for row in self.rows]

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary."""
        return {
            "name": self.name,
            "headers": self.headers,
            "rows": self.rows,
            "row_count": self.row_count,
            "column_count": self.column_count,
            "metadata": self.metadata,
            "header_confidence": self.header_confidence,
            "column_stats": [
                {
                    "name": s.name,
                    "data_type": s.data_type.value,
                    "non_empty_count": s.non_empty_count,
                    "fill_rate": s.fill_rate,
                }
                for s in self.column_stats
            ] if self.column_stats else [],
        }

    def iter_rows(self) -> Iterator[Dict[str, Any]]:
        """Iterate over rows as dictionaries."""
        for row in self.rows:
            yield {
                self.headers[i]: row[i] if i < len(row) else None
                for i in range(len(self.headers))
            }

@dataclass
class ExcelExtractionResult:
    """Result of Excel extraction."""
    sheets: List[ExcelSheet]
    filename: str
    format: str
    errors: List[str] = field(default_factory=list)
    metadata: Dict[str, Any] = field(default_factory=dict)
    extraction_time_ms: float = 0.0

    @property
    def total_rows(self) -> int:
        """Total rows across all sheets."""
        return sum(s.row_count for s in self.sheets)

    @property
    def has_data(self) -> bool:
        """Check if any data was extracted."""
        return any(s.row_count > 0 for s in self.sheets)

    def get_sheet(self, name: str) -> Optional[ExcelSheet]:
        """Get sheet by name."""
        for sheet in self.sheets:
            if sheet.name == name:
                return sheet
        return None

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary."""
        return {
            "sheets": [s.to_dict() for s in self.sheets],
            "filename": self.filename,
            "format": self.format,
            "errors": self.errors,
            "metadata": self.metadata,
            "extraction_time_ms": self.extraction_time_ms,
            "total_rows": self.total_rows,
            "has_data": self.has_data,
        }

def _detect_data_type(value: Any) -> DataType:
    """Detect the data type of a value."""
    if value is None or (isinstance(value, str) and not value.strip()):
        return DataType.EMPTY

    if isinstance(value, bool):
        return DataType.BOOLEAN

    if isinstance(value, int):
        return DataType.INTEGER

    if isinstance(value, float):
        return DataType.FLOAT

    if isinstance(value, datetime):
        return DataType.DATETIME

    if not isinstance(value, str):
        return DataType.STRING

    text = value.strip()

    # Check for boolean
    if text.lower() in ("true", "false", "yes", "no", "1", "0"):
        return DataType.BOOLEAN

    # Check for currency
    if re.match(r'^[$€£¥₹]?\s*-?\d{1,3}(,\d{3})*(\.\d{2})?$', text):
        return DataType.CURRENCY

    # Check for percentage
    if re.match(r'^-?\d+(\.\d+)?%$', text):
        return DataType.PERCENTAGE

    # Check for integer
    if re.match(r'^-?\d+$', text):
        return DataType.INTEGER

    # Check for float
    if re.match(r'^-?\d+\.\d+$', text):
        return DataType.FLOAT

    # Check for date patterns
    date_patterns = [
        r'^\d{4}-\d{2}-\d{2}$',
        r'^\d{2}/\d{2}/\d{4}$',
        r'^\d{2}-\d{2}-\d{4}$',
    ]
    for pattern in date_patterns:
        if re.match(pattern, text):
            return DataType.DATE

    # Check for datetime patterns
    if re.match(r'^\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}', text):
        return DataType.DATETIME

    return DataType.STRING

def _infer_column_type(values: List[Any]) -> DataType:
    """Infer the predominant data type for a column."""
    type_counts: Dict[DataType, int] = {}

    for value in values:
        dtype = _detect_data_type(value)
        if dtype != DataType.EMPTY:
            type_counts[dtype] = type_counts.get(dtype, 0) + 1

    if not type_counts:
        return DataType.EMPTY

    # Get most common type
    most_common = max(type_counts.items(), key=lambda x: x[1])
    total_non_empty = sum(type_counts.values())

    # If most common type covers >80% of values, use it
    if most_common[1] / total_non_empty >= 0.8:
        return most_common[0]

    return DataType.MIXED

def _compute_column_stats(
    header: str,
    values: List[Any],
    config: ExtractionConfig,
) -> ColumnStats:
    """Compute statistics for a column."""
    non_empty = [v for v in values if v is not None and (not isinstance(v, str) or v.strip())]
    empty_count = len(values) - len(non_empty)

    # Infer type
    data_type = _infer_column_type(values) if config.infer_types else DataType.STRING

    # Calculate unique count (sample if too large)
    sample_for_unique = non_empty[:1000] if len(non_empty) > 1000 else non_empty
    try:
        unique_count = len(set(str(v) for v in sample_for_unique))
    except Exception as e:
        logger.debug(f"Could not compute unique count for column '{header}': {e}")
        unique_count = 0

    # Get min/max for numeric types
    min_val = None
    max_val = None
    if data_type in (DataType.INTEGER, DataType.FLOAT, DataType.CURRENCY):
        numeric_vals = []
        for v in non_empty:
            try:
                if isinstance(v, (int, float)):
                    numeric_vals.append(v)
                elif isinstance(v, str):
                    cleaned = v.replace("$", "").replace(",", "").replace("%", "").strip()
                    numeric_vals.append(float(cleaned))
            except (ValueError, TypeError):
                # Expected for non-numeric values, skip silently
                pass
        if numeric_vals:
            min_val = min(numeric_vals)
            max_val = max(numeric_vals)

    # Sample values
    sample_values = non_empty[:5] if non_empty else []

    return ColumnStats(
        name=header,
        data_type=data_type,
        non_empty_count=len(non_empty),
        empty_count=empty_count,
        unique_count=unique_count,
        min_value=min_val,
        max_value=max_val,
        sample_values=sample_values,
    )

def _calculate_header_confidence(row: List[Any], data_rows: List[List[Any]]) -> float:
    """
    Calculate confidence that a row is a header row.

    Uses multiple heuristics:
    - Headers are usually text, not numbers
    - Headers are often shorter than data
    - Headers have distinct patterns
    """
    if not row:
        return 0.0

    confidence = 0.5  # Start neutral

    # Check if row contains mostly text
    text_count = sum(1 for cell in row if isinstance(cell, str) and cell.strip())
    num_count = sum(1 for cell in row if isinstance(cell, (int, float)))
    non_empty = text_count + num_count

    if non_empty == 0:
        return 0.0

    text_ratio = text_count / non_empty
    if text_ratio > 0.8:
        confidence += 0.2
    elif text_ratio < 0.3:
        confidence -= 0.2

    # Check if data rows have different patterns
    if data_rows:
        data_text_ratios = []
        for data_row in data_rows[:10]:
            data_text = sum(1 for cell in data_row if isinstance(cell, str) and cell.strip())
            data_num = sum(1 for cell in data_row if isinstance(cell, (int, float)))
            data_total = data_text + data_num
            if data_total > 0:
                data_text_ratios.append(data_text / data_total)

        if data_text_ratios:
            avg_data_text_ratio = sum(data_text_ratios) / len(data_text_ratios)
            if text_ratio > avg_data_text_ratio + 0.3:
                confidence += 0.2

    # Check for common header patterns
    header_keywords = {
        "id", "name", "date", "time", "amount", "total", "qty", "quantity",
        "price", "description", "status", "type", "category", "email", "phone",
        "address", "city", "country", "code", "number", "#", "no", "count",
    }
    keyword_matches = sum(
        1 for cell in row
        if isinstance(cell, str) and cell.strip().lower() in header_keywords
    )
    if keyword_matches > 0:
        confidence += min(0.2, keyword_matches * 0.05)

    # Check for title case or all caps
    title_or_caps = sum(
        1 for cell in row
        if isinstance(cell, str) and (cell.istitle() or cell.isupper())
    )
    if title_or_caps > len(row) / 2:
        confidence += 0.1

    return min(1.0, max(0.0, confidence))

class ExcelExtractor:
    """
    Excel/spreadsheet data extractor.

    Supports multiple formats and handles:
    - Multiple sheets
    - Intelligent header detection
    - Data type inference and preservation
    - Large file handling with streaming
    - Column statistics
    - Multiple encoding support
    """

    def __init__(
        self,
        config: Optional[ExtractionConfig] = None,
        max_rows: int = 10000,
        max_sheets: int = 20,
        detect_headers: bool = True,
    ):
        self.config = config or ExtractionConfig(
            max_rows=max_rows,
            max_sheets=max_sheets,
            detect_headers=detect_headers,
        )

    def extract(
        self,
        file_path: Union[str, Path],
        sheet_names: Optional[List[str]] = None,
    ) -> ExcelExtractionResult:
        """Extract data from an Excel file."""
        start_time = time.time()
        file_path = Path(file_path)
        suffix = file_path.suffix.lower()

        result: ExcelExtractionResult

        if suffix == ".csv":
            result = self._extract_csv(file_path)
        elif suffix == ".tsv":
            result = self._extract_csv(file_path, delimiter="\t")
        elif suffix in (".xlsx", ".xlsm"):
            result = self._extract_xlsx(file_path, sheet_names)
        elif suffix == ".xls":
            result = self._extract_xls(file_path, sheet_names)
        elif suffix == ".ods":
            result = self._extract_ods(file_path, sheet_names)
        else:
            result = ExcelExtractionResult(
                sheets=[],
                filename=file_path.name,
                format="unknown",
                errors=[f"Unsupported format: {suffix}"],
            )

        result.extraction_time_ms = (time.time() - start_time) * 1000
        return result

    def _extract_ods(
        self,
        file_path: Path,
        sheet_names: Optional[List[str]] = None,
    ) -> ExcelExtractionResult:
        """Extract from ODS (LibreOffice/OpenOffice) files."""
        try:
            import pandas as pd
        except ImportError:
            return ExcelExtractionResult(
                sheets=[],
                filename=file_path.name,
                format="ods",
                errors=["pandas and odfpy required. Run: pip install pandas odfpy"],
            )

        sheets: List[ExcelSheet] = []
        errors: List[str] = []

        try:
            # Read all sheets
            ods_data = pd.read_excel(file_path, engine="odf", sheet_name=None)

            sheets_to_process = sheet_names or list(ods_data.keys())[:self.config.max_sheets]

            for sheet_name in sheets_to_process:
                if sheet_name not in ods_data:
                    errors.append(f"Sheet '{sheet_name}' not found")
                    continue

                df = ods_data[sheet_name]
                if df.empty:
                    continue

                # Truncate if needed
                if len(df) > self.config.max_rows:
                    errors.append(f"Sheet '{sheet_name}' truncated at {self.config.max_rows} rows")
                    df = df.head(self.config.max_rows)

                # Convert to lists
                headers = [str(col) for col in df.columns.tolist()]
                rows = df.fillna("").values.tolist()

                # Compute stats if enabled
                column_stats = []
                if self.config.compute_stats:
                    for i, header in enumerate(headers):
                        col_values = [row[i] if i < len(row) else None for row in rows]
                        column_stats.append(_compute_column_stats(header, col_values, self.config))

                sheets.append(ExcelSheet(
                    name=sheet_name,
                    headers=headers,
                    rows=rows,
                    row_count=len(rows),
                    column_count=len(headers),
                    column_stats=column_stats,
                ))

            return ExcelExtractionResult(
                sheets=sheets,
                filename=file_path.name,
                format="ods",
                errors=errors,
            )

        except Exception as e:
            logger.error(f"ODS extraction failed: {e}")
            return ExcelExtractionResult(
                sheets=[],
                filename=file_path.name,
                format="ods",
                errors=["Extraction failed"],
            )

    def _extract_xlsx(
        self,
        file_path: Path,
        sheet_names: Optional[List[str]] = None,
    ) -> ExcelExtractionResult:
        """Extract from .xlsx/.xlsm files using openpyxl."""
        try:
            import openpyxl
        except ImportError:
            return ExcelExtractionResult(
                sheets=[],
                filename=file_path.name,
                format="xlsx",
                errors=["openpyxl not installed. Run: pip install openpyxl"],
            )

        sheets: List[ExcelSheet] = []
        errors: List[str] = []

        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

            sheets_to_process = sheet_names or workbook.sheetnames[:self.config.max_sheets]

            if len(workbook.sheetnames) > self.config.max_sheets and not sheet_names:
                errors.append(f"File has {len(workbook.sheetnames)} sheets, processing first {self.config.max_sheets}")

            for sheet_name in sheets_to_process:
                if sheet_name not in workbook.sheetnames:
                    errors.append(f"Sheet '{sheet_name}' not found")
                    continue

                sheet = workbook[sheet_name]

                # Read data
                rows_data: List[List[Any]] = []
                for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                    if row_idx >= self.config.max_rows:
                        errors.append(f"Sheet '{sheet_name}' truncated at {self.config.max_rows} rows")
                        break

                    # Convert row to list and clean values
                    cleaned_row = []
                    for cell in row:
                        if cell is None:
                            cleaned_row.append("")
                        elif isinstance(cell, datetime):
                            cleaned_row.append(cell)
                        elif isinstance(cell, (int, float)):
                            cleaned_row.append(cell)
                        else:
                            cleaned_row.append(str(cell))
                    rows_data.append(cleaned_row)

                if not rows_data:
                    continue

                # Calculate header confidence and detect headers
                header_confidence = 1.0
                if self.config.detect_headers and rows_data:
                    header_confidence = _calculate_header_confidence(rows_data[0], rows_data[1:])

                    if header_confidence >= 0.5:
                        headers = [str(h) if h else f"Column_{i+1}" for i, h in enumerate(rows_data[0])]
                        data_rows = rows_data[1:]
                    else:
                        headers = [f"Column_{i+1}" for i in range(len(rows_data[0]))]
                        data_rows = rows_data
                else:
                    headers = [f"Column_{i+1}" for i in range(len(rows_data[0]))]
                    data_rows = rows_data

                # Normalize row lengths
                max_cols = min(len(headers), self.config.max_columns)
                headers = headers[:max_cols]
                normalized_rows = []
                for row in data_rows:
                    normalized = list(row)
                    while len(normalized) < max_cols:
                        normalized.append("")
                    normalized_rows.append(normalized[:max_cols])

                # Compute column statistics if enabled
                column_stats = []
                if self.config.compute_stats and normalized_rows:
                    for i, header in enumerate(headers):
                        col_values = [row[i] if i < len(row) else None for row in normalized_rows]
                        column_stats.append(_compute_column_stats(header, col_values, self.config))

                sheets.append(ExcelSheet(
                    name=sheet_name,
                    headers=headers,
                    rows=normalized_rows,
                    row_count=len(normalized_rows),
                    column_count=len(headers),
                    header_confidence=header_confidence,
                    column_stats=column_stats,
                ))

            workbook.close()

            return ExcelExtractionResult(
                sheets=sheets,
                filename=file_path.name,
                format="xlsx",
                errors=errors,
                metadata={"total_sheets": len(workbook.sheetnames)},
            )

        except Exception as e:
            logger.error(f"XLSX extraction failed: {e}")
            return ExcelExtractionResult(
                sheets=[],
                filename=file_path.name,
                format="xlsx",
                errors=["Extraction failed"],
            )

    def _extract_xls(
        self,
        file_path: Path,
        sheet_names: Optional[List[str]] = None,
    ) -> ExcelExtractionResult:
        """Extract from .xls files using xlrd."""
        try:
            import xlrd
        except ImportError:
            return ExcelExtractionResult(
                sheets=[],
                filename=file_path.name,
                format="xls",
                errors=["xlrd not installed. Run: pip install xlrd"],
            )

        sheets: List[ExcelSheet] = []
        errors: List[str] = []

        try:
            workbook = xlrd.open_workbook(file_path)

            sheet_list = sheet_names or workbook.sheet_names()[:self.max_sheets]

            for sheet_name in sheet_list:
                try:
                    sheet = workbook.sheet_by_name(sheet_name)
                except xlrd.XLRDError:
                    errors.append(f"Sheet '{sheet_name}' not found")
                    continue

                rows_data: List[List[Any]] = []
                for row_idx in range(min(sheet.nrows, self.max_rows)):
                    row = []
                    for col_idx in range(sheet.ncols):
                        cell = sheet.cell(row_idx, col_idx)
                        if cell.ctype == xlrd.XL_CELL_EMPTY:
                            row.append("")
                        elif cell.ctype == xlrd.XL_CELL_NUMBER:
                            row.append(cell.value)
                        else:
                            row.append(str(cell.value))
                    rows_data.append(row)

                if not rows_data:
                    continue

                # Detect headers
                if self.detect_headers and rows_data:
                    headers = [str(h) if h else f"Column_{i+1}" for i, h in enumerate(rows_data[0])]
                    data_rows = rows_data[1:]
                else:
                    headers = [f"Column_{i+1}" for i in range(len(rows_data[0]))]
                    data_rows = rows_data

                sheets.append(ExcelSheet(
                    name=sheet_name,
                    headers=headers,
                    rows=data_rows,
                    row_count=len(data_rows),
                    column_count=len(headers),
                ))

            return ExcelExtractionResult(
                sheets=sheets,
                filename=file_path.name,
                format="xls",
                errors=errors,
            )

        except Exception as e:
            logger.error(f"XLS extraction failed: {e}")
            return ExcelExtractionResult(
                sheets=[],
                filename=file_path.name,
                format="xls",
                errors=["Extraction failed"],
            )

    def _detect_encoding(self, file_path: Path) -> str:
        """Detect file encoding by trying multiple encodings."""
        # Try chardet if available
        try:
            import chardet
            with open(file_path, 'rb') as f:
                raw = f.read(10000)
            result = chardet.detect(raw)
            if result['confidence'] > 0.7:
                return result['encoding']
        except ImportError:
            pass

        # Fallback: try each encoding
        for encoding in self.config.encodings_to_try:
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    f.read(1000)
                return encoding
            except (UnicodeDecodeError, LookupError):
                continue

        return 'utf-8'

    def _extract_csv(
        self,
        file_path: Path,
        delimiter: str = ",",
    ) -> ExcelExtractionResult:
        """Extract from CSV/TSV files with smart encoding detection."""
        errors: List[str] = []

        # Detect encoding
        encoding = self._detect_encoding(file_path)

        try:
            with open(file_path, 'r', encoding=encoding, errors='replace') as f:
                # Detect delimiter
                sample = f.read(8192)
                f.seek(0)

                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=',\t;|')
                    delimiter = dialect.delimiter
                except csv.Error:
                    pass

                reader = csv.reader(f, delimiter=delimiter)
                rows_data = list(reader)

            if not rows_data:
                return ExcelExtractionResult(
                    sheets=[],
                    filename=file_path.name,
                    format="csv",
                    errors=["CSV file is empty"],
                )

            # Truncate if needed
            if len(rows_data) > self.config.max_rows:
                errors.append(f"CSV truncated at {self.config.max_rows} rows")
                rows_data = rows_data[:self.config.max_rows]

            # Calculate header confidence
            header_confidence = 1.0
            if self.config.detect_headers and rows_data:
                header_confidence = _calculate_header_confidence(rows_data[0], rows_data[1:])

                if header_confidence >= 0.5:
                    headers = [str(h) if h else f"Column_{i+1}" for i, h in enumerate(rows_data[0])]
                    data_rows = rows_data[1:]
                else:
                    max_cols = max(len(row) for row in rows_data) if rows_data else 0
                    headers = [f"Column_{i+1}" for i in range(max_cols)]
                    data_rows = rows_data
            else:
                max_cols = max(len(row) for row in rows_data) if rows_data else 0
                headers = [f"Column_{i+1}" for i in range(max_cols)]
                data_rows = rows_data

            # Normalize row lengths
            max_cols = min(len(headers), self.config.max_columns)
            headers = headers[:max_cols]
            normalized_rows = []
            for row in data_rows:
                normalized = list(row)
                while len(normalized) < max_cols:
                    normalized.append("")
                normalized_rows.append(normalized[:max_cols])

            # Compute column statistics if enabled
            column_stats = []
            if self.config.compute_stats and normalized_rows:
                for i, header in enumerate(headers):
                    col_values = [row[i] if i < len(row) else None for row in normalized_rows]
                    column_stats.append(_compute_column_stats(header, col_values, self.config))

            sheets = [ExcelSheet(
                name=file_path.stem,
                headers=headers,
                rows=normalized_rows,
                row_count=len(normalized_rows),
                column_count=len(headers),
                header_confidence=header_confidence,
                column_stats=column_stats,
                metadata={"encoding": encoding, "delimiter": delimiter},
            )]

            return ExcelExtractionResult(
                sheets=sheets,
                filename=file_path.name,
                format="csv",
                errors=errors,
                metadata={"encoding": encoding},
            )

        except Exception as e:
            logger.error(f"CSV extraction failed: {e}")
            return ExcelExtractionResult(
                sheets=[],
                filename=file_path.name,
                format="csv",
                errors=["Extraction failed"],
            )

def extract_excel_data(
    file_path: Union[str, Path],
    sheet_names: Optional[List[str]] = None,
    max_rows: int = 10000,
) -> ExcelExtractionResult:
    """Quick function to extract data from Excel/CSV files."""
    extractor = ExcelExtractor(max_rows=max_rows)
    return extractor.extract(file_path, sheet_names)
