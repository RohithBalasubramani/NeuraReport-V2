from __future__ import annotations

"""Merged LLM module: support + client."""

# mypy: ignore-errors
"""LLM: config, providers, agents, text-to-sql, RAG, document extraction, vision, client."""

import os
import logging
from dataclasses import dataclass, field
from enum import Enum
from typing import Any, Dict, List, Optional

logger = logging.getLogger("neura.llm.config")

class LLMProvider(str, Enum):
    """Supported LLM providers."""
    OPENAI_COMPAT = "openai_compat"  # OpenAI-compatible API (LiteLLM proxy, vLLM, etc.)
    CLAUDE_CODE = "claude_code"      # Claude Code CLI (subprocess-based, fallback)

# Default model for each provider
DEFAULT_MODELS: Dict[LLMProvider, str] = {
    LLMProvider.OPENAI_COMPAT: "qwen",
    LLMProvider.CLAUDE_CODE: "qwen",
}

# Vision-capable models
VISION_MODELS: Dict[LLMProvider, List[str]] = {
    LLMProvider.OPENAI_COMPAT: ["qwen"],
    LLMProvider.CLAUDE_CODE: ["qwen"],
}

# Recommended models for document analysis / code generation (same as defaults)
DOCUMENT_ANALYSIS_MODELS = DEFAULT_MODELS
CODE_GENERATION_MODELS = DEFAULT_MODELS

@dataclass
class LLMConfig:
    """Configuration for the LLM provider."""

    provider: LLMProvider = LLMProvider.OPENAI_COMPAT
    model: str = "qwen"

    # OpenAI-compatible API settings
    api_base: str = "http://localhost:4000"
    api_key: str = "none"

    # Request settings
    timeout_seconds: float = 240.0
    max_retries: int = 3
    retry_delay: float = 1.5
    retry_multiplier: float = 2.0

    # Model-specific settings
    temperature: Optional[float] = None
    max_tokens: Optional[int] = 8192

    # Extra body fields merged into every chat completion request
    # Used for vLLM-specific params like {"chat_template_kwargs": {"enable_thinking": True}}
    extra_body: Optional[dict] = None

    # Feature flags
    supports_vision: bool = True
    supports_function_calling: bool = True
    supports_streaming: bool = True

    # Vision model configuration (separate endpoint for multimodal tasks)
    vision_model: Optional[str] = None
    vision_api_base: Optional[str] = None
    vision_api_key: Optional[str] = None
    vision_enabled: bool = False
    is_vision_model: bool = False  # True only for dedicated vision provider instances

    # Additional options
    extra_options: Dict[str, Any] = field(default_factory=dict)

    def __post_init__(self):
        """Validate and finalize configuration."""
        vision_models = VISION_MODELS.get(self.provider, [])
        self.supports_vision = self.model in vision_models or bool(vision_models)

        logger.info(
            "llm_config_initialized",
            extra={
                "event": "llm_config_initialized",
                "provider": self.provider.value,
                "model": self.model,
                "api_base": self.api_base,
                "supports_vision": self.supports_vision,
            }
        )

    @classmethod
    def from_env(cls) -> "LLMConfig":
        """Create configuration from environment variables."""
        # Provider selection
        provider_str = os.getenv("LLM_PROVIDER", "openai_compat").lower()
        if provider_str == "claude_code":
            provider = LLMProvider.CLAUDE_CODE
        else:
            provider = LLMProvider.OPENAI_COMPAT

        # Model selection
        model = (
            os.getenv("LLM_MODEL") or
            DEFAULT_MODELS.get(provider, "qwen")
        )

        # API settings for OpenAI-compatible provider
        api_base = os.getenv("LLM_API_BASE", "http://localhost:8200/v1")
        api_key = os.getenv("LLM_API_KEY", "none")

        # Request settings
        timeout = float(os.getenv("LLM_TIMEOUT_SECONDS", "240"))
        max_retries = int(os.getenv("LLM_MAX_RETRIES", "3"))
        retry_delay = float(os.getenv("LLM_RETRY_DELAY", "1.5"))
        retry_multiplier = float(os.getenv("LLM_RETRY_MULTIPLIER", "2.0"))

        # Optional settings
        temperature = os.getenv("LLM_TEMPERATURE")
        max_tokens = os.getenv("LLM_MAX_TOKENS")

        # Extra body fields for vLLM (e.g. enable Qwen thinking mode)
        # Auto-enable when connecting directly to vLLM (not via LiteLLM on port 4000)
        import json as _json
        _extra_body_env = os.getenv("LLM_EXTRA_BODY")
        if _extra_body_env:
            extra_body = _json.loads(_extra_body_env)
        elif ":4000" not in api_base:
            # Direct vLLM connection — enable Qwen thinking mode
            extra_body = {"chat_template_kwargs": {"enable_thinking": True}}
        else:
            extra_body = {}

        # Vision model settings
        vision_enabled = os.getenv("VISION_LLM_ENABLED", "false").lower() in ("1", "true", "yes")
        vision_model = os.getenv("VISION_LLM_MODEL", "glm-ocr") if vision_enabled else None
        vision_api_base = os.getenv("VISION_LLM_API_BASE", "http://localhost:11434/v1") if vision_enabled else None
        vision_api_key = os.getenv("VISION_LLM_API_KEY", "ollama") if vision_enabled else None

        return cls(
            provider=provider,
            model=model,
            api_base=api_base,
            api_key=api_key,
            timeout_seconds=timeout,
            max_retries=max_retries,
            retry_delay=retry_delay,
            retry_multiplier=retry_multiplier,
            temperature=float(temperature) if temperature else None,
            max_tokens=int(max_tokens) if max_tokens else 8192,
            vision_model=vision_model,
            vision_api_base=vision_api_base,
            vision_api_key=vision_api_key,
            vision_enabled=vision_enabled,
            extra_body=extra_body,
        )

    def get_vision_config(self) -> "LLMConfig":
        """Return a separate LLMConfig for the vision endpoint (GLM-OCR)."""
        if not self.vision_enabled or not self.vision_model:
            return self
        return LLMConfig(
            provider=LLMProvider.OPENAI_COMPAT,
            model=self.vision_model,
            api_base=self.vision_api_base or self.api_base,
            api_key=self.vision_api_key or self.api_key,
            timeout_seconds=180.0,
            is_vision_model=True,
        )

    def get_vision_model(self) -> str:
        """Get the recommended vision model."""
        if self.vision_enabled and self.vision_model:
            return self.vision_model
        return self.model

    def get_document_analysis_model(self) -> str:
        return self.model

    def get_code_generation_model(self) -> str:
        return self.model

# Global cached config
_config: Optional[LLMConfig] = None

def get_llm_config(force_reload: bool = False) -> LLMConfig:
    """Get the global LLM configuration."""
    global _config
    if _config is None or force_reload:
        _config = LLMConfig.from_env()
    return _config


def get_model() -> str:
    """Return the configured LLM model name. Single source of truth.

    Use this instead of os.getenv("LLM_MODEL") — the env var may be
    mutated at runtime by dependency imports.
    """
    return get_llm_config().model


# ── Providers ──
import base64
import subprocess
import sys
import time
from abc import ABC, abstractmethod

from backend.app.utils import AppError

logger = logging.getLogger("neura.llm.providers")

def _no_window_kwargs() -> dict:
    """Return subprocess kwargs to suppress console windows on Windows."""
    if sys.platform == "win32":
        si = subprocess.STARTUPINFO()
        si.dwFlags |= subprocess.STARTF_USESHOWWINDOW
        si.wShowWindow = 0  # SW_HIDE
        return {"startupinfo": si, "creationflags": subprocess.CREATE_NO_WINDOW}
    return {}

# Patterns that may appear in exception messages containing secrets
import re as _re

_SECRET_PATTERNS = _re.compile(
    r"(sk-[A-Za-z0-9]{8,}|Bearer\s+\S{8,}|api[_-]?key[=:]\s*\S+)",
    _re.IGNORECASE,
)

def _sanitize_error(exc: Exception) -> str:
    """Return error string with possible API keys/tokens redacted."""
    msg = str(exc)
    return _SECRET_PATTERNS.sub("[REDACTED]", msg)

class BaseProvider(ABC):
    """Abstract base class for LLM providers."""

    def __init__(self, config: LLMConfig):
        self.config = config
        self._client: Any = None

    @abstractmethod
    def get_client(self) -> Any:
        """Get or create the provider client."""
        pass

    @abstractmethod
    def chat_completion(
        self,
        messages: List[Dict[str, Any]],
        model: Optional[str] = None,
        **kwargs: Any,
    ) -> Dict[str, Any]:
        """Execute a chat completion request."""
        pass

    @abstractmethod
    def chat_completion_stream(
        self,
        messages: List[Dict[str, Any]],
        model: Optional[str] = None,
        **kwargs: Any,
    ) -> Iterator[Dict[str, Any]]:
        """Execute a streaming chat completion request."""
        pass

    @abstractmethod
    def list_models(self) -> List[str]:
        """List available models."""
        pass

    @abstractmethod
    def health_check(self) -> bool:
        """Check if the provider is available."""
        pass

    def supports_vision(self, model: Optional[str] = None) -> bool:
        """Check if the model supports vision inputs."""
        return self.config.supports_vision

    def prepare_vision_message(
        self,
        text: str,
        images: List[Union[str, bytes, Path]],
        detail: str = "auto",
    ) -> Dict[str, Any]:
        """Prepare a message with vision content."""
        content: List[Dict[str, Any]] = [{"type": "text", "text": text}]

        for image in images:
            if isinstance(image, Path):
                image_data = base64.b64encode(image.read_bytes()).decode("utf-8")
                media_type = "image/png" if image.suffix.lower() == ".png" else "image/jpeg"
                image_url = f"data:{media_type};base64,{image_data}"
            elif isinstance(image, bytes):
                image_data = base64.b64encode(image).decode("utf-8")
                image_url = f"data:image/png;base64,{image_data}"
            else:
                # Assume it's already a URL or base64 string
                if image.startswith("data:") or image.startswith("http"):
                    image_url = image
                else:
                    image_url = f"data:image/png;base64,{image}"

            content.append({
                "type": "image_url",
                "image_url": {"url": image_url, "detail": detail}
            })

        return {"role": "user", "content": content}

class OpenAICompatProvider(BaseProvider):
    """
    OpenAI-compatible API provider.

    Uses the `openai` Python SDK to call any OpenAI-compatible endpoint:
    - LiteLLM proxy (routes to vLLM, Ollama, etc.)
    - vLLM directly
    - Any OpenAI-compatible server

    Default: LiteLLM proxy at http://localhost:4000 → Qwen3-32B-FP8
    """

    def __init__(self, config: "LLMConfig"):
        super().__init__(config)
        self._openai_client = None

    def get_client(self):
        """Get or create the OpenAI client."""
        if self._openai_client is None:
            from openai import OpenAI
            self._openai_client = OpenAI(
                base_url=self.config.api_base,
                api_key=self.config.api_key,
                timeout=self.config.timeout_seconds,
            )
        return self._openai_client

    def _strip_images_from_messages(self, messages: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Strip image content from messages for text-only models.

        Replaces image_url parts with OCR-extracted text when possible,
        or a placeholder note otherwise.
        """
        cleaned = []
        for msg in messages:
            content = msg.get("content", "")
            if isinstance(content, list):
                new_parts = []
                for part in content:
                    if isinstance(part, dict):
                        ptype = part.get("type", "")
                        if ptype in ("image_url", "image"):
                            # Try to OCR the image if it's base64
                            ocr_text = self._try_ocr_image(part)
                            if ocr_text:
                                new_parts.append({
                                    "type": "text",
                                    "text": f"[OCR extracted text from image]:\n{ocr_text}",
                                })
                            else:
                                new_parts.append({
                                    "type": "text",
                                    "text": "[An image was provided but this model does not support vision. Use the schema and text hints to complete the task.]",
                                })
                        else:
                            new_parts.append(part)
                    else:
                        new_parts.append(part)
                cleaned.append({**msg, "content": new_parts})
            else:
                cleaned.append(msg)
        return cleaned

    def _try_ocr_image(self, part: Dict[str, Any]) -> Optional[str]:
        """Try to extract text from a base64 image using GLM-OCR."""
        try:
            url = ""
            if part.get("type") == "image_url":
                img_info = part.get("image_url", {})
                url = img_info.get("url", "") if isinstance(img_info, dict) else ""
            elif part.get("type") == "image":
                source = part.get("source", {})
                if source.get("type") == "base64":
                    b64_data = source.get("data", "")
                    if b64_data:
                        url = f"data:image/png;base64,{b64_data}"

            if not url or not url.startswith("data:image/"):
                return None

            _, b64_data = url.split(",", 1)
            img_bytes = base64.b64decode(b64_data)

            # Try GLM-OCR first (best quality for industrial documents)
            try:
                from backend.app.services.infra_services import ocr_extract
                ocr_text = ocr_extract(img_bytes)
                if ocr_text:
                    return ocr_text
            except Exception:
                pass

            # Fallback: PyMuPDF (works for PDFs embedded as images)
            try:
                import fitz
                doc = fitz.open(stream=img_bytes, filetype="png")
                text = ""
                for page in doc:
                    text += page.get_text()
                doc.close()
                if text.strip():
                    return text.strip()
            except Exception:
                pass

            return None
        except Exception:
            return None

    def chat_completion(
        self,
        messages: List[Dict[str, Any]],
        model: Optional[str] = None,
        **kwargs: Any,
    ) -> Dict[str, Any]:
        """Execute a chat completion via OpenAI-compatible API."""
        client = self.get_client()
        model = model or self.config.model

        # Strip images from messages if model doesn't support vision
        has_images = any(
            isinstance(m.get("content"), list) and any(
                isinstance(p, dict) and p.get("type") in ("image_url", "image")
                for p in m["content"]
            )
            for m in messages
        )
        if has_images and not self.config.supports_vision:
            messages = self._strip_images_from_messages(messages)

        # Resolve adaptive timeout based on operation description
        from backend.app.services.llm import get_adaptive_timeout
        description = kwargs.pop("description", "")
        adaptive_timeout = get_adaptive_timeout(description, default=self.config.timeout_seconds)
        effective_timeout = max(adaptive_timeout, self.config.timeout_seconds)

        # Build request parameters
        request_params: Dict[str, Any] = {
            "model": model,
            "messages": messages,
        }

        # Add max_tokens if specified
        max_tokens = kwargs.pop("max_tokens", None) or self.config.max_tokens
        if max_tokens:
            request_params["max_tokens"] = max_tokens

        # Add temperature if specified
        temperature = kwargs.pop("temperature", None) or self.config.temperature
        if temperature is not None:
            request_params["temperature"] = temperature

        # Pop timeout from kwargs (we handle it separately via effective_timeout)
        caller_timeout = kwargs.pop("timeout", None)
        if caller_timeout and caller_timeout > effective_timeout:
            effective_timeout = caller_timeout

        # Merge config-level extra_body (e.g. vLLM chat_template_kwargs)
        # Auto-disable thinking for structured JSON operations — Qwen wastes
        # entire token budget on reasoning otherwise, producing no JSON.
        _NO_THINKING_OPS = {"llm_call_3", "mapping", "contract", "v3_df", "ops_refiner", "simulation"}
        _needs_thinking = not any(op in description for op in _NO_THINKING_OPS)

        caller_extra = kwargs.pop("extra_body", {}) or {}
        if self.config.extra_body:
            merged = {**self.config.extra_body, **caller_extra}
            if not _needs_thinking and "chat_template_kwargs" not in caller_extra:
                merged["chat_template_kwargs"] = {"enable_thinking": False}
            request_params["extra_body"] = merged
        elif caller_extra:
            request_params["extra_body"] = caller_extra

        logger.debug(
            "thinking_mode_decision",
            extra={
                "description": description,
                "needs_thinking": _needs_thinking,
                "extra_body": request_params.get("extra_body"),
            },
        )

        # Pass through any extra kwargs (but not internal ones)
        _internal_keys = {"use_cache", "cache_ttl", "timeout", "description"}
        for k, v in kwargs.items():
            if k not in _internal_keys:
                request_params[k] = v

        logger.info(
            "openai_compat_call",
            extra={
                "event": "openai_compat_call",
                "model": model,
                "api_base": self.config.api_base,
                "message_count": len(messages),
                "description": description,
            }
        )

        start_time = time.time()
        try:
            response = client.chat.completions.create(
                **request_params,
                timeout=effective_timeout,
            )

            elapsed = time.time() - start_time

            # Convert to dict format
            content = response.choices[0].message.content or "" if response.choices else ""
            usage = response.usage

            result = {
                "id": response.id or f"openai-compat-{int(time.time())}",
                "model": response.model or model,
                "choices": [{
                    "index": 0,
                    "message": {
                        "role": "assistant",
                        "content": content,
                    },
                    "finish_reason": response.choices[0].finish_reason if response.choices else "stop",
                }],
                "usage": {
                    "prompt_tokens": usage.prompt_tokens if usage else 0,
                    "completion_tokens": usage.completion_tokens if usage else 0,
                    "total_tokens": usage.total_tokens if usage else 0,
                }
            }

            logger.info(
                "openai_compat_success",
                extra={
                    "event": "openai_compat_success",
                    "model": model,
                    "elapsed_seconds": round(elapsed, 2),
                    "output_length": len(content),
                    "input_tokens": result["usage"]["prompt_tokens"],
                    "output_tokens": result["usage"]["completion_tokens"],
                }
            )

            return result

        except Exception as e:
            elapsed = time.time() - start_time
            error_msg = _sanitize_error(e)
            logger.error(
                "openai_compat_error",
                extra={
                    "event": "openai_compat_error",
                    "model": model,
                    "error": error_msg[:500],
                    "elapsed_seconds": round(elapsed, 2),
                }
            )
            raise AppError(
                code="llm_call_failed",
                message=f"LLM request failed ({description or 'default'}): {error_msg[:200]}",
                status_code=502,
                detail=error_msg[:500],
            )

    def chat_completion_stream(
        self,
        messages: List[Dict[str, Any]],
        model: Optional[str] = None,
        **kwargs: Any,
    ) -> Iterator[Dict[str, Any]]:
        """Execute a streaming chat completion via OpenAI-compatible API."""
        client = self.get_client()
        model = model or self.config.model
        kwargs.pop("description", None)

        max_tokens = kwargs.pop("max_tokens", None) or self.config.max_tokens
        temperature = kwargs.pop("temperature", None) or self.config.temperature

        request_params: Dict[str, Any] = {
            "model": model,
            "messages": messages,
            "stream": True,
        }
        if max_tokens:
            request_params["max_tokens"] = max_tokens
        if temperature is not None:
            request_params["temperature"] = temperature

        stream = client.chat.completions.create(**request_params)

        for chunk in stream:
            delta_content = ""
            finish_reason = None
            if chunk.choices:
                delta = chunk.choices[0].delta
                delta_content = delta.content or "" if delta else ""
                finish_reason = chunk.choices[0].finish_reason

            yield {
                "id": chunk.id or f"stream-{int(time.time())}",
                "model": chunk.model or model,
                "choices": [{
                    "index": 0,
                    "delta": {"content": delta_content},
                    "finish_reason": finish_reason,
                }]
            }

    def list_models(self) -> List[str]:
        """List available models from the API."""
        try:
            client = self.get_client()
            models = client.models.list()
            return [m.id for m in models.data]
        except Exception:
            return [self.config.model]

    def health_check(self) -> bool:
        """Check if the OpenAI-compatible API is available."""
        try:
            import urllib.request
            req = urllib.request.Request(f"{self.config.api_base}/health", method="GET")
            with urllib.request.urlopen(req, timeout=5) as resp:
                return resp.status == 200
        except Exception:
            return False

class ClaudeCodeCLIProvider(BaseProvider):
    """
    Claude Code CLI provider (fallback).

    Uses the `claude` CLI tool as a subprocess to get LLM completions.
    Set LLM_PROVIDER=claude_code to use this.

    Model names: qwen (default)
    """

    def __init__(self, config: "LLMConfig"):
        super().__init__(config)
        self._available: Optional[bool] = None
        self._claude_bin: str = "claude"

    def get_client(self) -> Any:
        """Check CLI availability on first use."""
        if self._available is None:
            self._available = self._check_cli_available()
        if not self._available:
            raise AppError(
                code="llm_unavailable",
                message="Claude Code CLI is not installed. Set LLM_PROVIDER=openai_compat to use the LiteLLM proxy instead.",
                status_code=503,
            )
        return True

    def _check_cli_available(self) -> bool:
        """Check if claude CLI is available."""
        import shutil
        claude_bin = shutil.which("claude")
        if not claude_bin:
            common_paths = [
                Path.home() / ".local" / "bin" / "claude",
                Path("/usr/local/bin/claude"),
                Path.home() / ".npm-global" / "bin" / "claude",
            ]
            for p in common_paths:
                if p.is_file():
                    claude_bin = str(p)
                    break
        if not claude_bin:
            return False
        self._claude_bin = claude_bin
        try:
            result = subprocess.run(
                [claude_bin, "--version"],
                capture_output=True, text=True, timeout=10,
                **_no_window_kwargs(),
            )
            return result.returncode == 0
        except Exception:
            return False

    def _extract_images_from_messages(self, messages: List[Dict[str, Any]]) -> List[str]:
        """Extract base64 images from messages and save to temp files."""
        import tempfile
        image_files = []
        for msg in messages:
            content = msg.get("content", "")
            if isinstance(content, list):
                for part in content:
                    if isinstance(part, dict):
                        if part.get("type") == "image_url":
                            image_url = part.get("image_url", {})
                            url = image_url.get("url", "") if isinstance(image_url, dict) else ""
                            if url.startswith("data:image/"):
                                try:
                                    header, b64_data = url.split(",", 1)
                                    ext = ".png" if "png" in header else ".jpg"
                                    img_bytes = base64.b64decode(b64_data)
                                    with tempfile.NamedTemporaryFile(mode='wb', suffix=ext, delete=False) as f:
                                        f.write(img_bytes)
                                        image_files.append(f.name)
                                except Exception as e:
                                    logger.warning(f"Failed to extract image: {e}")
                        elif part.get("type") == "image":
                            source = part.get("source", {})
                            if source.get("type") == "base64":
                                try:
                                    b64_data = source.get("data", "")
                                    media_type = source.get("media_type", "image/png")
                                    ext = ".png" if "png" in media_type else ".jpg"
                                    img_bytes = base64.b64decode(b64_data)
                                    with tempfile.NamedTemporaryFile(mode='wb', suffix=ext, delete=False) as f:
                                        f.write(img_bytes)
                                        image_files.append(f.name)
                                except Exception as e:
                                    logger.warning(f"Failed to extract image: {e}")
        return image_files

    def _build_prompt(self, messages: List[Dict[str, Any]]) -> str:
        """Convert message format to a single prompt string."""
        parts = []
        for msg in messages:
            role = msg.get("role", "user")
            content = msg.get("content", "")
            if isinstance(content, list):
                text_parts = []
                for part in content:
                    if isinstance(part, dict) and part.get("type") == "text":
                        text_parts.append(part.get("text", ""))
                content = "\n".join(text_parts)
            if role == "system":
                parts.append(f"<system>\n{content}\n</system>")
            elif role == "assistant":
                parts.append(f"<assistant>\n{content}\n</assistant>")
            else:
                parts.append(content)
        return "\n\n".join(parts)

    def _has_images(self, messages: List[Dict[str, Any]]) -> bool:
        """Check if any message contains image content."""
        for msg in messages:
            content = msg.get("content", "")
            if isinstance(content, list):
                for part in content:
                    if isinstance(part, dict) and part.get("type") in ("image_url", "image"):
                        return True
        return False

    def _call_litellm_direct(self, messages: List[Dict[str, Any]], **kwargs: Any) -> Dict[str, Any]:
        """
        Direct HTTP call to LiteLLM /v1/messages endpoint.
        Used for vision calls — Qwen 3.5 27B handles images natively via vLLM.
        Claude CLI can't pass images in --bare mode so we bypass it for vision.
        """
        import requests as _requests

        api_base = self.config.api_base.rstrip('/')
        url = f"{api_base}/v1/messages"

        # Convert OpenAI-style image_url to messages-API image blocks
        anthropic_messages = []
        for msg in messages:
            content = msg.get("content", "")
            if isinstance(content, list):
                parts = []
                for part in content:
                    if isinstance(part, dict):
                        if part.get("type") == "text":
                            parts.append({"type": "text", "text": part.get("text", "")})
                        elif part.get("type") == "image_url":
                            image_url = part.get("image_url", {})
                            url_str = image_url.get("url", "") if isinstance(image_url, dict) else ""
                            if url_str.startswith("data:"):
                                header, b64_data = url_str.split(",", 1)
                                media_type = header.split(";")[0].split(":")[1] if ":" in header else "image/png"
                                parts.append({"type": "image", "source": {"type": "base64", "media_type": media_type, "data": b64_data}})
                        elif part.get("type") == "image":
                            parts.append(part)
                    elif isinstance(part, str):
                        parts.append({"type": "text", "text": part})
                anthropic_messages.append({"role": msg.get("role", "user"), "content": parts})
            else:
                anthropic_messages.append({"role": msg.get("role", "user"), "content": str(content)})

        start_time = time.time()
        logger.info("litellm_vision_call", extra={"event": "litellm_vision_call", "message_count": len(messages)})

        try:
            resp = _requests.post(url, json={
                "model": self.config.model,
                "max_tokens": kwargs.get("max_tokens", 8192),
                "messages": anthropic_messages,
            }, headers={
                "x-api-key": self.config.api_key or "dummy",
                "anthropic-version": "2023-06-01",
            }, timeout=self.config.timeout_seconds)

            data = resp.json()
            if "error" in data:
                raise AppError(code="llm_call_failed", message=f"Vision call failed: {data['error']}", status_code=502)

            content = data.get("content", [{}])[0].get("text", "")
            elapsed = time.time() - start_time
            usage = data.get("usage", {})
            logger.info("litellm_vision_success", extra={"event": "litellm_vision_success", "model": data.get("model"), "elapsed_seconds": round(elapsed, 2)})

            return {
                "id": data.get("id", f"litellm-{int(time.time())}"),
                "model": data.get("model", "qwen"),
                "choices": [{"index": 0, "message": {"role": "assistant", "content": content}, "finish_reason": "stop"}],
                "usage": {"prompt_tokens": usage.get("input_tokens", 0), "completion_tokens": usage.get("output_tokens", 0), "total_tokens": usage.get("total_tokens", 0)},
            }
        except AppError:
            raise
        except Exception as e:
            logger.error("litellm_vision_failed", extra={"error": _sanitize_error(e)})
            raise AppError(code="llm_call_failed", message="Vision call failed.", status_code=502, detail=_sanitize_error(e))

    def _call_claude_cli(self, prompt: str, **kwargs: Any) -> Dict[str, Any]:
        """
        Text-only call via Claude CLI subprocess.
        Claude CLI → ANTHROPIC_BASE_URL (LiteLLM) → vLLM → Qwen.
        """
        import tempfile

        from backend.app.services.llm import get_adaptive_timeout
        description = kwargs.pop("description", "")
        adaptive_timeout = get_adaptive_timeout(description, default=self.config.timeout_seconds)
        effective_timeout = max(adaptive_timeout, self.config.timeout_seconds)

        cmd = [self._claude_bin, "-p", "--bare", "--model", self.config.model]

        logger.info("claude_code_cli_call", extra={"event": "claude_code_cli_call", "prompt_length": len(prompt)})

        start_time = time.time()
        try:
            with tempfile.NamedTemporaryFile(mode='w', suffix='.txt', delete=False, encoding='utf-8') as f:
                f.write(prompt)
                prompt_file = f.name

            import os as _env_os
            env = _env_os.environ.copy()
            env.pop('CLAUDECODE', None)
            env['ANTHROPIC_BASE_URL'] = self.config.api_base.rstrip('/v1').rstrip('/')
            env['ANTHROPIC_API_KEY'] = self.config.api_key or 'dummy'

            with open(prompt_file, 'r', encoding='utf-8') as pf:
                result = subprocess.run(cmd, stdin=pf, capture_output=True, text=True, timeout=effective_timeout, env=env, **_no_window_kwargs())

            import os as _os
            _os.unlink(prompt_file)

            if result.returncode != 0:
                error_msg = ((result.stderr or "").strip() or (result.stdout or "").strip())[:500]
                raise AppError(code="llm_call_failed", message="Claude CLI returned an error.", status_code=502, detail=error_msg)

            content = result.stdout.strip()
            elapsed = time.time() - start_time
            input_tokens = len(prompt) // 4
            output_tokens = len(content) // 4

            logger.info("claude_code_cli_success", extra={"event": "claude_code_cli_success", "elapsed_seconds": round(elapsed, 2), "output_length": len(content)})

            return {
                "id": f"claude-cli-{int(time.time())}",
                "model": "qwen",
                "choices": [{"index": 0, "message": {"role": "assistant", "content": content}, "finish_reason": "stop"}],
                "usage": {"prompt_tokens": input_tokens, "completion_tokens": output_tokens, "total_tokens": input_tokens + output_tokens}
            }

        except subprocess.TimeoutExpired:
            raise AppError(code="llm_timeout", message=f"Timed out after {effective_timeout}s (op={description or 'default'}).", status_code=504)
        except AppError:
            raise
        except Exception as e:
            raise AppError(code="llm_error", message="AI request failed.", status_code=502, detail=_sanitize_error(e))

    def chat_completion(
        self,
        messages: List[Dict[str, Any]],
        model: Optional[str] = None,
        **kwargs: Any,
    ) -> Dict[str, Any]:
        """
        Hybrid provider: Claude CLI for text, direct LiteLLM for vision.

        - Messages with images → direct HTTP to LiteLLM /v1/messages
          (LiteLLM routes to Qwen on vLLM)
        - Text-only messages → Claude CLI subprocess with --bare
          (ANTHROPIC_BASE_URL redirects to LiteLLM → Qwen)

        Both paths end at Qwen via vLLM. No external API is ever hit.
        """
        self.get_client()

        if self._has_images(messages):
            return self._call_litellm_direct(messages, **kwargs)
        else:
            prompt = self._build_prompt(messages)
            return self._call_claude_cli(prompt, **kwargs)

    def chat_completion_stream(self, messages: List[Dict[str, Any]], model: Optional[str] = None, **kwargs: Any) -> Iterator[Dict[str, Any]]:
        """Streaming not natively supported - yield full response as single chunk."""
        response = self.chat_completion(messages, model, **kwargs)
        content = response["choices"][0]["message"]["content"]
        yield {"id": response["id"], "model": response["model"], "choices": [{"index": 0, "delta": {"content": content}, "finish_reason": "stop"}]}

    def list_models(self) -> List[str]:
        return ["qwen"]

    def health_check(self) -> bool:
        return self._check_cli_available()

# For backwards compatibility
LiteLLMProvider = OpenAICompatProvider

def get_provider(config: LLMConfig) -> BaseProvider:
    """Get the appropriate LLM provider based on configuration."""
    if config.provider == LLMProvider.CLAUDE_CODE:
        logger.info("provider_selected", extra={"provider": "claude_code"})
        return ClaudeCodeCLIProvider(config)
    else:
        logger.info("provider_selected", extra={"provider": "openai_compat", "api_base": config.api_base})
        return OpenAICompatProvider(config)

# ── Semantic Cache ──
import hashlib
from collections import OrderedDict

logger = logging.getLogger("neura.llm.semantic_cache")

def _cosine_similarity(a: List[float], b: List[float]) -> float:
    """Compute cosine similarity between two vectors."""
    if len(a) != len(b) or not a:
        return 0.0
    dot = sum(x * y for x, y in zip(a, b))
    norm_a = sum(x * x for x in a) ** 0.5
    norm_b = sum(x * x for x in b) ** 0.5
    if norm_a == 0 or norm_b == 0:
        return 0.0
    return dot / (norm_a * norm_b)

def _parse_json_from_llm(raw_content: str, default: Dict[str, Any]) -> Dict[str, Any]:
    """Extract JSON from LLM response (code block or raw)."""
    json_match = re.search(r"```(?:json)?\s*([\s\S]*?)```", raw_content)
    json_str = json_match.group(1).strip() if json_match else raw_content.strip()

    start = json_str.find("{")
    if start == -1:
        return default

    depth = 0
    in_string = False
    escape_next = False
    for i, char in enumerate(json_str[start:], start):
        if escape_next:
            escape_next = False
            continue
        if char == "\\":
            escape_next = True
            continue
        if char == '"' and not escape_next:
            in_string = not in_string
            continue
        if in_string:
            continue
        if char == "{":
            depth += 1
        elif char == "}":
            depth -= 1
            if depth == 0:
                json_str = json_str[start:i + 1]
                break
    try:
        return json.loads(json_str)
    except json.JSONDecodeError:
        return default

class SemanticCacheL2:
    """
    Embedding-based semantic cache for LLM responses.

    Uses cosine similarity of prompt embeddings to find cached
    responses for semantically similar (but not identical) prompts.
    """

    def __init__(self, max_size: int = 500, similarity_threshold: float = 0.92) -> None:
        self._max_size = max_size
        self._threshold = similarity_threshold
        self._cache: OrderedDict[str, Dict[str, Any]] = OrderedDict()
        self._embeddings: Dict[str, List[float]] = {}
        self._hits = 0
        self._misses = 0
        self._embed_fn = None

    def _get_embed_fn(self):
        """Lazy-load embedding function."""
        if self._embed_fn is None:
            try:
                from backend.app.services.knowledge_service import get_embedding_client
                client = get_embedding_client()
                self._embed_fn = client.embed
            except ImportError:
                # Fallback: simple hash-based "embedding" (no semantic similarity)
                import hashlib

                def _hash_embed(text: str) -> List[float]:
                    h = hashlib.sha256(text.encode()).hexdigest()
                    return [int(c, 16) / 15.0 for c in h[:64]]

                self._embed_fn = _hash_embed
        return self._embed_fn

    def lookup(self, prompt: str) -> Optional[Any]:
        """
        Look up a semantically similar cached response.

        Returns cached response if cosine similarity >= threshold,
        otherwise returns None.
        """
        embed_fn = self._get_embed_fn()
        try:
            query_embedding = embed_fn(prompt)
        except Exception:
            self._misses += 1
            return None

        best_score = 0.0
        best_key = None

        for key, embedding in self._embeddings.items():
            score = _cosine_similarity(query_embedding, embedding)
            if score > best_score:
                best_score = score
                best_key = key

        if best_score >= self._threshold and best_key is not None:
            entry = self._cache.get(best_key)
            if entry:
                self._hits += 1
                # Move to end (LRU)
                self._cache.move_to_end(best_key)
                logger.debug("L2 cache hit: similarity=%.3f", best_score)
                return entry["response"]

        self._misses += 1
        return None

    def store(self, prompt: str, response: Any) -> None:
        """Store a prompt-response pair in the cache."""
        embed_fn = self._get_embed_fn()
        try:
            embedding = embed_fn(prompt)
        except Exception:
            return

        key = hashlib.sha256(prompt.encode()).hexdigest()[:16]

        # Evict if at capacity
        while len(self._cache) >= self._max_size:
            evicted_key, _ = self._cache.popitem(last=False)
            self._embeddings.pop(evicted_key, None)

        self._cache[key] = {
            "prompt": prompt[:200],  # Truncate for memory
            "response": response,
            "timestamp": time.time(),
        }
        self._embeddings[key] = embedding

    def get_stats(self) -> Dict[str, Any]:
        """Return cache statistics."""
        total = self._hits + self._misses
        return {
            "size": len(self._cache),
            "max_size": self._max_size,
            "hits": self._hits,
            "misses": self._misses,
            "hit_rate": self._hits / total if total > 0 else 0.0,
            "threshold": self._threshold,
        }

    def clear(self) -> None:
        """Clear the cache."""
        self._cache.clear()
        self._embeddings.clear()

    # Singleton
    _instance: Optional["SemanticCacheL2"] = None

    @classmethod
    def get_instance(cls) -> "SemanticCacheL2":
        if cls._instance is None:
            from backend.app.services.infra_services import get_v2_config
            cfg = get_v2_config()
            cls._instance = cls(
                max_size=cfg.cache_l2_max_size,
                similarity_threshold=cfg.cache_l2_threshold,
            )
        return cls._instance

# ── Cost Tracker (with daily rollups — from prodo) ──
import threading
from collections import defaultdict
from dataclasses import dataclass, field as dc_field
from datetime import datetime, timezone

# Approximate costs per 1M tokens (USD) for common models
_MODEL_COSTS: Dict[str, Dict[str, float]] = {
    "default": {"input": 0.0, "output": 0.0},
}

_singleton_lock = threading.Lock()
_singleton: "CostTracker | None" = None


@dataclass
class OperationCost:
    """Accumulated cost for a single operation category."""
    operation: str
    input_tokens: int = 0
    output_tokens: int = 0
    estimated_cost_usd: float = 0.0
    call_count: int = 0
    error_count: int = 0
    total_latency_ms: float = 0.0
    first_call: Optional[float] = None
    last_call: Optional[float] = None

    @property
    def avg_latency_ms(self) -> float:
        return self.total_latency_ms / self.call_count if self.call_count > 0 else 0.0

    @property
    def error_rate(self) -> float:
        total = self.call_count + self.error_count
        return self.error_count / total if total > 0 else 0.0

    def to_dict(self) -> Dict[str, Any]:
        return {
            "operation": self.operation,
            "input_tokens": self.input_tokens,
            "output_tokens": self.output_tokens,
            "total_tokens": self.input_tokens + self.output_tokens,
            "estimated_cost_usd": round(self.estimated_cost_usd, 6),
            "call_count": self.call_count,
            "error_count": self.error_count,
            "error_rate": round(self.error_rate, 4),
            "avg_latency_ms": round(self.avg_latency_ms, 2),
            "first_call": datetime.fromtimestamp(self.first_call, tz=timezone.utc).isoformat() if self.first_call else None,
            "last_call": datetime.fromtimestamp(self.last_call, tz=timezone.utc).isoformat() if self.last_call else None,
        }


@dataclass
class DailyRollup:
    """Daily aggregation of costs."""
    date: str  # YYYY-MM-DD
    input_tokens: int = 0
    output_tokens: int = 0
    estimated_cost_usd: float = 0.0
    call_count: int = 0
    error_count: int = 0


class CostTracker:
    """Tracks per-operation LLM cost estimates, error counts, and daily rollups."""

    def __init__(self, max_daily_history: int = 30) -> None:
        self._lock = threading.Lock()
        # Legacy dict-based ops (kept for backward compat with existing callers)
        self._ops: Dict[str, Dict[str, Any]] = defaultdict(
            lambda: {
                "calls": 0,
                "errors": 0,
                "input_tokens": 0,
                "output_tokens": 0,
                "total_latency_ms": 0.0,
                "estimated_cost_usd": 0.0,
            }
        )
        # Enhanced per-operation tracking (from prodo)
        self._operations: Dict[str, OperationCost] = {}
        self._daily: Dict[str, DailyRollup] = {}
        self._max_daily_history = max_daily_history

    def record(
        self,
        operation: str,
        *,
        model: str = "",
        input_tokens: int = 0,
        output_tokens: int = 0,
        latency_ms: float = 0.0,
    ) -> None:
        costs = _MODEL_COSTS.get(model, _MODEL_COSTS["default"])
        est = (input_tokens * costs["input"] + output_tokens * costs["output"]) / 1_000_000
        now = time.time()
        today = datetime.fromtimestamp(now, tz=timezone.utc).strftime("%Y-%m-%d")

        with self._lock:
            # Legacy dict
            op = self._ops[operation]
            op["calls"] += 1
            op["input_tokens"] += input_tokens
            op["output_tokens"] += output_tokens
            op["total_latency_ms"] += latency_ms
            op["estimated_cost_usd"] += est

            # Enhanced OperationCost
            if operation not in self._operations:
                self._operations[operation] = OperationCost(operation=operation, first_call=now)
            oc = self._operations[operation]
            oc.input_tokens += input_tokens
            oc.output_tokens += output_tokens
            oc.estimated_cost_usd += est
            oc.call_count += 1
            oc.total_latency_ms += latency_ms
            oc.last_call = now

            # Daily rollup
            if today not in self._daily:
                self._daily[today] = DailyRollup(date=today)
                self._prune_daily()
            day = self._daily[today]
            day.input_tokens += input_tokens
            day.output_tokens += output_tokens
            day.estimated_cost_usd += est
            day.call_count += 1

    def record_error(self, operation: str) -> None:
        now = time.time()
        today = datetime.fromtimestamp(now, tz=timezone.utc).strftime("%Y-%m-%d")
        with self._lock:
            self._ops[operation]["errors"] += 1
            if operation not in self._operations:
                self._operations[operation] = OperationCost(operation=operation, first_call=now)
            self._operations[operation].error_count += 1
            if today not in self._daily:
                self._daily[today] = DailyRollup(date=today)
            self._daily[today].error_count += 1

    def get_stats(self) -> Dict[str, Any]:
        with self._lock:
            total_cost = sum(oc.estimated_cost_usd for oc in self._operations.values())
            total_calls = sum(oc.call_count for oc in self._operations.values())
            total_errors = sum(oc.error_count for oc in self._operations.values())
            total_tokens = sum(oc.input_tokens + oc.output_tokens for oc in self._operations.values())

            return {
                "summary": {
                    "total_cost_usd": round(total_cost, 6),
                    "total_calls": total_calls,
                    "total_errors": total_errors,
                    "total_tokens": total_tokens,
                },
                "by_operation": {
                    name: oc.to_dict() for name, oc in sorted(self._operations.items())
                },
                "daily": [
                    {
                        "date": d.date,
                        "cost_usd": round(d.estimated_cost_usd, 6),
                        "calls": d.call_count,
                        "errors": d.error_count,
                        "tokens": d.input_tokens + d.output_tokens,
                    }
                    for d in sorted(self._daily.values(), key=lambda x: x.date, reverse=True)
                ],
            }

    def get_operation_stats(self, operation: str) -> Optional[Dict[str, Any]]:
        """Get stats for a specific operation."""
        with self._lock:
            oc = self._operations.get(operation)
            return oc.to_dict() if oc else None

    def reset(self) -> None:
        """Reset all tracking data."""
        with self._lock:
            self._ops.clear()
            self._operations.clear()
            self._daily.clear()

    def _prune_daily(self) -> None:
        """Remove old daily entries beyond max_daily_history."""
        if len(self._daily) > self._max_daily_history:
            sorted_dates = sorted(self._daily.keys())
            for old_date in sorted_dates[:len(self._daily) - self._max_daily_history]:
                del self._daily[old_date]

def get_cost_tracker() -> CostTracker:
    global _singleton
    if _singleton is None:
        with _singleton_lock:
            if _singleton is None:
                _singleton = CostTracker()
    return _singleton

# ── Agents / Crew ──
import json

def _lazy_get_llm_client():
    from backend.app.services.llm import get_llm_client
    return get_llm_client()

logger = logging.getLogger("neura.llm.agents")

class AgentRole(str, Enum):
    """Predefined agent roles."""
    DOCUMENT_ANALYZER = "document_analyzer"
    DATA_EXTRACTOR = "data_extractor"
    SQL_GENERATOR = "sql_generator"
    CHART_SUGGESTER = "chart_suggester"
    TEMPLATE_MAPPER = "template_mapper"
    REPORT_GENERATOR = "report_generator"
    QUALITY_REVIEWER = "quality_reviewer"
    COORDINATOR = "coordinator"

@dataclass
class AgentConfig:
    """Configuration for an agent."""
    role: str
    goal: str
    backstory: str
    model: Optional[str] = None
    temperature: float = 0.7
    max_tokens: Optional[int] = 8192
    tools: List[str] = field(default_factory=list)
    allow_delegation: bool = False
    verbose: bool = False

@dataclass
class Task:
    """A task to be executed by an agent."""
    description: str
    agent_role: str
    expected_output: str
    context: Dict[str, Any] = field(default_factory=dict)
    dependencies: List[str] = field(default_factory=list)
    tools: List[str] = field(default_factory=list)

@dataclass
class TaskResult:
    """Result of a task execution."""
    task_id: str
    agent_role: str
    output: Any
    success: bool
    error: Optional[str] = None
    execution_time: float = 0.0
    token_usage: Dict[str, int] = field(default_factory=dict)

class Tool(ABC):
    """Base class for agent tools."""
    name: str
    description: str

    @abstractmethod
    def execute(self, **kwargs: Any) -> Any:
        """Execute the tool with given arguments."""
        pass

class Agent:
    """
    An AI agent with a specific role and capabilities.

    Agents can:
    - Execute tasks based on their role
    - Use tools to accomplish tasks
    - Delegate to other agents (if allowed)
    """

    def __init__(
        self,
        config: AgentConfig,
        client: Optional[LLMClient] = None,
        tools: Optional[Dict[str, Tool]] = None,
    ):
        self.config = config
        self.client = client or _lazy_get_llm_client()
        self.tools = tools or {}
        self._conversation_history: List[Dict[str, Any]] = []

    @property
    def role(self) -> str:
        return self.config.role

    def execute_task(
        self,
        task: Task,
        context: Optional[Dict[str, Any]] = None,
    ) -> TaskResult:
        """Execute a task and return the result."""
        start_time = time.time()
        task_id = f"{self.role}_{int(start_time)}"

        try:
            # Build the prompt
            prompt = self._build_task_prompt(task, context)

            # Execute the task
            response = self.client.complete(
                messages=[
                    {"role": "system", "content": self._build_system_prompt()},
                    {"role": "user", "content": prompt},
                ],
                model=self.config.model,
                description=f"agent_{self.role}_{task_id}",
                temperature=self.config.temperature,
                max_tokens=self.config.max_tokens,
            )

            output = response["choices"][0]["message"]["content"]
            token_usage = response.get("usage", {})

            # Check if tool use is requested
            output = self._process_tool_calls(output, task)

            # Store in conversation history
            self._conversation_history.append({
                "task": task.description,
                "output": output,
            })

            execution_time = time.time() - start_time

            logger.info(
                "agent_task_complete",
                extra={
                    "event": "agent_task_complete",
                    "agent_role": self.role,
                    "task_id": task_id,
                    "execution_time": execution_time,
                }
            )

            return TaskResult(
                task_id=task_id,
                agent_role=self.role,
                output=output,
                success=True,
                execution_time=execution_time,
                token_usage=token_usage,
            )

        except Exception as e:
            logger.error(
                "agent_task_failed",
                extra={
                    "event": "agent_task_failed",
                    "agent_role": self.role,
                    "task_id": task_id,
                    "error": str(e),
                }
            )
            return TaskResult(
                task_id=task_id,
                agent_role=self.role,
                output=None,
                success=False,
                error=str(e),
                execution_time=time.time() - start_time,
            )

    def _build_system_prompt(self) -> str:
        """Build the system prompt for the agent."""
        prompt = f"""You are a {self.config.role} agent.

GOAL: {self.config.goal}

BACKSTORY: {self.config.backstory}

GUIDELINES:
- Focus on your specific role and expertise
- Provide clear, structured outputs
- If you need to use a tool, indicate it clearly
- Be thorough but concise
"""

        if self.tools:
            prompt += "\n\nAVAILABLE TOOLS:\n"
            for name, tool in self.tools.items():
                prompt += f"- {name}: {tool.description}\n"
            prompt += "\nTo use a tool, respond with: TOOL_CALL: tool_name(arg1=value1, arg2=value2)"

        return prompt

    def _build_task_prompt(
        self,
        task: Task,
        context: Optional[Dict[str, Any]],
    ) -> str:
        """Build the task prompt."""
        prompt = f"""TASK: {task.description}

EXPECTED OUTPUT: {task.expected_output}
"""

        if task.context:
            prompt += f"\nTASK CONTEXT:\n{json.dumps(task.context, indent=2)}\n"

        if context:
            prompt += f"\nPREVIOUS RESULTS:\n{json.dumps(context, indent=2)}\n"

        return prompt

    def _process_tool_calls(self, output: str, task: Task) -> str:
        """Process any tool calls in the output."""
        import re

        tool_pattern = r"TOOL_CALL:\s*(\w+)\((.*?)\)"
        matches = re.findall(tool_pattern, output)

        for tool_name, args_str in matches:
            if tool_name in self.tools:
                try:
                    # Parse arguments
                    args = {}
                    if args_str:
                        for arg in args_str.split(","):
                            if "=" in arg:
                                key, value = arg.split("=", 1)
                                args[key.strip()] = value.strip().strip("'\"")

                    # Execute tool
                    result = self.tools[tool_name].execute(**args)

                    # Replace tool call with result
                    tool_call = f"TOOL_CALL: {tool_name}({args_str})"
                    output = output.replace(tool_call, f"TOOL_RESULT ({tool_name}): {result}")

                except Exception as e:
                    logger.warning(
                        "agent_tool_call_failed",
                        extra={
                            "tool": tool_name,
                            "error": str(e),
                        }
                    )

        return output

class Crew:
    """
    A crew of agents working together on tasks.

    Manages agent coordination and task execution flow.
    """

    def __init__(
        self,
        agents: List[Agent],
        tasks: List[Task],
        verbose: bool = False,
    ):
        self.agents = {agent.role: agent for agent in agents}
        self.tasks = tasks
        self.verbose = verbose
        self._results: Dict[str, TaskResult] = {}

    def kickoff(self, inputs: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """Start the crew's work on all tasks."""
        context = inputs or {}

        logger.info(
            "crew_kickoff",
            extra={
                "event": "crew_kickoff",
                "num_agents": len(self.agents),
                "num_tasks": len(self.tasks),
            }
        )

        for task in self.tasks:
            # Check dependencies
            for dep in task.dependencies:
                if dep not in self._results or not self._results[dep].success:
                    logger.warning(
                        "crew_task_dependency_not_met",
                        extra={
                            "task": task.description[:50],
                            "dependency": dep,
                        }
                    )
                    continue

            # Get the agent for this task
            agent = self.agents.get(task.agent_role)
            if not agent:
                logger.error(
                    "crew_agent_not_found",
                    extra={
                        "agent_role": task.agent_role,
                        "available_agents": list(self.agents.keys()),
                    }
                )
                continue

            # Build context from previous results
            task_context = {**context}
            for dep in task.dependencies:
                if dep in self._results:
                    task_context[dep] = self._results[dep].output

            # Execute the task
            result = agent.execute_task(task, task_context)
            self._results[task.description[:50]] = result

            if result.success:
                context[task.description[:50]] = result.output

            if self.verbose:
                print(f"[{task.agent_role}] {task.description[:50]}: {'SUCCESS' if result.success else 'FAILED'}")

        return {
            "results": {k: v.output for k, v in self._results.items() if v.success},
            "errors": {k: v.error for k, v in self._results.items() if not v.success},
            "execution_summary": self._get_execution_summary(),
        }

    def _get_execution_summary(self) -> Dict[str, Any]:
        """Get a summary of the execution."""
        total_time = sum(r.execution_time for r in self._results.values())
        total_tokens = sum(
            r.token_usage.get("total_tokens", 0)
            for r in self._results.values()
        )

        return {
            "total_tasks": len(self.tasks),
            "successful_tasks": sum(1 for r in self._results.values() if r.success),
            "failed_tasks": sum(1 for r in self._results.values() if not r.success),
            "total_execution_time": total_time,
            "total_tokens_used": total_tokens,
        }

# Predefined agent configs: role -> (goal, backstory, temperature)
_AGENT_PRESETS: Dict[str, tuple] = {
    AgentRole.DOCUMENT_ANALYZER.value: (
        "Analyze documents to extract structure, content, and metadata",
        "Expert document analyst skilled at PDFs, spreadsheets, and reports.", 0.3,
    ),
    AgentRole.DATA_EXTRACTOR.value: (
        "Extract structured data from documents accurately",
        "Meticulous data extraction specialist. Accuracy is top priority.", 0.2,
    ),
    AgentRole.SQL_GENERATOR.value: (
        "Generate accurate and efficient SQL queries",
        "Database expert proficient in DuckDB, SQLite, and standard SQL.", 0.1,
    ),
    AgentRole.CHART_SUGGESTER.value: (
        "Suggest optimal visualizations for data",
        "Data visualization expert who presents data effectively.", 0.5,
    ),
    AgentRole.TEMPLATE_MAPPER.value: (
        "Map data fields to template placeholders accurately",
        "Expert at understanding document templates and mapping data fields.", 0.2,
    ),
    AgentRole.QUALITY_REVIEWER.value: (
        "Review outputs for accuracy and quality",
        "Meticulous QA specialist who catches errors others miss.", 0.3,
    ),
}

def _create_preset_agent(role: str, client: Optional[LLMClient] = None) -> Agent:
    """Create an agent from presets."""
    goal, backstory, temp = _AGENT_PRESETS[role]
    return Agent(AgentConfig(role=role, goal=goal, backstory=backstory, temperature=temp), client)

def create_document_analyzer_agent(client: Optional[LLMClient] = None) -> Agent:
    return _create_preset_agent(AgentRole.DOCUMENT_ANALYZER.value, client)

def create_data_extractor_agent(client: Optional[LLMClient] = None) -> Agent:
    return _create_preset_agent(AgentRole.DATA_EXTRACTOR.value, client)

def create_sql_generator_agent(client: Optional[LLMClient] = None) -> Agent:
    return _create_preset_agent(AgentRole.SQL_GENERATOR.value, client)

def create_chart_suggester_agent(client: Optional[LLMClient] = None) -> Agent:
    return _create_preset_agent(AgentRole.CHART_SUGGESTER.value, client)

def create_template_mapper_agent(client: Optional[LLMClient] = None) -> Agent:
    return _create_preset_agent(AgentRole.TEMPLATE_MAPPER.value, client)

def create_quality_reviewer_agent(client: Optional[LLMClient] = None) -> Agent:
    return _create_preset_agent(AgentRole.QUALITY_REVIEWER.value, client)

# Pre-built crews for common workflows

def create_document_processing_crew(
    client: Optional[LLMClient] = None,
    verbose: bool = False,
) -> Crew:
    """Create a crew for document processing workflow."""
    agents = [
        create_document_analyzer_agent(client),
        create_data_extractor_agent(client),
        create_quality_reviewer_agent(client),
    ]

    tasks = [
        Task(
            description="Analyze the document structure and identify key sections",
            agent_role=AgentRole.DOCUMENT_ANALYZER.value,
            expected_output="JSON with document structure, sections, and content overview",
        ),
        Task(
            description="Extract all tables and structured data from the document",
            agent_role=AgentRole.DATA_EXTRACTOR.value,
            expected_output="JSON with extracted tables, their headers, and row data",
            dependencies=["Analyze the document structure and identify key sections"[:50]],
        ),
        Task(
            description="Review extracted data for accuracy and completeness",
            agent_role=AgentRole.QUALITY_REVIEWER.value,
            expected_output="Quality report with any issues found and suggestions",
            dependencies=["Extract all tables and structured data from the document"[:50]],
        ),
    ]

    return Crew(agents, tasks, verbose)

def create_report_generation_crew(
    client: Optional[LLMClient] = None,
    verbose: bool = False,
) -> Crew:
    """Create a crew for report generation workflow."""
    agents = [
        create_data_extractor_agent(client),
        create_sql_generator_agent(client),
        create_template_mapper_agent(client),
        create_chart_suggester_agent(client),
    ]

    tasks = [
        Task(
            description="Extract and prepare data for the report",
            agent_role=AgentRole.DATA_EXTRACTOR.value,
            expected_output="Clean, structured data ready for report generation",
        ),
        Task(
            description="Generate SQL queries for data retrieval",
            agent_role=AgentRole.SQL_GENERATOR.value,
            expected_output="DuckDB-compatible SQL queries for report data",
            dependencies=["Extract and prepare data for the report"[:50]],
        ),
        Task(
            description="Map data fields to template placeholders",
            agent_role=AgentRole.TEMPLATE_MAPPER.value,
            expected_output="Field mapping configuration JSON",
            dependencies=["Extract and prepare data for the report"[:50]],
        ),
        Task(
            description="Suggest visualizations for the report data",
            agent_role=AgentRole.CHART_SUGGESTER.value,
            expected_output="Chart recommendations with configurations",
            dependencies=["Extract and prepare data for the report"[:50]],
        ),
    ]

    return Crew(agents, tasks, verbose)

# ── Text-to-SQL ──
import re

logger = logging.getLogger("neura.llm.text_to_sql")

@dataclass
class TableSchema:
    """Schema information for a database table."""
    name: str
    columns: List[Dict[str, str]]  # [{"name": "col", "type": "TEXT", "description": "..."}]
    primary_key: Optional[str] = None
    foreign_keys: List[Dict[str, str]] = field(default_factory=list)
    sample_values: Dict[str, List[Any]] = field(default_factory=dict)
    description: Optional[str] = None

@dataclass
class SQLGenerationResult:
    """Result of SQL generation."""
    sql: str
    explanation: str
    confidence: float
    dialect: str
    warnings: List[str] = field(default_factory=list)
    raw_response: str = ""

class TextToSQL:
    """
    Text-to-SQL generation using LLM with SQLCoder-style prompting.

    Supports:
    - Natural language to SQL conversion
    - Schema-aware generation
    - Multi-table joins
    - Aggregations and grouping
    - DuckDB and SQLite dialects
    """

    # Recommended models for SQL generation per provider
    SQL_MODELS = {
        LLMProvider.CLAUDE_CODE: "qwen",  # Local Qwen model default
    }

    def __init__(
        self,
        client: Optional[LLMClient] = None,
        dialect: str = "duckdb",
        model: Optional[str] = None,
    ):
        self.client = client or _lazy_get_llm_client()
        self.dialect = dialect.lower()
        self._model = model
        self._schemas: Dict[str, TableSchema] = {}

    @property
    def model(self) -> str:
        """Get the model for SQL generation."""
        if self._model:
            return self._model
        provider = self.client.config.provider
        return self.SQL_MODELS.get(provider, self.client.config.model)

    def add_table_schema(self, schema: TableSchema) -> None:
        """Add a table schema for context."""
        self._schemas[schema.name] = schema

    def add_schemas_from_catalog(self, catalog: Dict[str, Any]) -> None:
        """Add schemas from a database catalog dictionary."""
        for table_name, table_info in catalog.items():
            columns = []
            for col in table_info.get("columns", []):
                if isinstance(col, dict):
                    columns.append({
                        "name": col.get("name", ""),
                        "type": col.get("type", "TEXT"),
                        "description": col.get("description", ""),
                    })
                elif isinstance(col, str):
                    columns.append({"name": col, "type": "TEXT", "description": ""})

            schema = TableSchema(
                name=table_name,
                columns=columns,
                primary_key=table_info.get("primary_key"),
                foreign_keys=table_info.get("foreign_keys", []),
                sample_values=table_info.get("sample_values", {}),
                description=table_info.get("description"),
            )
            self._schemas[table_name] = schema

    def generate_sql(
        self,
        question: str,
        tables: Optional[List[str]] = None,
        context: Optional[str] = None,
    ) -> SQLGenerationResult:
        """Generate SQL from a natural language question."""
        # Build the prompt
        prompt = self._build_sqlcoder_prompt(question, tables, context)

        # Generate SQL
        response = self.client.complete(
            messages=[{"role": "user", "content": prompt}],
            model=self.model,
            description="text_to_sql",
            temperature=0.0,  # Deterministic for SQL
        )

        raw_content = response["choices"][0]["message"]["content"]
        return self._parse_sql_response(raw_content)

    def generate_sql_with_decomposition(
        self,
        question: str,
        tables: Optional[List[str]] = None,
    ) -> SQLGenerationResult:
        """
        Generate SQL using query decomposition for complex questions.

        This approach breaks down complex questions into simpler sub-queries
        before combining them.
        """
        # First, analyze the question complexity
        analysis_prompt = f"""Analyze this question and determine if it needs to be decomposed:

Question: {question}

Respond in JSON format:
```json
{{
  "is_complex": true/false,
  "sub_questions": ["sub-question 1", "sub-question 2"],
  "combination_strategy": "join|union|subquery|none"
}}
```"""

        analysis_response = self.client.complete(
            messages=[{"role": "user", "content": analysis_prompt}],
            model=self.model,
            description="sql_decomposition_analysis",
            temperature=0.0,
        )

        analysis = self._parse_json_response(
            analysis_response["choices"][0]["message"]["content"],
            {"is_complex": False, "sub_questions": [], "combination_strategy": "none"}
        )

        if not analysis.get("is_complex") or not analysis.get("sub_questions"):
            # Simple question, generate directly
            return self.generate_sql(question, tables)

        # Generate SQL for each sub-question
        sub_queries = []
        for sub_q in analysis["sub_questions"]:
            result = self.generate_sql(sub_q, tables)
            sub_queries.append({
                "question": sub_q,
                "sql": result.sql,
            })

        # Combine sub-queries
        combination_prompt = f"""Combine these sub-queries to answer the original question.

Original Question: {question}

Sub-queries:
{json.dumps(sub_queries, indent=2)}

Combination Strategy: {analysis.get('combination_strategy', 'join')}

Generate the final combined SQL query for {self.dialect.upper()}.
Return ONLY the SQL, no explanation."""

        final_response = self.client.complete(
            messages=[{"role": "user", "content": combination_prompt}],
            model=self.model,
            description="sql_combination",
            temperature=0.0,
        )

        final_sql = self._extract_sql(final_response["choices"][0]["message"]["content"])

        return SQLGenerationResult(
            sql=final_sql,
            explanation=f"Combined from {len(sub_queries)} sub-queries",
            confidence=0.8,
            dialect=self.dialect,
            warnings=["Query was decomposed and combined"],
            raw_response=final_response["choices"][0]["message"]["content"],
        )

    def validate_and_fix_sql(
        self,
        sql: str,
        error_message: Optional[str] = None,
    ) -> SQLGenerationResult:
        """Validate SQL and attempt to fix any errors."""
        schema_context = self._build_schema_context()

        prompt = f"""Review and fix this SQL query.

Schema:
{schema_context}

Original SQL:
```sql
{sql}
```

{f"Error encountered: {error_message}" if error_message else ""}

Tasks:
1. Check for syntax errors
2. Verify table and column names match the schema
3. Fix any issues found
4. Ensure the query is valid {self.dialect.upper()}

Return the corrected SQL in a code block, followed by an explanation of changes made."""

        response = self.client.complete(
            messages=[{"role": "user", "content": prompt}],
            model=self.model,
            description="sql_validation",
            temperature=0.0,
        )

        raw_content = response["choices"][0]["message"]["content"]
        result = self._parse_sql_response(raw_content)

        if error_message:
            result.warnings.append(f"Fixed error: {error_message}")

        return result

    def explain_sql(self, sql: str) -> str:
        """Generate a natural language explanation of SQL."""
        prompt = f"""Explain this SQL query in simple terms:

```sql
{sql}
```

Provide a clear, non-technical explanation of:
1. What data is being retrieved
2. What filters/conditions are applied
3. How the results are organized
4. Any calculations or aggregations performed"""

        response = self.client.complete(
            messages=[{"role": "user", "content": prompt}],
            model=self.model,
            description="sql_explanation",
            temperature=0.3,
        )

        return response["choices"][0]["message"]["content"]

    def _build_sqlcoder_prompt(
        self,
        question: str,
        tables: Optional[List[str]],
        context: Optional[str],
    ) -> str:
        """Build a SQLCoder-style prompt."""
        schema_context = self._build_schema_context(tables)

        # SQLCoder-style prompt structure
        prompt = f"""### Task
Generate a SQL query to answer [QUESTION]{question}[/QUESTION]

### Database Schema
The query will run on a database with the following schema:
{schema_context}

### SQL Dialect
Use {self.dialect.upper()} syntax.

"""
        if context:
            prompt += f"""### Additional Context
{context}

"""

        prompt += """### Guidelines
- Use proper table aliases
- Handle NULL values appropriately
- Use appropriate JOIN types
- Include only necessary columns
- Use appropriate aggregation functions
- Ensure the query is efficient

### Answer
Given the database schema, here is the SQL query that answers [QUESTION]{question}[/QUESTION]:
```sql
"""
        return prompt

    def _build_schema_context(
        self,
        tables: Optional[List[str]] = None,
    ) -> str:
        """Build schema context string for the prompt."""
        if not self._schemas:
            return "No schema information available."

        schemas_to_use = (
            {name: self._schemas[name] for name in tables if name in self._schemas}
            if tables
            else self._schemas
        )

        context_parts = []

        for table_name, schema in schemas_to_use.items():
            # Format CREATE TABLE statement
            columns_def = []
            for col in schema.columns:
                col_def = f"  {col['name']} {col.get('type', 'TEXT')}"
                if col.get('description'):
                    col_def += f" -- {col['description']}"
                columns_def.append(col_def)

            table_def = f"CREATE TABLE {table_name} (\n"
            table_def += ",\n".join(columns_def)

            if schema.primary_key:
                table_def += f",\n  PRIMARY KEY ({schema.primary_key})"

            for fk in schema.foreign_keys:
                table_def += f",\n  FOREIGN KEY ({fk.get('column')}) REFERENCES {fk.get('references')}"

            table_def += "\n);"

            if schema.description:
                table_def = f"-- {schema.description}\n{table_def}"

            context_parts.append(table_def)

            # Add sample values if available
            if schema.sample_values:
                samples = []
                for col, values in schema.sample_values.items():
                    if values:
                        samples.append(f"  {col}: {', '.join(str(v) for v in values[:3])}")
                if samples:
                    context_parts.append(f"-- Sample values for {table_name}:\n" + "\n".join(samples))

        return "\n\n".join(context_parts)

    def _parse_sql_response(self, raw_content: str) -> SQLGenerationResult:
        """Parse the LLM response to extract SQL."""
        sql = self._extract_sql(raw_content)

        # Extract explanation (text after the SQL block)
        explanation = ""
        sql_end = raw_content.rfind("```")
        if sql_end != -1:
            explanation = raw_content[sql_end + 3:].strip()

        confidence = 0.9
        warnings = []

        if not sql:
            confidence = 0.0
            warnings.append("No SQL found in response")
        elif "SELECT *" in sql.upper():
            confidence -= 0.1
            warnings.append("Using SELECT * may be inefficient")
        if "-- TODO" in sql or "-- FIXME" in sql:
            confidence -= 0.2
            warnings.append("Query contains TODO/FIXME comments")

        return SQLGenerationResult(
            sql=sql,
            explanation=explanation,
            confidence=min(max(confidence, 0.0), 1.0),
            dialect=self.dialect,
            warnings=warnings,
            raw_response=raw_content,
        )

    def _extract_sql(self, content: str) -> str:
        """Extract SQL from LLM response."""
        # Try to find SQL in code blocks
        sql_match = re.search(r"```(?:sql)?\s*([\s\S]*?)```", content, re.IGNORECASE)
        if sql_match:
            return sql_match.group(1).strip()

        # Try to find SELECT/INSERT/UPDATE/DELETE statement
        statement_match = re.search(
            r"(SELECT|INSERT|UPDATE|DELETE|WITH|CREATE|ALTER|DROP)\b[\s\S]+?(?:;|$)",
            content,
            re.IGNORECASE
        )
        if statement_match:
            return statement_match.group(0).strip().rstrip(";") + ";"

        return content.strip()

    def _parse_json_response(self, raw_content: str, default: Dict[str, Any]) -> Dict[str, Any]:
        return _parse_json_from_llm(raw_content, default)

# Convenience functions

def get_text_to_sql(dialect: str = "duckdb") -> TextToSQL:
    """Get a TextToSQL instance."""
    return TextToSQL(dialect=dialect)

def generate_sql(
    question: str,
    schema: Dict[str, Any],
    dialect: str = "duckdb",
) -> str:
    """Quick function to generate SQL from a question."""
    t2sql = TextToSQL(dialect=dialect)
    t2sql.add_schemas_from_catalog(schema)
    result = t2sql.generate_sql(question)
    return result.sql

# ── RAG Retrieval ──
import math
from collections import Counter

logger = logging.getLogger("neura.llm.rag")

@dataclass
class Document:
    """A document chunk for retrieval."""
    id: str
    content: str
    metadata: Dict[str, Any] = field(default_factory=dict)
    embedding: Optional[List[float]] = None

@dataclass
class RetrievalResult:
    """Result of document retrieval."""
    documents: List[Document]
    scores: List[float]
    query: str
    method: str

class BM25Index:
    """
    BM25 keyword-based retrieval index.

    Efficient for exact and fuzzy keyword matching.
    """

    def __init__(self, k1: float = 1.5, b: float = 0.75):
        self.k1 = k1
        self.b = b
        self._documents: List[Document] = []
        self._doc_freqs: Dict[str, int] = Counter()
        self._doc_lens: List[int] = []
        self._avg_doc_len: float = 0.0
        self._inverted_index: Dict[str, List[Tuple[int, int]]] = {}  # term -> [(doc_idx, term_freq)]

    def add_documents(self, documents: List[Document]) -> None:
        """Add documents to the index."""
        for doc in documents:
            self._add_document(doc)
        self._avg_doc_len = sum(self._doc_lens) / len(self._doc_lens) if self._doc_lens else 0

    def _add_document(self, doc: Document) -> None:
        """Add a single document to the index."""
        doc_idx = len(self._documents)
        self._documents.append(doc)

        # Tokenize
        tokens = self._tokenize(doc.content)
        self._doc_lens.append(len(tokens))

        # Count term frequencies
        term_freqs = Counter(tokens)

        # Update inverted index and document frequencies
        for term, freq in term_freqs.items():
            if term not in self._inverted_index:
                self._inverted_index[term] = []
            self._inverted_index[term].append((doc_idx, freq))
            self._doc_freqs[term] += 1

    def search(self, query: str, top_k: int = 5) -> List[Tuple[Document, float]]:
        """Search the index and return top-k documents with scores."""
        if not self._documents:
            return []

        query_tokens = self._tokenize(query)
        scores: Dict[int, float] = {}
        n_docs = len(self._documents)

        for term in query_tokens:
            if term not in self._inverted_index:
                continue

            # Calculate IDF
            df = self._doc_freqs[term]
            idf = math.log((n_docs - df + 0.5) / (df + 0.5) + 1)

            # Score each document containing the term
            for doc_idx, tf in self._inverted_index[term]:
                doc_len = self._doc_lens[doc_idx]
                # BM25 scoring formula
                numerator = tf * (self.k1 + 1)
                denominator = tf + self.k1 * (1 - self.b + self.b * doc_len / self._avg_doc_len)
                score = idf * numerator / denominator

                scores[doc_idx] = scores.get(doc_idx, 0) + score

        # Sort by score and return top-k
        sorted_results = sorted(scores.items(), key=lambda x: x[1], reverse=True)[:top_k]
        return [(self._documents[idx], score) for idx, score in sorted_results]

    def _tokenize(self, text: str) -> List[str]:
        """Simple tokenization."""
        # Lowercase and split on non-alphanumeric
        tokens = re.findall(r'\b\w+\b', text.lower())
        # Remove very short tokens and stopwords
        stopwords = {'the', 'a', 'an', 'is', 'are', 'was', 'were', 'be', 'been',
                     'being', 'have', 'has', 'had', 'do', 'does', 'did', 'will',
                     'would', 'could', 'should', 'may', 'might', 'must', 'can',
                     'this', 'that', 'these', 'those', 'i', 'you', 'he', 'she',
                     'it', 'we', 'they', 'what', 'which', 'who', 'whom', 'whose',
                     'where', 'when', 'why', 'how', 'all', 'each', 'every', 'both',
                     'few', 'more', 'most', 'other', 'some', 'such', 'no', 'nor',
                     'not', 'only', 'own', 'same', 'so', 'than', 'too', 'very',
                     'just', 'and', 'but', 'or', 'if', 'because', 'as', 'until',
                     'while', 'of', 'at', 'by', 'for', 'with', 'about', 'against',
                     'between', 'into', 'through', 'during', 'before', 'after',
                     'above', 'below', 'to', 'from', 'up', 'down', 'in', 'out',
                     'on', 'off', 'over', 'under', 'again', 'further', 'then', 'once'}
        return [t for t in tokens if len(t) > 2 and t not in stopwords]

class SimpleVectorStore:
    """
    Simple in-memory vector store using cosine similarity.

    For production, consider using a proper vector database like
    Chroma, Weaviate, or Pinecone.
    """

    def __init__(self):
        self._documents: List[Document] = []

    def add_documents(self, documents: List[Document]) -> None:
        """Add documents with embeddings to the store."""
        for doc in documents:
            if doc.embedding is not None:
                self._documents.append(doc)

    def search(
        self,
        query_embedding: List[float],
        top_k: int = 5,
    ) -> List[Tuple[Document, float]]:
        """Search for similar documents using cosine similarity."""
        if not self._documents or not query_embedding:
            return []

        scores = []
        for doc in self._documents:
            if doc.embedding:
                scores.append((doc, _cosine_similarity(query_embedding, doc.embedding)))

        scores.sort(key=lambda x: x[1], reverse=True)
        return scores[:top_k]

class RAGRetriever:
    """
    RAG Retriever combining keyword and semantic search.

    Features:
    - BM25 keyword search
    - Optional vector similarity search
    - Hybrid retrieval with score fusion
    - Context window management
    """

    def __init__(
        self,
        client: Optional[LLMClient] = None,
        use_embeddings: bool = False,
        max_context_tokens: int = 4000,
    ):
        self.client = client or _lazy_get_llm_client()
        self.use_embeddings = use_embeddings
        self.max_context_tokens = max_context_tokens

        self._bm25_index = BM25Index()
        self._vector_store = SimpleVectorStore() if use_embeddings else None
        self._documents: Dict[str, Document] = {}

    def add_document(
        self,
        content: str,
        doc_id: Optional[str] = None,
        metadata: Optional[Dict[str, Any]] = None,
        chunk_size: int = 500,
        chunk_overlap: int = 50,
    ) -> List[str]:
        """Add a document to the retriever."""
        if not doc_id:
            doc_id = hashlib.md5(content.encode()).hexdigest()[:12]

        # Chunk the document
        chunks = self._chunk_text(content, chunk_size, chunk_overlap)
        chunk_ids = []

        for i, chunk_content in enumerate(chunks):
            chunk_id = f"{doc_id}_chunk_{i}"
            chunk_metadata = {
                **(metadata or {}),
                "parent_doc_id": doc_id,
                "chunk_index": i,
                "total_chunks": len(chunks),
            }

            doc = Document(
                id=chunk_id,
                content=chunk_content,
                metadata=chunk_metadata,
            )

            # Generate embedding if enabled
            if self.use_embeddings:
                doc.embedding = self._get_embedding(chunk_content)

            self._documents[chunk_id] = doc
            chunk_ids.append(chunk_id)

        # Rebuild indices
        self._rebuild_indices()

        return chunk_ids

    def add_documents_bulk(
        self,
        documents: List[Dict[str, Any]],
        chunk_size: int = 500,
    ) -> None:
        """Add multiple documents in bulk."""
        for doc in documents:
            self.add_document(
                content=doc["content"],
                doc_id=doc.get("id"),
                metadata=doc.get("metadata"),
                chunk_size=chunk_size,
            )

    def retrieve(
        self,
        query: str,
        top_k: int = 5,
        method: str = "hybrid",
    ) -> RetrievalResult:
        """Retrieve relevant documents for a query."""
        if method == "bm25" or not self.use_embeddings:
            results = self._bm25_index.search(query, top_k)
            return RetrievalResult(
                documents=[doc for doc, _ in results],
                scores=[score for _, score in results],
                query=query,
                method="bm25",
            )

        if method == "vector" and self._vector_store:
            query_embedding = self._get_embedding(query)
            results = self._vector_store.search(query_embedding, top_k)
            return RetrievalResult(
                documents=[doc for doc, _ in results],
                scores=[score for _, score in results],
                query=query,
                method="vector",
            )

        # Hybrid: combine BM25 and vector scores
        if method == "hybrid" and self._vector_store:
            bm25_results = self._bm25_index.search(query, top_k * 2)
            query_embedding = self._get_embedding(query)
            vector_results = self._vector_store.search(query_embedding, top_k * 2)

            # Reciprocal Rank Fusion
            fused_scores = self._reciprocal_rank_fusion(
                [r[0].id for r in bm25_results],
                [r[0].id for r in vector_results],
            )

            # Get top-k documents
            sorted_ids = sorted(fused_scores.items(), key=lambda x: x[1], reverse=True)[:top_k]
            documents = [self._documents[doc_id] for doc_id, _ in sorted_ids if doc_id in self._documents]
            scores = [score for _, score in sorted_ids[:len(documents)]]

            return RetrievalResult(
                documents=documents,
                scores=scores,
                query=query,
                method="hybrid",
            )

        # Fallback to BM25
        results = self._bm25_index.search(query, top_k)
        return RetrievalResult(
            documents=[doc for doc, _ in results],
            scores=[score for _, score in results],
            query=query,
            method="bm25",
        )

    def query_with_context(
        self,
        question: str,
        top_k: int = 5,
        include_sources: bool = True,
    ) -> Dict[str, Any]:
        """Answer a question using retrieved context."""
        # Retrieve relevant documents
        retrieval = self.retrieve(question, top_k, method="hybrid")

        if not retrieval.documents:
            return {
                "answer": "I couldn't find relevant information to answer this question.",
                "sources": [],
                "context_used": False,
            }

        # Build context
        context = self._build_context(retrieval.documents)

        # Generate answer with context
        prompt = f"""Answer the following question based on the provided context.

Context:
{context}

Question: {question}

Instructions:
- Answer based ONLY on the information provided in the context
- If the context does not contain enough information, say so
- Be concise but complete
- Cite specific parts of the context when relevant

Answer:"""

        response = self.client.complete(
            messages=[{"role": "user", "content": prompt}],
            description="rag_query",
            temperature=0.3,
        )

        answer = response["choices"][0]["message"]["content"]

        sources = []
        if include_sources:
            for doc in retrieval.documents:
                sources.append({
                    "id": doc.id,
                    "content_preview": doc.content[:200] + "..." if len(doc.content) > 200 else doc.content,
                    "metadata": doc.metadata,
                })

        return {
            "answer": answer,
            "sources": sources,
            "context_used": True,
            "documents_retrieved": len(retrieval.documents),
        }

    def _chunk_text(
        self,
        text: str,
        chunk_size: int,
        chunk_overlap: int,
    ) -> List[str]:
        """Split text into overlapping chunks."""
        if len(text) <= chunk_size:
            return [text]

        chunks = []
        start = 0

        while start < len(text):
            end = start + chunk_size

            # Try to break at sentence boundary
            if end < len(text):
                # Look for sentence end near the boundary
                for punct in ['. ', '! ', '? ', '\n\n', '\n']:
                    boundary = text.rfind(punct, start + chunk_size // 2, end)
                    if boundary != -1:
                        end = boundary + len(punct)
                        break

            chunks.append(text[start:end].strip())
            start = end - chunk_overlap

        return [c for c in chunks if c]  # Filter empty chunks

    def _build_context(self, documents: List[Document]) -> str:
        """Build context string from documents, respecting token limit."""
        context_parts = []
        estimated_tokens = 0

        for i, doc in enumerate(documents):
            # Rough token estimation (4 chars per token)
            doc_tokens = len(doc.content) // 4

            if estimated_tokens + doc_tokens > self.max_context_tokens:
                # Truncate if needed
                remaining_tokens = self.max_context_tokens - estimated_tokens
                truncated_content = doc.content[:remaining_tokens * 4]
                context_parts.append(f"[Document {i + 1}]\n{truncated_content}...")
                break

            context_parts.append(f"[Document {i + 1}]\n{doc.content}")
            estimated_tokens += doc_tokens

        return "\n\n".join(context_parts)

    def _rebuild_indices(self) -> None:
        """Rebuild search indices."""
        documents = list(self._documents.values())
        self._bm25_index = BM25Index()
        self._bm25_index.add_documents(documents)

        if self._vector_store:
            self._vector_store = SimpleVectorStore()
            self._vector_store.add_documents(documents)

    def _get_embedding(self, text: str) -> List[float]:
        """
        Get embedding for text.

        Uses LLM to generate a simple embedding (not ideal but works without
        external embedding service).
        """
        # For production, use a proper embedding model like:
        # - OpenAI text-embedding-ada-002
        # - Sentence transformers
        # - Ollama embeddings

        # Simple hash-based embedding as fallback
        words = re.findall(r'\b\w+\b', text.lower())
        word_hashes = [hash(word) % 1000 / 1000.0 for word in words[:256]]

        # Pad or truncate to fixed size
        embedding_size = 256
        if len(word_hashes) < embedding_size:
            word_hashes.extend([0.0] * (embedding_size - len(word_hashes)))

        return word_hashes[:embedding_size]

    def _reciprocal_rank_fusion(
        self,
        ranking1: List[str],
        ranking2: List[str],
        k: int = 60,
    ) -> Dict[str, float]:
        """Combine two rankings using Reciprocal Rank Fusion."""
        fused_scores: Dict[str, float] = {}

        for rank, doc_id in enumerate(ranking1):
            fused_scores[doc_id] = fused_scores.get(doc_id, 0) + 1 / (k + rank + 1)

        for rank, doc_id in enumerate(ranking2):
            fused_scores[doc_id] = fused_scores.get(doc_id, 0) + 1 / (k + rank + 1)

        return fused_scores

# Convenience functions

def create_retriever(use_embeddings: bool = False) -> RAGRetriever:
    """Create a RAG retriever instance."""
    return RAGRetriever(use_embeddings=use_embeddings)

def quick_rag_query(
    question: str,
    documents: List[str],
    top_k: int = 3,
) -> str:
    """Quick RAG query over a list of documents."""
    retriever = RAGRetriever(use_embeddings=False)

    for i, doc in enumerate(documents):
        retriever.add_document(doc, doc_id=f"doc_{i}")

    result = retriever.query_with_context(question, top_k=top_k)
    return result["answer"]

# ── Document Extraction ──
logger = logging.getLogger("neura.llm.document_extractor")

@dataclass
class ExtractedTable:
    """Extracted table from a document."""
    id: str
    title: Optional[str]
    headers: List[str]
    rows: List[List[Any]]
    metadata: Dict[str, Any] = field(default_factory=dict)
    confidence: float = 1.0

@dataclass
class ExtractedContent:
    """Complete extracted content from a document."""
    text: str
    tables: List[ExtractedTable]
    metadata: Dict[str, Any]
    structure: Dict[str, Any]
    warnings: List[str] = field(default_factory=list)

@dataclass
class FieldSchema:
    """Schema for an extracted field."""
    name: str
    data_type: str  # text, numeric, datetime, boolean
    sample_values: List[Any]
    nullable: bool = True
    description: Optional[str] = None

class EnhancedDocumentExtractor:
    """
    Enhanced document extractor with AI-powered understanding.

    Combines:
    - Traditional PDF/Excel extraction
    - Layout analysis
    - VLM for complex documents
    - Automatic schema inference
    """

    def __init__(
        self,
        use_vlm: bool = True,
        use_ocr: bool = True,
        max_pages: int = 50,
        max_tables: int = 100,
    ):
        self.use_vlm = use_vlm
        self.use_ocr = use_ocr
        self.max_pages = max_pages
        self.max_tables = max_tables

        # Lazy load VLM
        self._vlm = None

    @property
    def vlm(self):
        """Get VLM instance (lazy loaded)."""
        if self._vlm is None and self.use_vlm:
            try:
                from .vision import get_vlm
                self._vlm = get_vlm()
            except Exception as e:
                logger.warning(f"VLM not available: {e}")
                self._vlm = None
        return self._vlm

    def extract(
        self,
        file_path: Union[str, Path],
        extraction_mode: str = "auto",
    ) -> ExtractedContent:
        """Extract content from a document."""
        file_path = Path(file_path)

        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        suffix = file_path.suffix.lower()

        if suffix == ".pdf":
            return self._extract_pdf(file_path, extraction_mode)
        elif suffix in (".xlsx", ".xls", ".xlsm"):
            return self._extract_excel(file_path, extraction_mode)
        elif suffix == ".csv":
            return self._extract_csv(file_path)
        elif suffix in (".png", ".jpg", ".jpeg", ".tiff", ".bmp"):
            return self._extract_image(file_path, extraction_mode)
        else:
            raise ValueError(f"Unsupported file format: {suffix}")

    def _extract_pdf(
        self,
        file_path: Path,
        mode: str,
    ) -> ExtractedContent:
        """Extract content from PDF."""
        try:
            import fitz  # PyMuPDF
        except ImportError:
            raise RuntimeError("PyMuPDF is required for PDF extraction. Install with: pip install pymupdf")

        doc = fitz.open(file_path)
        warnings = []

        if doc.page_count > self.max_pages:
            warnings.append(f"PDF has {doc.page_count} pages, only processing first {self.max_pages}")

        text_content = []
        tables: List[ExtractedTable] = []
        metadata = {
            "filename": file_path.name,
            "page_count": doc.page_count,
            "format": "pdf",
        }

        for page_num, page in enumerate(doc):
            if page_num >= self.max_pages:
                break

            # Extract text
            page_text = page.get_text("text")
            text_content.append(f"--- Page {page_num + 1} ---\n{page_text}")

            # Extract tables using built-in method
            page_tables = self._extract_pdf_tables(page, page_num)
            tables.extend(page_tables)

            # Use VLM for complex layouts if enabled
            if self.vlm and mode in ("auto", "comprehensive"):
                try:
                    pix = page.get_pixmap(dpi=150)
                    img_bytes = pix.tobytes("png")
                    vlm_result = self.vlm.extract_tables(img_bytes)

                    # Merge VLM-extracted tables
                    for i, vt in enumerate(vlm_result.tables):
                        table_id = f"page{page_num + 1}_vlm_{i + 1}"
                        if not self._table_exists(tables, vt.get("headers", [])):
                            tables.append(ExtractedTable(
                                id=table_id,
                                title=vt.get("title"),
                                headers=vt.get("headers", []),
                                rows=vt.get("rows", []),
                                metadata={"source": "vlm", "page": page_num + 1},
                                confidence=vlm_result.confidence,
                            ))
                except Exception as e:
                    logger.warning(f"VLM extraction failed for page {page_num + 1}: {e}")

            if len(tables) >= self.max_tables:
                warnings.append(f"Table limit ({self.max_tables}) reached")
                break

        doc.close()

        # Analyze document structure
        structure = self._analyze_structure("\n\n".join(text_content))

        return ExtractedContent(
            text="\n\n".join(text_content),
            tables=tables,
            metadata=metadata,
            structure=structure,
            warnings=warnings,
        )

    def _extract_pdf_tables(
        self,
        page,
        page_num: int,
    ) -> List[ExtractedTable]:
        """Extract tables from a PDF page."""
        tables = []

        try:
            # Use PyMuPDF's table detection
            page_tables = page.find_tables()

            for i, table in enumerate(page_tables):
                if table.row_count == 0:
                    continue

                data = table.extract()
                if not data or len(data) < 2:
                    continue

                # First row as headers
                headers = [str(cell or "").strip() for cell in data[0]]

                # Remaining rows as data
                rows = []
                for row in data[1:]:
                    normalized_row = []
                    for j, cell in enumerate(row):
                        if j < len(headers):
                            normalized_row.append(str(cell or "").strip())
                    # Pad row to match headers
                    while len(normalized_row) < len(headers):
                        normalized_row.append("")
                    rows.append(normalized_row)

                table_id = f"page{page_num + 1}_table_{i + 1}"
                tables.append(ExtractedTable(
                    id=table_id,
                    title=None,
                    headers=headers,
                    rows=rows,
                    metadata={"source": "pymupdf", "page": page_num + 1},
                    confidence=0.9,
                ))

        except Exception as e:
            logger.warning(f"Table extraction failed for page {page_num + 1}: {e}")

        return tables

    def _extract_excel(
        self,
        file_path: Path,
        mode: str,
    ) -> ExtractedContent:
        """Extract content from Excel file."""
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError("openpyxl is required for Excel extraction. Install with: pip install openpyxl")

        workbook = openpyxl.load_workbook(file_path, data_only=True)
        warnings = []

        tables: List[ExtractedTable] = []
        text_parts = []
        metadata = {
            "filename": file_path.name,
            "sheet_count": len(workbook.sheetnames),
            "format": "excel",
            "sheets": workbook.sheetnames,
        }

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            if sheet.max_row == 0 or sheet.max_column == 0:
                continue

            # Convert sheet to table
            headers = []
            rows = []

            for row_num, row in enumerate(sheet.iter_rows(values_only=True)):
                # Clean row
                cleaned_row = [str(cell) if cell is not None else "" for cell in row]

                if row_num == 0:
                    headers = cleaned_row
                else:
                    # Skip empty rows
                    if any(cell.strip() for cell in cleaned_row):
                        # Normalize row length
                        while len(cleaned_row) < len(headers):
                            cleaned_row.append("")
                        rows.append(cleaned_row[:len(headers)])

            if headers and rows:
                tables.append(ExtractedTable(
                    id=f"sheet_{sheet_name}",
                    title=sheet_name,
                    headers=headers,
                    rows=rows,
                    metadata={"source": "openpyxl", "sheet": sheet_name},
                    confidence=1.0,
                ))

            # Build text representation
            text_parts.append(f"=== Sheet: {sheet_name} ===")
            text_parts.append("\t".join(headers))
            for row in rows[:10]:  # First 10 rows for text preview
                text_parts.append("\t".join(row))
            if len(rows) > 10:
                text_parts.append(f"... ({len(rows) - 10} more rows)")

        workbook.close()

        return ExtractedContent(
            text="\n".join(text_parts),
            tables=tables,
            metadata=metadata,
            structure={"type": "spreadsheet", "sheets": workbook.sheetnames},
            warnings=warnings,
        )

    def _extract_csv(self, file_path: Path) -> ExtractedContent:
        """Extract content from CSV file."""
        import csv

        tables = []
        warnings = []

        with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
            # Detect delimiter
            sample = f.read(4096)
            f.seek(0)

            try:
                dialect = csv.Sniffer().sniff(sample)
            except csv.Error:
                dialect = csv.excel

            reader = csv.reader(f, dialect)
            rows_list = list(reader)

        if not rows_list:
            return ExtractedContent(
                text="",
                tables=[],
                metadata={"filename": file_path.name, "format": "csv"},
                structure={"type": "empty"},
                warnings=["CSV file is empty"],
            )

        headers = rows_list[0] if rows_list else []
        rows = rows_list[1:] if len(rows_list) > 1 else []

        # Normalize row lengths
        max_cols = max(len(row) for row in rows_list) if rows_list else 0
        while len(headers) < max_cols:
            headers.append(f"Column_{len(headers) + 1}")

        normalized_rows = []
        for row in rows:
            normalized = list(row)
            while len(normalized) < len(headers):
                normalized.append("")
            normalized_rows.append(normalized[:len(headers)])

        tables.append(ExtractedTable(
            id="csv_table",
            title=file_path.stem,
            headers=headers,
            rows=normalized_rows,
            metadata={"source": "csv"},
            confidence=1.0,
        ))

        # Text representation
        text_parts = ["\t".join(headers)]
        for row in normalized_rows[:20]:
            text_parts.append("\t".join(row))
        if len(normalized_rows) > 20:
            text_parts.append(f"... ({len(normalized_rows) - 20} more rows)")

        return ExtractedContent(
            text="\n".join(text_parts),
            tables=tables,
            metadata={
                "filename": file_path.name,
                "format": "csv",
                "row_count": len(normalized_rows),
                "column_count": len(headers),
            },
            structure={"type": "tabular"},
            warnings=warnings,
        )

    def _extract_image(
        self,
        file_path: Path,
        mode: str,
    ) -> ExtractedContent:
        """Extract content from image using VLM."""
        if not self.vlm:
            raise RuntimeError("VLM is required for image extraction but not available")

        # Use VLM for comprehensive extraction
        result = self.vlm.analyze_document(file_path, analysis_type=mode)

        tables = []
        for i, table_data in enumerate(result.tables):
            tables.append(ExtractedTable(
                id=f"image_table_{i + 1}",
                title=table_data.get("title"),
                headers=table_data.get("headers", []),
                rows=table_data.get("rows", []),
                metadata={"source": "vlm"},
                confidence=0.8,
            ))

        return ExtractedContent(
            text=result.text_content,
            tables=tables,
            metadata={
                "filename": file_path.name,
                "format": "image",
                **result.metadata,
            },
            structure=result.structure,
            warnings=[],
        )

    def _analyze_structure(self, text: str) -> Dict[str, Any]:
        """Analyze document structure from text."""
        structure = {
            "type": "document",
            "has_headers": False,
            "has_lists": False,
            "has_tables": False,
            "sections": [],
        }

        # Detect headers (lines that look like titles)
        header_pattern = re.compile(r'^[A-Z][A-Za-z\s]+:?\s*$', re.MULTILINE)
        headers = header_pattern.findall(text)
        structure["has_headers"] = len(headers) > 0
        structure["sections"] = [h.strip().rstrip(':') for h in headers[:10]]

        # Detect lists
        list_pattern = re.compile(r'^[\s]*[-•*]\s+.+$', re.MULTILINE)
        structure["has_lists"] = bool(list_pattern.search(text))

        # Detect table-like content
        table_pattern = re.compile(r'\|.+\|', re.MULTILINE)
        structure["has_tables"] = bool(table_pattern.search(text))

        return structure

    def _table_exists(
        self,
        tables: List[ExtractedTable],
        headers: List[str],
    ) -> bool:
        """Check if a table with similar headers already exists."""
        if not headers:
            return True

        for table in tables:
            if len(table.headers) == len(headers):
                # Check if headers match (case-insensitive)
                if all(
                    h1.lower().strip() == h2.lower().strip()
                    for h1, h2 in zip(table.headers, headers)
                ):
                    return True
        return False

    def infer_schema(
        self,
        table: ExtractedTable,
    ) -> List[FieldSchema]:
        """Infer schema for a table's columns."""
        schemas = []

        for col_idx, header in enumerate(table.headers):
            # Collect sample values
            sample_values = []
            for row in table.rows[:100]:
                if col_idx < len(row) and row[col_idx]:
                    sample_values.append(row[col_idx])

            # Infer data type
            data_type = self._infer_column_type(sample_values)

            schemas.append(FieldSchema(
                name=header,
                data_type=data_type,
                sample_values=sample_values[:5],
                nullable=any(not v for v in sample_values),
            ))

        return schemas

    def _infer_column_type(self, values: List[Any]) -> str:
        """Infer the data type of a column from sample values."""
        if not values:
            return "text"

        # Count type matches
        numeric_count = 0
        date_count = 0
        bool_count = 0

        date_patterns = [
            r'^\d{4}-\d{2}-\d{2}$',
            r'^\d{2}/\d{2}/\d{4}$',
            r'^\d{2}-\d{2}-\d{4}$',
            r'^\d{1,2}/\d{1,2}/\d{2,4}$',
        ]

        for value in values:
            value_str = str(value).strip()

            if not value_str:
                continue

            # Check numeric
            try:
                cleaned = re.sub(r'[$,% ]', '', value_str)
                float(cleaned)
                numeric_count += 1
                continue
            except (ValueError, TypeError):
                pass

            # Check date
            for pattern in date_patterns:
                if re.match(pattern, value_str):
                    date_count += 1
                    break

            # Check boolean
            if value_str.lower() in ('true', 'false', 'yes', 'no', '1', '0'):
                bool_count += 1

        total = len([v for v in values if str(v).strip()])
        if total == 0:
            return "text"

        # Require 70% match for type classification
        threshold = 0.7

        if date_count / total >= threshold:
            return "datetime"
        if numeric_count / total >= threshold:
            return "numeric"
        if bool_count / total >= threshold:
            return "boolean"

        return "text"

# Convenience functions

def extract_document(
    file_path: Union[str, Path],
    use_vlm: bool = True,
) -> ExtractedContent:
    """Quick function to extract content from a document."""
    extractor = EnhancedDocumentExtractor(use_vlm=use_vlm)
    return extractor.extract(file_path)

def extract_tables(
    file_path: Union[str, Path],
) -> List[ExtractedTable]:
    """Quick function to extract tables from a document."""
    extractor = EnhancedDocumentExtractor(use_vlm=True)
    result = extractor.extract(file_path, extraction_mode="tables_only")
    return result.tables

# ── Vision Language Model ──
def _lazy_get_vlm_client():
    from backend.app.services.llm import get_llm_client
    return get_llm_client()

logger = logging.getLogger("neura.llm.vision")

@dataclass
class DocumentAnalysisResult:
    """Result of VLM document analysis."""
    text_content: str
    tables: List[Dict[str, Any]]
    structure: Dict[str, Any]
    metadata: Dict[str, Any]
    raw_response: str

@dataclass
class TableExtractionResult:
    """Result of table extraction from image."""
    tables: List[Dict[str, Any]]
    confidence: float
    raw_response: str

class VisionLanguageModel:
    """
    Vision-Language Model service for document understanding.

    Provides high-level methods for:
    - Document OCR with structure understanding
    - Table extraction from images
    - Form field extraction
    - Chart/graph analysis
    """

    def __init__(
        self,
        client: Optional[LLMClient] = None,
        model: Optional[str] = None,
    ):
        self.client = client or _lazy_get_vlm_client()
        self._model = model

    @property
    def model(self) -> str:
        """Get the vision model to use."""
        if self._model:
            return self._model
        return self.client.config.get_vision_model()

    def analyze_document(
        self,
        image: Union[str, bytes, Path],
        analysis_type: str = "comprehensive",
        language: str = "auto",
    ) -> DocumentAnalysisResult:
        """Analyze a document image using VLM."""
        prompt = self._build_document_analysis_prompt(analysis_type, language)

        response = self.client.complete_with_vision(
            text=prompt,
            images=[image],
            model=self.model,
            description=f"vlm_document_analysis_{analysis_type}",
        )

        raw_content = response["choices"][0]["message"]["content"]
        return self._parse_document_analysis(raw_content, analysis_type)

    def extract_tables(
        self,
        image: Union[str, bytes, Path],
        expected_columns: Optional[List[str]] = None,
    ) -> TableExtractionResult:
        """Extract tables from a document image."""
        prompt = self._build_table_extraction_prompt(expected_columns)

        response = self.client.complete_with_vision(
            text=prompt,
            images=[image],
            model=self.model,
            description="vlm_table_extraction",
        )

        raw_content = response["choices"][0]["message"]["content"]
        return self._parse_table_extraction(raw_content)

    def extract_text_with_layout(
        self,
        image: Union[str, bytes, Path],
        preserve_formatting: bool = True,
    ) -> Dict[str, Any]:
        """Extract text from image while preserving layout structure."""
        prompt = f"""Analyze this document image and extract all text content.

{"Preserve the original formatting, spacing, and layout as much as possible." if preserve_formatting else "Extract the text content in reading order."}

Return your response in the following JSON format:
```json
{{
  "title": "Document title if present",
  "sections": [
    {{
      "type": "header|paragraph|list|table|footer",
      "level": 1,
      "content": "Text content",
      "formatting": {{
        "bold": false,
        "italic": false,
        "alignment": "left|center|right"
      }}
    }}
  ],
  "page_number": null,
  "reading_order_text": "Full text in reading order"
}}
```

Analyze the image carefully and extract all visible text."""

        response = self.client.complete_with_vision(
            text=prompt,
            images=[image],
            model=self.model,
            description="vlm_text_extraction",
        )

        raw_content = response["choices"][0]["message"]["content"]
        return self._parse_json_response(raw_content, {
            "title": None,
            "sections": [],
            "page_number": None,
            "reading_order_text": raw_content,
        })

    def analyze_chart(
        self,
        image: Union[str, bytes, Path],
    ) -> Dict[str, Any]:
        """Analyze a chart/graph image and extract data."""
        prompt = """Analyze this chart/graph image and extract the data.

Return your response in the following JSON format:
```json
{
  "chart_type": "bar|line|pie|scatter|area|other",
  "title": "Chart title if visible",
  "x_axis": {
    "label": "X axis label",
    "values": ["value1", "value2"]
  },
  "y_axis": {
    "label": "Y axis label",
    "min": 0,
    "max": 100
  },
  "data_series": [
    {
      "name": "Series name",
      "values": [10, 20, 30],
      "color": "blue"
    }
  ],
  "legend": ["Item 1", "Item 2"],
  "insights": "Brief description of what the chart shows"
}
```

Extract as much data as you can accurately determine from the image."""

        response = self.client.complete_with_vision(
            text=prompt,
            images=[image],
            model=self.model,
            description="vlm_chart_analysis",
        )

        raw_content = response["choices"][0]["message"]["content"]
        return self._parse_json_response(raw_content, {
            "chart_type": "unknown",
            "data_series": [],
            "insights": raw_content,
        })

    def extract_form_fields(
        self,
        image: Union[str, bytes, Path],
    ) -> Dict[str, Any]:
        """Extract form fields and their values from an image."""
        prompt = """Analyze this form image and extract all form fields with their values.

Return your response in the following JSON format:
```json
{
  "form_title": "Form title if visible",
  "fields": [
    {
      "label": "Field label",
      "value": "Filled value or null if empty",
      "type": "text|checkbox|radio|date|signature|other",
      "required": true
    }
  ],
  "sections": [
    {
      "name": "Section name",
      "fields": ["field_label_1", "field_label_2"]
    }
  ]
}
```

Extract all visible form fields, whether filled or empty."""

        response = self.client.complete_with_vision(
            text=prompt,
            images=[image],
            model=self.model,
            description="vlm_form_extraction",
        )

        raw_content = response["choices"][0]["message"]["content"]
        return self._parse_json_response(raw_content, {
            "form_title": None,
            "fields": [],
        })

    def compare_documents(
        self,
        image1: Union[str, bytes, Path],
        image2: Union[str, bytes, Path],
    ) -> Dict[str, Any]:
        """Compare two document images and identify differences."""
        prompt = """Compare these two document images and identify any differences.

The first image is the reference, and the second is the version to compare.

Return your response in the following JSON format:
```json
{
  "identical": false,
  "similarity_score": 0.95,
  "differences": [
    {
      "type": "text_change|layout_change|missing_element|added_element",
      "location": "Description of where the difference is",
      "reference_content": "Content in first image",
      "compared_content": "Content in second image"
    }
  ],
  "summary": "Brief summary of the comparison"
}
```

Be thorough but focus on meaningful differences, not minor formatting variations."""

        response = self.client.complete_with_vision(
            text=prompt,
            images=[image1, image2],
            model=self.model,
            description="vlm_document_comparison",
        )

        raw_content = response["choices"][0]["message"]["content"]
        return self._parse_json_response(raw_content, {
            "identical": False,
            "differences": [],
            "summary": raw_content,
        })

    def _build_document_analysis_prompt(
        self,
        analysis_type: str,
        language: str,
    ) -> str:
        """Build the prompt for document analysis."""
        base_prompt = """Analyze this document image comprehensively.

"""
        if analysis_type == "text_only":
            base_prompt += """Focus on extracting all text content accurately.

Return your response in JSON format:
```json
{
  "text_content": "Full extracted text",
  "paragraphs": ["paragraph 1", "paragraph 2"],
  "headers": ["header 1"],
  "language": "detected language"
}
```"""
        elif analysis_type == "tables_only":
            base_prompt += """Focus on extracting any tables present in the document.

Return your response in JSON format:
```json
{
  "tables": [
    {
      "id": 1,
      "title": "Table title if present",
      "headers": ["Column 1", "Column 2"],
      "rows": [
        ["Value 1", "Value 2"],
        ["Value 3", "Value 4"]
      ]
    }
  ]
}
```"""
        elif analysis_type == "structure":
            base_prompt += """Focus on understanding the document structure and layout.

Return your response in JSON format:
```json
{
  "document_type": "invoice|report|form|letter|other",
  "sections": [
    {"name": "Section name", "type": "header|body|footer|sidebar"}
  ],
  "has_tables": true,
  "has_images": false,
  "layout": "single_column|multi_column|mixed"
}
```"""
        else:  # comprehensive
            base_prompt += """Extract all content including text, tables, and structure.

Return your response in JSON format:
```json
{
  "document_type": "invoice|report|form|letter|other",
  "title": "Document title if present",
  "text_content": "Full text content in reading order",
  "tables": [
    {
      "id": 1,
      "title": "Table title",
      "headers": ["Col1", "Col2"],
      "rows": [["Val1", "Val2"]]
    }
  ],
  "structure": {
    "sections": ["Section 1", "Section 2"],
    "has_headers": true,
    "has_footers": true
  },
  "metadata": {
    "language": "en",
    "date_found": "2024-01-01",
    "page_number": 1
  }
}
```"""

        if language != "auto":
            base_prompt += f"\n\nThe document is in {language}."

        return base_prompt

    def _build_table_extraction_prompt(
        self,
        expected_columns: Optional[List[str]],
    ) -> str:
        """Build the prompt for table extraction."""
        prompt = """Extract all tables from this document image.

Return your response in JSON format:
```json
{
  "tables": [
    {
      "id": 1,
      "title": "Table title if visible",
      "headers": ["Column 1", "Column 2", "Column 3"],
      "rows": [
        ["Row 1 Col 1", "Row 1 Col 2", "Row 1 Col 3"],
        ["Row 2 Col 1", "Row 2 Col 2", "Row 2 Col 3"]
      ],
      "notes": "Any footnotes or notes about the table"
    }
  ],
  "confidence": 0.95
}
```

Be accurate with the data extraction. If a cell is empty, use an empty string.
If you cannot read a value clearly, indicate it with "[unclear]".
"""

        if expected_columns:
            prompt += f"\n\nExpected columns: {', '.join(expected_columns)}"
            prompt += "\nMap extracted columns to these expected names if they match."

        return prompt

    def _parse_document_analysis(
        self,
        raw_content: str,
        analysis_type: str,
    ) -> DocumentAnalysisResult:
        """Parse the document analysis response."""
        parsed = self._parse_json_response(raw_content, {})

        text_content = parsed.get("text_content", "")
        if not text_content and "paragraphs" in parsed:
            text_content = "\n\n".join(parsed.get("paragraphs", []))

        tables = parsed.get("tables", [])
        structure = parsed.get("structure", {})

        if "document_type" in parsed:
            structure["document_type"] = parsed["document_type"]
        if "title" in parsed:
            structure["title"] = parsed["title"]

        metadata = parsed.get("metadata", {})
        if "language" in parsed:
            metadata["language"] = parsed["language"]

        return DocumentAnalysisResult(
            text_content=text_content or raw_content,
            tables=tables,
            structure=structure,
            metadata=metadata,
            raw_response=raw_content,
        )

    def _parse_table_extraction(
        self,
        raw_content: str,
    ) -> TableExtractionResult:
        """Parse the table extraction response."""
        parsed = self._parse_json_response(raw_content, {"tables": [], "confidence": 0.5})

        return TableExtractionResult(
            tables=parsed.get("tables", []),
            confidence=parsed.get("confidence", 0.5),
            raw_response=raw_content,
        )

    def _parse_json_response(self, raw_content: str, default: Dict[str, Any]) -> Dict[str, Any]:
        return _parse_json_from_llm(raw_content, default)

# Convenience functions

def get_vlm(model: Optional[str] = None) -> VisionLanguageModel:
    """Get a VisionLanguageModel instance."""
    return VisionLanguageModel(model=model)

def analyze_document_image(
    image: Union[str, bytes, Path],
    analysis_type: str = "comprehensive",
) -> DocumentAnalysisResult:
    """Quick function to analyze a document image."""
    vlm = get_vlm()
    return vlm.analyze_document(image, analysis_type)

def extract_tables_from_image(
    image: Union[str, bytes, Path],
) -> TableExtractionResult:
    """Quick function to extract tables from an image."""
    vlm = get_vlm()
    return vlm.extract_tables(image)

# ── Client ──
from collections import deque
from datetime import datetime, timezone

from pydantic import BaseModel

def _append_raw_output(*args, **kwargs):
    """Lazy import to break circular dependency with infra_services."""
    from backend.app.services.infra_services import append_raw_llm_output
    return append_raw_llm_output(*args, **kwargs)

logger = logging.getLogger("neura.llm.client")

_MAX_LOG_PREVIEW = 2000

# ── Error Classification ──

class LLMErrorCategory(str, Enum):
    """Structured error categories for LLM failures."""
    RATE_LIMIT = "rate_limit"
    CONTEXT_OVERFLOW = "context_overflow"
    CONTENT_FILTER = "content_filter"
    MODEL_OVERLOAD = "model_overload"
    TIMEOUT = "timeout"
    PARSE_FAILURE = "parse_failure"
    INVALID_REQUEST = "invalid_request"
    AUTH_FAILURE = "auth_failure"
    PROVIDER_UNAVAILABLE = "provider_unavailable"
    UNKNOWN = "unknown"

_ERROR_PATTERNS: Dict[str, tuple] = {
    LLMErrorCategory.RATE_LIMIT: ("rate_limit", "ratelimit", "429", "too many requests"),
    LLMErrorCategory.CONTEXT_OVERFLOW: ("context_length_exceeded", "maximum context length", "token limit", "too many tokens", "context window", "input too long"),
    LLMErrorCategory.CONTENT_FILTER: ("content_filter", "content_policy", "safety", "blocked"),
    LLMErrorCategory.MODEL_OVERLOAD: ("overloaded", "capacity", "503", "service unavailable"),
    LLMErrorCategory.TIMEOUT: ("timeout", "timed out", "504"),
    LLMErrorCategory.PARSE_FAILURE: ("json", "parse", "decode", "invalid json", "malformed"),
    LLMErrorCategory.INVALID_REQUEST: ("invalid_request", "bad request", "400"),
    LLMErrorCategory.AUTH_FAILURE: ("auth", "unauthorized", "forbidden", "api_key", "401", "403", "insufficient_quota", "quota"),
    LLMErrorCategory.PROVIDER_UNAVAILABLE: ("unavailable", "connection", "refused", "unreachable"),
}

def classify_error(exc: BaseException) -> LLMErrorCategory:
    """Classify an LLM exception into a structured category."""
    detail = str(exc).lower()
    for category, patterns in _ERROR_PATTERNS.items():
        if any(p in detail for p in patterns):
            return category
    return LLMErrorCategory.UNKNOWN

# ── Adaptive Timeouts (operation prefix → seconds) ──
OPERATION_TIMEOUTS: Dict[str, float] = {
    "template-verify": 480.0,
    "template-initial": 480.0,
    "template-fix": 480.0,
    "template-initial-html": 480.0,
    "mapping-preview": 360.0,
    "mapping-approve": 360.0,
    "mapping-inline": 360.0,
    "llm-call-3": 360.0,
    "contract-build": 480.0,
    "edit-template": 240.0,
    "v3-df": 480.0,
    "report-generate": 300.0,
    "agent-research": 240.0,
    "agent-analysis": 180.0,
    "agent-email": 60.0,
    "agent-content": 120.0,
    "agent-proofread": 60.0,
    "agent-report": 180.0,
    "nl2sql": 30.0,
    "docqa": 60.0,
    "chart-suggest": 30.0,
    "vision": 120.0,
    "template-initial-vision": 240.0,
    "vision-text-extraction": 180.0,
    "vlm-text-extraction": 180.0,
    "vlm-document-analysis": 240.0,
    "assistant-chat": 30.0,
}

def get_adaptive_timeout(description: str, default: float = 120.0) -> float:
    """Get timeout for an operation based on its description prefix."""
    desc_lower = description.lower().replace("_", "-")
    for prefix, timeout in OPERATION_TIMEOUTS.items():
        if desc_lower.startswith(prefix):
            return timeout
    return default

# ── Call History ──

@dataclass
class CallRecord:
    """Single LLM call record for debugging and observability."""
    timestamp: float
    description: str
    model: str
    success: bool
    latency_ms: float
    input_tokens: int
    output_tokens: int
    error_category: Optional[str] = None
    error_message: Optional[str] = None
    prompt_preview: str = ""
    response_preview: str = ""
    cached: bool = False
    circuit_name: Optional[str] = None

    def to_dict(self) -> Dict[str, Any]:
        return {
            "timestamp": self.timestamp,
            "time": datetime.fromtimestamp(self.timestamp, tz=timezone.utc).isoformat(),
            "description": self.description,
            "model": self.model,
            "success": self.success,
            "latency_ms": round(self.latency_ms, 2),
            "input_tokens": self.input_tokens,
            "output_tokens": self.output_tokens,
            "error_category": self.error_category,
            "error_message": self.error_message,
            "prompt_preview": self.prompt_preview[:200],
            "response_preview": self.response_preview[:200],
            "cached": self.cached,
            "circuit_name": self.circuit_name,
        }

def _summarize_messages(messages: List[Dict[str, Any]]) -> str:
    """Build a concise summary of messages for the LLM log."""
    parts: list[str] = []
    for msg in messages:
        role = msg.get("role", "?")
        content = msg.get("content", "")
        if isinstance(content, list):
            # Multimodal: extract text parts
            text_parts = [p.get("text", "") for p in content if isinstance(p, dict) and p.get("type") == "text"]
            content = " ".join(text_parts) or f"[{len(content)} content blocks]"
        text = str(content)
        if len(text) > _MAX_LOG_PREVIEW:
            text = text[:_MAX_LOG_PREVIEW] + f"... ({len(text)} chars total)"
        parts.append(f"  [{role}] {text}")
    return "\n".join(parts)

def _extract_response_text(response: Dict[str, Any]) -> str:
    """Extract the assistant text from an OpenAI-compatible response dict."""
    try:
        choices = response.get("choices") or []
        if choices:
            return choices[0].get("message", {}).get("content", "") or ""
    except (IndexError, KeyError, TypeError, AttributeError):
        pass
    return ""

# ── Circuit Breaker ──

class CircuitState(Enum):
    """Circuit breaker states."""
    CLOSED = "closed"  # Normal operation
    OPEN = "open"  # Failing, reject requests
    HALF_OPEN = "half_open"  # Testing if service recovered

@dataclass
class CircuitBreakerConfig:
    """Configuration for circuit breaker."""
    failure_threshold: int = 5  # Failures before opening
    success_threshold: int = 2  # Successes in half-open before closing
    timeout_seconds: float = 60.0  # Time before moving from open to half-open
    failure_window_seconds: float = 120.0  # Window to count failures

class CircuitBreaker:
    """
    Circuit breaker for fault tolerance.

    Prevents cascading failures by stopping requests to failing services.
    Based on the pattern from resilience4j and Netflix Hystrix.
    """

    def __init__(self, name: str, config: Optional[CircuitBreakerConfig] = None):
        self.name = name
        self.config = config or CircuitBreakerConfig()
        self._state = CircuitState.CLOSED
        self._failure_count = 0
        self._success_count = 0
        self._last_failure_time: Optional[float] = None
        self._state_changed_at = time.time()
        self._failure_timestamps: deque = deque()
        self._lock = threading.Lock()

    @property
    def state(self) -> CircuitState:
        """Get current circuit state, potentially transitioning from OPEN to HALF_OPEN."""
        with self._lock:
            if self._state == CircuitState.OPEN:
                if time.time() - self._state_changed_at >= self.config.timeout_seconds:
                    self._transition_to(CircuitState.HALF_OPEN)
            return self._state

    def _transition_to(self, new_state: CircuitState) -> None:
        """Transition to a new state."""
        old_state = self._state
        self._state = new_state
        self._state_changed_at = time.time()

        if new_state == CircuitState.CLOSED:
            self._failure_count = 0
            self._success_count = 0
            self._failure_timestamps.clear()
        elif new_state == CircuitState.HALF_OPEN:
            self._success_count = 0

        logger.info(
            "circuit_breaker_state_change",
            extra={
                "event": "circuit_breaker_state_change",
                "breaker_name": self.name,
                "old_state": old_state.value,
                "new_state": new_state.value,
            }
        )

    def allow_request(self) -> bool:
        """Check if request should be allowed."""
        current_state = self.state  # This may trigger OPEN -> HALF_OPEN

        if current_state == CircuitState.CLOSED:
            return True
        elif current_state == CircuitState.OPEN:
            return False
        else:  # HALF_OPEN
            return True  # Allow test requests

    def record_success(self) -> None:
        """Record a successful request."""
        with self._lock:
            if self._state == CircuitState.HALF_OPEN:
                self._success_count += 1
                if self._success_count >= self.config.success_threshold:
                    self._transition_to(CircuitState.CLOSED)
            elif self._state == CircuitState.CLOSED:
                # Clean old failures from window
                self._clean_old_failures()

    def record_failure(self) -> None:
        """Record a failed request."""
        with self._lock:
            now = time.time()
            self._last_failure_time = now

            if self._state == CircuitState.HALF_OPEN:
                self._transition_to(CircuitState.OPEN)
            elif self._state == CircuitState.CLOSED:
                self._failure_timestamps.append(now)
                self._clean_old_failures()

                if len(self._failure_timestamps) >= self.config.failure_threshold:
                    self._transition_to(CircuitState.OPEN)

    def _clean_old_failures(self) -> None:
        """Remove failures outside the failure window."""
        cutoff = time.time() - self.config.failure_window_seconds
        while self._failure_timestamps and self._failure_timestamps[0] < cutoff:
            self._failure_timestamps.popleft()

    def get_stats(self) -> Dict[str, Any]:
        """Get circuit breaker statistics."""
        with self._lock:
            return {
                "breaker_name": self.name,
                "state": self._state.value,
                "failure_count": len(self._failure_timestamps),
                "success_count": self._success_count,
                "last_failure": self._last_failure_time,
                "state_changed_at": self._state_changed_at,
            }

class CircuitBreakerRegistry:
    """
    Registry of per-operation circuit breakers.

    Each operation (e.g., 'template-verify', 'agent-research') gets its own
    circuit breaker so that failures in one operation don't block others.
    Inspired by BFI pipeline_v45's per-stage resilience pattern.
    """

    def __init__(self, default_config: Optional[CircuitBreakerConfig] = None):
        self._breakers: Dict[str, CircuitBreaker] = {}
        self._lock = threading.Lock()
        self._default_config = default_config or CircuitBreakerConfig()

    def get(self, operation: str) -> CircuitBreaker:
        """Get or create a circuit breaker for the given operation."""
        with self._lock:
            if operation not in self._breakers:
                self._breakers[operation] = CircuitBreaker(
                    name=f"llm_{operation}",
                    config=self._default_config,
                )
            return self._breakers[operation]

    def get_all_stats(self) -> Dict[str, Dict[str, Any]]:
        """Get stats for all registered circuit breakers."""
        with self._lock:
            return {name: cb.get_stats() for name, cb in self._breakers.items()}

    def reset(self, operation: Optional[str] = None) -> None:
        """Reset one or all circuit breakers."""
        with self._lock:
            if operation and operation in self._breakers:
                del self._breakers[operation]
            elif operation is None:
                self._breakers.clear()

# Global circuit breaker registry
_circuit_registry = CircuitBreakerRegistry()

# ── Response Cache ──

@dataclass
class CacheEntry:
    """Cache entry with metadata."""
    response: Dict[str, Any]
    created_at: float
    expires_at: float
    hit_count: int = 0
    request_hash: str = ""

class ResponseCache:
    """
    LRU cache for LLM responses with disk persistence.

    Features:
    - Memory cache with LRU eviction
    - Optional disk persistence for long-term caching
    - TTL-based expiration
    - Cache key based on request content hash
    """

    def __init__(
        self,
        max_memory_items: int = 100,
        max_disk_items: int = 1000,
        default_ttl_seconds: float = 3600.0,
        cache_dir: Optional[Path] = None,
    ):
        self.max_memory_items = max_memory_items
        self.max_disk_items = max_disk_items
        self.default_ttl_seconds = default_ttl_seconds
        self.cache_dir = cache_dir

        self._memory_cache: Dict[str, CacheEntry] = {}
        self._access_order: deque = deque()
        self._lock = threading.Lock()
        self._stats = {"hits": 0, "misses": 0, "evictions": 0}

        if cache_dir:
            cache_dir.mkdir(parents=True, exist_ok=True)

    def _compute_key(
        self,
        messages: List[Dict[str, Any]],
        model: str,
        **kwargs: Any,
    ) -> str:
        """Compute cache key from request parameters."""
        # Create deterministic hash of request
        key_data = {
            "messages": messages,
            "model": model,
            "kwargs": {k: v for k, v in sorted(kwargs.items()) if k not in ("stream",)},
        }
        key_json = json.dumps(key_data, sort_keys=True, default=str)
        return hashlib.sha256(key_json.encode()).hexdigest()[:32]

    def get(
        self,
        messages: List[Dict[str, Any]],
        model: str,
        **kwargs: Any,
    ) -> Optional[Dict[str, Any]]:
        """Get cached response if available."""
        key = self._compute_key(messages, model, **kwargs)

        with self._lock:
            # Check memory cache
            entry = self._memory_cache.get(key)
            if entry:
                if time.time() < entry.expires_at:
                    entry.hit_count += 1
                    self._stats["hits"] += 1
                    # Move to end of access order (most recent)
                    if key in self._access_order:
                        self._access_order.remove(key)
                    self._access_order.append(key)
                    return entry.response
                else:
                    # Expired
                    del self._memory_cache[key]
                    if key in self._access_order:
                        self._access_order.remove(key)

            # Check disk cache
            if self.cache_dir:
                disk_response = self._read_from_disk(key)
                if disk_response:
                    self._stats["hits"] += 1
                    # Promote to memory cache
                    self._set_memory(key, disk_response)
                    return disk_response

            self._stats["misses"] += 1
            return None

    def set(
        self,
        messages: List[Dict[str, Any]],
        model: str,
        response: Dict[str, Any],
        ttl_seconds: Optional[float] = None,
        **kwargs: Any,
    ) -> None:
        """Cache a response."""
        key = self._compute_key(messages, model, **kwargs)
        ttl = ttl_seconds or self.default_ttl_seconds

        with self._lock:
            self._set_memory(key, response, ttl)

            # Also write to disk for persistence
            if self.cache_dir:
                self._write_to_disk(key, response, ttl)

    def _set_memory(
        self,
        key: str,
        response: Dict[str, Any],
        ttl_seconds: Optional[float] = None,
    ) -> None:
        """Set entry in memory cache."""
        ttl = ttl_seconds or self.default_ttl_seconds
        now = time.time()

        # Evict if at capacity
        while len(self._memory_cache) >= self.max_memory_items and self._access_order:
            oldest_key = self._access_order.popleft()
            if oldest_key in self._memory_cache:
                del self._memory_cache[oldest_key]
                self._stats["evictions"] += 1

        self._memory_cache[key] = CacheEntry(
            response=response,
            created_at=now,
            expires_at=now + ttl,
            request_hash=key,
        )
        self._access_order.append(key)

    def _read_from_disk(self, key: str) -> Optional[Dict[str, Any]]:
        """Read cached response from disk."""
        if not self.cache_dir:
            return None

        cache_file = self.cache_dir / f"{key}.json"
        if not cache_file.exists():
            return None

        try:
            data = json.loads(cache_file.read_text(encoding="utf-8"))
            if time.time() < data.get("expires_at", 0):
                return data.get("response")
            else:
                # Expired, delete file
                cache_file.unlink(missing_ok=True)
                return None
        except Exception:
            return None

    def _write_to_disk(
        self,
        key: str,
        response: Dict[str, Any],
        ttl_seconds: float,
    ) -> None:
        """Write cached response to disk."""
        if not self.cache_dir:
            return

        cache_file = self.cache_dir / f"{key}.json"
        try:
            data = {
                "response": response,
                "created_at": time.time(),
                "expires_at": time.time() + ttl_seconds,
            }
            cache_file.write_text(json.dumps(data), encoding="utf-8")
        except Exception as e:
            logger.debug(f"Failed to write cache to disk: {e}")

    def clear(self) -> None:
        """Clear all cached entries."""
        with self._lock:
            self._memory_cache.clear()
            self._access_order.clear()
            self._stats = {"hits": 0, "misses": 0, "evictions": 0}

            if self.cache_dir:
                for f in self.cache_dir.glob("*.json"):
                    try:
                        f.unlink()
                    except Exception:
                        pass

    def get_stats(self) -> Dict[str, Any]:
        """Get cache statistics."""
        with self._lock:
            total = self._stats["hits"] + self._stats["misses"]
            hit_rate = self._stats["hits"] / total if total > 0 else 0
            return {
                **self._stats,
                "hit_rate": hit_rate,
                "memory_size": len(self._memory_cache),
            }

# ── Token Counter / Cost Estimator ──
TOKEN_COSTS: Dict[str, Dict[str, float]] = {
    "qwen": {"input": 0.0, "output": 0.0},
}

def estimate_tokens(text: str) -> int:
    """
    Estimate token count for text.

    Uses a simple heuristic: ~4 characters per token for English text.
    For more accurate counting, use tiktoken library.
    """
    try:
        import tiktoken
        enc = tiktoken.get_encoding("cl100k_base")
        return len(enc.encode(text))
    except ImportError:
        # Fallback to heuristic
        return max(1, len(text) // 4)

def estimate_cost(
    model: str,
    input_tokens: int,
    output_tokens: int,
) -> float:
    """Estimate cost for a completion."""
    costs = TOKEN_COSTS.get(model, TOKEN_COSTS.get("qwen", {"input": 0.0, "output": 0.0}))
    input_cost = (input_tokens / 1000) * costs["input"]
    output_cost = (output_tokens / 1000) * costs["output"]
    return input_cost + output_cost

@dataclass
class UsageTracker:
    """Track token usage and costs."""
    total_input_tokens: int = 0
    total_output_tokens: int = 0
    total_cost: float = 0.0
    request_count: int = 0
    _lock: threading.Lock = field(default_factory=threading.Lock)

    def record(self, model: str, input_tokens: int, output_tokens: int) -> None:
        """Record token usage from a request."""
        with self._lock:
            self.total_input_tokens += input_tokens
            self.total_output_tokens += output_tokens
            self.total_cost += estimate_cost(model, input_tokens, output_tokens)
            self.request_count += 1

    def get_stats(self) -> Dict[str, Any]:
        """Get usage statistics."""
        with self._lock:
            return {
                "total_input_tokens": self.total_input_tokens,
                "total_output_tokens": self.total_output_tokens,
                "total_tokens": self.total_input_tokens + self.total_output_tokens,
                "estimated_cost_usd": round(self.total_cost, 4),
                "request_count": self.request_count,
            }

    def reset(self) -> None:
        """Reset usage statistics."""
        with self._lock:
            self.total_input_tokens = 0
            self.total_output_tokens = 0
            self.total_cost = 0.0
            self.request_count = 0

# Global usage tracker
_usage_tracker = UsageTracker()

# Raw output logging is handled by backend.app.services.utils.llm

class LLMClient:
    """
    Unified LLM client supporting multiple providers.

    Features:
    - Per-operation circuit breakers for fault isolation
    - Response caching (memory and disk)
    - Token usage tracking + per-operation cost tracking
    - Call history recording for debugging and observability
    - Adaptive timeouts per operation category
    - Structured error categorization
    - Automatic retry with exponential backoff
    - Fallback to secondary provider

    Usage:
        client = LLMClient()
        response = client.complete(
            messages=[{"role": "user", "content": "Hello"}],
            description="template-verify"
        )

        # Access call history
        for record in client.call_history:
            print(record.to_dict())

        # Access per-operation costs
        print(client.cost_tracker.get_stats())

        # Check circuit breaker states
        print(client.circuit_registry.get_all_stats())
    """

    def __init__(
        self,
        config: Optional[LLMConfig] = None,
        provider: Optional[BaseProvider] = None,
        enable_cache: bool = True,
        enable_circuit_breaker: bool = True,
        cache_dir: Optional[Path] = None,
    ):
        self.config = config or get_llm_config()
        self._provider = provider or get_provider(self.config)
        self._fallback_provider: Optional[BaseProvider] = None

        # Per-operation circuit breaker registry (V2 enhancement)
        self._use_per_op_circuits = enable_circuit_breaker
        self._circuit_registry_ref = _circuit_registry

        # Legacy single circuit breaker (kept for backwards compat)
        self._circuit_breaker: Optional[CircuitBreaker] = None
        if enable_circuit_breaker:
            self._circuit_breaker = CircuitBreaker(
                name=f"llm_{self.config.provider.value}",
                config=CircuitBreakerConfig(
                    failure_threshold=self.config.max_retries + 2,
                    timeout_seconds=60.0,
                )
            )

        # Call history ring buffer (V2 enhancement)
        self._call_history: deque[CallRecord] = deque(maxlen=500)
        self._call_history_lock = threading.Lock()

        # Per-operation cost tracker (V2 enhancement)
        from backend.app.services.llm import get_cost_tracker
        self._cost_tracker = get_cost_tracker()

        self._cache: Optional[ResponseCache] = None
        if isinstance(self._provider, (LiteLLMProvider,)):
            enable_cache = os.getenv("LLM_CACHE_ENABLED", "false").lower() in {"1", "true", "yes"}

        if enable_cache:
            default_cache_dir = cache_dir
            if not default_cache_dir and os.getenv("LLM_CACHE_DIR"):
                _raw_dir = Path(os.getenv("LLM_CACHE_DIR", ""))
                # Reject paths with traversal components
                if ".." not in str(_raw_dir):
                    default_cache_dir = _raw_dir
                else:
                    logger.warning("llm_cache_dir_rejected", extra={"reason": "path contains '..'"})
            try:
                max_items = int(os.getenv("LLM_CACHE_MAX_ITEMS", "100"))
            except (ValueError, TypeError):
                max_items = 100
            try:
                ttl = float(os.getenv("LLM_CACHE_TTL_SECONDS", "3600"))
            except (ValueError, TypeError):
                ttl = 3600.0
            _max_items = min(max(1, max_items), 10000)
            self._cache = ResponseCache(
                max_memory_items=_max_items,
                default_ttl_seconds=ttl,
                cache_dir=default_cache_dir,
            )

        self._usage_tracker = UsageTracker()

        # Fallback provider (not used for Claude Code CLI - single provider only)
        self._fallback_provider = None

    @property
    def provider(self) -> BaseProvider:
        """Get the current provider."""
        return self._provider

    @property
    def model(self) -> str:
        """Get the current model name."""
        return self.config.model

    @property
    def call_history(self) -> List[CallRecord]:
        """Get the call history as a list (most recent last)."""
        with self._call_history_lock:
            return list(self._call_history)

    @property
    def cost_tracker(self):
        """Get the per-operation cost tracker."""
        return self._cost_tracker

    @property
    def circuit_registry(self) -> CircuitBreakerRegistry:
        """Get the per-operation circuit breaker registry."""
        return self._circuit_registry_ref

    def _record_call(self, record: CallRecord) -> None:
        """Add a call record to history."""
        with self._call_history_lock:
            self._call_history.append(record)

    def _get_op_circuit(self, description: str) -> Optional[CircuitBreaker]:
        """Get the per-operation circuit breaker for a description."""
        if not self._use_per_op_circuits:
            return None
        # Derive operation key from description prefix
        op_key = description.lower().replace("_", "-").split(".")[0]
        # Only use per-op circuits for known operations
        for prefix in OPERATION_TIMEOUTS:
            if op_key.startswith(prefix):
                return self._circuit_registry_ref.get(prefix)
        return None

    def get_call_history_stats(self) -> Dict[str, Any]:
        """Get summary statistics from call history."""
        with self._call_history_lock:
            records = list(self._call_history)
        if not records:
            return {"total_calls": 0}

        successes = [r for r in records if r.success]
        failures = [r for r in records if not r.success]
        cached = [r for r in records if r.cached]

        # Error category breakdown
        error_cats: Dict[str, int] = {}
        for r in failures:
            cat = r.error_category or "unknown"
            error_cats[cat] = error_cats.get(cat, 0) + 1

        return {
            "total_calls": len(records),
            "successes": len(successes),
            "failures": len(failures),
            "cached_hits": len(cached),
            "success_rate": len(successes) / len(records) if records else 0,
            "avg_latency_ms": sum(r.latency_ms for r in successes) / len(successes) if successes else 0,
            "error_categories": error_cats,
            "recent_calls": [r.to_dict() for r in records[-10:]],
        }

    def _agent_retry_strategy(
        self,
        error: Exception,
        attempt: int,
        operation: str,
        error_cat: "LLMErrorCategory",
        current_delay: float,
    ) -> dict:
        """Agent decides retry strategy based on error analysis.

        Called for retryable errors that don't match known non-retryable
        categories.  Returns {"action": ..., "delay_seconds": ...}.
        Falls back to default strategy on any failure.
        """
        try:
            # Avoid recursion: don't use the agent to fix agent retry calls
            if operation in ("agent_retry_strategy", "json_self_repair"):
                return {"action": "retry", "delay_seconds": current_delay}

            prompt = (
                f"LLM call '{operation}' failed on attempt {attempt}:\n"
                f"Error type: {type(error).__name__}\n"
                f"Error category: {error_cat.value}\n"
                f"Error: {str(error)[:500]}\n\n"
                "Should we:\n"
                "- retry: try again with a delay\n"
                "- abort: stop retrying, the error is permanent\n\n"
                "Also suggest an appropriate delay in seconds (1-60).\n"
                'Return ONLY JSON: {"action": "retry" or "abort", '
                '"delay_seconds": N, "reason": "..."}'
            )

            resp = self._provider.chat_completion(
                messages=[{"role": "user", "content": prompt}],
                model=self.config.model,
                max_tokens=128,
            )

            text = _extract_response_text(resp)
            import json as _json
            result = _json.loads(text.strip())

            action = str(result.get("action", "retry")).lower()
            delay = float(result.get("delay_seconds", current_delay))
            delay = max(1.0, min(delay, 60.0))  # Hard bounds

            logger.info(
                "agent_retry_strategy",
                extra={
                    "event": "agent_retry_strategy",
                    "operation": operation,
                    "action": action,
                    "delay": delay,
                    "reason": str(result.get("reason", ""))[:200],
                },
            )
            return {"action": action, "delay_seconds": delay}
        except Exception:
            logger.debug("agent_retry_strategy_failed", exc_info=True)
            return {"action": "retry", "delay_seconds": current_delay}

    def complete(
        self,
        messages: List[Dict[str, Any]],
        model: Optional[str] = None,
        description: str = "llm_call",
        use_cache: bool = True,
        cache_ttl: Optional[float] = None,
        **kwargs: Any,
    ) -> Dict[str, Any]:
        """Execute a chat completion with retries, caching, and fallback."""
        # Resolve model: if using OpenAI-compatible provider, ignore Claude
        # shorthand model names passed by legacy callers.
        _claude_shorthands = {"qwen"}
        if model and model in _claude_shorthands and self.config.provider.value == "openai_compat":
            model = self.config.model  # Use the configured model (e.g., Qwen)
        else:
            model = model or self.config.model
        delay = self.config.retry_delay
        last_exc: Optional[Exception] = None
        call_start = time.time()

        # Build prompt preview for call history
        prompt_preview = ""
        if messages:
            last_msg = messages[-1].get("content", "")
            if isinstance(last_msg, list):
                last_msg = " ".join(p.get("text", "") for p in last_msg if isinstance(p, dict) and p.get("type") == "text")
            prompt_preview = str(last_msg)[:200]

        # Check cache first
        if use_cache and self._cache and not kwargs.get("stream"):
            cached = self._cache.get(messages, model, **kwargs)
            if cached:
                logger.debug(
                    "llm_cache_hit",
                    extra={
                        "event": "llm_cache_hit",
                        "description": description,
                        "model": model,
                    }
                )
                # Record cache hit in call history
                self._record_call(CallRecord(
                    timestamp=call_start,
                    description=description,
                    model=model,
                    success=True,
                    latency_ms=0.0,
                    input_tokens=0,
                    output_tokens=0,
                    prompt_preview=prompt_preview,
                    response_preview=_extract_response_text(cached)[:200],
                    cached=True,
                ))
                return cached

        # V2: L2 Semantic cache — embedding-based similarity
        # If L1 cache miss, check L2 semantic similarity
        try:
            from backend.app.services.infra_services import get_v2_config
            _v2_cfg = get_v2_config()
            if _v2_cfg.enable_semantic_cache and _v2_cfg.cache_l2_enabled:
                from backend.app.services.llm import SemanticCacheL2
                _l2 = SemanticCacheL2.get_instance()
                # Build prompt text from messages for embedding lookup
                _prompt_text = "\n".join(
                    str(m.get("content", "")) for m in messages
                )
                _l2_result = _l2.lookup(_prompt_text)
                if _l2_result is not None:
                    logger.debug("L2 cache hit (semantic)")
                    # Record L2 cache hit in call history
                    self._record_call(CallRecord(
                        timestamp=call_start,
                        description=description,
                        model=model,
                        success=True,
                        latency_ms=0.0,
                        input_tokens=0,
                        output_tokens=0,
                        prompt_preview=prompt_preview,
                        response_preview=_extract_response_text(_l2_result)[:200] if isinstance(_l2_result, dict) else "",
                        cached=True,
                    ))
                    return _l2_result
        except Exception:
            pass  # L2 cache is non-critical

        # Check per-operation circuit breaker first, then global
        op_circuit = self._get_op_circuit(description)
        if op_circuit and not op_circuit.allow_request():
            logger.warning(
                "llm_op_circuit_open",
                extra={
                    "event": "llm_op_circuit_open",
                    "description": description,
                    "circuit": op_circuit.name,
                }
            )
            self._record_call(CallRecord(
                timestamp=call_start,
                description=description,
                model=model,
                success=False,
                latency_ms=0.0,
                input_tokens=0,
                output_tokens=0,
                error_category=LLMErrorCategory.PROVIDER_UNAVAILABLE.value,
                error_message=f"Circuit breaker open for {op_circuit.name}",
                prompt_preview=prompt_preview,
                circuit_name=op_circuit.name,
            ))
            self._cost_tracker.record_error(description)
            raise RuntimeError(
                f"AI service is temporarily unavailable for {description} due to repeated failures. "
                "Please try again in a few minutes."
            )

        # Fallback to global circuit breaker
        if self._circuit_breaker and not self._circuit_breaker.allow_request():
            logger.warning(
                "llm_circuit_open",
                extra={
                    "event": "llm_circuit_open",
                    "description": description,
                    "provider": self.config.provider.value,
                }
            )
            # Try fallback immediately if circuit is open
            if self._fallback_provider:
                return self._try_fallback(messages, model, description, **kwargs)
            self._record_call(CallRecord(
                timestamp=call_start,
                description=description,
                model=model,
                success=False,
                latency_ms=0.0,
                input_tokens=0,
                output_tokens=0,
                error_category=LLMErrorCategory.PROVIDER_UNAVAILABLE.value,
                error_message="Global circuit breaker open",
                prompt_preview=prompt_preview,
            ))
            self._cost_tracker.record_error(description)
            raise RuntimeError(
                "AI service is temporarily unavailable due to repeated failures. "
                "Please try again in a few minutes. If the problem persists, check your API configuration."
            )

        _llm_logger = logging.getLogger("neura.llm")

        for attempt in range(1, self.config.max_retries + 1):
            try:
                logger.info(
                    "llm_call_start",
                    extra={
                        "event": "llm_call_start",
                        "description": description,
                        "attempt": attempt,
                        "model": model,
                        "provider": self.config.provider.value,
                    }
                )

                # Log the input prompt/messages
                _llm_logger.info(
                    "LLM_INPUT [%s] model=%s messages=%d\n%s",
                    description,
                    model,
                    len(messages),
                    _summarize_messages(messages),
                )

                start_time = time.time()
                response = self._provider.chat_completion(
                    messages=messages,
                    model=model,
                    description=description,
                    **kwargs
                )
                latency_ms = (time.time() - start_time) * 1000

                # Validate response has actual content (empty responses are useless)
                _resp_content = _extract_response_text(response)
                if not _resp_content or not _resp_content.strip():
                    raise ValueError("LLM returned empty response content")

                _append_raw_output(description, response)

                # Record success with both circuit breakers
                if self._circuit_breaker:
                    self._circuit_breaker.record_success()
                if op_circuit:
                    op_circuit.record_success()

                # Track usage
                usage = response.get("usage", {})
                input_tokens = usage.get("prompt_tokens", 0)
                output_tokens = usage.get("completion_tokens", 0)
                self._usage_tracker.record(model, input_tokens, output_tokens)
                _usage_tracker.record(model, input_tokens, output_tokens)

                # Track per-operation costs (V2 enhancement)
                self._cost_tracker.record(
                    description,
                    model=model,
                    input_tokens=input_tokens,
                    output_tokens=output_tokens,
                    latency_ms=latency_ms,
                )

                # Log the output response content
                output_text = _extract_response_text(response)
                _llm_logger.info(
                    "LLM_OUTPUT [%s] model=%s latency=%.0fms tokens=%d/%d\n%s",
                    description,
                    model,
                    latency_ms,
                    input_tokens,
                    output_tokens,
                    output_text[:4000] if output_text else "(empty)",
                )

                logger.info(
                    "llm_call_success",
                    extra={
                        "event": "llm_call_success",
                        "description": description,
                        "attempt": attempt,
                        "model": model,
                        "provider": self.config.provider.value,
                        "latency_ms": round(latency_ms, 2),
                        "input_tokens": input_tokens,
                        "output_tokens": output_tokens,
                    }
                )

                # Record in call history (V2 enhancement)
                self._record_call(CallRecord(
                    timestamp=start_time,
                    description=description,
                    model=model,
                    success=True,
                    latency_ms=latency_ms,
                    input_tokens=input_tokens,
                    output_tokens=output_tokens,
                    prompt_preview=prompt_preview,
                    response_preview=output_text[:200] if output_text else "",
                    circuit_name=op_circuit.name if op_circuit else None,
                ))

                # Cache successful response
                if use_cache and self._cache and not kwargs.get("stream"):
                    self._cache.set(messages, model, response, cache_ttl, **kwargs)

                # V2: Populate L2 semantic cache
                try:
                    from backend.app.services.infra_services import get_v2_config
                    _v2_cfg = get_v2_config()
                    if _v2_cfg.enable_semantic_cache and _v2_cfg.cache_l2_enabled:
                        from backend.app.services.llm import SemanticCacheL2
                        _l2 = SemanticCacheL2.get_instance()
                        _prompt_text = "\n".join(
                            str(m.get("content", "")) for m in messages
                        )
                        _l2.store(_prompt_text, response)
                except Exception:
                    pass  # L2 cache population is non-critical

                return response

            except Exception as exc:
                last_exc = exc
                error_cat = classify_error(exc)
                exc_latency = (time.time() - start_time) * 1000 if 'start_time' in dir() else 0

                _llm_logger.error(
                    "LLM_ERROR [%s] model=%s attempt=%d category=%s error=%s",
                    description,
                    model,
                    attempt,
                    error_cat.value,
                    _sanitize_error(exc),
                )

                # Record failure with both circuit breakers
                if self._circuit_breaker:
                    self._circuit_breaker.record_failure()
                if op_circuit:
                    op_circuit.record_failure()

                # Record error in cost tracker
                self._cost_tracker.record_error(description)

                # Record in call history
                self._record_call(CallRecord(
                    timestamp=time.time(),
                    description=description,
                    model=model,
                    success=False,
                    latency_ms=exc_latency,
                    input_tokens=0,
                    output_tokens=0,
                    error_category=error_cat.value,
                    error_message=_sanitize_error(exc)[:200],
                    prompt_preview=prompt_preview,
                    circuit_name=op_circuit.name if op_circuit else None,
                ))

                # Check for non-retryable errors by category
                if error_cat == LLMErrorCategory.AUTH_FAILURE:
                    logger.warning(
                        "llm_quota_exceeded",
                        extra={
                            "event": "llm_quota_exceeded",
                            "description": description,
                            "provider": self.config.provider.value,
                            "error_category": error_cat.value,
                        }
                    )
                    break

                if error_cat == LLMErrorCategory.CONTEXT_OVERFLOW:
                    logger.warning(
                        "llm_context_length_exceeded",
                        extra={
                            "event": "llm_context_length_exceeded",
                            "description": description,
                            "model": model,
                            "error_category": error_cat.value,
                        }
                    )
                    break  # Don't retry, won't help

                if error_cat == LLMErrorCategory.CONTENT_FILTER:
                    break  # Content policy violation, retrying won't help

                # Check for temperature errors (some models don't support it)
                if "temperature" in kwargs and _is_temperature_unsupported_error(exc):
                    logger.info(
                        "llm_temperature_override_removed",
                        extra={
                            "event": "llm_temperature_override_removed",
                            "description": description,
                            "model": model,
                        }
                    )
                    kwargs.pop("temperature", None)
                    continue

                logger.warning(
                    "llm_call_retry",
                    extra={
                        "event": "llm_call_retry",
                        "description": description,
                        "attempt": attempt,
                        "max_attempts": self.config.max_retries,
                        "retry_in": delay if attempt < self.config.max_retries else None,
                        "error": _sanitize_error(exc),
                        "error_type": type(exc).__name__,
                        "error_category": error_cat.value,
                    }
                )

                if attempt >= self.config.max_retries:
                    break

                # Agent-driven retry strategy — may adjust delay or abort
                strategy = self._agent_retry_strategy(
                    exc, attempt, description, error_cat, delay,
                )
                if strategy.get("action") == "abort":
                    logger.info(
                        "agent_retry_abort",
                        extra={"event": "agent_retry_abort", "description": description},
                    )
                    break
                delay = strategy.get("delay_seconds", delay)

                time.sleep(delay)
                delay = min(delay * self.config.retry_multiplier, 30.0)

        # Try fallback provider if available
        fallback_exc: Optional[Exception] = None
        if self._fallback_provider and last_exc:
            try:
                return self._try_fallback(messages, model, description, **kwargs)
            except Exception as fb_exc:
                fallback_exc = fb_exc
                logger.warning(
                    "llm_fallback_also_failed",
                    extra={
                        "event": "llm_fallback_also_failed",
                        "description": description,
                        "primary_error": _sanitize_error(last_exc),
                        "fallback_error": _sanitize_error(fb_exc),
                    }
                )

        # All attempts failed
        assert last_exc is not None
        fallback_attempted = self._fallback_provider is not None
        logger.error(
            "llm_call_failed",
            extra={
                "event": "llm_call_failed",
                "description": description,
                "attempts": self.config.max_retries,
                "model": model,
                "error_type": type(last_exc).__name__,
                "fallback_attempted": fallback_attempted,
                "fallback_error": _sanitize_error(fallback_exc) if fallback_exc else None,
            },
            exc_info=last_exc,
        )

        if _is_quota_exceeded_error(last_exc):
            raise RuntimeError(
                "AI service quota exceeded. Please check your API plan and billing details, "
                "or wait for the rate limit to reset."
            ) from last_exc

        if _is_context_length_error(last_exc):
            raise RuntimeError(
                "The document is too large for the AI to process. "
                "Please try with a smaller document or fewer pages."
            ) from last_exc

        # Include fallback info in error message if fallback was attempted
        error_msg = "AI processing failed. Please try again. If the problem persists, check your API configuration or contact support."
        if fallback_exc:
            error_msg = f"AI processing failed (primary and fallback providers both failed). Please try again. If the problem persists, check your API configuration or contact support."

        raise RuntimeError(error_msg) from last_exc

    def complete_structured(
        self,
        messages: List[Dict[str, Any]],
        response_model: type[BaseModel],
        model: Optional[str] = None,
        description: str = "llm_structured",
        **kwargs: Any,
    ) -> BaseModel:
        """Execute a structured completion using Instructor-compatible models."""
        try:
            import instructor
        except ImportError as exc:
            raise RuntimeError("instructor package is required. Install with: pip install instructor") from exc

        model = model or self.config.model
        try:
            if isinstance(self._provider, LiteLLMProvider):
                litellm = self._provider.get_client()
                instructor_client = instructor.from_litellm(litellm.completion)
                return instructor_client.chat.completions.create(
                    model=model,
                    messages=messages,
                    response_model=response_model,
                    **kwargs,
                )

            base_client = self._provider.get_client()
            instructor_client = instructor.from_openai(base_client)
            return instructor_client.chat.completions.create(
                model=model,
                messages=messages,
                response_model=response_model,
                **kwargs,
            )
        except Exception:
            # Fallback to manual parsing of JSON output.
            response = self.complete(
                messages=messages,
                model=model,
                description=description,
                **kwargs,
            )
            content = (
                response.get("choices", [{}])[0]
                .get("message", {})
                .get("content", "")
            )
            try:
                return response_model.model_validate_json(content)
            except AttributeError:
                return response_model.parse_raw(content)

    def _try_fallback(
        self,
        messages: List[Dict[str, Any]],
        model: str,
        description: str,
        **kwargs: Any,
    ) -> Dict[str, Any]:
        """Try fallback provider."""
        logger.info(
            "llm_fallback_attempt",
            extra={
                "event": "llm_fallback_attempt",
                "description": description,
                "fallback_provider": self.config.fallback_provider.value if self.config.fallback_provider else None,
            }
        )
        try:
            response = self._fallback_provider.chat_completion(
                messages=messages,
                model=self.config.fallback_model,
                **kwargs
            )
            _append_raw_output(f"{description}_fallback", response)

            # Track usage for fallback
            usage = response.get("usage", {})
            input_tokens = usage.get("prompt_tokens", 0)
            output_tokens = usage.get("completion_tokens", 0)
            fallback_model = self.config.fallback_model or model
            self._usage_tracker.record(fallback_model, input_tokens, output_tokens)
            _usage_tracker.record(fallback_model, input_tokens, output_tokens)

            return response
        except Exception as fallback_exc:
            logger.error(
                "llm_fallback_failed",
                extra={
                    "event": "llm_fallback_failed",
                    "description": description,
                    "error": _sanitize_error(fallback_exc),
                }
            )
            raise

    def complete_with_vision(
        self,
        text: str,
        images: List[Union[str, bytes, Path]],
        description: str = "vision_call",
        model: Optional[str] = None,
        detail: str = "auto",
        **kwargs: Any,
    ) -> Dict[str, Any]:
        """Execute a chat completion with vision/image inputs."""
        model = model or self.config.get_vision_model()

        vision_message = self._provider.prepare_vision_message(
            text=text,
            images=images,
            detail=detail,
        )

        return self.complete(
            messages=[vision_message],
            model=model,
            description=description,
            **kwargs
        )

    def stream(
        self,
        messages: List[Dict[str, Any]],
        model: Optional[str] = None,
        description: str = "llm_stream",
        **kwargs: Any,
    ) -> Iterator[Dict[str, Any]]:
        """Execute a streaming chat completion."""
        model = model or self.config.model

        logger.info(
            "llm_stream_start",
            extra={
                "event": "llm_stream_start",
                "description": description,
                "model": model,
                "provider": self.config.provider.value,
            }
        )

        try:
            for chunk in self._provider.chat_completion_stream(
                messages=messages,
                model=model,
                **kwargs
            ):
                yield chunk

            logger.info(
                "llm_stream_complete",
                extra={
                    "event": "llm_stream_complete",
                    "description": description,
                    "model": model,
                }
            )
        except Exception as exc:
            logger.error(
                "llm_stream_failed",
                extra={
                    "event": "llm_stream_failed",
                    "description": description,
                    "error": _sanitize_error(exc),
                }
            )
            raise

    def list_models(self) -> List[str]:
        """List available models from the current provider."""
        return self._provider.list_models()

    def health_check(self) -> bool:
        """Check if the provider is available."""
        return self._provider.health_check()

# Global client instance
_client: Optional[LLMClient] = None
_client_lock = threading.Lock()

def get_llm_client(force_new: bool = False) -> LLMClient:
    """Get the global LLM client instance."""
    global _client
    with _client_lock:
        if _client is None or force_new:
            _client = LLMClient()
    return _client

def call_completion(
    client: Any,  # Can be LLMClient or OpenAI client for backwards compatibility
    *,
    model: str,
    messages: List[Dict[str, Any]],
    description: str,
    timeout: Optional[float] = None,
    **kwargs: Any,
) -> Any:
    """Execute a chat completion - backwards compatible with existing code."""
    if isinstance(client, LLMClient):
        return client.complete(
            messages=messages,
            model=model,
            description=description,
            **kwargs
        )

    # Backwards compatibility: use existing OpenAI client
    # Import the old implementation
    from ..utils.llm import call_chat_completion as legacy_call
    return legacy_call(
        client,
        model=model,
        messages=messages,
        description=description,
        timeout=timeout,
        **kwargs
    )

def call_completion_with_vision(
    client: Any,
    *,
    text: str,
    images: List[Union[str, bytes, Path]],
    model: str,
    description: str,
    detail: str = "auto",
    **kwargs: Any,
) -> Any:
    """Execute a chat completion with vision inputs."""
    if isinstance(client, LLMClient):
        return client.complete_with_vision(
            text=text,
            images=images,
            model=model,
            description=description,
            detail=detail,
            **kwargs
        )

    # Backwards compatibility: build vision message manually
    import base64

    content: List[Dict[str, Any]] = [{"type": "text", "text": text}]

    for image in images:
        if isinstance(image, Path):
            image_data = base64.b64encode(image.read_bytes()).decode("utf-8")
            media_type = "image/png" if image.suffix.lower() == ".png" else "image/jpeg"
            image_url = f"data:{media_type};base64,{image_data}"
        elif isinstance(image, bytes):
            image_data = base64.b64encode(image).decode("utf-8")
            image_url = f"data:image/png;base64,{image_data}"
        else:
            image_url = image if image.startswith(("data:", "http")) else f"data:image/png;base64,{image}"

        content.append({
            "type": "image_url",
            "image_url": {"url": image_url, "detail": detail}
        })

    messages = [{"role": "user", "content": content}]

    from ..utils.llm import call_chat_completion as legacy_call
    return legacy_call(
        client,
        model=model,
        messages=messages,
        description=description,
        **kwargs
    )

def get_available_models() -> List[str]:
    """Get list of available models from the current provider."""
    client = get_llm_client()
    return client.list_models()

def health_check() -> Dict[str, Any]:
    """Check health of the LLM provider."""
    client = get_llm_client()
    config = client.config

    result = {
        "provider": config.provider.value,
        "model": config.model,
        "healthy": False,
        "fallback_available": config.fallback_provider is not None,
    }

    try:
        result["healthy"] = client.health_check()
        if result["healthy"]:
            result["available_models"] = client.list_models()[:5]  # First 5 models
    except Exception as e:
        logger.warning("llm_health_check_failed", extra={"error": str(e)})
        result["error"] = "Health check failed"

    return result

# Helper functions
# NOTE: _append_raw_output is imported from backend.app.services.utils.llm

def _is_quota_exceeded_error(exc: BaseException) -> bool:
    return classify_error(exc) in (LLMErrorCategory.RATE_LIMIT, LLMErrorCategory.AUTH_FAILURE)

def _is_temperature_unsupported_error(exc: BaseException) -> bool:
    detail = str(exc).lower()
    return "temperature" in detail and "unsupported" in detail

def _is_context_length_error(exc: BaseException) -> bool:
    return classify_error(exc) == LLMErrorCategory.CONTEXT_OVERFLOW

def _is_invalid_request_error(exc: BaseException) -> bool:
    return classify_error(exc) == LLMErrorCategory.INVALID_REQUEST

def get_global_usage_stats() -> Dict[str, Any]:
    """Get global token usage statistics."""
    return _usage_tracker.get_stats()

def reset_global_usage_stats() -> None:
    """Reset global token usage statistics."""
    _usage_tracker.reset()

# ── V2 Public API ──

def get_call_history_stats() -> Dict[str, Any]:
    """Get call history statistics from the global client."""
    client = get_llm_client()
    return client.get_call_history_stats()

def get_circuit_breaker_stats() -> Dict[str, Dict[str, Any]]:
    """Get per-operation circuit breaker stats."""
    return _circuit_registry.get_all_stats()

def get_cost_stats() -> Dict[str, Any]:
    """Get per-operation cost statistics."""
    from backend.app.services.llm import get_cost_tracker
    return get_cost_tracker().get_stats()

def get_full_observability() -> Dict[str, Any]:
    """Get combined observability data: usage, costs, circuits, call history."""
    client = get_llm_client()
    from backend.app.services.llm import get_cost_tracker
    return {
        "usage": _usage_tracker.get_stats(),
        "costs": get_cost_tracker().get_stats(),
        "circuits": _circuit_registry.get_all_stats(),
        "cache": client._cache.get_stats() if client._cache else {},
        "call_history": client.get_call_history_stats(),
    }

# __all__ kept for public API surface
__all__ = [
    "LLMClient", "get_llm_client", "call_completion", "call_completion_with_vision",
    "get_available_models", "health_check",
    "LLMConfig", "LLMProvider", "get_llm_config",
    "BaseProvider", "ClaudeCodeCLIProvider", "LiteLLMProvider", "get_provider",
    "VisionLanguageModel", "DocumentAnalysisResult", "TableExtractionResult",
    "get_vlm", "analyze_document_image", "extract_tables_from_image",
    "Agent", "AgentConfig", "AgentRole", "Task", "TaskResult", "Crew", "Tool",
    "create_document_analyzer_agent", "create_data_extractor_agent",
    "create_sql_generator_agent", "create_chart_suggester_agent",
    "create_template_mapper_agent", "create_quality_reviewer_agent",
    "create_document_processing_crew", "create_report_generation_crew",
    "TextToSQL", "TableSchema", "SQLGenerationResult", "get_text_to_sql", "generate_sql",
    "RAGRetriever", "Document", "RetrievalResult", "BM25Index",
    "create_retriever", "quick_rag_query",
    "EnhancedDocumentExtractor", "ExtractedContent", "ExtractedTable",
    "FieldSchema", "extract_document", "extract_tables",
]
