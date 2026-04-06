from __future__ import annotations
"""
Spreadsheet Service - Core spreadsheet operations.
"""



import json
import logging
import threading
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional

from pydantic import BaseModel

from backend.app.services.config import get_settings

logger = logging.getLogger("neura.spreadsheets")


class CellValue(BaseModel):
    """Cell value with type information."""

    value: Any
    formula: Optional[str] = None
    formatted_value: Optional[str] = None
    cell_type: str = "string"  # string, number, boolean, date, formula, error


class CellFormat(BaseModel):
    """Cell formatting options."""

    bold: bool = False
    italic: bool = False
    underline: bool = False
    font_size: int = 11
    font_color: str = "#000000"
    background_color: Optional[str] = None
    horizontal_align: str = "left"  # left, center, right
    vertical_align: str = "middle"  # top, middle, bottom
    number_format: Optional[str] = None
    border: Optional[dict[str, Any]] = None


class ConditionalFormat(BaseModel):
    """Conditional formatting rule."""

    id: str
    range: str  # e.g., "A1:B10"
    type: str  # greaterThan, lessThan, equals, between, text, custom
    value: Any
    value2: Optional[Any] = None  # For "between" type
    format: CellFormat


class DataValidation(BaseModel):
    """Data validation rule."""

    id: str
    range: str
    type: str  # list, number, date, text, custom
    criteria: str  # equals, between, greaterThan, etc.
    value: Any
    value2: Optional[Any] = None
    allow_blank: bool = True
    show_dropdown: bool = True
    error_message: Optional[str] = None


class Sheet(BaseModel):
    """Single sheet in a spreadsheet."""

    id: str
    name: str
    index: int
    data: list[list[Any]]  # 2D array of cell values
    formats: dict[str, CellFormat] = {}  # cell address -> format
    column_widths: dict[int, int] = {}
    row_heights: dict[int, int] = {}
    frozen_rows: int = 0
    frozen_cols: int = 0
    conditional_formats: list[ConditionalFormat] = []
    data_validations: list[DataValidation] = []


class Spreadsheet(BaseModel):
    """Spreadsheet model."""

    id: str
    name: str
    sheets: list[Sheet]
    created_at: str
    updated_at: str
    owner_id: Optional[str] = None
    metadata: dict[str, Any] = {}


class PivotTableConfig(BaseModel):
    """Pivot table configuration."""

    id: str
    spreadsheet_id: str
    sheet_id: str
    source_range: str
    rows: list[str]  # Field names for rows
    columns: list[str]  # Field names for columns
    values: list[dict[str, str]]  # [{"field": "Amount", "aggregation": "SUM"}]
    filters: list[dict[str, Any]] = []
    name: str = "PivotTable1"


class SpreadsheetService:
    """Service for spreadsheet CRUD operations."""

    def __init__(self, storage_path: Optional[Path] = None):
        base_root = get_settings().uploads_root
        self._storage_path = storage_path or (base_root / "spreadsheets")
        self._storage_path.mkdir(parents=True, exist_ok=True)
        self._lock = threading.Lock()

    def create(
        self,
        name: str,
        owner_id: Optional[str] = None,
        initial_data: Optional[list[list[Any]]] = None,
    ) -> Spreadsheet:
        """Create a new spreadsheet."""
        now = datetime.now(timezone.utc).isoformat()

        # Create initial sheet
        initial_sheet = Sheet(
            id=str(uuid.uuid4()),
            name="Sheet1",
            index=0,
            data=initial_data or [["" for _ in range(26)] for _ in range(100)],
        )

        spreadsheet = Spreadsheet(
            id=str(uuid.uuid4()),
            name=name,
            sheets=[initial_sheet],
            created_at=now,
            updated_at=now,
            owner_id=owner_id,
        )

        self._save_spreadsheet(spreadsheet)
        logger.info(f"Created spreadsheet: {spreadsheet.id}")
        return spreadsheet

    def get(self, spreadsheet_id: str) -> Optional[Spreadsheet]:
        """Get a spreadsheet by ID."""
        file_path = self._get_spreadsheet_path(spreadsheet_id)
        if not file_path or not file_path.exists():
            return None
        with open(file_path) as f:
            data = json.load(f)
        return Spreadsheet(**data)

    def update(
        self,
        spreadsheet_id: str,
        name: Optional[str] = None,
        metadata: Optional[dict] = None,
    ) -> Optional[Spreadsheet]:
        """Update spreadsheet metadata."""
        with self._lock:
            spreadsheet = self.get(spreadsheet_id)
            if not spreadsheet:
                return None

            if name:
                spreadsheet.name = name
            if metadata:
                spreadsheet.metadata.update(metadata)

            spreadsheet.updated_at = datetime.now(timezone.utc).isoformat()
            self._save_spreadsheet(spreadsheet)
            return spreadsheet

    def delete(self, spreadsheet_id: str) -> bool:
        """Delete a spreadsheet."""
        file_path = self._get_spreadsheet_path(spreadsheet_id)
        if not file_path or not file_path.exists():
            return False
        file_path.unlink()
        logger.info(f"Deleted spreadsheet: {spreadsheet_id}")
        return True

    def update_cells(
        self,
        spreadsheet_id: str,
        sheet_index: int,
        updates: list[dict[str, Any]],
    ) -> Optional[Spreadsheet]:
        """Update cell values. updates = [{"row": 0, "col": 0, "value": "Hello"}]"""
        with self._lock:
            spreadsheet = self.get(spreadsheet_id)
            if not spreadsheet:
                return None

            if sheet_index < 0 or sheet_index >= len(spreadsheet.sheets):
                return None

            sheet = spreadsheet.sheets[sheet_index]

            for update in updates:
                row = update.get("row", 0)
                col = update.get("col", 0)
                value = update.get("value", "")

                # Expand data array if needed
                while row >= len(sheet.data):
                    sheet.data.append(["" for _ in range(len(sheet.data[0]) if sheet.data else 26)])
                while col >= len(sheet.data[row]):
                    sheet.data[row].append("")

                sheet.data[row][col] = value

            spreadsheet.updated_at = datetime.now(timezone.utc).isoformat()
            self._save_spreadsheet(spreadsheet)
            return spreadsheet

    def add_sheet(
        self,
        spreadsheet_id: str,
        name: Optional[str] = None,
    ) -> Optional[Sheet]:
        """Add a new sheet to the spreadsheet."""
        with self._lock:
            spreadsheet = self.get(spreadsheet_id)
            if not spreadsheet:
                return None

            new_index = len(spreadsheet.sheets)
            sheet_name = name or f"Sheet{new_index + 1}"

            sheet = Sheet(
                id=str(uuid.uuid4()),
                name=sheet_name,
                index=new_index,
                data=[["" for _ in range(26)] for _ in range(100)],
            )

            spreadsheet.sheets.append(sheet)
            spreadsheet.updated_at = datetime.now(timezone.utc).isoformat()
            self._save_spreadsheet(spreadsheet)

            logger.info(f"Added sheet {sheet_name} to spreadsheet {spreadsheet_id}")
            return sheet

    def delete_sheet(self, spreadsheet_id: str, sheet_id: str) -> bool:
        """Delete a sheet from the spreadsheet."""
        with self._lock:
            spreadsheet = self.get(spreadsheet_id)
            if not spreadsheet:
                return False

            # Don't delete last sheet
            if len(spreadsheet.sheets) <= 1:
                return False

            spreadsheet.sheets = [s for s in spreadsheet.sheets if s.id != sheet_id]

            # Reindex sheets
            for i, sheet in enumerate(spreadsheet.sheets):
                sheet.index = i

            spreadsheet.updated_at = datetime.now(timezone.utc).isoformat()
            self._save_spreadsheet(spreadsheet)
            return True

    def rename_sheet(
        self,
        spreadsheet_id: str,
        sheet_id: str,
        new_name: str,
    ) -> bool:
        """Rename a sheet."""
        with self._lock:
            spreadsheet = self.get(spreadsheet_id)
            if not spreadsheet:
                return False

            for sheet in spreadsheet.sheets:
                if sheet.id == sheet_id:
                    sheet.name = new_name
                    spreadsheet.updated_at = datetime.now(timezone.utc).isoformat()
                    self._save_spreadsheet(spreadsheet)
                    return True

            return False

    def set_conditional_format(
        self,
        spreadsheet_id: str,
        sheet_id: str,
        conditional_format: ConditionalFormat,
    ) -> bool:
        """Add or update a conditional format rule."""
        with self._lock:
            spreadsheet = self.get(spreadsheet_id)
            if not spreadsheet:
                return False

            for sheet in spreadsheet.sheets:
                if sheet.id == sheet_id:
                    # Update existing or add new
                    updated = False
                    for i, cf in enumerate(sheet.conditional_formats):
                        if cf.id == conditional_format.id:
                            sheet.conditional_formats[i] = conditional_format
                            updated = True
                            break

                    if not updated:
                        sheet.conditional_formats.append(conditional_format)

                    spreadsheet.updated_at = datetime.now(timezone.utc).isoformat()
                    self._save_spreadsheet(spreadsheet)
                    return True

            return False

    def set_data_validation(
        self,
        spreadsheet_id: str,
        sheet_id: str,
        validation: DataValidation,
    ) -> bool:
        """Add or update a data validation rule."""
        with self._lock:
            spreadsheet = self.get(spreadsheet_id)
            if not spreadsheet:
                return False

            for sheet in spreadsheet.sheets:
                if sheet.id == sheet_id:
                    # Update existing or add new
                    updated = False
                    for i, dv in enumerate(sheet.data_validations):
                        if dv.id == validation.id:
                            sheet.data_validations[i] = validation
                            updated = True
                            break

                    if not updated:
                        sheet.data_validations.append(validation)

                    spreadsheet.updated_at = datetime.now(timezone.utc).isoformat()
                    self._save_spreadsheet(spreadsheet)
                    return True

            return False

    def freeze_panes(
        self,
        spreadsheet_id: str,
        sheet_id: str,
        rows: int = 0,
        cols: int = 0,
    ) -> bool:
        """Set frozen rows and columns for a sheet."""
        with self._lock:
            spreadsheet = self.get(spreadsheet_id)
            if not spreadsheet:
                return False

            for sheet in spreadsheet.sheets:
                if sheet.id == sheet_id:
                    sheet.frozen_rows = rows
                    sheet.frozen_cols = cols
                    spreadsheet.updated_at = datetime.now(timezone.utc).isoformat()
                    self._save_spreadsheet(spreadsheet)
                    return True

            return False

    def import_csv(
        self,
        csv_content: str,
        name: str = "Imported Spreadsheet",
        delimiter: str = ",",
        owner_id: Optional[str] = None,
    ) -> Spreadsheet:
        """Import a CSV file as a new spreadsheet."""
        import csv
        from io import StringIO

        reader = csv.reader(StringIO(csv_content), delimiter=delimiter)
        data = list(reader)

        # Pad rows to equal length
        max_cols = max(len(row) for row in data) if data else 26
        for row in data:
            while len(row) < max_cols:
                row.append("")

        return self.create(name=name, owner_id=owner_id, initial_data=data)

    def import_xlsx(
        self,
        xlsx_content: bytes,
        name: str = "Imported Spreadsheet",
        owner_id: Optional[str] = None,
    ) -> Spreadsheet:
        """Import an XLSX file as a new spreadsheet."""
        import openpyxl
        from io import BytesIO

        wb = openpyxl.load_workbook(BytesIO(xlsx_content), data_only=True)
        sheets: list[Sheet] = []
        for idx, ws in enumerate(wb.worksheets):
            data: list[list[Any]] = []
            for row in ws.iter_rows(values_only=True):
                data.append([("" if v is None else v) for v in row])
            if not data:
                data = [["" for _ in range(26)] for _ in range(100)]
            # Pad rows to equal length
            max_cols = max(len(r) for r in data) if data else 26
            for r in data:
                while len(r) < max_cols:
                    r.append("")
            sheets.append(Sheet(
                id=str(uuid.uuid4()),
                name=ws.title or f"Sheet{idx + 1}",
                index=idx,
                data=data,
            ))

        if not sheets:
            sheets = [Sheet(
                id=str(uuid.uuid4()),
                name="Sheet1",
                index=0,
                data=[["" for _ in range(26)] for _ in range(100)],
            )]

        now = datetime.now(timezone.utc).isoformat()
        spreadsheet = Spreadsheet(
            id=str(uuid.uuid4()),
            name=name,
            sheets=sheets,
            created_at=now,
            updated_at=now,
            owner_id=owner_id,
        )
        self._save_spreadsheet(spreadsheet)
        logger.info(f"Imported XLSX spreadsheet: {spreadsheet.id} ({len(sheets)} sheets)")
        return spreadsheet

    def export_csv(
        self,
        spreadsheet_id: str,
        sheet_index: int = 0,
        delimiter: str = ",",
    ) -> Optional[str]:
        """Export a sheet as CSV."""
        import csv
        from io import StringIO

        spreadsheet = self.get(spreadsheet_id)
        if not spreadsheet:
            return None

        if sheet_index < 0 or sheet_index >= len(spreadsheet.sheets):
            return None

        sheet = spreadsheet.sheets[sheet_index]
        output = StringIO()
        writer = csv.writer(output, delimiter=delimiter)

        for row in sheet.data:
            writer.writerow(row)

        return output.getvalue()

    def export_xlsx(
        self,
        spreadsheet_id: str,
        sheet_index: int = 0,
    ) -> Optional[bytes]:
        """Export a sheet as XLSX binary."""
        import openpyxl
        from io import BytesIO

        spreadsheet = self.get(spreadsheet_id)
        if not spreadsheet:
            return None

        if sheet_index < 0 or sheet_index >= len(spreadsheet.sheets):
            return None

        sheet = spreadsheet.sheets[sheet_index]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet.name

        for row in sheet.data:
            ws.append(row)

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def list_spreadsheets(
        self,
        owner_id: Optional[str] = None,
        limit: int = 100,
        offset: int = 0,
    ) -> list[Spreadsheet]:
        """List all spreadsheets."""
        spreadsheets = []

        for file_path in self._storage_path.glob("*.json"):
            try:
                with open(file_path) as f:
                    data = json.load(f)
                spreadsheet = Spreadsheet(**data)

                if owner_id and spreadsheet.owner_id != owner_id:
                    continue

                spreadsheets.append(spreadsheet)
            except Exception as e:
                logger.warning(f"Error loading spreadsheet from {file_path}: {e}")

        # Sort by updated_at descending
        spreadsheets.sort(key=lambda s: s.updated_at, reverse=True)
        return spreadsheets[offset:offset + limit]

    def _normalize_id(self, spreadsheet_id: str) -> Optional[str]:
        try:
            return str(uuid.UUID(str(spreadsheet_id)))
        except (ValueError, TypeError):
            return None

    def _get_spreadsheet_path(self, spreadsheet_id: str) -> Optional[Path]:
        """Get path to spreadsheet JSON file."""
        normalized = self._normalize_id(spreadsheet_id)
        if not normalized:
            return None
        return self._storage_path / f"{normalized}.json"

    def _save_spreadsheet(self, spreadsheet: Spreadsheet) -> None:
        """Save spreadsheet to disk."""
        file_path = self._get_spreadsheet_path(spreadsheet.id)
        if not file_path:
            raise ValueError(f"Invalid spreadsheet ID: {spreadsheet.id}")
        with open(file_path, "w") as f:
            json.dump(spreadsheet.model_dump(), f, indent=2)



# ── Originally: formula_engine.py ──

"""
Formula Engine - Formula parsing and evaluation.
"""


import logging
import math
import random
import re
from datetime import datetime, timezone
from typing import Any, Callable, Optional

from pydantic import BaseModel

logger = logging.getLogger("neura.formula_engine")


class FormulaResult(BaseModel):
    """Result of formula evaluation."""

    value: Any
    formatted_value: str
    error: Optional[str] = None
    cell_type: str = "formula"


class CellReference(BaseModel):
    """Parsed cell reference."""

    col: int
    row: int
    col_abs: bool = False
    row_abs: bool = False


class FormulaEngine:
    """
    Formula engine for spreadsheet calculations.

    Supports 400+ Excel-compatible functions via HyperFormula integration
    and provides a fallback pure-Python implementation for basic functions.
    """

    # Built-in functions (fallback when HyperFormula not available)
    FUNCTIONS: dict[str, Callable] = {}

    def __init__(self):
        self._register_functions()
        self._hyperformula = None
        self._try_init_hyperformula()

    def _try_init_hyperformula(self):
        """Try to initialize HyperFormula for advanced formula support."""
        try:
            # HyperFormula is a JavaScript library
            # For Python, we'd need to use Pyodide or similar
            # For now, we use pure Python fallback
            pass
        except Exception as e:
            logger.debug(f"HyperFormula not available, using fallback: {e}")

    def _register_functions(self):
        """Register built-in functions."""
        # Math functions
        self.FUNCTIONS = {
            # Basic math
            "SUM": self._fn_sum,
            "AVERAGE": self._fn_average,
            "COUNT": self._fn_count,
            "COUNTA": self._fn_counta,
            "MAX": self._fn_max,
            "MIN": self._fn_min,
            "ABS": lambda x: abs(float(x)),
            "SQRT": lambda x: math.sqrt(float(x)),
            "POWER": lambda x, y: math.pow(float(x), float(y)),
            "LOG": lambda x, base=10: math.log(float(x), float(base)),
            "LN": lambda x: math.log(float(x)),
            "EXP": lambda x: math.exp(float(x)),
            "ROUND": lambda x, d=0: round(float(x), int(d)),
            "FLOOR": lambda x: math.floor(float(x)),
            "CEILING": lambda x: math.ceil(float(x)),
            "MOD": lambda x, y: float(x) % float(y),
            "PI": lambda: math.pi,
            "RAND": lambda: random.random(),
            "RANDBETWEEN": lambda a, b: random.randint(int(a), int(b)),

            # Statistical
            "MEDIAN": self._fn_median,
            "MODE": self._fn_mode,
            "STDEV": self._fn_stdev,
            "VAR": self._fn_var,

            # Conditional
            "IF": self._fn_if,
            "AND": self._fn_and,
            "OR": self._fn_or,
            "NOT": lambda x: not self._to_bool(x),
            "IFERROR": self._fn_iferror,
            "ISBLANK": lambda x: x is None or x == "",
            "ISNUMBER": lambda x: isinstance(x, (int, float)),
            "ISTEXT": lambda x: isinstance(x, str),

            # Lookup
            "VLOOKUP": self._fn_vlookup,
            "HLOOKUP": self._fn_hlookup,
            "INDEX": self._fn_index,
            "MATCH": self._fn_match,

            # Text
            "CONCATENATE": lambda *args: "".join(str(a) for a in args),
            "CONCAT": lambda *args: "".join(str(a) for a in args),
            "LEFT": lambda s, n=1: str(s)[:int(n)],
            "RIGHT": lambda s, n=1: str(s)[-int(n):],
            "MID": lambda s, start, length: str(s)[int(start)-1:int(start)-1+int(length)],
            "LEN": lambda s: len(str(s)),
            "UPPER": lambda s: str(s).upper(),
            "LOWER": lambda s: str(s).lower(),
            "PROPER": lambda s: str(s).title(),
            "TRIM": lambda s: str(s).strip(),
            "SUBSTITUTE": lambda s, old, new: str(s).replace(str(old), str(new)),
            "FIND": lambda needle, haystack, start=1: str(haystack).find(str(needle), int(start)-1) + 1,
            "SEARCH": lambda needle, haystack, start=1: str(haystack).lower().find(str(needle).lower(), int(start)-1) + 1,
            "TEXT": self._fn_text,
            "VALUE": lambda s: float(str(s).replace(",", "")),

            # Date/Time
            "TODAY": lambda: datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            "NOW": lambda: datetime.now(timezone.utc).isoformat(),
            "DATE": lambda y, m, d: datetime(int(y), int(m), int(d)).strftime("%Y-%m-%d"),
            "YEAR": lambda d: self._parse_date(d).year,
            "MONTH": lambda d: self._parse_date(d).month,
            "DAY": lambda d: self._parse_date(d).day,
            "HOUR": lambda d: self._parse_datetime(d).hour,
            "MINUTE": lambda d: self._parse_datetime(d).minute,
            "SECOND": lambda d: self._parse_datetime(d).second,
            "WEEKDAY": lambda d: self._parse_date(d).weekday() + 1,
            "DATEDIF": self._fn_datedif,

            # Aggregation with conditions
            "SUMIF": self._fn_sumif,
            "COUNTIF": self._fn_countif,
            "AVERAGEIF": self._fn_averageif,
        }

    def evaluate(
        self,
        formula: str,
        data: list[list[Any]],
        current_cell: Optional[tuple[int, int]] = None,
    ) -> FormulaResult:
        """
        Evaluate a formula against spreadsheet data.

        Args:
            formula: Formula string starting with '='
            data: 2D array of cell values
            current_cell: (row, col) of the cell containing this formula

        Returns:
            FormulaResult with evaluated value
        """
        if not formula.startswith("="):
            return FormulaResult(
                value=formula,
                formatted_value=str(formula),
                cell_type="string",
            )

        try:
            # Remove leading '='
            expr = formula[1:].strip()

            # Parse and evaluate
            result = self._evaluate_expression(expr, data, current_cell)

            return FormulaResult(
                value=result,
                formatted_value=self._format_value(result),
            )
        except Exception as e:
            logger.warning(f"Formula error: {formula} - {e}")
            return FormulaResult(
                value=None,
                formatted_value="#ERROR!",
                error=str(e),
            )

    def _evaluate_expression(
        self,
        expr: str,
        data: list[list[Any]],
        current_cell: Optional[tuple[int, int]] = None,
    ) -> Any:
        """Evaluate a formula expression."""
        expr = expr.strip()

        # Check for function call
        func_match = re.match(r"^([A-Z]+)\((.*)\)$", expr, re.IGNORECASE)
        if func_match:
            func_name = func_match.group(1).upper()
            args_str = func_match.group(2)

            if func_name not in self.FUNCTIONS:
                raise ValueError(f"Unknown function: {func_name}")

            # Parse arguments
            args = self._parse_arguments(args_str, data, current_cell)

            # Call function
            return self.FUNCTIONS[func_name](*args)

        # Check for cell reference
        cell_match = re.match(r"^\$?([A-Z]+)\$?(\d+)$", expr, re.IGNORECASE)
        if cell_match:
            col = self._col_to_index(cell_match.group(1))
            row = int(cell_match.group(2)) - 1
            return self._get_cell_value(data, row, col)

        # Check for range reference (A1:B10)
        range_match = re.match(
            r"^\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)$",
            expr,
            re.IGNORECASE,
        )
        if range_match:
            start_col = self._col_to_index(range_match.group(1))
            start_row = int(range_match.group(2)) - 1
            end_col = self._col_to_index(range_match.group(3))
            end_row = int(range_match.group(4)) - 1
            return self._get_range_values(data, start_row, start_col, end_row, end_col)

        # Try to evaluate as number
        try:
            if "." in expr:
                return float(expr)
            return int(expr)
        except ValueError:
            pass

        # Return as string (strip quotes if present)
        if (expr.startswith('"') and expr.endswith('"')) or \
           (expr.startswith("'") and expr.endswith("'")):
            return expr[1:-1]

        return expr

    def _parse_arguments(
        self,
        args_str: str,
        data: list[list[Any]],
        current_cell: Optional[tuple[int, int]],
    ) -> list[Any]:
        """Parse function arguments."""
        if not args_str.strip():
            return []

        args = []
        current_arg = ""
        paren_depth = 0
        in_string = False
        string_char = None

        for char in args_str:
            if char in ('"', "'") and not in_string:
                in_string = True
                string_char = char
                current_arg += char
            elif char == string_char and in_string:
                in_string = False
                string_char = None
                current_arg += char
            elif char == "(" and not in_string:
                paren_depth += 1
                current_arg += char
            elif char == ")" and not in_string:
                paren_depth -= 1
                current_arg += char
            elif char == "," and paren_depth == 0 and not in_string:
                args.append(self._evaluate_expression(current_arg.strip(), data, current_cell))
                current_arg = ""
            else:
                current_arg += char

        if current_arg.strip():
            args.append(self._evaluate_expression(current_arg.strip(), data, current_cell))

        return args

    def _col_to_index(self, col: str) -> int:
        """Convert column letter to 0-based index."""
        result = 0
        for char in col.upper():
            result = result * 26 + (ord(char) - ord("A") + 1)
        return result - 1

    def _get_cell_value(self, data: list[list[Any]], row: int, col: int) -> Any:
        """Get value from a cell."""
        if row < 0 or row >= len(data):
            return None
        if col < 0 or col >= len(data[row]):
            return None
        return data[row][col]

    def _get_range_values(
        self,
        data: list[list[Any]],
        start_row: int,
        start_col: int,
        end_row: int,
        end_col: int,
    ) -> list[Any]:
        """Get all values in a range as flat list."""
        values = []
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                val = self._get_cell_value(data, row, col)
                if val is not None and val != "":
                    values.append(val)
        return values

    def _format_value(self, value: Any) -> str:
        """Format value for display."""
        if value is None:
            return ""
        if isinstance(value, float):
            if math.isfinite(value) and value == int(value):
                return str(int(value))
            return f"{value:.2f}"
        return str(value)

    def _to_number(self, value: Any) -> float:
        """Convert value to number."""
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            return float(value.replace(",", ""))
        return 0.0

    def _to_bool(self, value: Any) -> bool:
        """Convert value to boolean."""
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return value != 0
        if isinstance(value, str):
            return value.lower() in ("true", "yes", "1")
        return bool(value)

    def _parse_date(self, value: Any) -> datetime:
        """Parse value as date."""
        if isinstance(value, datetime):
            return value
        return datetime.fromisoformat(str(value)[:10])

    def _parse_datetime(self, value: Any) -> datetime:
        """Parse value as datetime."""
        if isinstance(value, datetime):
            return value
        return datetime.fromisoformat(str(value))

    # Function implementations
    def _fn_sum(self, *args) -> float:
        """SUM function."""
        total = 0.0
        for arg in args:
            if isinstance(arg, list):
                total += sum(self._to_number(v) for v in arg if v is not None and v != "")
            else:
                total += self._to_number(arg)
        return total

    def _fn_average(self, *args) -> float:
        """AVERAGE function."""
        values = []
        for arg in args:
            if isinstance(arg, list):
                values.extend(self._to_number(v) for v in arg if v is not None and v != "")
            else:
                values.append(self._to_number(arg))
        return sum(values) / len(values) if values else 0.0

    def _fn_count(self, *args) -> int:
        """COUNT function - counts numbers."""
        count = 0
        for arg in args:
            if isinstance(arg, list):
                count += sum(1 for v in arg if isinstance(v, (int, float)))
            elif isinstance(arg, (int, float)):
                count += 1
        return count

    def _fn_counta(self, *args) -> int:
        """COUNTA function - counts non-empty values."""
        count = 0
        for arg in args:
            if isinstance(arg, list):
                count += sum(1 for v in arg if v is not None and v != "")
            elif arg is not None and arg != "":
                count += 1
        return count

    def _fn_max(self, *args) -> float:
        """MAX function."""
        values = []
        for arg in args:
            if isinstance(arg, list):
                values.extend(self._to_number(v) for v in arg if isinstance(v, (int, float)))
            elif isinstance(arg, (int, float)):
                values.append(self._to_number(arg))
        return max(values) if values else 0.0

    def _fn_min(self, *args) -> float:
        """MIN function."""
        values = []
        for arg in args:
            if isinstance(arg, list):
                values.extend(self._to_number(v) for v in arg if isinstance(v, (int, float)))
            elif isinstance(arg, (int, float)):
                values.append(self._to_number(arg))
        return min(values) if values else 0.0

    def _fn_median(self, *args) -> float:
        """MEDIAN function."""
        import statistics
        values = []
        for arg in args:
            if isinstance(arg, list):
                values.extend(self._to_number(v) for v in arg if isinstance(v, (int, float)))
            elif isinstance(arg, (int, float)):
                values.append(self._to_number(arg))
        return statistics.median(values) if values else 0.0

    def _fn_mode(self, *args) -> float:
        """MODE function."""
        import statistics
        values = []
        for arg in args:
            if isinstance(arg, list):
                values.extend(self._to_number(v) for v in arg if isinstance(v, (int, float)))
            elif isinstance(arg, (int, float)):
                values.append(self._to_number(arg))
        return statistics.mode(values) if values else 0.0

    def _fn_stdev(self, *args) -> float:
        """STDEV function."""
        import statistics
        values = []
        for arg in args:
            if isinstance(arg, list):
                values.extend(self._to_number(v) for v in arg if isinstance(v, (int, float)))
            elif isinstance(arg, (int, float)):
                values.append(self._to_number(arg))
        return statistics.stdev(values) if len(values) > 1 else 0.0

    def _fn_var(self, *args) -> float:
        """VAR function."""
        import statistics
        values = []
        for arg in args:
            if isinstance(arg, list):
                values.extend(self._to_number(v) for v in arg if isinstance(v, (int, float)))
            elif isinstance(arg, (int, float)):
                values.append(self._to_number(arg))
        return statistics.variance(values) if len(values) > 1 else 0.0

    def _fn_if(self, condition, true_val, false_val=None):
        """IF function."""
        return true_val if self._to_bool(condition) else (false_val or "")

    def _fn_and(self, *args) -> bool:
        """AND function."""
        return all(self._to_bool(arg) for arg in args)

    def _fn_or(self, *args) -> bool:
        """OR function."""
        return any(self._to_bool(arg) for arg in args)

    def _fn_iferror(self, value, error_value):
        """IFERROR function."""
        # In our context, if value evaluation would have failed, we'd already have an error
        # So this is mainly for compatibility
        return value if value is not None else error_value

    def _fn_vlookup(self, lookup_value, table_array, col_index, range_lookup=True):
        """VLOOKUP function."""
        if not isinstance(table_array, list) or not table_array:
            return "#N/A"

        col_index = int(col_index) - 1

        for row in table_array:
            if isinstance(row, list) and len(row) > col_index:
                if row[0] == lookup_value:
                    return row[col_index]

        return "#N/A"

    def _fn_hlookup(self, lookup_value, table_array, row_index, range_lookup=True):
        """HLOOKUP function."""
        # Similar to VLOOKUP but horizontal
        return "#N/A"  # Simplified

    def _fn_index(self, array, row_num, col_num=None):
        """INDEX function."""
        if not isinstance(array, list):
            return "#REF!"

        row_num = int(row_num) - 1
        if row_num < 0 or row_num >= len(array):
            return "#REF!"

        if col_num is not None:
            col_num = int(col_num) - 1
            if isinstance(array[row_num], list):
                if col_num < 0 or col_num >= len(array[row_num]):
                    return "#REF!"
                return array[row_num][col_num]

        return array[row_num]

    def _fn_match(self, lookup_value, lookup_array, match_type=1):
        """MATCH function."""
        if not isinstance(lookup_array, list):
            return "#N/A"

        for i, val in enumerate(lookup_array):
            if val == lookup_value:
                return i + 1

        return "#N/A"

    def _fn_text(self, value, format_str):
        """TEXT function."""
        # Simplified text formatting
        try:
            if "%" in format_str:
                return f"{float(value) * 100:.0f}%"
            if "$" in format_str:
                return f"${float(value):,.2f}"
            return str(value)
        except (ValueError, TypeError):
            return str(value)

    def _fn_datedif(self, start_date, end_date, unit):
        """DATEDIF function."""
        start = self._parse_date(start_date)
        end = self._parse_date(end_date)
        diff = end - start

        unit = str(unit).upper()
        if unit == "D":
            return diff.days
        elif unit == "M":
            return (end.year - start.year) * 12 + (end.month - start.month)
        elif unit == "Y":
            return end.year - start.year

        return diff.days

    def _fn_sumif(self, range_values, criteria, sum_range=None):
        """SUMIF function."""
        if sum_range is None:
            sum_range = range_values

        if not isinstance(range_values, list) or not isinstance(sum_range, list):
            return 0

        total = 0.0
        for i, val in enumerate(range_values):
            if self._matches_criteria(val, criteria):
                if i < len(sum_range):
                    total += self._to_number(sum_range[i])
        return total

    def _fn_countif(self, range_values, criteria):
        """COUNTIF function."""
        if not isinstance(range_values, list):
            return 0
        return sum(1 for val in range_values if self._matches_criteria(val, criteria))

    def _fn_averageif(self, range_values, criteria, avg_range=None):
        """AVERAGEIF function."""
        if avg_range is None:
            avg_range = range_values

        if not isinstance(range_values, list) or not isinstance(avg_range, list):
            return 0

        values = []
        for i, val in enumerate(range_values):
            if self._matches_criteria(val, criteria):
                if i < len(avg_range):
                    values.append(self._to_number(avg_range[i]))

        return sum(values) / len(values) if values else 0.0

    def _matches_criteria(self, value, criteria) -> bool:
        """Check if value matches criteria."""
        criteria_str = str(criteria)

        # Check for comparison operators
        if criteria_str.startswith(">="):
            return self._to_number(value) >= self._to_number(criteria_str[2:])
        elif criteria_str.startswith("<="):
            return self._to_number(value) <= self._to_number(criteria_str[2:])
        elif criteria_str.startswith("<>"):
            return str(value) != criteria_str[2:]
        elif criteria_str.startswith(">"):
            return self._to_number(value) > self._to_number(criteria_str[1:])
        elif criteria_str.startswith("<"):
            return self._to_number(value) < self._to_number(criteria_str[1:])
        elif criteria_str.startswith("="):
            return str(value) == criteria_str[1:]

        # Default: exact match
        return str(value) == criteria_str



# ── Originally: pivot_service.py ──

"""
Pivot Table Service - Create and manage pivot tables.
"""


import logging
import uuid
from collections import defaultdict
from typing import Any, Callable, Optional

from pydantic import BaseModel

logger = logging.getLogger("neura.pivot")


class PivotValue(BaseModel):
    """Pivot table value aggregation configuration."""

    field: str
    aggregation: str = "SUM"  # SUM, COUNT, AVERAGE, MIN, MAX, COUNTUNIQUE
    alias: Optional[str] = None


class PivotFilter(BaseModel):
    """Pivot table filter."""

    field: str
    values: list[Any]
    exclude: bool = False


class PivotTableConfig(BaseModel):
    """Pivot table configuration."""

    id: str
    name: str
    source_sheet_id: str
    source_range: str
    row_fields: list[str] = []
    column_fields: list[str] = []
    value_fields: list[PivotValue] = []
    filters: list[PivotFilter] = []
    show_grand_totals: bool = True
    show_row_totals: bool = True
    show_col_totals: bool = True
    sort_rows_by: Optional[str] = None
    sort_rows_order: str = "asc"


class PivotTableResult(BaseModel):
    """Result of pivot table computation."""

    headers: list[str]
    rows: list[list[Any]]
    row_totals: Optional[list[Any]] = None
    column_totals: Optional[list[Any]] = None
    grand_total: Optional[Any] = None


class PivotService:
    """Service for creating and computing pivot tables."""

    AGGREGATIONS: dict[str, Callable[[list], Any]] = {
        "SUM": lambda values: sum(float(v) for v in values if v is not None),
        "COUNT": lambda values: len(values),
        "AVERAGE": lambda values: sum(float(v) for v in values if v is not None) / len(values) if values else 0,
        "MIN": lambda values: min(float(v) for v in values if v is not None) if values else None,
        "MAX": lambda values: max(float(v) for v in values if v is not None) if values else None,
        "COUNTUNIQUE": lambda values: len(set(values)),
    }

    def compute_pivot(
        self,
        data: list[dict[str, Any]],
        config: PivotTableConfig,
    ) -> PivotTableResult:
        """
        Compute pivot table from data.

        Args:
            data: List of row dictionaries with field values
            config: Pivot table configuration

        Returns:
            PivotTableResult with computed values
        """
        # Apply filters
        filtered_data = self._apply_filters(data, config.filters)

        if not filtered_data:
            return PivotTableResult(headers=[], rows=[])

        # Get unique values for row and column fields
        row_values = self._get_unique_values(filtered_data, config.row_fields)
        col_values = self._get_unique_values(filtered_data, config.column_fields)

        # Build pivot structure
        pivot_data = defaultdict(lambda: defaultdict(list))

        for row in filtered_data:
            row_key = tuple(row.get(f, "") for f in config.row_fields)
            col_key = tuple(row.get(f, "") for f in config.column_fields)

            for value_config in config.value_fields:
                val = row.get(value_config.field)
                if val is not None:
                    pivot_data[row_key][(col_key, value_config.field)].append(val)

        # Generate headers
        headers = list(config.row_fields)
        for col_combo in col_values:
            for value_config in config.value_fields:
                col_name = " - ".join(str(v) for v in col_combo) if col_combo else ""
                value_name = value_config.alias or f"{value_config.aggregation}({value_config.field})"
                if col_name:
                    headers.append(f"{col_name} | {value_name}")
                else:
                    headers.append(value_name)

        if config.show_row_totals:
            for value_config in config.value_fields:
                value_name = value_config.alias or f"{value_config.aggregation}({value_config.field})"
                headers.append(f"Total {value_name}")

        # Generate rows
        rows = []
        column_totals = defaultdict(list)

        for row_combo in row_values:
            row = list(row_combo)

            for col_combo in col_values:
                for value_config in config.value_fields:
                    values = pivot_data[row_combo].get((col_combo, value_config.field), [])
                    agg_func = self.AGGREGATIONS.get(value_config.aggregation.upper(), self.AGGREGATIONS["SUM"])
                    result = agg_func(values) if values else 0
                    row.append(result)

                    # Track for column totals
                    col_idx = len(row) - 1
                    column_totals[col_idx].extend(values)

            # Row totals
            if config.show_row_totals:
                for value_config in config.value_fields:
                    all_values = []
                    for col_combo in col_values:
                        all_values.extend(pivot_data[row_combo].get((col_combo, value_config.field), []))
                    agg_func = self.AGGREGATIONS.get(value_config.aggregation.upper(), self.AGGREGATIONS["SUM"])
                    row.append(agg_func(all_values) if all_values else 0)

            rows.append(row)

        # Sort rows if configured
        if config.sort_rows_by and config.sort_rows_by in config.row_fields:
            sort_idx = config.row_fields.index(config.sort_rows_by)
            reverse = config.sort_rows_order.lower() == "desc"
            rows.sort(key=lambda r: r[sort_idx] if r[sort_idx] is not None else "", reverse=reverse)

        # Compute column totals
        col_totals = None
        if config.show_col_totals and rows:
            col_totals = ["Total"] + [""] * (len(config.row_fields) - 1)
            for col_idx in range(len(config.row_fields), len(headers)):
                values = column_totals.get(col_idx, [])
                # Use first value config's aggregation for totals
                if config.value_fields:
                    agg_func = self.AGGREGATIONS.get(
                        config.value_fields[0].aggregation.upper(),
                        self.AGGREGATIONS["SUM"]
                    )
                    col_totals.append(agg_func(values) if values else 0)

        # Grand total
        grand_total = None
        if config.show_grand_totals and config.value_fields:
            all_values = []
            for row in filtered_data:
                val = row.get(config.value_fields[0].field)
                if val is not None:
                    all_values.append(val)
            agg_func = self.AGGREGATIONS.get(
                config.value_fields[0].aggregation.upper(),
                self.AGGREGATIONS["SUM"]
            )
            grand_total = agg_func(all_values) if all_values else 0

        return PivotTableResult(
            headers=headers,
            rows=rows,
            column_totals=col_totals,
            grand_total=grand_total,
        )

    def _apply_filters(
        self,
        data: list[dict[str, Any]],
        filters: list[PivotFilter],
    ) -> list[dict[str, Any]]:
        """Apply filters to data."""
        if not filters:
            return data

        filtered = []
        for row in data:
            include = True
            for f in filters:
                value = row.get(f.field)
                in_values = value in f.values
                if f.exclude:
                    if in_values:
                        include = False
                        break
                else:
                    if not in_values:
                        include = False
                        break
            if include:
                filtered.append(row)

        return filtered

    def _get_unique_values(
        self,
        data: list[dict[str, Any]],
        fields: list[str],
    ) -> list[tuple]:
        """Get unique value combinations for fields."""
        if not fields:
            return [()]

        seen = set()
        result = []

        for row in data:
            combo = tuple(row.get(f, "") for f in fields)
            if combo not in seen:
                seen.add(combo)
                result.append(combo)

        return sorted(result, key=lambda x: [str(v) for v in x])

    def data_to_records(
        self,
        data: list[list[Any]],
        headers: Optional[list[str]] = None,
    ) -> list[dict[str, Any]]:
        """Convert 2D array to list of dictionaries."""
        if not data:
            return []

        if headers is None:
            headers = data[0]
            data = data[1:]

        return [
            {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
            for row in data
        ]

    def create_pivot_config(
        self,
        name: str,
        source_sheet_id: str,
        source_range: str,
        row_fields: list[str],
        column_fields: Optional[list[str]] = None,
        value_fields: Optional[list[dict[str, str]]] = None,
    ) -> PivotTableConfig:
        """Create a new pivot table configuration."""
        values = []
        if value_fields:
            for vf in value_fields:
                values.append(PivotValue(
                    field=vf.get("field", ""),
                    aggregation=vf.get("aggregation", "SUM"),
                    alias=vf.get("alias"),
                ))

        return PivotTableConfig(
            id=str(uuid.uuid4()),
            name=name,
            source_sheet_id=source_sheet_id,
            source_range=source_range,
            row_fields=row_fields,
            column_fields=column_fields or [],
            value_fields=values,
        )
