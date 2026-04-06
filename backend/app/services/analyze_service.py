from __future__ import annotations
"""Document analysis — extraction, engines, orchestration."""


# EXTRACTION

# mypy: ignore-errors
"""
Enhanced Extraction Service - Intelligent data extraction from documents.

Features:
1.1 Smart Table Detection & Normalization
1.2 Entity & Metric Extraction
1.3 Form & Invoice Intelligence
"""

import asyncio
import contextlib
import json
import logging
import math
import os
import re
import threading
import time
import uuid
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import Any, AsyncGenerator, Dict, List, Optional

from backend.app.schemas import (
    EnhancedExtractedTable,
    EntityType,
    ExtractedContract,
    ExtractedEntity,
    ExtractedInvoice,
    ExtractedMetric,
    FormField,
    InvoiceLineItem,
    MetricType,
    TableRelationship,
)
from backend.app.services.infra_services import (
    call_chat_completion,
    extract_json_from_llm_response,
    extract_json_array_from_llm_response,
)
from backend.app.services.llm import get_llm_client

logger = logging.getLogger("neura.analyze.extraction")


# ENTITY EXTRACTION

ENTITY_PATTERNS = {
    EntityType.EMAIL: r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b',
    EntityType.PHONE: r'\b(?:\+?1[-.\s]?)?\(?[0-9]{3}\)?[-.\s]?[0-9]{3}[-.\s]?[0-9]{4}\b',
    EntityType.URL: r'https?://(?:[-\w.]|(?:%[\da-fA-F]{2}))+[/\w\-.?=%&]*',
    EntityType.PERCENTAGE: r'\b\d+(?:\.\d+)?%\b',
    EntityType.MONEY: r'(?:\$|€|£|¥|USD|EUR|GBP)\s*\d{1,3}(?:,\d{3})*(?:\.\d{2})?\b|\b\d{1,3}(?:,\d{3})*(?:\.\d{2})?\s*(?:dollars?|euros?|pounds?|USD|EUR|GBP)\b',
    EntityType.DATE: r'\b(?:\d{1,2}[-/]\d{1,2}[-/]\d{2,4}|\d{4}[-/]\d{1,2}[-/]\d{1,2}|(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+\d{1,2},?\s+\d{4}|Q[1-4]\s*\d{4}|FY\d{2,4})\b',
}


def extract_entities_regex(text: str) -> List[ExtractedEntity]:
    """Extract entities using regex patterns."""
    entities = []
    seen = set()

    for entity_type, pattern in ENTITY_PATTERNS.items():
        for match in re.finditer(pattern, text, re.IGNORECASE):
            value = match.group().strip()
            key = f"{entity_type}:{value.lower()}"

            if key not in seen:
                seen.add(key)
                entities.append(ExtractedEntity(
                    id=f"ent_{uuid.uuid4().hex[:8]}",
                    type=entity_type,
                    value=value,
                    confidence=0.9,
                    position={"start": match.start(), "end": match.end()},
                    context=text[max(0, match.start() - 50):match.end() + 50],
                ))

    return entities


def extract_entities_llm(text: str, client: Any) -> List[ExtractedEntity]:
    """Extract entities using LLM for named entity recognition."""
    prompt = f"""Extract all named entities from the following text. Identify:
- PERSON: Names of people
- ORGANIZATION: Company names, institutions
- LOCATION: Cities, countries, addresses
- PRODUCT: Product or service names
- DATE: Dates, time periods (normalize to ISO format when possible)
- MONEY: Currency amounts (normalize to number + currency code)
- PERCENTAGE: Percentage values

Text:
{text[:8000]}

Return JSON array:
```json
[
  {{"type": "PERSON", "value": "John Smith", "normalized": "John Smith", "context": "CEO John Smith announced..."}},
  {{"type": "MONEY", "value": "$1.5M", "normalized": 1500000, "currency": "USD", "context": "revenue of $1.5M"}},
  {{"type": "DATE", "value": "Q3 2025", "normalized": "2025-07-01/2025-09-30", "context": "in Q3 2025"}}
]
```

Extract ALL entities found. Be thorough."""

    try:
        response = call_chat_completion(
            client,
            model=None,
            messages=[{"role": "user", "content": prompt}],
            description="entity_extraction",
            temperature=0.1,
        )

        raw_text = response.choices[0].message.content or ""
        data = extract_json_array_from_llm_response(raw_text, default=[])
        if data:
            entities = []
            for item in data:
                entity_type = item.get("type", "").upper()
                try:
                    etype = EntityType[entity_type]
                except KeyError:
                    etype = EntityType.CUSTOM

                entities.append(ExtractedEntity(
                    id=f"ent_{uuid.uuid4().hex[:8]}",
                    type=etype,
                    value=item.get("value", ""),
                    normalized_value=str(item.get("normalized", "")),
                    confidence=0.85,
                    context=item.get("context"),
                    metadata={"currency": item.get("currency")} if item.get("currency") else {},
                ))
            return entities
    except Exception as e:
        logger.warning(f"LLM entity extraction failed: {e}")

    return []


def extract_all_entities(text: str, use_llm: bool = True) -> List[ExtractedEntity]:
    """Extract entities using both regex and LLM."""
    # Start with regex extraction (fast, high precision)
    entities = extract_entities_regex(text)
    entity_values = {e.value.lower() for e in entities}

    # Add LLM extraction for semantic entities
    if use_llm:
        try:
            client = get_llm_client()
            llm_entities = extract_entities_llm(text, client)

            # Merge, avoiding duplicates
            for ent in llm_entities:
                if ent.value.lower() not in entity_values:
                    entities.append(ent)
                    entity_values.add(ent.value.lower())
        except Exception as e:
            logger.warning(f"LLM extraction skipped: {e}")

    return entities


# METRIC EXTRACTION

def extract_metrics_llm(text: str, tables: List[EnhancedExtractedTable]) -> List[ExtractedMetric]:
    """Extract key metrics and KPIs using LLM."""
    # Build context from tables
    table_context = ""
    for table in tables[:5]:
        table_context += f"\nTable: {table.title or table.id}\n"
        table_context += f"Headers: {', '.join(table.headers[:10])}\n"
        if table.rows:
            table_context += f"Sample row: {table.rows[0][:10]}\n"

    prompt = f"""Analyze this document and extract ALL key metrics, KPIs, and important numerical data.

Text excerpt:
{text[:6000]}

Tables found:
{table_context}

Extract metrics with context. Return JSON array:
```json
[
  {{
    "name": "Revenue",
    "value": 1500000,
    "raw_value": "$1.5M",
    "metric_type": "currency",
    "unit": null,
    "currency": "USD",
    "period": "Q3 2025",
    "change": 15.5,
    "change_direction": "increase",
    "comparison_base": "vs Q3 2024",
    "context": "Revenue reached $1.5M in Q3 2025, up 15.5% YoY",
    "importance": 0.9
  }},
  {{
    "name": "Customer Count",
    "value": 50000,
    "raw_value": "50,000",
    "metric_type": "count",
    "period": "2025",
    "importance": 0.7
  }}
]
```

Metric types: currency, percentage, count, ratio, duration, quantity, score, rate
Extract ALL significant numbers with their context. Focus on KPIs."""

    try:
        client = get_llm_client()
        response = call_chat_completion(
            client,
            model=None,
            messages=[{"role": "user", "content": prompt}],
            description="metric_extraction",
            temperature=0.1,
        )

        raw_text = response.choices[0].message.content or ""
        data = extract_json_array_from_llm_response(raw_text, default=[])
        if data:
            metrics = []
            for item in data:
                try:
                    mtype = MetricType[item.get("metric_type", "count").upper()]
                except KeyError:
                    mtype = MetricType.COUNT

                raw_val = item.get("value")
                if raw_val is None:
                    raw_val = item.get("raw_value", "0")
                metrics.append(ExtractedMetric(
                    id=f"met_{uuid.uuid4().hex[:8]}",
                    name=item.get("name", "Unknown"),
                    value=raw_val if raw_val is not None else "0",
                    raw_value=str(item.get("raw_value", "")),
                    metric_type=mtype,
                    unit=item.get("unit"),
                    currency=item.get("currency"),
                    period=item.get("period"),
                    change=item.get("change"),
                    change_direction=item.get("change_direction"),
                    comparison_base=item.get("comparison_base"),
                    confidence=0.85,
                    context=item.get("context"),
                    importance_score=float(item.get("importance", 0.5)),
                ))
            return metrics
    except Exception as e:
        logger.warning(f"Metric extraction failed: {e}")

    return []


# FORM FIELD EXTRACTION

def extract_form_fields(text: str, tables: List[EnhancedExtractedTable]) -> List[FormField]:
    """Extract form fields from document."""
    prompt = f"""Analyze this document and identify if it contains a form. If so, extract all form fields.

Text:
{text[:6000]}

If this is a form, return JSON:
```json
{{
  "is_form": true,
  "form_title": "Application Form",
  "fields": [
    {{
      "label": "Full Name",
      "value": "John Doe",
      "type": "text",
      "required": true,
      "section": "Personal Information"
    }},
    {{
      "label": "Date of Birth",
      "value": "1990-05-15",
      "type": "date",
      "required": true
    }},
    {{
      "label": "Agree to Terms",
      "value": "checked",
      "type": "checkbox",
      "required": true
    }}
  ]
}}
```

Field types: text, checkbox, radio, date, signature, dropdown, number, email, phone
If not a form, return {{"is_form": false, "fields": []}}"""

    try:
        client = get_llm_client()
        response = call_chat_completion(
            client,
            model=None,
            messages=[{"role": "user", "content": prompt}],
            description="form_extraction",
            temperature=0.1,
        )

        raw_text = response.choices[0].message.content or ""
        data = extract_json_from_llm_response(raw_text, default={})
        if data:
            if data.get("is_form"):
                fields = []
                for item in data.get("fields", []):
                    fields.append(FormField(
                        id=f"field_{uuid.uuid4().hex[:8]}",
                        label=item.get("label", ""),
                        value=item.get("value"),
                        field_type=item.get("type", "text"),
                        required=item.get("required", False),
                        section=item.get("section"),
                        is_filled=bool(item.get("value")),
                        confidence=0.85,
                    ))
                return fields
    except Exception as e:
        logger.warning(f"Form extraction failed: {e}")

    return []


# INVOICE EXTRACTION

def extract_invoice(text: str, tables: List[EnhancedExtractedTable]) -> Optional[ExtractedInvoice]:
    """Extract invoice data from document."""
    # Build table context
    table_context = ""
    for table in tables[:3]:
        table_context += f"\nTable: {table.title or 'Untitled'}\n"
        for row in table.rows[:10]:
            table_context += f"  {row}\n"

    prompt = f"""Analyze this document and determine if it's an invoice. If so, extract all invoice data.

Text:
{text[:5000]}

Tables:
{table_context}

If this is an invoice, return JSON:
```json
{{
  "is_invoice": true,
  "vendor_name": "Acme Corp",
  "vendor_address": "123 Main St, City, ST 12345",
  "vendor_tax_id": "12-3456789",
  "customer_name": "Client Inc",
  "customer_address": "456 Oak Ave",
  "invoice_number": "INV-2025-001",
  "invoice_date": "2025-01-15",
  "due_date": "2025-02-15",
  "purchase_order": "PO-12345",
  "line_items": [
    {{
      "description": "Consulting Services",
      "quantity": 10,
      "unit_price": 150.00,
      "total": 1500.00,
      "tax": 120.00
    }}
  ],
  "subtotal": 1500.00,
  "tax_total": 120.00,
  "discount_total": 0,
  "grand_total": 1620.00,
  "currency": "USD",
  "payment_terms": "Net 30",
  "notes": "Thank you for your business"
}}
```

If not an invoice, return {{"is_invoice": false}}"""

    try:
        client = get_llm_client()
        response = call_chat_completion(
            client,
            model=None,
            messages=[{"role": "user", "content": prompt}],
            description="invoice_extraction",
            temperature=0.1,
        )

        raw_text = response.choices[0].message.content or ""
        data = extract_json_from_llm_response(raw_text, default={})
        if data:
            if data.get("is_invoice"):
                line_items = []
                for item in data.get("line_items", []):
                    line_items.append(InvoiceLineItem(
                        id=f"li_{uuid.uuid4().hex[:8]}",
                        description=item.get("description", ""),
                        quantity=item.get("quantity"),
                        unit_price=item.get("unit_price"),
                        total=item.get("total"),
                        tax=item.get("tax"),
                        discount=item.get("discount"),
                        sku=item.get("sku"),
                        category=item.get("category"),
                    ))

                return ExtractedInvoice(
                    id=f"inv_{uuid.uuid4().hex[:8]}",
                    vendor_name=data.get("vendor_name"),
                    vendor_address=data.get("vendor_address"),
                    vendor_tax_id=data.get("vendor_tax_id"),
                    customer_name=data.get("customer_name"),
                    customer_address=data.get("customer_address"),
                    invoice_number=data.get("invoice_number"),
                    invoice_date=data.get("invoice_date"),
                    due_date=data.get("due_date"),
                    purchase_order=data.get("purchase_order"),
                    line_items=line_items,
                    subtotal=data.get("subtotal"),
                    tax_total=data.get("tax_total"),
                    discount_total=data.get("discount_total"),
                    grand_total=data.get("grand_total"),
                    currency=data.get("currency", "USD"),
                    payment_terms=data.get("payment_terms"),
                    notes=data.get("notes"),
                    confidence=0.85,
                )
    except Exception as e:
        logger.warning(f"Invoice extraction failed: {e}")

    return None


# CONTRACT EXTRACTION

def extract_contract(text: str) -> Optional[ExtractedContract]:
    """Extract contract data from document."""
    prompt = f"""Analyze this document and determine if it's a contract or legal agreement. If so, extract key information.

Text:
{text[:8000]}

If this is a contract, return JSON:
```json
{{
  "is_contract": true,
  "contract_type": "Service Agreement",
  "parties": [
    {{"name": "Company A", "role": "Provider"}},
    {{"name": "Company B", "role": "Client"}}
  ],
  "effective_date": "2025-01-01",
  "expiration_date": "2026-01-01",
  "auto_renewal": true,
  "renewal_terms": "Automatically renews for 1-year periods",
  "key_terms": [
    "Monthly payment of $5,000",
    "30-day termination notice required"
  ],
  "obligations": [
    {{"party": "Provider", "obligation": "Deliver services monthly"}},
    {{"party": "Client", "obligation": "Pay within 30 days"}}
  ],
  "termination_clauses": [
    "Either party may terminate with 30 days written notice",
    "Immediate termination for material breach"
  ],
  "governing_law": "State of California",
  "signatures": [
    {{"name": "John Doe", "title": "CEO", "date": "2025-01-01", "signed": true}}
  ]
}}
```

If not a contract, return {{"is_contract": false}}"""

    try:
        client = get_llm_client()
        response = call_chat_completion(
            client,
            model=None,
            messages=[{"role": "user", "content": prompt}],
            description="contract_extraction",
            temperature=0.1,
        )

        raw_text = response.choices[0].message.content or ""
        data = extract_json_from_llm_response(raw_text, default={})
        if data:
            if data.get("is_contract"):
                return ExtractedContract(
                    id=f"con_{uuid.uuid4().hex[:8]}",
                    contract_type=data.get("contract_type"),
                    parties=data.get("parties", []),
                    effective_date=data.get("effective_date"),
                    expiration_date=data.get("expiration_date"),
                    auto_renewal=data.get("auto_renewal", False),
                    renewal_terms=data.get("renewal_terms"),
                    key_terms=data.get("key_terms", []),
                    obligations=data.get("obligations", []),
                    termination_clauses=data.get("termination_clauses", []),
                    governing_law=data.get("governing_law"),
                    signatures=data.get("signatures", []),
                    confidence=0.8,
                )
    except Exception as e:
        logger.warning(f"Contract extraction failed: {e}")

    return None


# TABLE ENHANCEMENT

def enhance_table(table: Dict[str, Any], all_tables: List[Dict[str, Any]]) -> EnhancedExtractedTable:
    """Enhance a table with additional metadata and analysis."""
    headers = table.get("headers", [])
    rows = table.get("rows", [])

    # Infer data types
    data_types = []
    for col_idx in range(len(headers)):
        col_values = [row[col_idx] for row in rows if col_idx < len(row)]
        data_types.append(_infer_column_type(col_values))

    # Calculate statistics for numeric columns
    statistics = {}
    for col_idx, dtype in enumerate(data_types):
        if dtype == "numeric":
            values = []
            for row in rows:
                if col_idx < len(row):
                    try:
                        val = float(str(row[col_idx]).replace(",", "").replace("$", "").replace("%", ""))
                        values.append(val)
                    except (ValueError, TypeError):
                        pass

            if values:
                statistics[headers[col_idx]] = {
                    "min": min(values),
                    "max": max(values),
                    "mean": sum(values) / len(values),
                    "count": len(values),
                }

    # Check for totals row
    has_totals = False
    if rows:
        last_row = rows[-1]
        if any("total" in str(cell).lower() for cell in last_row):
            has_totals = True

    return EnhancedExtractedTable(
        id=table.get("id", f"table_{uuid.uuid4().hex[:8]}"),
        title=table.get("title"),
        headers=headers,
        rows=rows,
        data_types=data_types,
        source_page=table.get("source_page"),
        source_sheet=table.get("source_sheet"),
        confidence=table.get("confidence", 0.9),
        row_count=len(rows),
        column_count=len(headers),
        has_totals_row=has_totals,
        has_header_row=True,
        statistics=statistics,
    )


def detect_table_relationships(tables: List[EnhancedExtractedTable]) -> List[TableRelationship]:
    """Detect relationships between tables (e.g., continuation across pages)."""
    relationships = []

    for i, table1 in enumerate(tables):
        for j, table2 in enumerate(tables):
            if i >= j:
                continue

            # Check for continuation (same headers)
            if table1.headers == table2.headers:
                # Check if on consecutive pages
                if (table1.source_page and table2.source_page and
                        abs(table1.source_page - table2.source_page) == 1):
                    relationships.append(TableRelationship(
                        table1_id=table1.id,
                        table2_id=table2.id,
                        relationship_type="continuation",
                        confidence=0.9,
                    ))
            # Check for related tables (shared columns)
            elif set(table1.headers) & set(table2.headers):
                shared = len(set(table1.headers) & set(table2.headers))
                total = len(set(table1.headers) | set(table2.headers))
                if shared / total > 0.3:
                    relationships.append(TableRelationship(
                        table1_id=table1.id,
                        table2_id=table2.id,
                        relationship_type="related",
                        confidence=shared / total,
                    ))

    return relationships


def stitch_continuation_tables(
    tables: List[EnhancedExtractedTable],
    relationships: List[TableRelationship]
) -> List[EnhancedExtractedTable]:
    """Merge tables that are continuations of each other."""
    continuations = {r.table1_id: r.table2_id for r in relationships if r.relationship_type == "continuation"}

    if not continuations:
        return tables

    merged_ids = set()
    result = []

    for table in tables:
        if table.id in merged_ids:
            continue

        # Find all continuations
        current_id = table.id
        all_rows = list(table.rows)

        while current_id in continuations:
            next_id = continuations[current_id]
            merged_ids.add(next_id)

            # Find the continuation table
            for t in tables:
                if t.id == next_id:
                    all_rows.extend(t.rows)
                    break

            current_id = next_id

        # Create merged table
        merged = EnhancedExtractedTable(
            id=table.id,
            title=table.title,
            headers=table.headers,
            rows=all_rows,
            data_types=table.data_types,
            source_page=table.source_page,
            confidence=table.confidence,
            row_count=len(all_rows),
            column_count=len(table.headers),
            has_totals_row=table.has_totals_row,
            has_header_row=True,
            statistics=table.statistics,
            related_tables=list(merged_ids) if merged_ids else [],
        )
        result.append(merged)

    return result


def _infer_column_type(values: List[Any]) -> str:
    """Infer the data type of a column."""
    if not values:
        return "text"

    numeric_count = 0
    date_count = 0
    total_valid = 0

    date_patterns = [
        r'^\d{4}-\d{2}-\d{2}$',
        r'^\d{1,2}/\d{1,2}/\d{2,4}$',
        r'^\d{1,2}-\d{1,2}-\d{2,4}$',
    ]

    for value in values:
        value_str = str(value).strip()
        if not value_str:
            continue

        total_valid += 1

        # Check numeric
        try:
            cleaned = re.sub(r'[$,% ]', '', value_str)
            float(cleaned)
            numeric_count += 1
            continue
        except (ValueError, TypeError):
            pass

        # Check date
        for pattern in date_patterns:
            if re.match(pattern, value_str):
                date_count += 1
                break

    if total_valid == 0:
        return "text"

    if numeric_count / total_valid >= 0.7:
        return "numeric"
    if date_count / total_valid >= 0.7:
        return "datetime"

    return "text"


# MAIN EXTRACTION ORCHESTRATOR

class EnhancedExtractionService:
    """Orchestrates all intelligent extraction operations."""

    def __init__(self, use_llm: bool = True):
        self.use_llm = use_llm
        self._client = None

    @property
    def client(self):
        if self._client is None:
            self._client = get_llm_client()
        return self._client

    def extract_all(
        self,
        text: str,
        raw_tables: List[Dict[str, Any]],
    ) -> Dict[str, Any]:
        """
        Perform all extraction operations.

        Uses regex for fast entity extraction + ONE consolidated LLM call
        for semantic extraction (entities, metrics) instead of 5 separate calls.
        """
        # Enhance tables (no LLM)
        enhanced_tables = [enhance_table(t, raw_tables) for t in raw_tables]

        # Detect relationships and stitch (no LLM)
        relationships = detect_table_relationships(enhanced_tables)
        stitched_tables = stitch_continuation_tables(enhanced_tables, relationships)

        # Fast regex-based entity extraction (no LLM)
        entities = extract_entities_regex(text)

        metrics: List[ExtractedMetric] = []
        forms: List[FormField] = []
        invoices: List[ExtractedInvoice] = []
        contracts: List[ExtractedContract] = []

        if self.use_llm:
            try:
                llm_result = self._extract_with_llm(text, stitched_tables)

                # Merge LLM entities with regex entities (dedup)
                entity_values = {e.value.lower() for e in entities}
                for ent in llm_result.get("entities", []):
                    if ent.value.lower() not in entity_values:
                        entities.append(ent)
                        entity_values.add(ent.value.lower())

                metrics = llm_result.get("metrics", [])
                forms = llm_result.get("forms", [])
                invoices = llm_result.get("invoices", [])
                contracts = llm_result.get("contracts", [])

            except Exception as e:
                logger.error("LLM extraction failed: %s", e, exc_info=True)
                raise RuntimeError(f"Intelligent extraction failed: {e}") from e

        return {
            "tables": stitched_tables,
            "table_relationships": relationships,
            "entities": entities,
            "metrics": metrics,
            "forms": forms,
            "invoices": invoices,
            "contracts": contracts,
        }

    def _extract_with_llm(
        self,
        text: str,
        tables: List[EnhancedExtractedTable],
    ) -> Dict[str, Any]:
        """
        Single consolidated LLM call for semantic extraction.

        Combines entity NER + metric extraction + document type detection.
        """
        table_context = ""
        for table in tables[:5]:
            table_context += f"- {table.title or table.id}: columns={', '.join(table.headers[:8])}"
            if table.rows:
                table_context += f", sample={table.rows[0][:6]}"
            table_context += "\n"
        table_context = table_context or "(no tables)"

        prompt = f"""Extract structured data from this document in ONE JSON response.

DOCUMENT TEXT:
{text[:8000]}

TABLES:
{table_context}

Return a JSON object:
{{
  "entities": [
    {{"type": "PERSON|ORGANIZATION|LOCATION|PRODUCT|DATE|MONEY|PERCENTAGE", "value": "...", "context": "surrounding text"}}
  ],
  "metrics": [
    {{
      "name": "Metric name",
      "value": 1500000,
      "raw_value": "$1.5M",
      "metric_type": "currency|percentage|count|ratio|duration|other",
      "currency": "USD",
      "period": "Q3 2025",
      "change": 15.5,
      "change_direction": "increase|decrease|stable",
      "context": "Brief context sentence",
      "importance": 0.9
    }}
  ],
  "document_type": "invoice|contract|form|report|letter|other"
}}

RULES:
- Extract ALL named entities (people, companies, locations, products)
- Extract ALL key metrics, KPIs, and numerical data with context
- Be thorough - don't miss important data points
- Return ONLY valid JSON"""

        client = get_llm_client()
        response = call_chat_completion(
            client,
            model=None,
            messages=[{"role": "user", "content": prompt}],
            description="consolidated_extraction",
            temperature=0.1,
        )

        raw_text = response.choices[0].message.content or ""
        data = extract_json_from_llm_response(raw_text, default={})

        result: Dict[str, Any] = {"entities": [], "metrics": [], "forms": [],
                                   "invoices": [], "contracts": []}

        # Parse entities
        for item in data.get("entities", []):
            if isinstance(item, dict):
                entity_type = str(item.get("type", "")).upper()
                try:
                    etype = EntityType[entity_type]
                except KeyError:
                    etype = EntityType.CUSTOM
                result["entities"].append(ExtractedEntity(
                    id=f"ent_{uuid.uuid4().hex[:8]}",
                    type=etype,
                    value=item.get("value", ""),
                    confidence=0.85,
                    context=item.get("context"),
                ))

        # Parse metrics
        for item in data.get("metrics", []):
            if isinstance(item, dict):
                mtype_map = {
                    "currency": MetricType.CURRENCY, "percentage": MetricType.PERCENTAGE,
                    "count": MetricType.COUNT, "ratio": MetricType.RATIO,
                    "duration": MetricType.DURATION, "quantity": MetricType.QUANTITY,
                    "score": MetricType.SCORE, "rate": MetricType.RATE,
                }
                raw_val = item.get("value")
                if raw_val is None:
                    raw_val = item.get("raw_value", "0")
                result["metrics"].append(ExtractedMetric(
                    id=f"met_{uuid.uuid4().hex[:8]}",
                    name=item.get("name", ""),
                    value=raw_val if raw_val is not None else "0",
                    raw_value=str(item.get("raw_value", "")),
                    metric_type=mtype_map.get(str(item.get("metric_type", "count")).lower(), MetricType.COUNT),
                    currency=item.get("currency"),
                    period=item.get("period"),
                    change=item.get("change"),
                    change_direction=item.get("change_direction"),
                    context=item.get("context"),
                    importance_score=float(item.get("importance", 0.5)),
                ))

        return result


# EXTRACTION PIPELINE (merged from extraction_pipeline.py)

logger = logging.getLogger("neura.analyze.extraction")

try:
    import fitz  # PyMuPDF
except ImportError:  # pragma: no cover
    fitz = None  # type: ignore

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None  # type: ignore

_DEFAULT_MAX_PDF_PAGES = int(os.getenv("NEURA_ANALYSIS_PDF_MAX_PAGES", "50"))
_DEFAULT_MAX_PDF_SECONDS = int(os.getenv("NEURA_ANALYSIS_PDF_MAX_SECONDS", "20"))
_DEFAULT_MAX_EXCEL_ROWS = int(os.getenv("NEURA_ANALYSIS_EXCEL_MAX_ROWS", "500"))
_DEFAULT_MAX_EXCEL_SHEETS = int(os.getenv("NEURA_ANALYSIS_EXCEL_MAX_SHEETS", "10"))


@dataclass
class ExtractedContent:
    """Content extracted from a document before LLM processing."""

    document_type: str  # "pdf" | "excel"
    file_name: str
    page_count: int = 1
    text_content: str = ""
    tables_raw: list[dict[str, Any]] = field(default_factory=list)
    images: list[bytes] = field(default_factory=list)
    sheets: list[dict[str, Any]] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


def _stringify_cell(value: object) -> str:
    """Convert cell value to string."""
    if value is None:
        return ""
    try:
        text = str(value)
    except Exception as exc:
        logger.debug(f"Failed to stringify cell value: {exc}")
        return ""
    return text.strip()


def _row_has_values(values) -> bool:
    """Check if a row has any non-empty values."""
    for value in values:
        if value is None:
            continue
        if isinstance(value, str):
            if value.strip():
                return True
            continue
        return True
    return False


def _ensure_label(value: object, idx: int) -> str:
    """Ensure a column has a label."""
    if value not in (None, ""):
        text = str(value).strip()
        if text:
            return text
    return f"Column {idx + 1}"


def _infer_data_type_from_values(values: list[str]) -> str:
    """Infer data type from a list of string values."""
    if not values:
        return "text"

    date_patterns = [
        r"^\d{4}-\d{2}-\d{2}",
        r"^\d{2}/\d{2}/\d{4}",
        r"^\d{2}-\d{2}-\d{4}",
        r"^\d{1,2}/\d{1,2}/\d{2,4}",
    ]

    numeric_count = 0
    date_count = 0
    total_valid = 0

    for val in values[:30]:
        if not val or not val.strip():
            continue
        total_valid += 1
        str_val = val.strip()

        for pattern in date_patterns:
            if re.match(pattern, str_val):
                date_count += 1
                break
        else:
            try:
                cleaned = str_val.replace(",", "").replace("$", "").replace("%", "").replace(" ", "")
                float(cleaned)
                numeric_count += 1
            except (ValueError, TypeError):
                pass

    if total_valid == 0:
        return "text"
    if date_count >= total_valid * 0.7:
        return "datetime"
    if numeric_count >= total_valid * 0.7:
        return "numeric"
    return "text"


def extract_pdf_content(
    file_path: Path | str,
    file_bytes: bytes | None = None,
    max_pages: int | None = None,
    max_seconds: int | None = None,
) -> ExtractedContent:
    """Extract text, tables, and images from a PDF file."""
    if fitz is None:
        return ExtractedContent(
            document_type="pdf",
            file_name=str(file_path),
            errors=["PyMuPDF (fitz) is not installed. Cannot extract PDF content."],
        )

    file_name = Path(file_path).name if file_path else "document.pdf"

    try:
        if file_bytes:
            doc = fitz.open(stream=file_bytes, filetype="pdf")
        else:
            doc = fitz.open(str(file_path))
    except Exception as exc:
        return ExtractedContent(
            document_type="pdf",
            file_name=file_name,
            errors=[f"Failed to open PDF: {exc}"],
        )

    page_count = len(doc)
    max_pages = max_pages if max_pages is not None else _DEFAULT_MAX_PDF_PAGES
    max_seconds = max_seconds if max_seconds is not None else _DEFAULT_MAX_PDF_SECONDS
    text_parts: list[str] = []
    tables_raw: list[dict[str, Any]] = []
    images: list[bytes] = []
    errors: list[str] = []
    started = time.time()

    if max_pages and page_count > max_pages:
        errors.append(f"PDF has {page_count} pages; processed first {max_pages} pages")

    for page_num, page in enumerate(doc):
        if max_pages and page_num >= max_pages:
            break
        if max_seconds and (time.time() - started) > max_seconds:
            errors.append(f"PDF extraction exceeded time budget of {max_seconds}s; stopping early.")
            break
        try:
            text = page.get_text("text")
            if text.strip():
                text_parts.append(f"--- Page {page_num + 1} ---\n{text}")
        except Exception as exc:
            errors.append(f"Failed to extract text from page {page_num + 1}: {exc}")

        try:
            page_tables = page.find_tables()
            if page_tables and page_tables.tables:
                for table_idx, table in enumerate(page_tables.tables):
                    try:
                        extracted = table.extract()
                        if extracted and len(extracted) > 0:
                            headers = [_stringify_cell(c) or f"Col{i+1}" for i, c in enumerate(extracted[0])]
                            num_cols = len(headers)
                            rows = []
                            for row in extracted[1:]:
                                # Normalize row to match header length
                                normalized_row = [
                                    _stringify_cell(row[i] if i < len(row) else "")
                                    for i in range(num_cols)
                                ]
                                rows.append(normalized_row)

                            col_values: dict[int, list[str]] = {i: [] for i in range(len(headers))}
                            for row in rows[:30]:
                                for i, cell in enumerate(row):
                                    if i < len(headers):
                                        col_values[i].append(cell)

                            data_types = [_infer_data_type_from_values(col_values.get(i, [])) for i in range(len(headers))]

                            tables_raw.append({
                                "id": f"table_p{page_num + 1}_{table_idx + 1}",
                                "headers": headers,
                                "rows": rows,
                                "data_types": data_types,
                                "source_page": page_num + 1,
                            })
                    except Exception as table_exc:
                        errors.append(f"Failed to extract table {table_idx + 1} from page {page_num + 1}: {table_exc}")
        except Exception as exc:
            logger.debug(f"Table extraction not available for page {page_num + 1}: {exc}")

        try:
            image_list = page.get_images(full=True)
            for img_idx, img_info in enumerate(image_list[:3]):
                try:
                    xref = img_info[0]
                    pix = fitz.Pixmap(doc, xref)
                    if pix.n > 4:
                        pix = fitz.Pixmap(fitz.csRGB, pix)
                    img_bytes = pix.tobytes("png")
                    images.append(img_bytes)
                except Exception as img_exc:
                    logger.debug(f"Failed to extract image {img_idx + 1} from page {page_num + 1}: {img_exc}")
        except Exception as img_list_exc:
            logger.debug(f"Failed to get image list from page {page_num + 1}: {img_list_exc}")

    doc.close()

    return ExtractedContent(
        document_type="pdf",
        file_name=file_name,
        page_count=page_count,
        text_content="\n\n".join(text_parts),
        tables_raw=tables_raw,
        images=images[:10],
        errors=errors,
    )


def extract_excel_content(
    file_path: Path | str,
    file_bytes: bytes | None = None,
    max_rows: int | None = None,
    max_sheets: int | None = None,
) -> ExtractedContent:
    """Extract tables and data from an Excel file."""
    if openpyxl is None:
        return ExtractedContent(
            document_type="excel",
            file_name=str(file_path),
            errors=["openpyxl is not installed. Cannot extract Excel content."],
        )

    file_name = Path(file_path).name if file_path else "document.xlsx"

    try:
        if file_bytes:
            wb = openpyxl.load_workbook(filename=BytesIO(file_bytes), data_only=True, read_only=True)
        else:
            wb = openpyxl.load_workbook(filename=str(file_path), data_only=True, read_only=True)
    except Exception as exc:
        return ExtractedContent(
            document_type="excel",
            file_name=file_name,
            errors=[f"Failed to open Excel file: {exc}"],
        )

    sheet_count = len(wb.sheetnames)
    tables_raw: list[dict[str, Any]] = []
    sheets_info: list[dict[str, Any]] = []
    text_parts: list[str] = []
    errors: list[str] = []

    max_sheets = max_sheets if max_sheets is not None else _DEFAULT_MAX_EXCEL_SHEETS
    if sheet_count > max_sheets:
        logger.warning(f"Excel file has {sheet_count} sheets, processing only first {max_sheets}")
        errors.append(f"File has {sheet_count} sheets - only the first {max_sheets} were processed")

    for sheet_idx, sheet_name in enumerate(wb.sheetnames[:max_sheets]):
        try:
            sheet = wb[sheet_name]
            max_rows = max_rows if max_rows is not None else _DEFAULT_MAX_EXCEL_ROWS
            header_row = None
            header_index = -1
            headers: list[str] = []
            data_rows: list[list[str]] = []

            for idx, row in enumerate(sheet.iter_rows(values_only=True)):
                if header_row is None:
                    if _row_has_values(row):
                        header_row = row
                        header_index = idx
                        headers = [_ensure_label(v, i) for i, v in enumerate(header_row)]
                    continue

                if not _row_has_values(row):
                    continue
                if len(data_rows) < max_rows:
                    data_rows.append(
                        [_stringify_cell(row[i] if i < len(row) else "") for i in range(len(headers))]
                    )
                if len(data_rows) >= max_rows:
                    break

            if not data_rows:
                continue

            col_values: dict[int, list[str]] = {i: [] for i in range(len(headers))}
            for row in data_rows[:30]:
                for i, cell in enumerate(row):
                    if i < len(headers):
                        col_values[i].append(cell)

            data_types = [_infer_data_type_from_values(col_values.get(i, [])) for i in range(len(headers))]

            total_rows = max(0, (sheet.max_row or 0) - (header_index + 1))
            truncated = total_rows > max_rows if max_rows else False
            if truncated:
                logger.info(f"Sheet '{sheet_name}' truncated from {total_rows} to {max_rows} rows")

            table_id = f"table_sheet_{sheet_idx + 1}"
            tables_raw.append({
                "id": table_id,
                "title": sheet_name,
                "headers": headers,
                "rows": data_rows[:max_rows],
                "data_types": data_types,
                "source_sheet": sheet_name,
                "truncated": truncated,
                "total_row_count": total_rows,
            })

            sheets_info.append({
                "name": sheet_name,
                "row_count": total_rows,
                "column_count": len(headers),
                "headers": headers,
                "truncated": truncated,
            })

            preview_rows = min(5, len(data_rows))
            text_preview = f"--- Sheet: {sheet_name} ---\n"
            text_preview += f"Headers: {', '.join(headers)}\n"
            text_preview += f"Rows: {total_rows}\n"
            for i in range(preview_rows):
                text_preview += f"Row {i+1}: {', '.join(data_rows[i][:10])}\n"
            text_parts.append(text_preview)

        except Exception as exc:
            errors.append(f"Failed to process sheet '{sheet_name}': {exc}")

    wb.close()

    return ExtractedContent(
        document_type="excel",
        file_name=file_name,
        page_count=sheet_count,
        text_content="\n\n".join(text_parts),
        tables_raw=tables_raw,
        sheets=sheets_info,
        errors=errors,
    )


def extract_document_content(
    file_path: Path | str | None = None,
    file_bytes: bytes | None = None,
    file_name: str | None = None,
    max_pages: int | None = None,
    max_rows: int | None = None,
    max_seconds: int | None = None,
) -> ExtractedContent:
    """Extract content from a document, auto-detecting type from extension."""
    if file_name is None and file_path:
        file_name = Path(file_path).name

    if not file_name:
        file_name = "document"

    ext = Path(file_name).suffix.lower()

    if ext == ".pdf":
        return extract_pdf_content(
            file_path or file_name,
            file_bytes,
            max_pages=max_pages,
            max_seconds=max_seconds,
        )
    elif ext in (".xlsx", ".xls", ".xlsm"):
        return extract_excel_content(file_path or file_name, file_bytes, max_rows=max_rows)
    else:
        return ExtractedContent(
            document_type="unknown",
            file_name=file_name,
            errors=[f"Unsupported file type: {ext}. Only PDF and Excel files are supported."],
        )


def format_content_for_llm(content: ExtractedContent, max_chars: int = 50000) -> str:
    """Format extracted content into a string for LLM processing."""
    parts: list[str] = []

    if content.text_content:
        text_preview = content.text_content[:max_chars // 2]
        parts.append(f"TEXT CONTENT:\n{text_preview}")

    if content.tables_raw:
        parts.append(f"\nEXTRACTED TABLES ({len(content.tables_raw)} found):")
        for table in content.tables_raw[:10]:
            table_str = f"\n[Table: {table.get('id', 'unknown')}]"
            if table.get("title"):
                table_str += f" Title: {table['title']}"
            if table.get("source_page"):
                table_str += f" (Page {table['source_page']})"
            if table.get("source_sheet"):
                table_str += f" (Sheet: {table['source_sheet']})"

            headers = table.get("headers", [])
            table_str += f"\nHeaders: {', '.join(headers)}"

            rows = table.get("rows", [])
            table_str += f"\nRow count: {len(rows)}"

            for i, row in enumerate(rows[:5]):
                row_preview = [str(c)[:50] for c in row[:8]]
                table_str += f"\n  Row {i+1}: {' | '.join(row_preview)}"

            if len(rows) > 5:
                table_str += f"\n  ... and {len(rows) - 5} more rows"

            parts.append(table_str)

    if content.sheets:
        parts.append(f"\nSHEET SUMMARY ({len(content.sheets)} sheets):")
        for sheet in content.sheets:
            parts.append(f"  - {sheet['name']}: {sheet['row_count']} rows, {sheet['column_count']} columns")

    result = "\n".join(parts)
    if len(result) > max_chars:
        result = result[:max_chars] + "\n... (truncated)"

    return result


__all__ = [
    "ExtractedContent",
    "extract_pdf_content",
    "extract_excel_content",
    "extract_document_content",
    "format_content_for_llm",
]


# DOCUMENT ANALYSIS SERVICE (merged from document_analysis_service.py)


from backend.app.services.config import get_settings
from backend.app.schemas import (
    AnalysisResult,
    AnalysisSuggestChartsPayload,
    ExtractedDataPoint,
    ExtractedTable,
    FieldInfo,
    TimeSeriesCandidate,
)
from backend.app.schemas import ChartSpec
from backend.app.services.ai_services import (
    build_analysis_prompt,
    build_chart_suggestion_prompt,
    parse_analysis_response,
    strip_code_fences,
)
from backend.app.utils import write_json_atomic
from backend.app.services.infra_services import call_chat_completion, call_chat_completion_async

logger = logging.getLogger("neura.analyze.service")


@dataclass
class CacheEntry:
    """Cache entry with TTL support."""
    value: AnalysisResult
    created_at: float = field(default_factory=time.time)
    last_accessed: float = field(default_factory=time.time)


class TTLCache:
    """Thread-safe TTL cache with LRU eviction and size limits."""

    def __init__(self, max_items: int = 100, ttl_seconds: int = 3600):
        self.max_items = max_items
        self.ttl_seconds = ttl_seconds
        self._cache: dict[str, CacheEntry] = {}
        self._lock = threading.Lock()
        self._eviction_threshold = max(1, max_items * 8 // 10)  # Evict until at 80% capacity

    def _is_expired(self, entry: CacheEntry) -> bool:
        """Check if cache entry has expired based on TTL from creation time."""
        return time.time() - entry.created_at > self.ttl_seconds

    def _evict_stale(self) -> int:
        """Remove expired entries. Returns number of entries evicted."""
        now = time.time()
        expired_keys = [
            key for key, entry in self._cache.items()
            if now - entry.created_at > self.ttl_seconds
        ]
        for key in expired_keys:
            del self._cache[key]
        if expired_keys:
            logger.debug(f"Evicted {len(expired_keys)} stale cache entries")
        return len(expired_keys)

    def _evict_lru(self) -> int:
        """
        Evict least recently used entries if at or over capacity.
        Evicts until cache size is below 80% of max_items.
        Returns number of entries evicted.
        """
        if len(self._cache) < self.max_items:
            return 0

        # Sort by last_accessed (least recent first)
        sorted_keys = sorted(
            self._cache.keys(),
            key=lambda k: self._cache[k].last_accessed,
        )

        # Evict until below threshold
        num_to_evict = len(self._cache) - self._eviction_threshold
        num_to_evict = max(1, num_to_evict)  # At least evict 1

        for key in sorted_keys[:num_to_evict]:
            del self._cache[key]

        logger.debug(f"LRU evicted {num_to_evict} cache entries, size now {len(self._cache)}")
        return num_to_evict

    def get(self, key: str) -> Optional[AnalysisResult]:
        """Get value from cache, returns None if not found or expired."""
        with self._lock:
            entry = self._cache.get(key)
            if entry is None:
                return None
            if self._is_expired(entry):
                del self._cache[key]
                logger.debug(f"Cache entry {key} expired on access")
                return None
            # Update last accessed time
            entry.last_accessed = time.time()
            return entry.value

    def set(self, key: str, value: AnalysisResult) -> None:
        """Set value in cache with automatic eviction."""
        with self._lock:
            # Evict stale entries first
            self._evict_stale()
            # Then evict LRU if still at capacity
            self._evict_lru()
            # Add new entry
            self._cache[key] = CacheEntry(value=value)
            logger.debug(f"Cache set {key}, size now {len(self._cache)}")

    def __contains__(self, key: str) -> bool:
        """Check if key exists and is not expired."""
        return self.get(key) is not None

    def size(self) -> int:
        """Return current cache size."""
        with self._lock:
            return len(self._cache)

    def clear(self) -> None:
        """Clear all cache entries."""
        with self._lock:
            self._cache.clear()
            logger.debug("Cache cleared")


_ANALYSIS_CACHE: TTLCache | None = None
_ANALYSIS_SEMAPHORE: asyncio.Semaphore | None = None


def _analysis_cache() -> TTLCache:
    global _ANALYSIS_CACHE
    if _ANALYSIS_CACHE is None:
        settings = get_settings()
        _ANALYSIS_CACHE = TTLCache(
            max_items=settings.analysis_cache_max_items,
            ttl_seconds=settings.analysis_cache_ttl_seconds,
        )
    return _ANALYSIS_CACHE


def _analysis_size_limits() -> tuple[int, int]:
    override = os.getenv("ANALYZE_MAX_FILE_SIZE_MB")
    if override:
        try:
            mb = int(override)
        except ValueError:
            mb = None
        if mb and mb > 0:
            return mb * 1024 * 1024, mb
    max_bytes = get_settings().max_upload_bytes
    max_mb = max(1, int(max_bytes / (1024 * 1024)))
    return max_bytes, max_mb


def _analysis_persist_ttl_seconds() -> Optional[int]:
    raw = os.getenv("NEURA_ANALYSIS_PERSIST_TTL_SECONDS")
    if raw is None or raw == "":
        return None
    try:
        value = int(raw)
    except (TypeError, ValueError):
        return None
    return value if value > 0 else None


def _analysis_store_dir() -> Path:
    base = get_settings().state_dir / "analysis_cache"
    base.mkdir(parents=True, exist_ok=True)
    return base


def _analysis_store_path(analysis_id: str) -> Path:
    safe_id = str(analysis_id or "").strip()
    return _analysis_store_dir() / f"{safe_id}.json"


def _parse_analysis_result(payload: dict[str, Any]) -> AnalysisResult:
    if hasattr(AnalysisResult, "model_validate"):
        return AnalysisResult.model_validate(payload)
    return AnalysisResult.parse_obj(payload)


def _persist_analysis_result(result: AnalysisResult) -> None:
    try:
        path = _analysis_store_path(result.analysis_id)
        write_json_atomic(
            path,
            {
                "analysis_id": result.analysis_id,
                "created_at": time.time(),
                "result": result.model_dump(),
            },
            ensure_ascii=False,
            indent=2,
            step="analysis_store",
        )
    except Exception as exc:
        logger.warning("analysis_persist_failed", extra={"event": "analysis_persist_failed", "error": str(exc)})


def _load_persisted_analysis(analysis_id: str) -> Optional[AnalysisResult]:
    path = _analysis_store_path(analysis_id)
    if not path.exists():
        return None
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        logger.warning("analysis_persist_read_failed", extra={"event": "analysis_persist_read_failed", "error": str(exc)})
        return None

    created_at = payload.get("created_at") if isinstance(payload, dict) else None
    ttl = _analysis_persist_ttl_seconds()
    if ttl and isinstance(created_at, (int, float)) and time.time() - float(created_at) > ttl:
        with contextlib.suppress(Exception):
            path.unlink(missing_ok=True)
        return None

    result_payload = payload.get("result") if isinstance(payload, dict) else payload
    if not isinstance(result_payload, dict):
        return None
    try:
        return _parse_analysis_result(result_payload)
    except Exception as exc:
        logger.warning("analysis_persist_parse_failed", extra={"event": "analysis_persist_parse_failed", "error": str(exc)})
        return None


def _get_analysis_semaphore() -> asyncio.Semaphore:
    global _ANALYSIS_SEMAPHORE
    if _ANALYSIS_SEMAPHORE is None:
        settings = get_settings()
        _ANALYSIS_SEMAPHORE = asyncio.Semaphore(max(1, int(settings.analysis_max_concurrency or 1)))
    return _ANALYSIS_SEMAPHORE


def _generate_analysis_id() -> str:
    """Generate a unique analysis ID."""
    return f"ana_{uuid.uuid4().hex[:12]}"


def _attach_event_metadata(
    payload: dict[str, Any],
    analysis_id: str,
    correlation_id: Optional[str],
) -> dict[str, Any]:
    payload["analysis_id"] = analysis_id
    if correlation_id:
        payload["correlation_id"] = correlation_id
    return payload


def _convert_llm_tables_to_schema(llm_tables: list[dict[str, Any]]) -> list[ExtractedTable]:
    """Convert LLM-extracted tables to schema objects."""
    result: list[ExtractedTable] = []
    for idx, table in enumerate(llm_tables):
        try:
            result.append(ExtractedTable(
                id=table.get("id", f"table_{idx + 1}"),
                title=table.get("title"),
                headers=table.get("headers", []),
                rows=table.get("rows", []),
                data_types=table.get("data_types"),
                source_page=table.get("source_page"),
                source_sheet=table.get("source_sheet"),
            ))
        except Exception as exc:
            logger.warning(f"Failed to convert table {idx}: {exc}")
    return result


def _convert_llm_metrics_to_schema(llm_metrics: list[dict[str, Any]]) -> list[ExtractedDataPoint]:
    """Convert LLM-extracted metrics to schema objects."""
    result: list[ExtractedDataPoint] = []
    for metric in llm_metrics:
        try:
            result.append(ExtractedDataPoint(
                key=metric.get("name", "Unknown"),
                value=metric.get("value"),
                data_type="numeric" if isinstance(metric.get("value"), (int, float)) else "text",
                unit=metric.get("unit"),
                context=metric.get("context"),
            ))
        except Exception as exc:
            logger.warning(f"Failed to convert metric: {exc}")
    return result


def _convert_llm_time_series_to_schema(llm_ts: list[dict[str, Any]]) -> list[TimeSeriesCandidate]:
    """Convert LLM time series candidates to schema objects."""
    result: list[TimeSeriesCandidate] = []
    for ts in llm_ts:
        try:
            result.append(TimeSeriesCandidate(
                date_column=ts.get("date_column", ""),
                value_columns=ts.get("value_columns", []),
                frequency=ts.get("frequency"),
                table_id=ts.get("table_id"),
            ))
        except Exception as exc:
            logger.warning(f"Failed to convert time series: {exc}")
    return result


def _convert_llm_charts_to_schema(llm_charts: list[dict[str, Any]]) -> list[ChartSpec]:
    """Convert LLM chart recommendations to ChartSpec objects."""
    result: list[ChartSpec] = []
    for idx, chart in enumerate(llm_charts):
        try:
            chart_type = chart.get("type", "bar").lower()
            if chart_type not in ("line", "bar", "pie", "scatter"):
                chart_type = "bar"

            y_fields = chart.get("y_fields") or chart.get("yFields") or []
            if isinstance(y_fields, str):
                y_fields = [y_fields]

            result.append(ChartSpec(
                id=chart.get("id", f"chart_{idx + 1}"),
                type=chart_type,
                xField=chart.get("x_field") or chart.get("xField") or "",
                yFields=y_fields,
                groupField=chart.get("group_field") or chart.get("groupField"),
                aggregation=chart.get("aggregation"),
                title=chart.get("title"),
                description=chart.get("description") or chart.get("rationale"),
            ))
        except Exception as exc:
            logger.warning(f"Failed to convert chart {idx}: {exc}")
    return result


def _build_field_catalog(tables: list[ExtractedTable]) -> list[FieldInfo]:
    """Build field catalog from extracted tables."""
    fields: list[FieldInfo] = []
    seen_names: set[str] = set()

    for table in tables:
        for idx, header in enumerate(table.headers):
            if header in seen_names:
                continue
            seen_names.add(header)

            data_type = "text"
            if table.data_types and idx < len(table.data_types):
                data_type = table.data_types[idx]

            sample_values: list[Any] = []
            for row in table.rows[:5]:
                if idx < len(row):
                    sample_values.append(row[idx])

            fields.append(FieldInfo(
                name=header,
                type=data_type,
                sample_values=sample_values[:3] if sample_values else None,
            ))

    return fields


def _build_raw_data(tables: list[ExtractedTable], max_rows: int = 500) -> list[dict[str, Any]]:
    """Flatten extracted tables into raw data records."""
    raw_data: list[dict[str, Any]] = []

    for table in tables:
        for row in table.rows[:max_rows]:
            record: dict[str, Any] = {}
            for idx, header in enumerate(table.headers):
                if idx < len(row):
                    record[header] = row[idx]
            if record:
                raw_data.append(record)

        if len(raw_data) >= max_rows:
            break

    return raw_data[:max_rows]


def _merge_extracted_tables(
    content_tables: list[dict[str, Any]],
    llm_tables: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Merge tables from extraction pipeline with LLM-enhanced tables."""
    if not llm_tables:
        return content_tables

    merged_ids = {t.get("id") for t in content_tables}
    merged = list(content_tables)

    for llm_table in llm_tables:
        llm_id = llm_table.get("id")
        if llm_id and llm_id not in merged_ids:
            merged.append(llm_table)
            merged_ids.add(llm_id)

    return merged


async def analyze_document_streaming(
    file_name: str,
    file_bytes: bytes | None = None,
    file_path: Path | str | None = None,
    template_id: Optional[str] = None,
    connection_id: Optional[str] = None,
    correlation_id: Optional[str] = None,
) -> AsyncGenerator[dict[str, Any], None]:
    """Analyze a document with streaming progress updates."""
    analysis_id = _generate_analysis_id()
    started = time.time()
    max_bytes, max_mb = _analysis_size_limits()
    resolved_path = Path(file_path) if file_path else None

    if resolved_path and not file_name:
        file_name = resolved_path.name

    semaphore = _get_analysis_semaphore()
    async with semaphore:
        yield _attach_event_metadata(
            {"event": "stage", "stage": "uploading", "progress": 10},
            analysis_id,
            correlation_id,
        )

        if file_bytes is None and resolved_path is None:
            yield _attach_event_metadata(
                {"event": "error", "detail": "Empty file provided."},
                analysis_id,
                correlation_id,
            )
            return
        if file_bytes is not None:
            if len(file_bytes) > max_bytes:
                yield _attach_event_metadata(
                    {"event": "error", "detail": f"File too large. Maximum size is {max_mb}MB."},
                    analysis_id,
                    correlation_id,
                )
                return
        elif resolved_path is not None:
            try:
                size_bytes = resolved_path.stat().st_size
            except Exception as exc:
                yield _attach_event_metadata(
                    {"event": "error", "detail": f"Failed to read file size: {exc}"},
                    analysis_id,
                    correlation_id,
                )
                return
            if size_bytes > max_bytes:
                yield _attach_event_metadata(
                    {"event": "error", "detail": f"File too large. Maximum size is {max_mb}MB."},
                    analysis_id,
                    correlation_id,
                )
                return

        yield _attach_event_metadata(
            {"event": "stage", "stage": "parsing", "progress": 20, "detail": "Extracting content..."},
            analysis_id,
            correlation_id,
        )

        content = await asyncio.to_thread(
            extract_document_content,
            file_path=resolved_path,
            file_bytes=file_bytes,
            file_name=file_name,
        )

        if content.errors and not content.tables_raw and not content.text_content:
            yield _attach_event_metadata(
                {"event": "error", "detail": f"Failed to extract content: {'; '.join(content.errors)}"},
                analysis_id,
                correlation_id,
            )
            return

        yield _attach_event_metadata(
            {
                "event": "stage",
                "stage": "table_extraction",
                "progress": 40,
                "detail": f"Found {len(content.tables_raw)} tables",
            },
            analysis_id,
            correlation_id,
        )

        yield _attach_event_metadata(
            {"event": "stage", "stage": "llm_analysis", "progress": 60, "detail": "Analyzing with AI..."},
            analysis_id,
            correlation_id,
        )

        llm_result = {"tables": [], "key_metrics": [], "time_series_candidates": [], "chart_recommendations": []}

        try:
            client = get_llm_client()
            formatted_content = format_content_for_llm(content)

            prompt = build_analysis_prompt(
                document_type=content.document_type,
                file_name=content.file_name,
                page_count=content.page_count,
                extracted_content=formatted_content,
            )

            messages = [{"role": "user", "content": prompt}]

            response = await call_chat_completion_async(
                client,
                model=None,
                messages=messages,
                description="document_analysis",
                temperature=0.2,
            )

            raw_text = ""
            if hasattr(response, "choices") and response.choices:
                choice = response.choices[0]
                if hasattr(choice, "message") and hasattr(choice.message, "content"):
                    raw_text = choice.message.content or ""

            llm_result = parse_analysis_response(raw_text)

        except Exception as exc:
            logger.warning(f"LLM analysis failed: {exc}")
            yield _attach_event_metadata(
                {
                    "event": "stage",
                    "stage": "llm_analysis",
                    "progress": 70,
                    "detail": "AI analysis skipped (using extracted data)",
                },
                analysis_id,
                correlation_id,
            )

        yield _attach_event_metadata(
            {"event": "stage", "stage": "chart_generation", "progress": 80, "detail": "Generating visualizations..."},
            analysis_id,
            correlation_id,
        )

        merged_tables = _merge_extracted_tables(content.tables_raw, llm_result.get("tables", []))
        tables = _convert_llm_tables_to_schema(merged_tables)
        data_points = _convert_llm_metrics_to_schema(llm_result.get("key_metrics", []))
        time_series = _convert_llm_time_series_to_schema(llm_result.get("time_series_candidates", []))
        charts = _convert_llm_charts_to_schema(llm_result.get("chart_recommendations", []))

        if not charts and tables:
            charts = _generate_fallback_charts(tables)

        field_catalog = _build_field_catalog(tables)
        raw_data = _build_raw_data(tables)

        processing_time_ms = int((time.time() - started) * 1000)

        warnings: list[str] = []
        if content.errors:
            warnings.extend(content.errors)

        result = AnalysisResult(
            analysis_id=analysis_id,
            document_name=file_name,
            document_type=content.document_type,
            processing_time_ms=processing_time_ms,
            summary=llm_result.get("summary"),
            tables=tables,
            data_points=data_points,
            time_series_candidates=time_series,
            chart_suggestions=charts,
            raw_data=raw_data,
            field_catalog=field_catalog,
            template_id=template_id,
            warnings=warnings,
        )

        _analysis_cache().set(analysis_id, result)
        _persist_analysis_result(result)

        yield _attach_event_metadata(
            {"event": "stage", "stage": "complete", "progress": 100},
            analysis_id,
            correlation_id,
        )

        result_payload = {"event": "result", **result.model_dump()}
        if correlation_id:
            result_payload["correlation_id"] = correlation_id
        yield result_payload


def _generate_fallback_charts(tables: list[ExtractedTable]) -> list[ChartSpec]:
    """Generate basic chart suggestions when LLM doesn't provide any."""
    charts: list[ChartSpec] = []

    for table in tables[:3]:
        datetime_cols: list[str] = []
        numeric_cols: list[str] = []
        text_cols: list[str] = []

        for idx, header in enumerate(table.headers):
            data_type = table.data_types[idx] if table.data_types and idx < len(table.data_types) else "text"
            if data_type in ("datetime", "date"):
                datetime_cols.append(header)
            elif data_type == "numeric":
                numeric_cols.append(header)
            else:
                text_cols.append(header)

        if datetime_cols and numeric_cols:
            charts.append(ChartSpec(
                id=f"fallback_line_{table.id}",
                type="line",
                xField=datetime_cols[0],
                yFields=numeric_cols[:3],
                title=f"Time Series: {table.title or table.id}",
                description="Numeric values over time",
            ))

        if text_cols and numeric_cols:
            charts.append(ChartSpec(
                id=f"fallback_bar_{table.id}",
                type="bar",
                xField=text_cols[0],
                yFields=numeric_cols[:2],
                title=f"Comparison: {table.title or table.id}",
                description="Numeric values by category",
            ))

    return charts[:5]


def get_analysis(analysis_id: str) -> Optional[AnalysisResult]:
    """Retrieve a cached analysis result."""
    cached = _analysis_cache().get(analysis_id)
    if cached is not None:
        return cached
    persisted = _load_persisted_analysis(analysis_id)
    if persisted is not None:
        _analysis_cache().set(analysis_id, persisted)
    return persisted


def get_analysis_data(analysis_id: str) -> Optional[list[dict[str, Any]]]:
    """Get raw data for an analysis."""
    result = get_analysis(analysis_id)
    if result:
        return result.raw_data
    return None


def suggest_charts_for_analysis(
    analysis_id: str,
    payload: AnalysisSuggestChartsPayload,
) -> list[ChartSpec]:
    """Generate additional chart suggestions for an existing analysis."""
    result = get_analysis(analysis_id)
    if not result:
        return []

    try:
        client = get_llm_client()

        data_summary = f"Document: {result.document_name}\n"
        data_summary += f"Tables: {len(result.tables)}\n"
        for table in result.tables[:5]:
            data_summary += f"  - {table.title or table.id}: {len(table.rows)} rows, columns: {', '.join(table.headers[:5])}\n"

        field_catalog_str = "\n".join([
            f"  - {f.name}: {f.type}" + (f" (samples: {f.sample_values})" if f.sample_values else "")
            for f in result.field_catalog[:20]
        ])

        prompt = build_chart_suggestion_prompt(
            data_summary=data_summary,
            field_catalog=field_catalog_str,
            user_question=payload.question,
        )

        messages = [{"role": "user", "content": prompt}]

        response = call_chat_completion(
            client,
            model=None,
            messages=messages,
            description="chart_suggestion_analysis",
            temperature=0.3,
        )

        raw_text = ""
        if hasattr(response, "choices") and response.choices:
            choice = response.choices[0]
            if hasattr(choice, "message") and hasattr(choice.message, "content"):
                raw_text = choice.message.content or ""

        cleaned = strip_code_fences(raw_text)
        data = json.loads(cleaned)
        charts = data.get("charts", [])
        return _convert_llm_charts_to_schema(charts)

    except Exception as exc:
        logger.warning(f"Chart suggestion failed: {exc}")
        return _generate_fallback_charts(result.tables)


__all__ = [
    "analyze_document_streaming",
    "get_analysis",
    "get_analysis_data",
    "suggest_charts_for_analysis",
]


# ENHANCED ANALYSIS STORE (merged from enhanced_analysis_store.py)

from typing import Any, Optional

from backend.app.schemas import EnhancedAnalysisResult


class EnhancedAnalysisStore:
    """File-backed store for enhanced analysis artifacts."""

    def __init__(self, base_dir: Path) -> None:
        self.base_dir = base_dir
        self.results_dir = base_dir / "results"
        self.context_dir = base_dir / "context"
        self.comments_dir = base_dir / "comments"
        self.versions_dir = base_dir / "versions"
        self.shares_dir = base_dir / "shares"
        self._lock = threading.Lock()
        self._ensure_dirs()

    def _ensure_dirs(self) -> None:
        self.results_dir.mkdir(parents=True, exist_ok=True)
        self.context_dir.mkdir(parents=True, exist_ok=True)
        self.comments_dir.mkdir(parents=True, exist_ok=True)
        self.versions_dir.mkdir(parents=True, exist_ok=True)
        self.shares_dir.mkdir(parents=True, exist_ok=True)

    def _write_json(self, path: Path, payload: Any) -> None:
        tmp_path = path.with_suffix(path.suffix + ".tmp")
        data = json.dumps(payload, ensure_ascii=True, default=str)
        with self._lock:
            path.parent.mkdir(parents=True, exist_ok=True)
            tmp_path.write_text(data, encoding="utf-8")
            tmp_path.replace(path)

    def _read_json(self, path: Path, default: Any) -> Any:
        try:
            if not path.exists():
                return default
            raw = path.read_text(encoding="utf-8")
            return json.loads(raw)
        except Exception:
            return default

    def _result_path(self, analysis_id: str) -> Path:
        return self.results_dir / f"{analysis_id}.json"

    def _context_path(self, analysis_id: str) -> Path:
        return self.context_dir / f"{analysis_id}.json"

    def _comments_path(self, analysis_id: str) -> Path:
        return self.comments_dir / f"{analysis_id}.json"

    def _versions_path(self, analysis_id: str) -> Path:
        return self.versions_dir / f"{analysis_id}.json"

    def _share_path(self, share_id: str) -> Path:
        return self.shares_dir / f"{share_id}.json"

    # ---------------------------------------------------------------------
    # Results
    # ---------------------------------------------------------------------
    def save_result(self, result: EnhancedAnalysisResult) -> None:
        if hasattr(result, "model_dump"):
            payload = result.model_dump(mode="json")
        else:
            payload = json.loads(result.json())
        self._write_json(self._result_path(result.analysis_id), payload)

    def load_result(self, analysis_id: str) -> Optional[EnhancedAnalysisResult]:
        payload = self._read_json(self._result_path(analysis_id), default=None)
        if not payload:
            return None
        try:
            if hasattr(EnhancedAnalysisResult, "model_validate"):
                return EnhancedAnalysisResult.model_validate(payload)
            return EnhancedAnalysisResult.parse_obj(payload)
        except Exception:
            return None

    def save_context(self, analysis_id: str, text_content: str) -> None:
        if not analysis_id:
            return
        payload = {"analysis_id": analysis_id, "text": text_content or ""}
        self._write_json(self._context_path(analysis_id), payload)

    def load_context(self, analysis_id: str) -> str:
        payload = self._read_json(self._context_path(analysis_id), default=None)
        if isinstance(payload, dict):
            return str(payload.get("text") or "")
        return ""

    # ---------------------------------------------------------------------
    # Comments, versions, shares
    # ---------------------------------------------------------------------
    def save_comments(self, analysis_id: str, payload: list[dict[str, Any]]) -> None:
        self._write_json(self._comments_path(analysis_id), payload)

    def load_comments(self, analysis_id: str) -> list[dict[str, Any]]:
        return self._read_json(self._comments_path(analysis_id), default=[])

    def save_versions(self, analysis_id: str, payload: list[dict[str, Any]]) -> None:
        self._write_json(self._versions_path(analysis_id), payload)

    def load_versions(self, analysis_id: str) -> list[dict[str, Any]]:
        return self._read_json(self._versions_path(analysis_id), default=[])

    def save_share(self, payload: dict[str, Any]) -> None:
        share_id = str(payload.get("id") or "")
        if not share_id:
            return
        self._write_json(self._share_path(share_id), payload)

    def load_share(self, share_id: str) -> Optional[dict[str, Any]]:
        return self._read_json(self._share_path(share_id), default=None)

    def list_shares_for_analysis(self, analysis_id: str) -> list[dict[str, Any]]:
        shares: list[dict[str, Any]] = []
        try:
            for share_file in self.shares_dir.glob("*.json"):
                payload = self._read_json(share_file, default=None)
                if payload and payload.get("analysis_id") == analysis_id:
                    shares.append(payload)
        except Exception:
            return []
        return shares


_STORE: Optional[EnhancedAnalysisStore] = None


def get_analysis_store() -> EnhancedAnalysisStore:
    """Return singleton analysis store."""
    global _STORE
    if _STORE is None:
        settings = get_settings()
        base_dir = settings.state_dir / "analysis_v2"
        _STORE = EnhancedAnalysisStore(base_dir)
    return _STORE


# ENGINES

# mypy: ignore-errors
"""
AI-Powered Analysis Engines - Document summarization, sentiment, and statistical analysis.

Features:
2.1 Document Summarization Suite
2.2 Sentiment & Tone Analysis
2.3 Comparative Analysis
"""

import hashlib
from abc import ABC, abstractmethod
from collections import Counter
from datetime import datetime, timezone
from enum import Enum
from typing import Any, Dict, List, Optional, Tuple

from backend.app.schemas import (
    ActionItem,
    ComparativeAnalysis,
    DocumentSummary,
    EnhancedExtractedTable,
    ExtractedMetric,
    FinancialAnalysis,
    Insight,
    OpportunityItem,
    Priority,
    RiskItem,
    RiskLevel,
    SentimentAnalysis,
    SentimentLevel,
    StatisticalAnalysis,
    SummaryMode,
    TextAnalytics,
)
from backend.app.services.infra_services import call_chat_completion, extract_json_from_llm_response

logger = logging.getLogger("neura.analyze.engines")


# DOCUMENT SUMMARIZATION

SUMMARY_PROMPTS = {
    SummaryMode.EXECUTIVE: """Create a C-suite executive summary of this document.
- Maximum 3 bullet points
- Focus on key business decisions and bottom-line impact
- Use clear, decisive language
- Highlight the most critical number or outcome

Format:
Title: [Brief title]
Bullets:
- [Key point 1]
- [Key point 2]
- [Key point 3]
Key Figure: [The single most important number/metric]""",

    SummaryMode.DATA: """Create a data-focused summary of this document.
- List all key figures, metrics, and KPIs found
- Include trends and comparisons
- Note data quality or completeness issues

Format:
Title: Data Summary
Key Metrics:
- [Metric 1]: [Value] ([context])
- [Metric 2]: [Value] ([context])
Trends: [Notable trends]
Data Quality: [Any issues noted]""",

    SummaryMode.QUICK: """Provide a one-sentence summary capturing the essence of this document.
Keep it under 30 words. Focus on the main purpose and key outcome.""",

    SummaryMode.COMPREHENSIVE: """Create a comprehensive structured summary of this document.

Format:
Title: [Document title]
Overview: [2-3 sentence overview]
Key Sections:
1. [Section 1 name]: [Summary]
2. [Section 2 name]: [Summary]
Key Findings:
- [Finding 1]
- [Finding 2]
- [Finding 3]
Data Highlights:
- [Key metric/number 1]
- [Key metric/number 2]
Conclusions: [Main conclusions]
Limitations: [Any caveats or limitations noted]""",

    SummaryMode.ACTION_ITEMS: """Extract all action items, to-dos, and next steps from this document.

Format:
Title: Action Items Summary
Immediate Actions:
- [Action 1] (Priority: High/Medium/Low)
- [Action 2] (Priority: High/Medium/Low)
Follow-up Required:
- [Follow-up 1]
Deadlines Mentioned:
- [Deadline 1]: [Date]
Responsibilities:
- [Person/Team]: [Their action items]""",

    SummaryMode.RISKS: """Identify and summarize all risks, concerns, and potential issues mentioned in this document.

Format:
Title: Risk Summary
Critical Risks:
- [Risk 1]: [Description] - Impact: [High/Medium/Low]
Warnings/Concerns:
- [Concern 1]
Compliance Issues:
- [Any compliance or regulatory concerns]
Mitigation Mentioned:
- [Any risk mitigation strategies noted]
Overall Risk Level: [Low/Medium/High/Critical]""",

    SummaryMode.OPPORTUNITIES: """Identify opportunities, growth areas, and positive developments in this document.

Format:
Title: Opportunities Summary
Growth Opportunities:
- [Opportunity 1]: [Description] - Potential: [High/Medium/Low]
Positive Trends:
- [Trend 1]
Recommendations for Action:
- [Recommendation 1]
Quick Wins:
- [Any easily achievable improvements noted]""",
}


def generate_summary(
    text: str,
    mode: SummaryMode,
    tables: List[EnhancedExtractedTable] = None,
    metrics: List[ExtractedMetric] = None,
) -> DocumentSummary:
    """Generate a document summary in the specified mode."""
    # Build context
    context_parts = [f"Document text:\n{text[:8000]}"]

    if tables:
        table_info = "\n\nTables found:\n"
        for t in tables[:5]:
            table_info += f"- {t.title or t.id}: {t.row_count} rows, columns: {', '.join(t.headers[:5])}\n"
        context_parts.append(table_info)

    if metrics:
        metrics_info = "\n\nKey metrics extracted:\n"
        for m in metrics[:10]:
            metrics_info += f"- {m.name}: {m.raw_value}\n"
        context_parts.append(metrics_info)

    context = "\n".join(context_parts)
    prompt = f"{SUMMARY_PROMPTS[mode]}\n\nDocument:\n{context}"

    try:
        client = get_llm_client()
        response = call_chat_completion(
            client,
            model=None,
            messages=[{"role": "user", "content": prompt}],
            description=f"summary_{mode.value}",
            temperature=0.3,
        )

        content = response.choices[0].message.content or ""

        # Parse the response
        title = ""
        bullet_points = []
        key_figures = []

        # Extract title
        title_match = re.search(r'Title:\s*(.+?)(?:\n|$)', content)
        if title_match:
            title = title_match.group(1).strip()

        # Extract bullet points
        bullet_matches = re.findall(r'^[\s]*[-•*]\s*(.+?)$', content, re.MULTILINE)
        bullet_points = [b.strip() for b in bullet_matches if b.strip()]

        # Extract key figures
        figure_matches = re.findall(r'[\$€£¥]?\d[\d,]*(?:\.\d+)?[%]?', content)
        key_figures = [{"value": f, "context": ""} for f in figure_matches[:5]]

        # Word count
        words = len(content.split())
        reading_time = max(1, words / 200)  # 200 words per minute

        return DocumentSummary(
            mode=mode,
            title=title or f"{mode.value.title()} Summary",
            content=content,
            bullet_points=bullet_points[:10],
            key_figures=key_figures,
            word_count=words,
            reading_time_minutes=reading_time,
        )
    except Exception as e:
        logger.error("Summary generation failed: %s", e, exc_info=True)
        return DocumentSummary(
            mode=mode,
            title="Summary Generation Failed",
            content="Could not generate summary due to an internal error.",
        )


def generate_all_summaries(
    text: str,
    tables: List[EnhancedExtractedTable] = None,
    metrics: List[ExtractedMetric] = None,
) -> Dict[str, DocumentSummary]:
    """Generate all summary types."""
    summaries = {}
    for mode in SummaryMode:
        summaries[mode.value] = generate_summary(text, mode, tables, metrics)
    return summaries


# SENTIMENT ANALYSIS

def analyze_sentiment(text: str) -> SentimentAnalysis:
    """Analyze document sentiment and tone."""
    prompt = f"""Analyze the sentiment and tone of this document.

Document:
{text[:8000]}

Provide analysis in JSON format:
```json
{{
  "overall_sentiment": "positive|negative|neutral|very_positive|very_negative",
  "overall_score": 0.5,  // -1.0 (very negative) to 1.0 (very positive)
  "confidence": 0.85,
  "emotional_tone": "formal|casual|urgent|optimistic|pessimistic|neutral|analytical|persuasive",
  "urgency_level": "low|normal|high|critical",
  "section_sentiments": [
    {{"section": "Introduction", "sentiment": "positive", "score": 0.6}},
    {{"section": "Financial Results", "sentiment": "negative", "score": -0.3}}
  ],
  "positive_phrases": ["exceeded expectations", "strong growth"],
  "negative_phrases": ["challenges ahead", "declining margins"],
  "bias_indicators": ["overly optimistic language", "missing context for claims"]
}}
```

Be objective and thorough in your analysis."""

    try:
        client = get_llm_client()
        response = call_chat_completion(
            client,
            model=None,
            messages=[{"role": "user", "content": prompt}],
            description="sentiment_analysis",
            temperature=0.2,
        )

        raw_text = response.choices[0].message.content or ""
        data = extract_json_from_llm_response(raw_text, default=None)

        if data:
            sentiment_map = {
                "very_positive": SentimentLevel.VERY_POSITIVE,
                "positive": SentimentLevel.POSITIVE,
                "neutral": SentimentLevel.NEUTRAL,
                "negative": SentimentLevel.NEGATIVE,
                "very_negative": SentimentLevel.VERY_NEGATIVE,
            }

            return SentimentAnalysis(
                overall_sentiment=sentiment_map.get(
                    data.get("overall_sentiment", "neutral").lower(),
                    SentimentLevel.NEUTRAL
                ),
                overall_score=float(data.get("overall_score", 0)),
                confidence=float(data.get("confidence", 0.8)),
                section_sentiments=data.get("section_sentiments", []),
                emotional_tone=data.get("emotional_tone", "neutral"),
                urgency_level=data.get("urgency_level", "normal"),
                bias_indicators=data.get("bias_indicators", []),
                key_phrases={
                    "positive": data.get("positive_phrases", []),
                    "negative": data.get("negative_phrases", []),
                },
            )
    except Exception as e:
        logger.warning(f"Sentiment analysis failed: {e}")

    return SentimentAnalysis(
        overall_sentiment=SentimentLevel.NEUTRAL,
        overall_score=0.0,
        confidence=0.5,
    )


# TEXT ANALYTICS

def analyze_text(text: str) -> TextAnalytics:
    """Perform text analytics including readability and keyword extraction."""
    # Basic counts
    words = text.split()
    word_count = len(words)
    sentences = re.split(r'[.!?]+', text)
    sentence_count = len([s for s in sentences if s.strip()])
    paragraphs = text.split('\n\n')
    paragraph_count = len([p for p in paragraphs if p.strip()])

    avg_sentence_length = word_count / max(sentence_count, 1)

    # Flesch-Kincaid readability
    syllables = sum(_count_syllables(word) for word in words)
    if word_count > 0 and sentence_count > 0:
        flesch_score = 206.835 - 1.015 * (word_count / sentence_count) - 84.6 * (syllables / word_count)
        flesch_score = max(0, min(100, flesch_score))
    else:
        flesch_score = 50

    # Grade level
    if flesch_score >= 90:
        grade = "5th grade"
    elif flesch_score >= 80:
        grade = "6th grade"
    elif flesch_score >= 70:
        grade = "7th grade"
    elif flesch_score >= 60:
        grade = "8th-9th grade"
    elif flesch_score >= 50:
        grade = "10th-12th grade"
    elif flesch_score >= 30:
        grade = "College"
    else:
        grade = "College graduate"

    # Keyword extraction
    stopwords = {'the', 'a', 'an', 'is', 'are', 'was', 'were', 'be', 'been', 'being',
                 'have', 'has', 'had', 'do', 'does', 'did', 'will', 'would', 'could',
                 'should', 'may', 'might', 'must', 'can', 'this', 'that', 'these',
                 'those', 'i', 'you', 'he', 'she', 'it', 'we', 'they', 'and', 'or',
                 'but', 'if', 'because', 'as', 'of', 'at', 'by', 'for', 'with',
                 'to', 'from', 'in', 'on', 'not', 'no', 'so', 'than', 'too', 'very'}

    word_freq = Counter(w.lower() for w in re.findall(r'\b[a-zA-Z]{3,}\b', text)
                        if w.lower() not in stopwords)
    keywords = [
        {"word": word, "frequency": count, "importance": min(1.0, count / 50)}
        for word, count in word_freq.most_common(20)
    ]

    # Detect language (simple heuristic)
    language = "en"  # Default to English

    return TextAnalytics(
        word_count=word_count,
        sentence_count=sentence_count,
        paragraph_count=paragraph_count,
        avg_sentence_length=round(avg_sentence_length, 1),
        readability_score=round(flesch_score, 1),
        readability_grade=grade,
        keywords=keywords,
        language=language,
        language_confidence=0.95,
    )


def _count_syllables(word: str) -> int:
    """Count syllables in a word."""
    word = word.lower()
    vowels = "aeiouy"
    count = 0
    prev_vowel = False

    for char in word:
        is_vowel = char in vowels
        if is_vowel and not prev_vowel:
            count += 1
        prev_vowel = is_vowel

    # Adjust for silent e
    if word.endswith('e'):
        count -= 1

    return max(1, count)


# STATISTICAL ANALYSIS

def analyze_statistics(tables: List[EnhancedExtractedTable]) -> StatisticalAnalysis:
    """Perform statistical analysis on numeric data in tables."""
    column_stats = {}
    correlations = []
    outliers = []
    distributions = {}
    trends = []

    for table in tables:
        numeric_columns = {}

        # Extract numeric columns
        for col_idx, (header, dtype) in enumerate(zip(table.headers, table.data_types)):
            if dtype == "numeric":
                values = []
                for row_idx, row in enumerate(table.rows):
                    if col_idx < len(row):
                        try:
                            val = float(str(row[col_idx]).replace(",", "").replace("$", "").replace("%", ""))
                            values.append((row_idx, val))
                        except (ValueError, TypeError):
                            pass

                if len(values) >= 3:
                    numeric_columns[header] = values

        # Calculate statistics for each column
        for header, indexed_values in numeric_columns.items():
            values = [v for _, v in indexed_values]
            n = len(values)

            mean = sum(values) / n
            sorted_vals = sorted(values)
            median = sorted_vals[n // 2] if n % 2 == 1 else (sorted_vals[n // 2 - 1] + sorted_vals[n // 2]) / 2

            variance = sum((x - mean) ** 2 for x in values) / n
            std = math.sqrt(variance)

            # Percentiles
            p25 = sorted_vals[int(n * 0.25)]
            p75 = sorted_vals[int(n * 0.75)]

            column_stats[f"{table.id}.{header}"] = {
                "count": n,
                "mean": round(mean, 4),
                "median": round(median, 4),
                "std": round(std, 4),
                "min": min(values),
                "max": max(values),
                "p25": p25,
                "p75": p75,
            }

            # Detect outliers (values beyond 2 standard deviations)
            if std > 0:
                for row_idx, val in indexed_values:
                    zscore = abs((val - mean) / std)
                    if zscore > 2:
                        outliers.append({
                            "table": table.id,
                            "column": header,
                            "row_index": row_idx,
                            "value": val,
                            "zscore": round(zscore, 2),
                        })

            # Simple trend detection (for sequential data)
            if n >= 5:
                first_half = sum(values[:n // 2]) / (n // 2)
                second_half = sum(values[n // 2:]) / (n - n // 2)

                if second_half > first_half * 1.1:
                    trend_dir = "increasing"
                elif second_half < first_half * 0.9:
                    trend_dir = "decreasing"
                else:
                    trend_dir = "stable"

                trends.append({
                    "table": table.id,
                    "column": header,
                    "trend_direction": trend_dir,
                    "change_ratio": round(second_half / first_half, 4) if first_half != 0 else 0,
                })

        # Calculate correlations between numeric columns
        col_names = list(numeric_columns.keys())
        for i in range(len(col_names)):
            for j in range(i + 1, len(col_names)):
                col1, col2 = col_names[i], col_names[j]
                vals1 = [v for _, v in numeric_columns[col1]]
                vals2 = [v for _, v in numeric_columns[col2]]

                # Align by index
                min_len = min(len(vals1), len(vals2))
                if min_len >= 5:
                    corr = _pearson_correlation(vals1[:min_len], vals2[:min_len])
                    if abs(corr) > 0.3:  # Only report meaningful correlations
                        correlations.append({
                            "table": table.id,
                            "column1": col1,
                            "column2": col2,
                            "correlation": round(corr, 4),
                        })

    return StatisticalAnalysis(
        column_stats=column_stats,
        correlations=correlations,
        outliers=outliers[:20],  # Limit to top 20
        distributions=distributions,
        trends=trends,
    )


def _pearson_correlation(x: List[float], y: List[float]) -> float:
    """Calculate Pearson correlation coefficient."""
    n = len(x)
    if n < 2:
        return 0

    mean_x = sum(x) / n
    mean_y = sum(y) / n

    numerator = sum((xi - mean_x) * (yi - mean_y) for xi, yi in zip(x, y))
    denom_x = math.sqrt(sum((xi - mean_x) ** 2 for xi in x))
    denom_y = math.sqrt(sum((yi - mean_y) ** 2 for yi in y))

    if denom_x == 0 or denom_y == 0:
        return 0

    return numerator / (denom_x * denom_y)


# FINANCIAL ANALYSIS

def analyze_financials(
    text: str,
    metrics: List[ExtractedMetric],
    tables: List[EnhancedExtractedTable],
) -> FinancialAnalysis:
    """Perform financial analysis using LLM."""
    # Build context
    metrics_context = "\n".join([
        f"- {m.name}: {m.raw_value}" + (f" ({m.change}% {m.change_direction})" if m.change else "")
        for m in metrics[:20]
    ])

    prompt = f"""Analyze this document for financial insights. Calculate ratios where data is available.

Metrics found:
{metrics_context}

Document excerpt:
{text[:5000]}

Return JSON:
```json
{{
  "currency": "USD",
  "gross_margin": 0.35,
  "operating_margin": 0.20,
  "net_margin": 0.15,
  "revenue_growth": 0.12,
  "profit_growth": 0.08,
  "yoy_comparison": {{"revenue": {{"current": 1000000, "previous": 900000, "change": 0.11}}}},
  "variance_analysis": [
    {{"metric": "Revenue", "actual": 1000000, "budget": 950000, "variance": 50000, "variance_pct": 5.26}}
  ],
  "insights": [
    "Revenue grew 11% year-over-year, exceeding industry average of 8%",
    "Operating margin improved despite increased costs"
  ],
  "warnings": [
    "Debt-to-equity ratio increased significantly",
    "Cash reserves declining quarter-over-quarter"
  ]
}}
```

Only include metrics you can calculate or find. Use null for unavailable data."""

    try:
        client = get_llm_client()
        response = call_chat_completion(
            client,
            model=None,
            messages=[{"role": "user", "content": prompt}],
            description="financial_analysis",
            temperature=0.2,
        )

        raw_text = response.choices[0].message.content or ""
        data = extract_json_from_llm_response(raw_text, default=None)

        if data:
            return FinancialAnalysis(
                metrics_found=len(metrics),
                currency=data.get("currency", "USD"),
                gross_margin=data.get("gross_margin"),
                operating_margin=data.get("operating_margin"),
                net_margin=data.get("net_margin"),
                revenue_growth=data.get("revenue_growth"),
                profit_growth=data.get("profit_growth"),
                yoy_comparison=data.get("yoy_comparison", {}),
                variance_analysis=data.get("variance_analysis", []),
                insights=data.get("insights", []),
                warnings=data.get("warnings", []),
            )
    except Exception as e:
        logger.warning(f"Financial analysis failed: {e}")

    return FinancialAnalysis(metrics_found=len(metrics))


# INSIGHTS, RISKS & OPPORTUNITIES

def generate_insights(
    text: str,
    metrics: List[ExtractedMetric],
    tables: List[EnhancedExtractedTable],
    sentiment: SentimentAnalysis,
    stats: StatisticalAnalysis,
) -> Tuple[List[Insight], List[RiskItem], List[OpportunityItem], List[ActionItem]]:
    """Generate insights, risks, opportunities, and action items."""
    context = f"""Document analysis context:
- Sentiment: {sentiment.overall_sentiment.value} (score: {sentiment.overall_score})
- Urgency: {sentiment.urgency_level}
- Key metrics: {len(metrics)}
- Tables: {len(tables)}
- Outliers detected: {len(stats.outliers)}
- Trends: {[t['trend_direction'] for t in stats.trends]}

Metrics:
{chr(10).join([f"- {m.name}: {m.raw_value}" for m in metrics[:15]])}

Document excerpt:
{text[:4000]}"""

    prompt = f"""{context}

Analyze this document and generate:
1. Key insights (findings, trends, anomalies)
2. Risks and concerns
3. Opportunities
4. Recommended action items

Return JSON:
```json
{{
  "insights": [
    {{
      "type": "finding|trend|anomaly|recommendation|warning",
      "title": "Short title",
      "description": "Detailed description",
      "priority": "critical|high|medium|low",
      "confidence": 0.85,
      "supporting_data": ["Revenue: $1.5M", "Growth: 15%"],
      "actionable": true,
      "suggested_actions": ["Review pricing strategy"]
    }}
  ],
  "risks": [
    {{
      "title": "Declining Cash Reserves",
      "description": "Cash reserves have decreased 20% this quarter",
      "risk_level": "high",
      "category": "financial",
      "probability": 0.7,
      "impact": 0.8,
      "mitigation_suggestions": ["Reduce non-essential spending"]
    }}
  ],
  "opportunities": [
    {{
      "title": "Market Expansion",
      "description": "Emerging market shows 30% growth potential",
      "opportunity_type": "growth",
      "potential_value": "$500K annual revenue",
      "confidence": 0.75,
      "requirements": ["Localization", "Partner network"],
      "suggested_actions": ["Conduct market research"]
    }}
  ],
  "action_items": [
    {{
      "title": "Review Q3 budget",
      "description": "Budget variance exceeds 10% threshold",
      "priority": "high",
      "category": "financial",
      "expected_outcome": "Realigned budget for Q4"
    }}
  ]
}}
```

Be specific and actionable. Base insights on actual data found."""

    insights = []
    risks = []
    opportunities = []
    action_items = []

    try:
        client = get_llm_client()
        response = call_chat_completion(
            client,
            model=None,
            messages=[{"role": "user", "content": prompt}],
            description="insights_generation",
            temperature=0.3,
        )

        raw_text = response.choices[0].message.content or ""
        data = extract_json_from_llm_response(raw_text, default=None)

        if data:
            # Parse insights
            for item in data.get("insights", []):
                priority_map = {"critical": Priority.CRITICAL, "high": Priority.HIGH,
                               "medium": Priority.MEDIUM, "low": Priority.LOW}
                insights.append(Insight(
                    id=f"ins_{uuid.uuid4().hex[:8]}",
                    type=item.get("type", "finding"),
                    title=item.get("title", ""),
                    description=item.get("description", ""),
                    priority=priority_map.get(item.get("priority", "medium"), Priority.MEDIUM),
                    confidence=item.get("confidence", 0.8),
                    supporting_data=item.get("supporting_data", []),
                    actionable=item.get("actionable", False),
                    suggested_actions=item.get("suggested_actions", []),
                ))

            # Parse risks
            for item in data.get("risks", []):
                risk_map = {"critical": RiskLevel.CRITICAL, "high": RiskLevel.HIGH,
                           "medium": RiskLevel.MEDIUM, "low": RiskLevel.LOW, "minimal": RiskLevel.MINIMAL}
                risk_level = risk_map.get(item.get("risk_level", "medium"), RiskLevel.MEDIUM)
                prob = item.get("probability", 0.5)
                impact = item.get("impact", 0.5)

                risks.append(RiskItem(
                    id=f"risk_{uuid.uuid4().hex[:8]}",
                    title=item.get("title", ""),
                    description=item.get("description", ""),
                    risk_level=risk_level,
                    category=item.get("category", "general"),
                    probability=prob,
                    impact=impact,
                    risk_score=prob * impact,
                    mitigation_suggestions=item.get("mitigation_suggestions", []),
                ))

            # Parse opportunities
            for item in data.get("opportunities", []):
                opportunities.append(OpportunityItem(
                    id=f"opp_{uuid.uuid4().hex[:8]}",
                    title=item.get("title", ""),
                    description=item.get("description", ""),
                    opportunity_type=item.get("opportunity_type", "growth"),
                    potential_value=item.get("potential_value"),
                    confidence=item.get("confidence", 0.7),
                    requirements=item.get("requirements", []),
                    suggested_actions=item.get("suggested_actions", []),
                ))

            # Parse action items
            for item in data.get("action_items", []):
                priority_map = {"critical": Priority.CRITICAL, "high": Priority.HIGH,
                               "medium": Priority.MEDIUM, "low": Priority.LOW}
                action_items.append(ActionItem(
                    id=f"act_{uuid.uuid4().hex[:8]}",
                    title=item.get("title", ""),
                    description=item.get("description", ""),
                    priority=priority_map.get(item.get("priority", "medium"), Priority.MEDIUM),
                    category=item.get("category", "general"),
                    expected_outcome=item.get("expected_outcome"),
                ))

    except Exception as e:
        logger.warning(f"Insights generation failed: {e}")

    return insights, risks, opportunities, action_items


# COMPARATIVE ANALYSIS

def compare_documents(
    text1: str,
    text2: str,
    metrics1: List[ExtractedMetric] = None,
    metrics2: List[ExtractedMetric] = None,
) -> ComparativeAnalysis:
    """Compare two documents and identify differences."""
    metrics1 = metrics1 or []
    metrics2 = metrics2 or []

    # Build metrics comparison
    metrics1_dict = {m.name: m for m in metrics1}
    metrics2_dict = {m.name: m for m in metrics2}

    metric_changes = []
    all_metric_names = set(metrics1_dict.keys()) | set(metrics2_dict.keys())

    for name in all_metric_names:
        m1 = metrics1_dict.get(name)
        m2 = metrics2_dict.get(name)

        if m1 and m2:
            try:
                v1 = float(m1.value) if isinstance(m1.value, (int, float)) else 0
                v2 = float(m2.value) if isinstance(m2.value, (int, float)) else 0
                change = ((v2 - v1) / v1 * 100) if v1 != 0 else 0
                metric_changes.append({
                    "metric": name,
                    "value_doc1": m1.raw_value,
                    "value_doc2": m2.raw_value,
                    "change_pct": round(change, 2),
                })
            except (ValueError, TypeError):
                pass
        elif m1:
            metric_changes.append({
                "metric": name,
                "value_doc1": m1.raw_value,
                "value_doc2": None,
                "status": "removed",
            })
        elif m2:
            metric_changes.append({
                "metric": name,
                "value_doc1": None,
                "value_doc2": m2.raw_value,
                "status": "added",
            })

    prompt = f"""Compare these two document excerpts and identify key differences.

Document 1:
{text1[:3000]}

Document 2:
{text2[:3000]}

Return JSON:
```json
{{
  "similarity_score": 0.75,
  "additions": [
    {{"location": "Section 3", "content": "New paragraph about...", "significance": "high"}}
  ],
  "deletions": [
    {{"location": "Section 2", "content": "Removed reference to...", "significance": "medium"}}
  ],
  "modifications": [
    {{"location": "Section 1", "original": "Revenue was $1M", "modified": "Revenue was $1.2M", "significance": "high"}}
  ],
  "change_summary": "Brief summary of overall changes",
  "significant_changes": ["Key change 1", "Key change 2"]
}}
```"""

    try:
        client = get_llm_client()
        response = call_chat_completion(
            client,
            model=None,
            messages=[{"role": "user", "content": prompt}],
            description="document_comparison",
            temperature=0.2,
        )

        raw_text = response.choices[0].message.content or ""
        data = extract_json_from_llm_response(raw_text, default=None)

        if data:
            return ComparativeAnalysis(
                comparison_type="version_diff",
                documents_compared=["document_1", "document_2"],
                additions=data.get("additions", []),
                deletions=data.get("deletions", []),
                modifications=data.get("modifications", []),
                metric_changes=metric_changes,
                similarity_score=data.get("similarity_score", 0.5),
                change_summary=data.get("change_summary", ""),
                significant_changes=data.get("significant_changes", []),
            )
    except Exception as e:
        logger.warning(f"Document comparison failed: {e}")

    return ComparativeAnalysis(
        comparison_type="version_diff",
        documents_compared=["document_1", "document_2"],
        metric_changes=metric_changes,
    )


# ANALYSIS ENGINE ORCHESTRATOR

class AnalysisEngineService:
    """Orchestrates all analysis engines."""

    def run_all_analyses(
        self,
        text: str,
        tables: List[EnhancedExtractedTable],
        metrics: List[ExtractedMetric],
    ) -> Dict[str, Any]:
        """
        Run all analysis engines.

        Uses a single consolidated LLM call for summary + sentiment +
        financial + insights instead of 10+ separate calls.
        """
        # ---- Pure-Python analyses (no LLM) ----
        text_analytics = analyze_text(text)
        statistical = analyze_statistics(tables)

        # ---- Single consolidated LLM call ----
        llm_results = self._run_consolidated_analysis(text, tables, metrics)

        return {
            "summaries": llm_results.get("summaries", {}),
            "sentiment": llm_results.get("sentiment", SentimentAnalysis(
                overall_sentiment=SentimentLevel.NEUTRAL,
                overall_score=0.0, confidence=0.5,
            )),
            "text_analytics": text_analytics,
            "statistical_analysis": statistical,
            "financial_analysis": llm_results.get("financial_analysis", FinancialAnalysis(
                metrics_found=len(metrics),
            )),
            "insights": llm_results.get("insights", []),
            "risks": llm_results.get("risks", []),
            "opportunities": llm_results.get("opportunities", []),
            "action_items": llm_results.get("action_items", []),
        }

    def _run_consolidated_analysis(
        self,
        text: str,
        tables: List[EnhancedExtractedTable],
        metrics: List[ExtractedMetric],
    ) -> Dict[str, Any]:
        """
        Perform comprehensive analysis in ONE LLM call.

        Combines: executive summary, comprehensive summary, sentiment,
        financial analysis, insights, risks, opportunities, and action items.
        """
        # Build context
        metrics_context = "\n".join([
            f"- {m.name}: {m.raw_value}" + (f" ({m.change}% {m.change_direction})" if m.change else "")
            for m in metrics[:20]
        ]) or "(no metrics extracted)"

        table_info = ""
        for t in tables[:5]:
            table_info += f"- {t.title or t.id}: {t.row_count} rows, columns: {', '.join(t.headers[:8])}\n"
        table_info = table_info or "(no tables found)"

        prompt = f"""Analyze this document comprehensively. Return a SINGLE JSON object with all analyses.

DOCUMENT TEXT:
{text[:10000]}

TABLES FOUND:
{table_info}

METRICS EXTRACTED:
{metrics_context}

Return a JSON object with these keys:

{{
  "executive_summary": {{
    "title": "Brief title",
    "content": "2-3 paragraph executive summary for C-suite",
    "bullet_points": ["Key point 1", "Key point 2", "Key point 3"]
  }},
  "comprehensive_summary": {{
    "title": "Document title",
    "content": "Detailed structured summary covering all major sections",
    "bullet_points": ["Finding 1", "Finding 2", "Finding 3", "Finding 4", "Finding 5"]
  }},
  "sentiment": {{
    "overall_sentiment": "positive|negative|neutral|very_positive|very_negative",
    "overall_score": 0.5,
    "confidence": 0.85,
    "emotional_tone": "formal|casual|urgent|optimistic|pessimistic|neutral|analytical",
    "urgency_level": "low|normal|high|critical"
  }},
  "financial": {{
    "currency": "USD",
    "revenue_growth": null,
    "profit_growth": null,
    "insights": ["Financial insight 1"],
    "warnings": ["Financial warning 1"]
  }},
  "insights": [
    {{
      "type": "finding|trend|anomaly|recommendation|warning",
      "title": "Short title",
      "description": "Detailed description",
      "priority": "critical|high|medium|low",
      "confidence": 0.85,
      "actionable": true,
      "suggested_actions": ["Action 1"]
    }}
  ],
  "risks": [
    {{
      "title": "Risk title",
      "description": "Risk description",
      "risk_level": "critical|high|medium|low|minimal",
      "category": "financial|operational|compliance|market",
      "probability": 0.7,
      "impact": 0.8,
      "mitigation_suggestions": ["Suggestion 1"]
    }}
  ],
  "opportunities": [
    {{
      "title": "Opportunity title",
      "description": "Description",
      "opportunity_type": "growth|efficiency|cost_saving|innovation",
      "potential_value": "$500K",
      "confidence": 0.75,
      "suggested_actions": ["Action 1"]
    }}
  ],
  "action_items": [
    {{
      "title": "Action title",
      "description": "Description",
      "priority": "critical|high|medium|low",
      "category": "financial|operational|strategic",
      "expected_outcome": "Expected result"
    }}
  ]
}}

IMPORTANT:
- Be specific and actionable, grounded in the actual document content
- Include at least 3 insights, 2 risks, 2 opportunities, and 2 action items
- For financial fields, only include data you can calculate from the document (use null otherwise)
- Return ONLY valid JSON, no extra text"""

        try:
            client = get_llm_client()
            response = call_chat_completion(
                client,
                model=None,
                messages=[{"role": "user", "content": prompt}],
                description="consolidated_analysis",
                temperature=0.3,
            )

            raw_text = response.choices[0].message.content or ""
            data = extract_json_from_llm_response(raw_text, default=None)

            if not data:
                raise ValueError("LLM returned no parseable JSON for analysis")

            return self._parse_consolidated_result(data, metrics)

        except Exception as e:
            logger.error("Consolidated analysis failed: %s", e, exc_info=True)
            raise RuntimeError(f"AI analysis failed: {e}") from e

    def _parse_consolidated_result(
        self,
        data: Dict[str, Any],
        metrics: List[ExtractedMetric],
    ) -> Dict[str, Any]:
        """Parse the consolidated LLM response into typed structures."""
        result: Dict[str, Any] = {}

        # --- Summaries ---
        summaries: Dict[str, DocumentSummary] = {}
        for key, mode in [("executive_summary", SummaryMode.EXECUTIVE),
                          ("comprehensive_summary", SummaryMode.COMPREHENSIVE)]:
            s = data.get(key, {})
            if isinstance(s, dict):
                content = s.get("content", "")
                words = len(content.split())
                summaries[mode.value] = DocumentSummary(
                    mode=mode,
                    title=s.get("title", f"{mode.value.title()} Summary"),
                    content=content,
                    bullet_points=s.get("bullet_points", []),
                    word_count=words,
                    reading_time_minutes=max(1, words / 200),
                )
        result["summaries"] = summaries

        # --- Sentiment ---
        sent = data.get("sentiment", {})
        if isinstance(sent, dict):
            sentiment_map = {
                "very_positive": SentimentLevel.VERY_POSITIVE,
                "positive": SentimentLevel.POSITIVE,
                "neutral": SentimentLevel.NEUTRAL,
                "negative": SentimentLevel.NEGATIVE,
                "very_negative": SentimentLevel.VERY_NEGATIVE,
            }
            result["sentiment"] = SentimentAnalysis(
                overall_sentiment=sentiment_map.get(
                    str(sent.get("overall_sentiment", "neutral")).lower(),
                    SentimentLevel.NEUTRAL
                ),
                overall_score=float(sent.get("overall_score", 0)),
                confidence=float(sent.get("confidence", 0.8)),
                emotional_tone=sent.get("emotional_tone", "neutral"),
                urgency_level=sent.get("urgency_level", "normal"),
            )

        # --- Financial ---
        fin = data.get("financial", {})
        if isinstance(fin, dict):
            result["financial_analysis"] = FinancialAnalysis(
                metrics_found=len(metrics),
                currency=fin.get("currency", "USD"),
                gross_margin=fin.get("gross_margin"),
                operating_margin=fin.get("operating_margin"),
                net_margin=fin.get("net_margin"),
                revenue_growth=fin.get("revenue_growth"),
                profit_growth=fin.get("profit_growth"),
                yoy_comparison=fin.get("yoy_comparison", {}),
                variance_analysis=fin.get("variance_analysis", []),
                insights=fin.get("insights", []),
                warnings=fin.get("warnings", []),
            )

        # --- Insights ---
        priority_map = {"critical": Priority.CRITICAL, "high": Priority.HIGH,
                        "medium": Priority.MEDIUM, "low": Priority.LOW}
        result["insights"] = []
        for item in data.get("insights", []):
            if isinstance(item, dict):
                result["insights"].append(Insight(
                    id=f"ins_{uuid.uuid4().hex[:8]}",
                    type=item.get("type", "finding"),
                    title=item.get("title", ""),
                    description=item.get("description", ""),
                    priority=priority_map.get(str(item.get("priority", "medium")).lower(), Priority.MEDIUM),
                    confidence=float(item.get("confidence", 0.8)),
                    supporting_data=item.get("supporting_data", []),
                    actionable=bool(item.get("actionable", False)),
                    suggested_actions=item.get("suggested_actions", []),
                ))

        # --- Risks ---
        risk_map = {"critical": RiskLevel.CRITICAL, "high": RiskLevel.HIGH,
                    "medium": RiskLevel.MEDIUM, "low": RiskLevel.LOW, "minimal": RiskLevel.MINIMAL}
        result["risks"] = []
        for item in data.get("risks", []):
            if isinstance(item, dict):
                prob = float(item.get("probability", 0.5))
                impact = float(item.get("impact", 0.5))
                result["risks"].append(RiskItem(
                    id=f"risk_{uuid.uuid4().hex[:8]}",
                    title=item.get("title", ""),
                    description=item.get("description", ""),
                    risk_level=risk_map.get(str(item.get("risk_level", "medium")).lower(), RiskLevel.MEDIUM),
                    category=item.get("category", "general"),
                    probability=prob,
                    impact=impact,
                    risk_score=prob * impact,
                    mitigation_suggestions=item.get("mitigation_suggestions", []),
                ))

        # --- Opportunities ---
        result["opportunities"] = []
        for item in data.get("opportunities", []):
            if isinstance(item, dict):
                result["opportunities"].append(OpportunityItem(
                    id=f"opp_{uuid.uuid4().hex[:8]}",
                    title=item.get("title", ""),
                    description=item.get("description", ""),
                    opportunity_type=item.get("opportunity_type", "growth"),
                    potential_value=item.get("potential_value"),
                    confidence=float(item.get("confidence", 0.7)),
                    requirements=item.get("requirements", []),
                    suggested_actions=item.get("suggested_actions", []),
                ))

        # --- Action Items ---
        result["action_items"] = []
        for item in data.get("action_items", []):
            if isinstance(item, dict):
                result["action_items"].append(ActionItem(
                    id=f"act_{uuid.uuid4().hex[:8]}",
                    title=item.get("title", ""),
                    description=item.get("description", ""),
                    priority=priority_map.get(str(item.get("priority", "medium")).lower(), Priority.MEDIUM),
                    category=item.get("category", "general"),
                    expected_outcome=item.get("expected_outcome"),
                ))

        return result


# ADVANCED AI FEATURES (merged from advanced_ai_features.py)

logger = logging.getLogger("neura.analyze.advanced_ai")


# 8.1 MULTI-MODAL UNDERSTANDING

@dataclass
class ImageAnalysisResult:
    """Result of image analysis within a document."""
    image_type: str  # chart, diagram, photo, logo, signature, handwriting
    description: str
    extracted_data: Dict[str, Any] = field(default_factory=dict)
    confidence: float = 0.8
    location: Optional[str] = None  # page number or position


@dataclass
class ChartDataExtraction:
    """Extracted data from a chart image."""
    chart_type: str
    title: Optional[str]
    x_axis: Dict[str, Any]
    y_axis: Dict[str, Any]
    data_series: List[Dict[str, Any]]
    insights: List[str]


def analyze_document_images(
    images: List[Any],
    document_context: str = "",
) -> List[ImageAnalysisResult]:
    """Analyze images found in a document using VLM."""
    results = []
    if not images:
        return results

    vlm = None
    try:
        from backend.app.services.llm import VisionLanguageModel
        vlm = VisionLanguageModel()
    except Exception as exc:
        logger.warning(f"Vision model unavailable for image analysis: {exc}")

    # For each image, determine type and extract relevant data
    for img in images:
        img_data = None
        page = None
        if isinstance(img, dict):
            img_data = img.get("data") or img.get("bytes") or img.get("image")
            page = img.get("page")
        else:
            img_data = img

        if not img_data:
            continue

        prompt = f"""Analyze this image from a document.

Document context: {document_context[:500]}

Determine:
1. Image type (chart, diagram, photo, logo, signature, handwriting, table, other)
2. Detailed description of content
3. Any extractable data (for charts: data points, for forms: field values, etc.)

Return JSON:
```json
{{
  "image_type": "chart|diagram|photo|logo|signature|handwriting|table|other",
  "description": "Detailed description",
  "extracted_data": {{
    "chart_type": "bar",
    "data_points": [{{"label": "Q1", "value": 100}}],
    "title": "Revenue by Quarter"
  }},
  "confidence": 0.85,
  "key_information": ["Item 1", "Item 2"]
}}
```"""

        try:
            parsed: Dict[str, Any] = {}
            if vlm:
                response = vlm.client.complete_with_vision(
                    text=prompt,
                    images=[img_data],
                    model=vlm.model,
                    description="vlm_image_analysis",
                )
                raw_content = response["choices"][0]["message"]["content"]
                parsed = _extract_json_payload(raw_content, {})

            image_type = parsed.get("image_type", "unknown")
            description = parsed.get("description", "Analyzed image")
            extracted_data = parsed.get("extracted_data", {}) or {}
            confidence = float(parsed.get("confidence", 0.7) or 0.7)

            # Optional sub-analyses
            if image_type in ("chart", "diagram"):
                chart_data = extract_chart_data_from_image(img_data)
                if chart_data:
                    extracted_data["chart_data"] = chart_data.__dict__
            if image_type in ("handwriting", "signature"):
                extracted_data["handwriting"] = detect_handwriting(img_data)
            if image_type == "logo":
                extracted_data["logos"] = detect_logos(img_data)

            results.append(ImageAnalysisResult(
                image_type=image_type,
                description=description,
                extracted_data=extracted_data,
                confidence=confidence,
                location=f"Page {page}" if page else None,
            ))
        except Exception as e:
            logger.warning(f"Image analysis failed: {e}")

    return results


def extract_chart_data_from_image(image_data: bytes) -> Optional[ChartDataExtraction]:
    """Extract structured data from a chart image."""
    prompt = """Analyze this chart image and extract all data.

Return JSON:
```json
{
  "chart_type": "bar|line|pie|scatter|area|other",
  "title": "Chart title if visible",
  "x_axis": {
    "label": "X axis label",
    "values": ["Jan", "Feb", "Mar"],
    "type": "category|numeric|date"
  },
  "y_axis": {
    "label": "Y axis label",
    "min": 0,
    "max": 100,
    "type": "numeric"
  },
  "data_series": [
    {
      "name": "Series 1",
      "values": [10, 20, 30],
      "color": "blue"
    }
  ],
  "insights": [
    "Key observation 1",
    "Key observation 2"
  ]
}
```

Extract as much data as you can accurately determine from the image."""

    try:
        vlm = VisionLanguageModel()
        result = vlm.analyze_chart(image_data)
        return ChartDataExtraction(
            chart_type=result.get("chart_type", "unknown"),
            title=result.get("title"),
            x_axis=result.get("x_axis", {}),
            y_axis=result.get("y_axis", {}),
            data_series=result.get("data_series", []),
            insights=result.get("insights", []) if isinstance(result.get("insights"), list) else [str(result.get("insights"))],
        )
    except Exception as exc:
        logger.warning(f"Chart extraction failed: {exc}")
        return None


def detect_handwriting(image_data: bytes) -> Dict[str, Any]:
    """Detect and transcribe handwritten text."""
    prompt = """Analyze this image for handwritten text.

Return JSON:
```json
{
  "has_handwriting": true,
  "transcribed_text": "Full transcription of handwritten content",
  "confidence": 0.75,
  "words": [
    {"text": "word", "confidence": 0.8, "position": {"x": 100, "y": 50}}
  ],
  "is_signature": false
}
```

Transcribe all visible handwritten text accurately."""

    try:
        vlm = VisionLanguageModel()
        response = vlm.client.complete_with_vision(
            text=prompt,
            images=[image_data],
            model=vlm.model,
            description="vlm_handwriting_detection",
        )
        raw_content = response["choices"][0]["message"]["content"]
        parsed = _extract_json_payload(raw_content, {})
        if parsed:
            return parsed
    except Exception as exc:
        logger.warning(f"Handwriting detection failed: {exc}")
    return {
        "has_handwriting": False,
        "transcribed_text": "",
        "confidence": 0.0,
        "words": [],
        "is_signature": False,
    }


def detect_logos(image_data: bytes) -> List[Dict[str, Any]]:
    """Detect and identify logos in an image."""
    prompt = """Analyze this image and detect any logos or brand marks.

Return JSON:
```json
{
  "logos": [
    {
      "name": "Brand or company name if known",
      "confidence": 0.85,
      "description": "Brief description of the logo",
      "position": {"x": 0, "y": 0, "width": 0, "height": 0}
    }
  ]
}
```"""

    try:
        vlm = VisionLanguageModel()
        response = vlm.client.complete_with_vision(
            text=prompt,
            images=[image_data],
            model=vlm.model,
            description="vlm_logo_detection",
        )
        raw_content = response["choices"][0]["message"]["content"]
        parsed = _extract_json_payload(raw_content, {})
        logos = parsed.get("logos", [])
        return logos if isinstance(logos, list) else []
    except Exception as exc:
        logger.warning(f"Logo detection failed: {exc}")
        return []


def _extract_json_payload(raw_content: str, default: Dict[str, Any]) -> Dict[str, Any]:
    """Extract JSON from an LLM response (handles Claude's markdown code blocks)."""
    json_match = re.search(r"```(?:json)?\s*([\s\S]*?)```", raw_content)
    if json_match:
        json_str = json_match.group(1).strip()
    else:
        json_str = raw_content.strip()

    start = json_str.find("{")
    if start == -1:
        return default

    depth = 0
    in_string = False
    escape_next = False
    for i, char in enumerate(json_str[start:], start):
        if escape_next:
            escape_next = False
            continue
        if char == "\\":
            escape_next = True
            continue
        if char == '"' and not escape_next:
            in_string = not in_string
            continue
        if in_string:
            continue
        if char == "{":
            depth += 1
        elif char == "}":
            depth -= 1
            if depth == 0:
                json_str = json_str[start:i + 1]
                break

    try:
        return json.loads(json_str)
    except Exception:
        return default


# 8.2 CROSS-DOCUMENT INTELLIGENCE

@dataclass
class KnowledgeGraphNode:
    """A node in the knowledge graph."""
    id: str
    label: str
    type: str  # entity, concept, metric, document
    properties: Dict[str, Any] = field(default_factory=dict)


@dataclass
class KnowledgeGraphEdge:
    """An edge in the knowledge graph."""
    source_id: str
    target_id: str
    relationship: str
    weight: float = 1.0
    properties: Dict[str, Any] = field(default_factory=dict)


@dataclass
class KnowledgeGraph:
    """A knowledge graph built from document analysis."""
    nodes: List[KnowledgeGraphNode] = field(default_factory=list)
    edges: List[KnowledgeGraphEdge] = field(default_factory=list)

    def add_node(self, node: KnowledgeGraphNode) -> None:
        if not any(n.id == node.id for n in self.nodes):
            self.nodes.append(node)

    def add_edge(self, edge: KnowledgeGraphEdge) -> None:
        self.edges.append(edge)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "nodes": [{"id": n.id, "label": n.label, "type": n.type, "properties": n.properties}
                      for n in self.nodes],
            "edges": [{"source": e.source_id, "target": e.target_id,
                       "relationship": e.relationship, "weight": e.weight}
                      for e in self.edges],
        }


def build_knowledge_graph(
    entities: List[ExtractedEntity],
    metrics: List[ExtractedMetric],
    document_id: str,
) -> KnowledgeGraph:
    """Build a knowledge graph from extracted entities and metrics."""
    graph = KnowledgeGraph()

    # Add document node
    doc_node = KnowledgeGraphNode(
        id=document_id,
        label=document_id,
        type="document",
    )
    graph.add_node(doc_node)

    # Add entity nodes
    for entity in entities:
        node = KnowledgeGraphNode(
            id=entity.id,
            label=entity.value,
            type=entity.type.value,
            properties={
                "normalized": entity.normalized_value,
                "confidence": entity.confidence,
            },
        )
        graph.add_node(node)

        # Connect to document
        graph.add_edge(KnowledgeGraphEdge(
            source_id=document_id,
            target_id=entity.id,
            relationship="contains",
            weight=entity.confidence,
        ))

    # Add metric nodes
    for metric in metrics:
        node = KnowledgeGraphNode(
            id=metric.id,
            label=metric.name,
            type="metric",
            properties={
                "value": metric.value,
                "raw_value": metric.raw_value,
                "metric_type": metric.metric_type.value,
                "period": metric.period,
            },
        )
        graph.add_node(node)

        # Connect to document
        graph.add_edge(KnowledgeGraphEdge(
            source_id=document_id,
            target_id=metric.id,
            relationship="reports",
            weight=metric.importance_score,
        ))

    # Find relationships between entities
    for i, e1 in enumerate(entities):
        for e2 in entities[i + 1:]:
            # If they appear in similar context, they might be related
            if e1.context and e2.context:
                # Simple proximity check
                if e1.value.lower() in (e2.context or "").lower() or e2.value.lower() in (e1.context or "").lower():
                    graph.add_edge(KnowledgeGraphEdge(
                        source_id=e1.id,
                        target_id=e2.id,
                        relationship="co_occurs",
                        weight=0.7,
                    ))

    return graph


def merge_knowledge_graphs(graphs: List[KnowledgeGraph]) -> KnowledgeGraph:
    """Merge multiple knowledge graphs into one."""
    merged = KnowledgeGraph()

    # Collect all nodes
    node_map: Dict[str, KnowledgeGraphNode] = {}
    for graph in graphs:
        for node in graph.nodes:
            # Merge by label for same-type nodes
            key = f"{node.type}:{node.label.lower()}"
            if key not in node_map:
                node_map[key] = node
            else:
                # Merge properties
                node_map[key].properties.update(node.properties)

    merged.nodes = list(node_map.values())

    # Collect all edges
    edge_set = set()
    for graph in graphs:
        for edge in graph.edges:
            edge_key = (edge.source_id, edge.target_id, edge.relationship)
            if edge_key not in edge_set:
                edge_set.add(edge_key)
                merged.add_edge(edge)

    return merged


@dataclass
class CitationLink:
    """A citation or reference link between documents."""
    source_doc_id: str
    target_doc_id: str
    citation_text: str
    citation_type: str  # reference, quote, data_source
    confidence: float = 0.8


def detect_citations(text: str, document_id: str) -> List[CitationLink]:
    """Detect citations and references in text."""
    citations = []

    # Common citation patterns
    patterns = [
        r'\[(\d+)\]',  # [1], [2]
        r'\(([A-Za-z]+(?:\s+et\s+al\.?)?,?\s*\d{4})\)',  # (Smith, 2023), (Smith et al., 2023)
        r'(?:Source|Reference|See|cf\.?):\s*(.+?)(?:\.|$)',  # Source: document name
        r'(?:According to|As stated in|Per)\s+(.+?)(?:,|\.)',  # According to X
    ]

    for pattern in patterns:
        for match in re.finditer(pattern, text, re.IGNORECASE):
            citations.append(CitationLink(
                source_doc_id=document_id,
                target_doc_id=f"ref_{uuid.uuid4().hex[:8]}",
                citation_text=match.group(0),
                citation_type="reference",
                confidence=0.7,
            ))

    return citations


@dataclass
class Contradiction:
    """A detected contradiction between statements or documents."""
    statement1: str
    statement2: str
    source1: str
    source2: str
    contradiction_type: str  # factual, numerical, temporal
    severity: str  # minor, moderate, major
    confidence: float = 0.7


def detect_contradictions(
    text1: str,
    text2: str,
    doc1_id: str = "doc1",
    doc2_id: str = "doc2",
) -> List[Contradiction]:
    """Detect contradictions between two texts using LLM."""
    prompt = f"""Compare these two texts and identify any contradictions, inconsistencies, or conflicting information.

Text 1:
{text1[:3000]}

Text 2:
{text2[:3000]}

Return JSON:
```json
{{
  "contradictions": [
    {{
      "statement1": "Quote or paraphrase from Text 1",
      "statement2": "Conflicting statement from Text 2",
      "type": "factual|numerical|temporal|logical",
      "severity": "minor|moderate|major",
      "explanation": "Why these are contradictory",
      "confidence": 0.8
    }}
  ],
  "overall_consistency": 0.85
}}
```

Only report genuine contradictions, not differences in scope or perspective."""

    try:
        client = get_llm_client()
        response = call_chat_completion(
            client,
            model=None,
            messages=[{"role": "user", "content": prompt}],
            description="contradiction_detection",
            temperature=0.2,
        )

        raw_text = response.choices[0].message.content or ""
        json_match = re.search(r'\{[\s\S]*\}', raw_text)

        if json_match:
            data = json.loads(json_match.group())
            contradictions = []

            for item in data.get("contradictions", []):
                contradictions.append(Contradiction(
                    statement1=item.get("statement1", ""),
                    statement2=item.get("statement2", ""),
                    source1=doc1_id,
                    source2=doc2_id,
                    contradiction_type=item.get("type", "factual"),
                    severity=item.get("severity", "moderate"),
                    confidence=item.get("confidence", 0.7),
                ))

            return contradictions

    except Exception as e:
        logger.warning(f"Contradiction detection failed: {e}")

    return []


# 8.3 PREDICTIVE ANALYTICS

@dataclass
class Forecast:
    """A forecast prediction."""
    metric_name: str
    current_value: float
    predictions: List[Dict[str, Any]]  # [{period, value, lower_bound, upper_bound}]
    trend: str  # increasing, decreasing, stable
    confidence: float
    method: str  # linear, exponential, seasonal
    factors: List[str] = field(default_factory=list)


@dataclass
class AnomalyPrediction:
    """Predicted anomaly or unusual pattern."""
    metric_name: str
    predicted_date: Optional[str]
    anomaly_type: str  # spike, dip, deviation
    probability: float
    expected_value: float
    threshold: float
    reasoning: str


@dataclass
class GrowthModel:
    """Growth model for a metric."""
    metric_name: str
    model_type: str  # linear, exponential, logistic, polynomial
    parameters: Dict[str, float]
    r_squared: float
    projected_values: List[Dict[str, Any]]
    saturation_point: Optional[float] = None


def forecast_time_series(
    data: List[Tuple[str, float]],  # [(date, value), ...]
    metric_name: str,
    periods: int = 6,
) -> Forecast:
    """Generate a forecast for time series data."""
    if len(data) < 3:
        return Forecast(
            metric_name=metric_name,
            current_value=data[-1][1] if data else 0,
            predictions=[],
            trend="unknown",
            confidence=0.3,
            method="insufficient_data",
        )

    values = [v for _, v in data]
    n = len(values)

    # Calculate trend using linear regression
    x = list(range(n))
    mean_x = sum(x) / n
    mean_y = sum(values) / n

    numerator = sum((xi - mean_x) * (yi - mean_y) for xi, yi in zip(x, values))
    denominator = sum((xi - mean_x) ** 2 for xi in x)

    if denominator == 0:
        slope = 0
        intercept = mean_y
    else:
        slope = numerator / denominator
        intercept = mean_y - slope * mean_x

    # Calculate R-squared
    ss_tot = sum((yi - mean_y) ** 2 for yi in values)
    ss_res = sum((yi - (slope * xi + intercept)) ** 2 for xi, yi in zip(x, values))
    r_squared = 1 - (ss_res / ss_tot) if ss_tot > 0 else 0

    # Determine trend
    if slope > 0.05 * mean_y:
        trend = "increasing"
    elif slope < -0.05 * mean_y:
        trend = "decreasing"
    else:
        trend = "stable"

    # Generate predictions
    predictions = []
    std_error = math.sqrt(ss_res / max(n - 2, 1)) if n > 2 else mean_y * 0.1

    for i in range(1, periods + 1):
        x_pred = n - 1 + i
        y_pred = slope * x_pred + intercept

        # Confidence interval (approximate)
        margin = 1.96 * std_error * math.sqrt(1 + 1 / n + (x_pred - mean_x) ** 2 / denominator) if denominator > 0 else y_pred * 0.2

        predictions.append({
            "period": i,
            "value": round(y_pred, 2),
            "lower_bound": round(y_pred - margin, 2),
            "upper_bound": round(y_pred + margin, 2),
        })

    confidence = min(0.95, max(0.3, r_squared))

    return Forecast(
        metric_name=metric_name,
        current_value=values[-1],
        predictions=predictions,
        trend=trend,
        confidence=round(confidence, 2),
        method="linear",
        factors=[f"Based on {n} historical data points", f"R² = {r_squared:.3f}"],
    )


def predict_anomalies(
    data: List[Tuple[str, float]],
    metric_name: str,
    sensitivity: float = 2.0,
) -> List[AnomalyPrediction]:
    """Predict potential anomalies based on historical patterns."""
    if len(data) < 10:
        return []

    values = [v for _, v in data]
    mean = sum(values) / len(values)
    std = math.sqrt(sum((v - mean) ** 2 for v in values) / len(values))

    if std == 0:
        return []

    # Calculate moving average and detect patterns
    window = min(5, len(values) // 2)
    predictions = []

    # Check for cyclic patterns (simplified)
    recent_values = values[-window:]
    recent_mean = sum(recent_values) / len(recent_values)

    # If recent trend differs significantly from overall
    if abs(recent_mean - mean) > std:
        if recent_mean > mean + std:
            predictions.append(AnomalyPrediction(
                metric_name=metric_name,
                predicted_date=None,
                anomaly_type="spike",
                probability=0.6,
                expected_value=recent_mean,
                threshold=mean + sensitivity * std,
                reasoning="Recent values significantly above historical average",
            ))
        else:
            predictions.append(AnomalyPrediction(
                metric_name=metric_name,
                predicted_date=None,
                anomaly_type="dip",
                probability=0.6,
                expected_value=recent_mean,
                threshold=mean - sensitivity * std,
                reasoning="Recent values significantly below historical average",
            ))

    # Check for increasing volatility
    recent_std = math.sqrt(sum((v - recent_mean) ** 2 for v in recent_values) / len(recent_values))
    if recent_std > std * 1.5:
        predictions.append(AnomalyPrediction(
            metric_name=metric_name,
            predicted_date=None,
            anomaly_type="deviation",
            probability=0.5,
            expected_value=recent_mean,
            threshold=recent_std,
            reasoning="Increased volatility in recent data",
        ))

    return predictions


def build_growth_model(
    data: List[Tuple[str, float]],
    metric_name: str,
    model_type: str = "auto",
) -> GrowthModel:
    """Build a growth model for a metric."""
    if len(data) < 5:
        return GrowthModel(
            metric_name=metric_name,
            model_type="insufficient_data",
            parameters={},
            r_squared=0,
            projected_values=[],
        )

    values = [v for _, v in data]
    n = len(values)
    x = list(range(n))

    # Try linear model
    mean_x = sum(x) / n
    mean_y = sum(values) / n

    num = sum((xi - mean_x) * (yi - mean_y) for xi, yi in zip(x, values))
    den = sum((xi - mean_x) ** 2 for xi in x)

    if den == 0:
        slope = 0
        intercept = mean_y
    else:
        slope = num / den
        intercept = mean_y - slope * mean_x

    # Calculate R-squared for linear
    ss_tot = sum((yi - mean_y) ** 2 for yi in values)
    ss_res_linear = sum((yi - (slope * xi + intercept)) ** 2 for xi, yi in zip(x, values))
    r2_linear = 1 - (ss_res_linear / ss_tot) if ss_tot > 0 else 0

    # Try exponential model (log-linear)
    positive_values = [max(0.01, v) for v in values]
    log_values = [math.log(v) for v in positive_values]
    mean_log_y = sum(log_values) / n

    num_exp = sum((xi - mean_x) * (yi - mean_log_y) for xi, yi in zip(x, log_values))
    if den > 0:
        exp_slope = num_exp / den
        exp_intercept = mean_log_y - exp_slope * mean_x

        predicted_exp = [math.exp(exp_slope * xi + exp_intercept) for xi in x]
        ss_res_exp = sum((yi - pi) ** 2 for yi, pi in zip(values, predicted_exp))
        r2_exp = 1 - (ss_res_exp / ss_tot) if ss_tot > 0 else 0
    else:
        r2_exp = 0
        exp_slope = 0
        exp_intercept = math.log(mean_y) if mean_y > 0 else 0

    # Choose best model
    if model_type == "auto":
        if r2_exp > r2_linear + 0.1 and exp_slope > 0:
            model_type = "exponential"
        else:
            model_type = "linear"

    if model_type == "exponential":
        parameters = {"growth_rate": exp_slope, "initial_value": math.exp(exp_intercept)}
        r_squared = r2_exp

        projected = []
        for i in range(1, 7):
            x_pred = n - 1 + i
            y_pred = math.exp(exp_slope * x_pred + exp_intercept)
            projected.append({"period": i, "value": round(y_pred, 2)})

        # Estimate saturation (if growth slowing)
        growth_rates = [values[i] / values[i - 1] if values[i - 1] > 0 else 1 for i in range(1, n)]
        if len(growth_rates) >= 3:
            recent_growth = sum(growth_rates[-3:]) / 3
            early_growth = sum(growth_rates[:3]) / 3
            if recent_growth < early_growth * 0.8:
                saturation = values[-1] * (1 / (1 - recent_growth)) if recent_growth < 1 else None
            else:
                saturation = None
        else:
            saturation = None

    else:  # linear
        parameters = {"slope": slope, "intercept": intercept}
        r_squared = r2_linear
        saturation = None

        projected = []
        for i in range(1, 7):
            x_pred = n - 1 + i
            y_pred = slope * x_pred + intercept
            projected.append({"period": i, "value": round(y_pred, 2)})

    return GrowthModel(
        metric_name=metric_name,
        model_type=model_type,
        parameters=parameters,
        r_squared=round(r_squared, 4),
        projected_values=projected,
        saturation_point=saturation,
    )


def generate_ai_predictions(
    metrics: List[ExtractedMetric],
    tables: List[EnhancedExtractedTable],
) -> Dict[str, Any]:
    """Generate AI-powered predictions using LLM."""
    # Build context
    metrics_context = "\n".join([
        f"- {m.name}: {m.raw_value}" + (f" ({m.change}% change)" if m.change else "")
        for m in metrics[:15]
    ])

    prompt = f"""Based on these metrics, provide strategic predictions and insights.

Metrics:
{metrics_context}

Return JSON:
```json
{{
  "predictions": [
    {{
      "metric": "Revenue",
      "prediction": "Expected to grow 15-20% based on current trajectory",
      "confidence": "medium",
      "timeframe": "next 6 months",
      "factors": ["Market expansion", "New product launch"]
    }}
  ],
  "strategic_insights": [
    "Key strategic observation 1",
    "Key strategic observation 2"
  ],
  "risk_indicators": [
    {{
      "indicator": "Declining margins",
      "severity": "moderate",
      "recommendation": "Review cost structure"
    }}
  ],
  "growth_opportunities": [
    {{
      "opportunity": "Market segment X",
      "potential": "20% revenue increase",
      "requirements": ["Investment needed", "Timeline"]
    }}
  ]
}}
```

Be specific and data-driven in your predictions."""

    try:
        client = get_llm_client()
        response = call_chat_completion(
            client,
            model=None,
            messages=[{"role": "user", "content": prompt}],
            description="ai_predictions",
            temperature=0.4,
        )

        raw_text = response.choices[0].message.content or ""
        json_match = re.search(r'\{[\s\S]*\}', raw_text)

        if json_match:
            return json.loads(json_match.group())

    except Exception as e:
        logger.warning(f"AI predictions failed: {e}")

    return {"predictions": [], "strategic_insights": [], "risk_indicators": [], "growth_opportunities": []}


# ADVANCED AI SERVICE ORCHESTRATOR

class AdvancedAIService:
    """Orchestrates all advanced AI features."""

    def analyze_images(
        self,
        images: List[Any],
        document_context: str = "",
    ) -> List[ImageAnalysisResult]:
        """Analyze images in the document."""
        return analyze_document_images(images, document_context)

    def build_knowledge_graph(
        self,
        entities: List[ExtractedEntity],
        metrics: List[ExtractedMetric],
        document_id: str,
    ) -> KnowledgeGraph:
        """Build a knowledge graph from extracted data."""
        return build_knowledge_graph(entities, metrics, document_id)

    def detect_citations(self, text: str, document_id: str) -> List[CitationLink]:
        """Detect citations in text."""
        return detect_citations(text, document_id)

    def detect_contradictions(
        self,
        text1: str,
        text2: str,
        doc1_id: str = "doc1",
        doc2_id: str = "doc2",
    ) -> List[Contradiction]:
        """Detect contradictions between texts."""
        return detect_contradictions(text1, text2, doc1_id, doc2_id)

    def generate_forecasts(
        self,
        tables: List[EnhancedExtractedTable],
    ) -> List[Forecast]:
        """Generate forecasts for time series data in tables."""
        forecasts = []

        for table in tables:
            # Find datetime and numeric column pairs
            datetime_cols = [i for i, d in enumerate(table.data_types) if d == "datetime"]
            numeric_cols = [i for i, d in enumerate(table.data_types) if d == "numeric"]

            if not datetime_cols or not numeric_cols:
                continue

            date_idx = datetime_cols[0]

            for num_idx in numeric_cols[:3]:  # Limit to 3 numeric columns
                # Extract time series data
                data = []
                for row in table.rows:
                    if date_idx < len(row) and num_idx < len(row):
                        date_val = str(row[date_idx])
                        try:
                            num_val = float(str(row[num_idx]).replace(",", "").replace("$", "").replace("%", ""))
                            data.append((date_val, num_val))
                        except (ValueError, TypeError):
                            pass

                if len(data) >= 3:
                    forecast = forecast_time_series(
                        data,
                        table.headers[num_idx],
                        periods=6,
                    )
                    forecasts.append(forecast)

        return forecasts

    def predict_anomalies(
        self,
        tables: List[EnhancedExtractedTable],
    ) -> List[AnomalyPrediction]:
        """Predict anomalies in the data."""
        all_predictions = []

        for table in tables:
            for col_idx, (header, dtype) in enumerate(zip(table.headers, table.data_types)):
                if dtype != "numeric":
                    continue

                # Extract values
                data = []
                for i, row in enumerate(table.rows):
                    if col_idx < len(row):
                        try:
                            val = float(str(row[col_idx]).replace(",", "").replace("$", "").replace("%", ""))
                            data.append((str(i), val))
                        except (ValueError, TypeError):
                            pass

                if len(data) >= 10:
                    predictions = predict_anomalies(data, header)
                    all_predictions.extend(predictions)

        return all_predictions

    def build_growth_models(
        self,
        tables: List[EnhancedExtractedTable],
    ) -> List[GrowthModel]:
        """Build growth models for metrics."""
        models = []

        for table in tables:
            numeric_cols = [(i, h) for i, (h, d) in enumerate(zip(table.headers, table.data_types)) if d == "numeric"]

            for col_idx, header in numeric_cols[:3]:
                data = []
                for i, row in enumerate(table.rows):
                    if col_idx < len(row):
                        try:
                            val = float(str(row[col_idx]).replace(",", "").replace("$", "").replace("%", ""))
                            data.append((str(i), val))
                        except (ValueError, TypeError):
                            pass

                if len(data) >= 5:
                    model = build_growth_model(data, header)
                    if model.r_squared > 0.5:  # Only include models with reasonable fit
                        models.append(model)

        return models

    def generate_ai_predictions(
        self,
        metrics: List[ExtractedMetric],
        tables: List[EnhancedExtractedTable],
    ) -> Dict[str, Any]:
        """Generate AI-powered strategic predictions."""
        return generate_ai_predictions(metrics, tables)

    def run_all_advanced_features(
        self,
        text: str,
        entities: List[ExtractedEntity],
        metrics: List[ExtractedMetric],
        tables: List[EnhancedExtractedTable],
        document_id: str,
        images: Optional[List[Any]] = None,
        document_context: str = "",
    ) -> Dict[str, Any]:
        """Run all advanced AI features."""
        results = {
            "knowledge_graph": self.build_knowledge_graph(entities, metrics, document_id).to_dict(),
            "citations": [c.__dict__ for c in self.detect_citations(text, document_id)],
            "forecasts": [f.__dict__ for f in self.generate_forecasts(tables)],
            "anomaly_predictions": [a.__dict__ for a in self.predict_anomalies(tables)],
            "growth_models": [m.__dict__ for m in self.build_growth_models(tables)],
            "ai_predictions": self.generate_ai_predictions(metrics, tables),
        }
        if images:
            results["image_analysis"] = [
                r.__dict__ for r in self.analyze_images(images, document_context=document_context)
            ]
        return results


# VISUALIZATION ENGINE (merged from visualization_engine.py)

logger = logging.getLogger("neura.analyze.visualization")


# DATA PATTERN DETECTION

class DataPattern:
    """Detected data pattern for chart recommendation."""
    def __init__(
        self,
        pattern_type: str,
        columns: List[str],
        recommended_charts: List[ChartType],
        confidence: float,
        description: str,
    ):
        self.pattern_type = pattern_type
        self.columns = columns
        self.recommended_charts = recommended_charts
        self.confidence = confidence
        self.description = description


def detect_data_patterns(table: EnhancedExtractedTable) -> List[DataPattern]:
    """Detect data patterns in a table for chart recommendations."""
    patterns = []

    datetime_cols = []
    numeric_cols = []
    categorical_cols = []

    for idx, (header, dtype) in enumerate(zip(table.headers, table.data_types)):
        if dtype == "datetime":
            datetime_cols.append(header)
        elif dtype == "numeric":
            numeric_cols.append(header)
        else:
            # Check if categorical (limited unique values)
            unique_vals = set()
            for row in table.rows[:100]:
                if idx < len(row):
                    unique_vals.add(str(row[idx]))
            if len(unique_vals) <= 20:
                categorical_cols.append(header)

    # Time series pattern
    if datetime_cols and numeric_cols:
        patterns.append(DataPattern(
            pattern_type="time_series",
            columns=datetime_cols + numeric_cols,
            recommended_charts=[ChartType.LINE, ChartType.AREA, ChartType.BAR],
            confidence=0.9,
            description=f"Time series data with {len(numeric_cols)} numeric variables over time",
        ))

    # Category comparison pattern
    if categorical_cols and numeric_cols:
        patterns.append(DataPattern(
            pattern_type="category_comparison",
            columns=categorical_cols + numeric_cols,
            recommended_charts=[ChartType.BAR, ChartType.PIE, ChartType.TREEMAP],
            confidence=0.85,
            description=f"Categorical data with {len(numeric_cols)} metrics per category",
        ))

    # Distribution pattern
    if numeric_cols and len(table.rows) >= 10:
        patterns.append(DataPattern(
            pattern_type="distribution",
            columns=numeric_cols,
            recommended_charts=[ChartType.HISTOGRAM, ChartType.BOX],
            confidence=0.75,
            description=f"Numeric distribution analysis for {len(numeric_cols)} variables",
        ))

    # Correlation pattern (multiple numeric columns)
    if len(numeric_cols) >= 2:
        patterns.append(DataPattern(
            pattern_type="correlation",
            columns=numeric_cols,
            recommended_charts=[ChartType.SCATTER, ChartType.BUBBLE, ChartType.HEATMAP],
            confidence=0.7,
            description=f"Potential correlations between {len(numeric_cols)} numeric variables",
        ))

    # Hierarchical pattern
    if len(categorical_cols) >= 2 and numeric_cols:
        patterns.append(DataPattern(
            pattern_type="hierarchy",
            columns=categorical_cols + numeric_cols[:1],
            recommended_charts=[ChartType.SUNBURST, ChartType.TREEMAP],
            confidence=0.65,
            description="Hierarchical categorical data",
        ))

    # Progress/funnel pattern
    if len(numeric_cols) >= 3 and len(table.rows) <= 10:
        patterns.append(DataPattern(
            pattern_type="funnel",
            columns=numeric_cols,
            recommended_charts=[ChartType.FUNNEL, ChartType.WATERFALL],
            confidence=0.6,
            description="Sequential stage data suitable for funnel visualization",
        ))

    return patterns


# AUTO CHART GENERATION

def auto_generate_charts(
    tables: List[EnhancedExtractedTable],
    max_charts: int = 10,
) -> List[EnhancedChartSpec]:
    """Automatically generate charts based on detected data patterns."""
    charts = []

    for table in tables:
        if len(charts) >= max_charts:
            break

        patterns = detect_data_patterns(table)

        for pattern in patterns:
            if len(charts) >= max_charts:
                break

            chart = _create_chart_from_pattern(table, pattern)
            if chart:
                charts.append(chart)

    return charts


def _create_chart_from_pattern(
    table: EnhancedExtractedTable,
    pattern: DataPattern,
) -> Optional[EnhancedChartSpec]:
    """Create a chart specification from a detected pattern."""
    if not pattern.recommended_charts:
        return None

    chart_type = pattern.recommended_charts[0]

    # Determine x and y fields based on pattern type
    datetime_cols = [h for h, d in zip(table.headers, table.data_types) if d == "datetime"]
    numeric_cols = [h for h, d in zip(table.headers, table.data_types) if d == "numeric"]
    categorical_cols = [h for h, d in zip(table.headers, table.data_types) if d == "text"]

    x_field = ""
    y_fields = []

    if pattern.pattern_type == "time_series":
        x_field = datetime_cols[0] if datetime_cols else (categorical_cols[0] if categorical_cols else table.headers[0])
        y_fields = numeric_cols[:3]
    elif pattern.pattern_type == "category_comparison":
        x_field = categorical_cols[0] if categorical_cols else table.headers[0]
        y_fields = numeric_cols[:2]
    elif pattern.pattern_type == "correlation":
        x_field = numeric_cols[0]
        y_fields = [numeric_cols[1]] if len(numeric_cols) > 1 else []
    elif pattern.pattern_type == "distribution":
        x_field = numeric_cols[0]
        y_fields = []
    else:
        x_field = table.headers[0]
        y_fields = numeric_cols[:2] if numeric_cols else []

    if not x_field:
        return None

    # Build data from table
    data = []
    for row in table.rows[:500]:
        record = {}
        for idx, header in enumerate(table.headers):
            if idx < len(row):
                record[header] = row[idx]
        data.append(record)

    # Generate title
    title = f"{pattern.pattern_type.replace('_', ' ').title()}: {table.title or table.id}"

    return EnhancedChartSpec(
        id=f"auto_{chart_type.value}_{uuid.uuid4().hex[:8]}",
        type=chart_type,
        title=title,
        description=pattern.description,
        x_field=x_field,
        y_fields=y_fields,
        data=data,
        x_axis_label=x_field,
        y_axis_label=y_fields[0] if y_fields else None,
        show_legend=len(y_fields) > 1,
        ai_insights=[],
        source_table_id=table.id,
        confidence=pattern.confidence,
        suggested_by_ai=True,
    )


# NATURAL LANGUAGE CHART GENERATION

def generate_chart_from_natural_language(
    query: str,
    tables: List[EnhancedExtractedTable],
    metrics: List[ExtractedMetric] = None,
) -> List[EnhancedChartSpec]:
    """Generate charts from natural language query."""
    # Build context about available data
    context_parts = []
    for table in tables[:5]:
        context_parts.append(f"""Table: {table.title or table.id}
Columns: {', '.join([f'{h} ({d})' for h, d in zip(table.headers, table.data_types)])}
Rows: {table.row_count}
Sample: {table.rows[0] if table.rows else 'N/A'}""")

    context = "\n\n".join(context_parts)

    prompt = f"""Generate chart specifications based on this request.

User request: "{query}"

Available data:
{context}

Generate 1-3 appropriate charts. Return JSON array:
```json
[
  {{
    "chart_type": "line|bar|pie|scatter|area|histogram|box|heatmap|treemap|funnel|radar|bubble|sunburst|waterfall|gauge",
    "title": "Chart title",
    "description": "What this chart shows",
    "x_field": "column_name",
    "y_fields": ["column1", "column2"],
    "group_field": null,
    "x_axis_label": "X axis label",
    "y_axis_label": "Y axis label",
    "show_legend": true,
    "source_table": "table_id",
    "rationale": "Why this visualization is appropriate"
  }}
]
```

Match column names exactly as provided. Choose the most appropriate chart type for the data and request."""

    try:
        client = get_llm_client()
        response = call_chat_completion(
            client,
            model=None,
            messages=[{"role": "user", "content": prompt}],
            description="nl_chart_generation",
            temperature=0.3,
        )

        raw_text = response.choices[0].message.content or ""
        json_match = re.search(r'\[[\s\S]*\]', raw_text)

        if json_match:
            specs = json.loads(json_match.group())
            charts = []

            for spec in specs:
                try:
                    chart_type = ChartType[spec.get("chart_type", "bar").upper()]
                except KeyError:
                    chart_type = ChartType.BAR

                # Find source table and get data
                source_table_id = spec.get("source_table")
                source_table = next((t for t in tables if t.id == source_table_id or t.title == source_table_id), None)

                if not source_table and tables:
                    source_table = tables[0]

                data = []
                if source_table:
                    for row in source_table.rows[:500]:
                        record = {}
                        for idx, header in enumerate(source_table.headers):
                            if idx < len(row):
                                record[header] = row[idx]
                        data.append(record)

                charts.append(EnhancedChartSpec(
                    id=f"nl_{chart_type.value}_{uuid.uuid4().hex[:8]}",
                    type=chart_type,
                    title=spec.get("title", "Generated Chart"),
                    description=spec.get("description"),
                    x_field=spec.get("x_field", ""),
                    y_fields=spec.get("y_fields", []),
                    group_field=spec.get("group_field"),
                    data=data,
                    x_axis_label=spec.get("x_axis_label"),
                    y_axis_label=spec.get("y_axis_label"),
                    show_legend=spec.get("show_legend", True),
                    source_table_id=source_table.id if source_table else None,
                    ai_insights=[spec.get("rationale", "")],
                    confidence=0.85,
                    suggested_by_ai=True,
                ))

            return charts

    except Exception as e:
        logger.warning(f"NL chart generation failed: {e}")

    return []


# CHART INTELLIGENCE (TRENDS, ANOMALIES, FORECASTS)

def add_trend_line(chart: EnhancedChartSpec) -> EnhancedChartSpec:
    """Add trend line to a chart."""
    if not chart.data or not chart.y_fields:
        return chart

    y_field = chart.y_fields[0]
    x_field = chart.x_field

    # Extract numeric values
    values = []
    for i, record in enumerate(chart.data):
        try:
            y_val = float(str(record.get(y_field, 0)).replace(",", "").replace("$", "").replace("%", ""))
            values.append((i, y_val))
        except (ValueError, TypeError):
            pass

    if len(values) < 3:
        return chart

    # Simple linear regression
    n = len(values)
    sum_x = sum(x for x, _ in values)
    sum_y = sum(y for _, y in values)
    sum_xy = sum(x * y for x, y in values)
    sum_xx = sum(x * x for x, _ in values)

    denominator = n * sum_xx - sum_x * sum_x
    if denominator == 0:
        return chart

    slope = (n * sum_xy - sum_x * sum_y) / denominator
    intercept = (sum_y - slope * sum_x) / n

    # Calculate R-squared
    mean_y = sum_y / n
    ss_tot = sum((y - mean_y) ** 2 for _, y in values)
    ss_res = sum((y - (slope * x + intercept)) ** 2 for x, y in values)
    r_squared = 1 - (ss_res / ss_tot) if ss_tot > 0 else 0

    chart.trend_line = {
        "type": "linear",
        "slope": round(slope, 4),
        "intercept": round(intercept, 4),
        "r_squared": round(r_squared, 4),
        "direction": "increasing" if slope > 0 else "decreasing" if slope < 0 else "stable",
    }

    # Add insight
    trend_desc = "increasing" if slope > 0 else "decreasing" if slope < 0 else "stable"
    chart.ai_insights.append(
        f"Trend is {trend_desc} (slope: {slope:.2f}, R²: {r_squared:.2f})"
    )

    return chart


def add_forecast(chart: EnhancedChartSpec, periods: int = 3) -> EnhancedChartSpec:
    """Add simple forecast to a chart."""
    if not chart.trend_line or not chart.data:
        return chart

    slope = chart.trend_line.get("slope", 0)
    intercept = chart.trend_line.get("intercept", 0)

    last_x = len(chart.data) - 1
    forecast_values = []

    for i in range(1, periods + 1):
        x = last_x + i
        y = slope * x + intercept
        forecast_values.append({
            "period": f"+{i}",
            "value": round(y, 2),
            "lower_bound": round(y * 0.9, 2),
            "upper_bound": round(y * 1.1, 2),
        })

    chart.forecast = {
        "periods": periods,
        "method": "linear_extrapolation",
        "values": forecast_values,
        "confidence": 0.7,
    }

    chart.ai_insights.append(
        f"Forecast: Next {periods} periods projected based on linear trend"
    )

    return chart


def detect_anomalies(chart: EnhancedChartSpec) -> EnhancedChartSpec:
    """Detect anomalies in chart data."""
    if not chart.data or not chart.y_fields:
        return chart

    y_field = chart.y_fields[0]

    # Extract values
    values = []
    for i, record in enumerate(chart.data):
        try:
            y_val = float(str(record.get(y_field, 0)).replace(",", "").replace("$", "").replace("%", ""))
            values.append((i, y_val, record))
        except (ValueError, TypeError):
            pass

    if len(values) < 5:
        return chart

    # Calculate mean and std
    y_vals = [y for _, y, _ in values]
    mean = sum(y_vals) / len(y_vals)
    std = math.sqrt(sum((y - mean) ** 2 for y in y_vals) / len(y_vals))

    if std == 0:
        return chart

    # Find anomalies (beyond 2 standard deviations)
    anomalies = []
    for idx, y_val, record in values:
        z_score = abs((y_val - mean) / std)
        if z_score > 2:
            anomaly_type = "spike" if y_val > mean else "dip"
            anomalies.append({
                "index": idx,
                "value": y_val,
                "z_score": round(z_score, 2),
                "type": anomaly_type,
                "x_value": record.get(chart.x_field),
            })

            # Add annotation
            chart.annotations.append(ChartAnnotation(
                type="point",
                label=f"Anomaly: {anomaly_type}",
                value=y_val,
                position={"index": idx},
                style={"color": "red" if anomaly_type == "spike" else "orange"},
            ))

    chart.anomalies = anomalies

    if anomalies:
        chart.ai_insights.append(
            f"Detected {len(anomalies)} anomal{'y' if len(anomalies) == 1 else 'ies'} in the data"
        )

    return chart


def enhance_chart_with_intelligence(chart: EnhancedChartSpec) -> EnhancedChartSpec:
    """Apply all chart intelligence features."""
    chart = add_trend_line(chart)
    chart = detect_anomalies(chart)
    return chart


# CHART SUGGESTIONS

def generate_chart_suggestions(
    tables: List[EnhancedExtractedTable],
    metrics: List[ExtractedMetric] = None,
) -> List[VisualizationSuggestion]:
    """Generate visualization suggestions with rationale."""
    suggestions = []

    for table in tables[:5]:
        patterns = detect_data_patterns(table)

        for pattern in patterns:
            for chart_type in pattern.recommended_charts[:2]:
                chart = _create_chart_from_pattern(table, pattern)
                if chart:
                    chart.type = chart_type
                    chart = enhance_chart_with_intelligence(chart)

                    suggestions.append(VisualizationSuggestion(
                        chart_spec=chart,
                        rationale=f"{pattern.description}. {chart_type.value.title()} chart is ideal for visualizing {pattern.pattern_type.replace('_', ' ')} patterns.",
                        relevance_score=pattern.confidence,
                        complexity="simple" if chart_type in [ChartType.BAR, ChartType.LINE, ChartType.PIE] else "moderate",
                        insights_potential=chart.ai_insights,
                    ))

    # Sort by relevance
    suggestions.sort(key=lambda s: s.relevance_score, reverse=True)

    return suggestions[:10]


# LLM-POWERED CHART ANALYSIS

def analyze_chart_with_llm(chart: EnhancedChartSpec) -> EnhancedChartSpec:
    """Use LLM to generate insights about a chart."""
    # Build data summary
    data_summary = f"Chart: {chart.title}\nType: {chart.type.value}\n"
    data_summary += f"X-axis: {chart.x_field}\nY-axis: {', '.join(chart.y_fields)}\n"

    if chart.data:
        data_summary += f"Data points: {len(chart.data)}\n"
        data_summary += f"Sample data: {chart.data[:5]}"

    if chart.trend_line:
        data_summary += f"\nTrend: {chart.trend_line}"

    if chart.anomalies:
        data_summary += f"\nAnomalies: {chart.anomalies[:3]}"

    prompt = f"""Analyze this chart data and provide 3-5 key insights.

{data_summary}

Return JSON:
```json
{{
  "insights": [
    "Clear, actionable insight 1",
    "Clear, actionable insight 2",
    "Clear, actionable insight 3"
  ],
  "key_finding": "The single most important observation",
  "recommended_actions": ["Action 1", "Action 2"]
}}
```

Focus on patterns, trends, outliers, and business implications."""

    try:
        client = get_llm_client()
        response = call_chat_completion(
            client,
            model=None,
            messages=[{"role": "user", "content": prompt}],
            description="chart_analysis",
            temperature=0.3,
        )

        raw_text = response.choices[0].message.content or ""
        json_match = re.search(r'\{[\s\S]*\}', raw_text)

        if json_match:
            data = json.loads(json_match.group())
            chart.ai_insights.extend(data.get("insights", []))

            if data.get("key_finding"):
                chart.annotations.append(ChartAnnotation(
                    type="text",
                    label=data["key_finding"],
                    position={"location": "top"},
                ))

    except Exception as e:
        logger.warning(f"Chart analysis failed: {e}")

    return chart


# VISUALIZATION ENGINE ORCHESTRATOR

class VisualizationEngine:
    """Orchestrates all visualization features."""

    def generate_all_visualizations(
        self,
        tables: List[EnhancedExtractedTable],
        metrics: List[ExtractedMetric] = None,
        max_charts: int = 10,
    ) -> Dict[str, Any]:
        """Generate all visualizations for the data."""
        # Auto-generate charts
        auto_charts = auto_generate_charts(tables, max_charts)

        # Enhance with intelligence
        enhanced_charts = [enhance_chart_with_intelligence(c) for c in auto_charts]

        # Generate suggestions
        suggestions = generate_chart_suggestions(tables, metrics)

        return {
            "charts": enhanced_charts,
            "suggestions": suggestions,
        }

    def generate_from_query(
        self,
        query: str,
        tables: List[EnhancedExtractedTable],
        metrics: List[ExtractedMetric] = None,
    ) -> List[EnhancedChartSpec]:
        """Generate charts from natural language query."""
        charts = generate_chart_from_natural_language(query, tables, metrics)
        return [enhance_chart_with_intelligence(c) for c in charts]

    def add_intelligence_to_chart(
        self,
        chart: EnhancedChartSpec,
        include_forecast: bool = False,
        forecast_periods: int = 3,
    ) -> EnhancedChartSpec:
        """Add intelligence features to a chart."""
        chart = add_trend_line(chart)
        chart = detect_anomalies(chart)

        if include_forecast:
            chart = add_forecast(chart, forecast_periods)

        return analyze_chart_with_llm(chart)


# INTEGRATIONS (merged from integrations.py)


# 10.1 DATA SOURCE CONNECTIONS

class DataSourceType(str, Enum):
    DATABASE = "database"
    REST_API = "rest_api"
    CLOUD_STORAGE = "cloud_storage"
    EMAIL = "email"
    WEBHOOK = "webhook"


@dataclass
class DataSourceConnection:
    """A data source connection configuration."""
    id: str
    name: str
    type: DataSourceType
    config: Dict[str, Any]
    created_at: datetime = field(default_factory=lambda: datetime.now(timezone.utc))
    last_used: Optional[datetime] = None
    is_active: bool = True


@dataclass
class FetchResult:
    """Result of fetching data from a source."""
    success: bool
    data: Optional[Any] = None
    error: Optional[str] = None
    metadata: Dict[str, Any] = field(default_factory=dict)


class DataSourceConnector(ABC):
    """Abstract base class for data source connectors."""

    @abstractmethod
    async def connect(self) -> bool:
        """Establish connection to the data source."""
        pass

    @abstractmethod
    async def fetch(self, query: Optional[str] = None) -> FetchResult:
        """Fetch data from the source."""
        pass

    @abstractmethod
    async def disconnect(self) -> None:
        """Disconnect from the data source."""
        pass


class DatabaseConnector(DataSourceConnector):
    """Connector for database sources."""

    def __init__(self, config: Dict[str, Any]):
        self.config = config
        self.connection = None

    async def connect(self) -> bool:
        """Connect to database."""
        db_type = self.config.get("type", "postgresql")
        host = self.config.get("host", "localhost")
        port = self.config.get("port", 5432)
        database = self.config.get("database", "")

        logger.info(f"Connecting to {db_type} database at {host}:{port}/{database}")

        # In a real implementation, you would use asyncpg, aiomysql, etc.
        # For now, this is a placeholder
        self.connection = {"connected": True, "config": self.config}
        return True

    async def fetch(self, query: Optional[str] = None) -> FetchResult:
        """Execute query and fetch results."""
        if not self.connection:
            return FetchResult(success=False, error="Not connected")

        try:
            # Placeholder - would execute actual query
            return FetchResult(
                success=True,
                data=[],
                metadata={"query": query, "rows_fetched": 0},
            )
        except Exception as e:
            logger.exception("Database fetch failed")
            return FetchResult(success=False, error="Database query failed")

    async def disconnect(self) -> None:
        """Disconnect from database."""
        self.connection = None


class RestAPIConnector(DataSourceConnector):
    """Connector for REST API sources."""

    def __init__(self, config: Dict[str, Any]):
        self.config = config
        self.base_url = config.get("base_url", "")
        self.headers = config.get("headers", {})
        self.auth_type = config.get("auth_type", "none")

    async def connect(self) -> bool:
        """Validate API connection."""
        # In real implementation, would make a test request
        return bool(self.base_url)

    async def fetch(self, query: Optional[str] = None) -> FetchResult:
        """Fetch data from API endpoint."""
        import aiohttp

        endpoint = query or self.config.get("endpoint", "/")
        url = f"{self.base_url.rstrip('/')}/{endpoint.lstrip('/')}"

        try:
            async with aiohttp.ClientSession() as session:
                async with session.get(url, headers=self.headers) as response:
                    if response.status == 200:
                        data = await response.json()
                        return FetchResult(
                            success=True,
                            data=data,
                            metadata={"url": url, "status": response.status},
                        )
                    else:
                        return FetchResult(
                            success=False,
                            error=f"HTTP {response.status}",
                            metadata={"url": url},
                        )
        except Exception as e:
            logger.exception("REST API fetch failed for endpoint %s", endpoint)
            return FetchResult(success=False, error="Failed to fetch data from API")

    async def disconnect(self) -> None:
        """No persistent connection to close."""
        pass


class CloudStorageConnector(DataSourceConnector):
    """Connector for cloud storage (S3, GCS, Azure Blob)."""

    def __init__(self, config: Dict[str, Any]):
        self.config = config
        self.provider = config.get("provider", "s3")  # s3, gcs, azure

    async def connect(self) -> bool:
        """Validate cloud storage credentials."""
        # Would validate credentials with the respective cloud provider
        return True

    async def fetch(self, query: Optional[str] = None) -> FetchResult:
        """Fetch file from cloud storage."""
        bucket = self.config.get("bucket", "")
        key = query or self.config.get("key", "")

        try:
            if self.provider == "s3":
                # Would use aioboto3
                pass
            elif self.provider == "gcs":
                # Would use google-cloud-storage
                pass
            elif self.provider == "azure":
                # Would use azure-storage-blob
                pass

            return FetchResult(
                success=True,
                data=b"",  # File bytes
                metadata={"bucket": bucket, "key": key, "provider": self.provider},
            )
        except Exception as e:
            logger.exception("Cloud storage fetch failed for provider %s", self.provider)
            return FetchResult(success=False, error="Failed to fetch file from cloud storage")

    async def disconnect(self) -> None:
        """No persistent connection to close."""
        pass


class DataSourceManager:
    """Manages data source connections."""

    def __init__(self):
        self._connections: Dict[str, DataSourceConnection] = {}
        self._connectors: Dict[str, DataSourceConnector] = {}

    def register_connection(
        self,
        name: str,
        source_type: DataSourceType,
        config: Dict[str, Any],
    ) -> DataSourceConnection:
        """Register a new data source connection."""
        conn_id = f"ds_{uuid.uuid4().hex[:12]}"

        connection = DataSourceConnection(
            id=conn_id,
            name=name,
            type=source_type,
            config=config,
        )

        self._connections[conn_id] = connection

        # Create appropriate connector
        if source_type == DataSourceType.DATABASE:
            self._connectors[conn_id] = DatabaseConnector(config)
        elif source_type == DataSourceType.REST_API:
            self._connectors[conn_id] = RestAPIConnector(config)
        elif source_type == DataSourceType.CLOUD_STORAGE:
            self._connectors[conn_id] = CloudStorageConnector(config)

        return connection

    async def fetch_data(
        self,
        connection_id: str,
        query: Optional[str] = None,
    ) -> FetchResult:
        """Fetch data from a registered connection."""
        connector = self._connectors.get(connection_id)
        if not connector:
            return FetchResult(success=False, error="Connection not found")

        connection = self._connections.get(connection_id)
        if connection:
            connection.last_used = datetime.now(timezone.utc)

        await connector.connect()
        result = await connector.fetch(query)
        await connector.disconnect()

        return result

    def list_connections(self) -> List[DataSourceConnection]:
        """List all registered connections."""
        return list(self._connections.values())


# 10.2 WORKFLOW AUTOMATION

@dataclass
class AnalysisTrigger:
    """A trigger for automated analysis."""
    id: str
    name: str
    trigger_type: str  # schedule, webhook, file_upload, api_call
    config: Dict[str, Any]
    action: str  # analysis_type to run
    enabled: bool = True
    last_triggered: Optional[datetime] = None


@dataclass
class AnalysisPipeline:
    """A multi-step analysis pipeline."""
    id: str
    name: str
    steps: List[Dict[str, Any]]  # [{type, config, depends_on}]
    created_at: datetime = field(default_factory=lambda: datetime.now(timezone.utc))
    enabled: bool = True


@dataclass
class PipelineExecution:
    """An execution of an analysis pipeline."""
    id: str
    pipeline_id: str
    started_at: datetime
    completed_at: Optional[datetime] = None
    status: str = "running"  # running, completed, failed, cancelled
    step_results: Dict[str, Any] = field(default_factory=dict)
    error: Optional[str] = None


class WorkflowAutomationService:
    """Manages workflow automation."""

    def __init__(self):
        self._triggers: Dict[str, AnalysisTrigger] = {}
        self._pipelines: Dict[str, AnalysisPipeline] = {}
        self._schedules: Dict[str, ScheduledAnalysis] = {}
        self._executions: Dict[str, PipelineExecution] = {}
        self._webhooks: Dict[str, WebhookConfig] = {}

    def create_trigger(
        self,
        name: str,
        trigger_type: str,
        config: Dict[str, Any],
        action: str,
    ) -> AnalysisTrigger:
        """Create a new analysis trigger."""
        trigger = AnalysisTrigger(
            id=f"trig_{uuid.uuid4().hex[:12]}",
            name=name,
            trigger_type=trigger_type,
            config=config,
            action=action,
        )
        self._triggers[trigger.id] = trigger
        return trigger

    def create_pipeline(
        self,
        name: str,
        steps: List[Dict[str, Any]],
    ) -> AnalysisPipeline:
        """Create an analysis pipeline."""
        pipeline = AnalysisPipeline(
            id=f"pipe_{uuid.uuid4().hex[:12]}",
            name=name,
            steps=steps,
        )
        self._pipelines[pipeline.id] = pipeline
        return pipeline

    async def execute_pipeline(
        self,
        pipeline_id: str,
        input_data: Dict[str, Any],
    ) -> PipelineExecution:
        """Execute an analysis pipeline."""
        pipeline = self._pipelines.get(pipeline_id)
        if not pipeline:
            raise ValueError(f"Pipeline not found: {pipeline_id}")

        execution = PipelineExecution(
            id=f"exec_{uuid.uuid4().hex[:12]}",
            pipeline_id=pipeline_id,
            started_at=datetime.now(timezone.utc),
        )
        self._executions[execution.id] = execution

        try:
            context = {"input": input_data}

            for i, step in enumerate(pipeline.steps):
                step_id = step.get("id", f"step_{i}")
                step_type = step.get("type")

                logger.info(f"Executing pipeline step: {step_id} ({step_type})")

                # Execute step based on type
                if step_type == "extract":
                    # Run extraction
                    result = {"extracted": True}
                elif step_type == "analyze":
                    # Run analysis
                    result = {"analyzed": True}
                elif step_type == "transform":
                    # Apply transformations
                    result = {"transformed": True}
                elif step_type == "export":
                    # Export results
                    result = {"exported": True}
                elif step_type == "notify":
                    # Send notifications
                    result = {"notified": True}
                else:
                    result = {"unknown_step": step_type}

                execution.step_results[step_id] = result
                context[step_id] = result

            execution.status = "completed"
            execution.completed_at = datetime.now(timezone.utc)

        except Exception as e:
            execution.status = "failed"
            execution.error = "Pipeline execution failed during processing"
            execution.completed_at = datetime.now(timezone.utc)
            logger.exception("Pipeline execution %s failed", execution.id)

        return execution

    def schedule_analysis(
        self,
        name: str,
        source_config: Dict[str, Any],
        schedule: str,  # Cron expression
        analysis_config: Dict[str, Any],
        notifications: List[str] = None,
    ) -> ScheduledAnalysis:
        """Schedule a recurring analysis."""
        scheduled = ScheduledAnalysis(
            id=f"sched_{uuid.uuid4().hex[:12]}",
            name=name,
            source_type=source_config.get("type", "upload"),
            source_config=source_config,
            schedule=schedule,
            notifications=notifications or [],
            enabled=True,
        )
        self._schedules[scheduled.id] = scheduled
        return scheduled

    def register_webhook(
        self,
        url: str,
        events: List[str],
        secret: Optional[str] = None,
    ) -> WebhookConfig:
        """Register a webhook for notifications."""
        webhook = WebhookConfig(
            url=url,
            events=events,
            secret=secret,
            enabled=True,
        )
        webhook_id = f"hook_{uuid.uuid4().hex[:12]}"
        self._webhooks[webhook_id] = webhook
        return webhook

    async def send_webhook(
        self,
        webhook_id: str,
        event: str,
        payload: Dict[str, Any],
    ) -> bool:
        """Send a webhook notification."""
        webhook = self._webhooks.get(webhook_id)
        if not webhook or not webhook.enabled:
            return False

        if event not in webhook.events:
            return False


        headers = {"Content-Type": "application/json"}

        # Add signature if secret is set
        if webhook.secret:
            signature = hashlib.sha256(
                f"{webhook.secret}{json.dumps(payload)}".encode()
            ).hexdigest()
            headers["X-Webhook-Signature"] = signature

        try:
            async with aiohttp.ClientSession() as session:
                async with session.post(
                    webhook.url,
                    json={"event": event, "data": payload},
                    headers=headers,
                ) as response:
                    return response.status < 400
        except Exception as e:
            logger.error(f"Webhook delivery failed: {e}")
            return False


# 10.3 EXTERNAL TOOLS INTEGRATION

class ExternalToolIntegration(ABC):
    """Abstract base class for external tool integrations."""

    @abstractmethod
    async def send_message(self, message: str, **kwargs) -> bool:
        """Send a message to the external tool."""
        pass

    @abstractmethod
    async def create_item(self, data: Dict[str, Any]) -> Optional[str]:
        """Create an item (task, ticket, etc.) in the external tool."""
        pass


class SlackIntegration(ExternalToolIntegration):
    """Slack integration."""

    def __init__(self, config: Dict[str, Any]):
        self.webhook_url = config.get("webhook_url", "")
        self.channel = config.get("channel", "")
        self.bot_token = config.get("bot_token", "")

    async def send_message(self, message: str, **kwargs) -> bool:
        """Send message to Slack channel."""

        payload = {
            "text": message,
            "channel": kwargs.get("channel", self.channel),
        }

        # Add blocks for rich formatting
        if "blocks" in kwargs:
            payload["blocks"] = kwargs["blocks"]

        try:
            async with aiohttp.ClientSession() as session:
                async with session.post(self.webhook_url, json=payload) as response:
                    return response.status == 200
        except Exception as e:
            logger.error(f"Slack message failed: {e}")
            return False

    async def create_item(self, data: Dict[str, Any]) -> Optional[str]:
        """Create a Slack reminder or scheduled message."""
        # Would use Slack API to create reminders
        return None

    def format_analysis_summary(self, result: EnhancedAnalysisResult) -> Dict[str, Any]:
        """Format analysis result as Slack blocks."""
        blocks = [
            {
                "type": "header",
                "text": {
                    "type": "plain_text",
                    "text": f"Analysis Complete: {result.document_name}",
                }
            },
            {
                "type": "section",
                "fields": [
                    {"type": "mrkdwn", "text": f"*Tables:* {result.total_tables}"},
                    {"type": "mrkdwn", "text": f"*Metrics:* {result.total_metrics}"},
                    {"type": "mrkdwn", "text": f"*Insights:* {len(result.insights)}"},
                    {"type": "mrkdwn", "text": f"*Risks:* {len(result.risks)}"},
                ]
            },
        ]

        # Add top insights
        if result.insights:
            insight_text = "\n".join([f"• {i.title}" for i in result.insights[:3]])
            blocks.append({
                "type": "section",
                "text": {"type": "mrkdwn", "text": f"*Key Insights:*\n{insight_text}"}
            })

        return {"blocks": blocks}


class TeamsIntegration(ExternalToolIntegration):
    """Microsoft Teams integration."""

    def __init__(self, config: Dict[str, Any]):
        self.webhook_url = config.get("webhook_url", "")

    async def send_message(self, message: str, **kwargs) -> bool:
        """Send message to Teams channel."""

        # Teams Adaptive Card format
        payload = {
            "@type": "MessageCard",
            "@context": "http://schema.org/extensions",
            "summary": message[:50],
            "themeColor": "0076D7",
            "title": kwargs.get("title", "NeuraReport Analysis"),
            "text": message,
        }

        try:
            async with aiohttp.ClientSession() as session:
                async with session.post(self.webhook_url, json=payload) as response:
                    return response.status == 200
        except Exception as e:
            logger.error(f"Teams message failed: {e}")
            return False

    async def create_item(self, data: Dict[str, Any]) -> Optional[str]:
        """Create Teams task or Planner item."""
        return None


class JiraIntegration(ExternalToolIntegration):
    """Jira integration."""

    def __init__(self, config: Dict[str, Any]):
        self.base_url = config.get("base_url", "")
        self.email = config.get("email", "")
        self.api_token = config.get("api_token", "")
        self.project_key = config.get("project_key", "")

    async def send_message(self, message: str, **kwargs) -> bool:
        """Add comment to a Jira issue."""
        issue_key = kwargs.get("issue_key")
        if not issue_key:
            return False

        from aiohttp import BasicAuth

        url = f"{self.base_url}/rest/api/3/issue/{issue_key}/comment"
        payload = {
            "body": {
                "type": "doc",
                "version": 1,
                "content": [{"type": "paragraph", "content": [{"type": "text", "text": message}]}]
            }
        }

        try:
            async with aiohttp.ClientSession() as session:
                async with session.post(
                    url,
                    json=payload,
                    auth=BasicAuth(self.email, self.api_token),
                ) as response:
                    return response.status < 400
        except Exception as e:
            logger.error(f"Jira comment failed: {e}")
            return False

    async def create_item(self, data: Dict[str, Any]) -> Optional[str]:
        """Create a Jira issue from analysis findings."""

        url = f"{self.base_url}/rest/api/3/issue"
        payload = {
            "fields": {
                "project": {"key": self.project_key},
                "summary": data.get("title", "Analysis Finding"),
                "description": {
                    "type": "doc",
                    "version": 1,
                    "content": [{"type": "paragraph", "content": [{"type": "text", "text": data.get("description", "")}]}]
                },
                "issuetype": {"name": data.get("issue_type", "Task")},
                "priority": {"name": data.get("priority", "Medium")},
            }
        }

        try:
            async with aiohttp.ClientSession() as session:
                async with session.post(
                    url,
                    json=payload,
                    auth=BasicAuth(self.email, self.api_token),
                ) as response:
                    if response.status < 400:
                        result = await response.json()
                        return result.get("key")
        except Exception as e:
            logger.error(f"Jira issue creation failed: {e}")

        return None


class EmailIntegration(ExternalToolIntegration):
    """Email integration for sending analysis reports."""

    def __init__(self, config: Dict[str, Any]):
        self.smtp_host = config.get("smtp_host", "")
        self.smtp_port = config.get("smtp_port", 587)
        self.username = config.get("username", "")
        self.password = config.get("password", "")
        self.from_email = config.get("from_email", "")

    async def send_message(self, message: str, **kwargs) -> bool:
        """Send email with analysis summary."""
        to_emails = kwargs.get("to", [])
        subject = kwargs.get("subject", "Analysis Report")
        html_content = kwargs.get("html", "")

        if not to_emails:
            return False

        import smtplib
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText

        try:
            msg = MIMEMultipart("alternative")
            msg["Subject"] = subject
            msg["From"] = self.from_email
            msg["To"] = ", ".join(to_emails)

            # Plain text
            msg.attach(MIMEText(message, "plain"))

            # HTML version
            if html_content:
                msg.attach(MIMEText(html_content, "html"))

            with smtplib.SMTP(self.smtp_host, self.smtp_port) as server:
                server.starttls()
                server.login(self.username, self.password)
                server.sendmail(self.from_email, to_emails, msg.as_string())

            return True
        except Exception as e:
            logger.error(f"Email send failed: {e}")
            return False

    async def create_item(self, data: Dict[str, Any]) -> Optional[str]:
        """N/A for email integration."""
        return None


# INTEGRATION SERVICE ORCHESTRATOR

class IntegrationService:
    """Orchestrates all integration capabilities."""

    def __init__(self):
        self.data_sources = DataSourceManager()
        self.workflows = WorkflowAutomationService()
        self._integrations: Dict[str, ExternalToolIntegration] = {}
        self._integration_types: Dict[str, str] = {}

    def register_integration(
        self,
        name: str,
        integration_type: str,
        config: Dict[str, Any],
    ) -> str:
        """Register an external tool integration."""
        integration_id = f"int_{uuid.uuid4().hex[:12]}"

        if integration_type == "slack":
            self._integrations[integration_id] = SlackIntegration(config)
        elif integration_type == "teams":
            self._integrations[integration_id] = TeamsIntegration(config)
        elif integration_type == "jira":
            self._integrations[integration_id] = JiraIntegration(config)
        elif integration_type == "email":
            self._integrations[integration_id] = EmailIntegration(config)
        else:
            raise ValueError(f"Unknown integration type: {integration_type}")

        self._integration_types[integration_id] = integration_type
        return integration_id

    def list_integrations(self) -> List[Dict[str, Any]]:
        """List registered integrations (without secrets)."""
        return [
            {"id": int_id, "type": self._integration_types.get(int_id, "unknown")}
            for int_id in self._integrations.keys()
        ]

    async def send_notification(
        self,
        integration_id: str,
        message: str,
        **kwargs,
    ) -> bool:
        """Send notification via integration."""
        integration = self._integrations.get(integration_id)
        if not integration:
            return False
        return await integration.send_message(message, **kwargs)

    async def create_external_item(
        self,
        integration_id: str,
        data: Dict[str, Any],
    ) -> Optional[str]:
        """Create item in external tool."""
        integration = self._integrations.get(integration_id)
        if not integration:
            return None
        return await integration.create_item(data)

    async def broadcast_analysis_complete(
        self,
        result: EnhancedAnalysisResult,
    ) -> Dict[str, bool]:
        """Broadcast analysis completion to all integrations."""
        results = {}

        message = f"Analysis complete: {result.document_name}\n"
        message += f"Found {result.total_tables} tables, {result.total_metrics} metrics\n"
        if result.insights:
            message += f"Top insight: {result.insights[0].title}"

        for int_id, integration in self._integrations.items():
            try:
                success = await integration.send_message(message)
                results[int_id] = success
            except Exception as e:
                logger.error(f"Broadcast to {int_id} failed: {e}")
                results[int_id] = False

        return results

    # Data source methods
    def register_data_source(self, *args, **kwargs) -> DataSourceConnection:
        return self.data_sources.register_connection(*args, **kwargs)

    async def fetch_from_source(self, *args, **kwargs) -> FetchResult:
        return await self.data_sources.fetch_data(*args, **kwargs)

    def list_data_sources(self) -> List[DataSourceConnection]:
        return self.data_sources.list_connections()

    # Workflow methods
    def create_trigger(self, *args, **kwargs) -> AnalysisTrigger:
        return self.workflows.create_trigger(*args, **kwargs)

    def create_pipeline(self, *args, **kwargs) -> AnalysisPipeline:
        return self.workflows.create_pipeline(*args, **kwargs)

    async def execute_pipeline(self, *args, **kwargs) -> PipelineExecution:
        return await self.workflows.execute_pipeline(*args, **kwargs)

    def schedule_analysis(self, *args, **kwargs) -> ScheduledAnalysis:
        return self.workflows.schedule_analysis(*args, **kwargs)

    def register_webhook(self, *args, **kwargs) -> WebhookConfig:
        return self.workflows.register_webhook(*args, **kwargs)

    async def send_webhook(self, *args, **kwargs) -> bool:
        return await self.workflows.send_webhook(*args, **kwargs)


# ORCHESTRATION

# mypy: ignore-errors
"""
Enhanced Analysis Orchestrator - Main service that orchestrates all AI-powered analysis features.

This service coordinates:
- Intelligent Data Extraction
- AI-Powered Analysis Engines
- Intelligent Visualization
- Data Transformation & Export
- Advanced AI Features
- User Experience Features
- Integration Capabilities
"""

import csv
import html as html_mod
import io
from typing import Any, AsyncGenerator, Callable, Dict, List, Optional

from backend.app.schemas import (
    AnalysisPreferences,
    DocumentType,
    EnhancedAnalysisResult,
    ExportConfiguration,
    ExportFormat,
    QuestionResponse,
)
from backend.app.services.llm import RAGRetriever
from backend.app.services.infra_services import call_chat_completion_async

logger = logging.getLogger("neura.analyze.orchestrator")


# In-memory cache for analysis results (bounded to prevent memory leaks)
_ANALYSIS_CACHE: Dict[str, EnhancedAnalysisResult] = {}
_ANALYSIS_CACHE_MAX = 500


def _cache_put(analysis_id: str, result: EnhancedAnalysisResult) -> None:
    """Add to cache with eviction when max size exceeded."""
    if len(_ANALYSIS_CACHE) >= _ANALYSIS_CACHE_MAX:
        # Evict oldest entry (first inserted)
        oldest_key = next(iter(_ANALYSIS_CACHE))
        del _ANALYSIS_CACHE[oldest_key]
    _ANALYSIS_CACHE[analysis_id] = result


def _generate_analysis_id() -> str:
    """Generate a unique analysis ID."""
    return f"eana_{uuid.uuid4().hex[:12]}"


def _detect_document_type(file_name: str) -> DocumentType:
    """Detect document type from file name."""
    name = file_name.lower()
    if name.endswith(".pdf"):
        return DocumentType.PDF
    elif name.endswith((".xlsx", ".xls", ".xlsm")):
        return DocumentType.EXCEL
    elif name.endswith(".csv"):
        return DocumentType.CSV
    elif name.endswith((".png", ".jpg", ".jpeg", ".tiff", ".bmp")):
        return DocumentType.IMAGE
    elif name.endswith((".doc", ".docx")):
        return DocumentType.WORD
    elif name.endswith(".txt"):
        return DocumentType.TEXT
    return DocumentType.UNKNOWN


class EnhancedAnalysisOrchestrator:
    """
    Main orchestrator for enhanced document analysis.

    Coordinates all AI-powered analysis features and provides
    a unified interface for the API layer.
    """

    def __init__(self):
        self.extraction_service = EnhancedExtractionService()
        self.analysis_engine = AnalysisEngineService()
        self.visualization_engine = VisualizationEngine()
        self.export_service = DataExportService()
        self.advanced_ai = AdvancedAIService()
        self.ux_service = UserExperienceService()
        self.integration_service = IntegrationService()
        self._rag_retrievers: Dict[str, RAGRetriever] = {}
        self._store = get_analysis_store()

    async def analyze_document_streaming(
        self,
        file_bytes: Optional[bytes],
        file_name: str,
        preferences: Optional[AnalysisPreferences] = None,
        correlation_id: Optional[str] = None,
        file_path: Optional[Path] = None,
        analysis_id: Optional[str] = None,
    ) -> AsyncGenerator[Dict[str, Any], None]:
        """
        Perform comprehensive document analysis with streaming progress updates.

        This is the main entry point for document analysis.
        """
        analysis_id = analysis_id or _generate_analysis_id()
        started = time.time()

        # Use default preferences if not provided
        if preferences is None:
            preferences = AnalysisPreferences()

        # Build configuration from preferences
        config = self.ux_service.build_configuration(preferences)

        # Create streaming session
        session = self.ux_service.create_streaming_session()

        warnings_list: List[str] = []

        try:
            # Stage 1: Upload validation
            yield self._event("stage", "Validating document...", 5, analysis_id, correlation_id)

            if not file_bytes and not file_path:
                yield self._event("error", "Empty file provided", 0, analysis_id, correlation_id)
                return

            document_type = _detect_document_type(file_name)

            # Stage 2: Content extraction (PDF/Excel parsing — no LLM)
            yield self._event("stage", "Extracting content from document...", 15, analysis_id, correlation_id)

            content = extract_document_content(
                file_bytes=file_bytes,
                file_path=file_path,
                file_name=file_name,
            )

            if content.errors and not content.tables_raw and not content.text_content:
                yield self._event("error", f"Could not extract content: {'; '.join(content.errors)}", 0, analysis_id, correlation_id)
                return

            if content.errors:
                warnings_list.extend(content.errors)

            text_len = len(content.text_content or "")
            yield self._event("stage", f"Extracted {text_len} chars, {len(content.tables_raw)} tables", 20, analysis_id, correlation_id)

            # Stage 3: Intelligent extraction (1 LLM call — runs in thread with progress ticks)
            yield self._event("stage", "Extracting entities & metrics with AI...", 30, analysis_id, correlation_id)

            try:
                extraction_task = asyncio.create_task(
                    asyncio.to_thread(
                        self.extraction_service.extract_all,
                        text=content.text_content,
                        raw_tables=content.tables_raw,
                    )
                )
                _pct = 30
                while not extraction_task.done():
                    await asyncio.sleep(2.0)
                    if extraction_task.done():
                        break
                    if _pct < 48:
                        _pct += 2
                        yield self._event("stage", "Extracting entities & metrics...", _pct, analysis_id, correlation_id)
                extraction_result = await extraction_task
            except Exception as exc:
                logger.error("Extraction stage failed: %s", exc, exc_info=True)
                yield self._event("stage", f"Extraction partially failed: {exc}", 35, analysis_id, correlation_id)
                warnings_list.append(f"AI extraction failed: {exc}")
                extraction_result = {
                    "tables": [], "entities": [], "metrics": [],
                    "forms": [], "invoices": [], "contracts": [],
                    "table_relationships": [],
                }

            tables = extraction_result["tables"]
            entities = extraction_result["entities"]
            metrics = extraction_result["metrics"]
            forms = extraction_result["forms"]
            invoices = extraction_result["invoices"]
            contracts = extraction_result["contracts"]
            table_relationships = extraction_result["table_relationships"]

            yield self._event("stage", f"Found {len(tables)} tables, {len(metrics)} metrics, {len(entities)} entities", 40, analysis_id, correlation_id)

            # Stage 4: AI Analysis — single consolidated LLM call (runs in thread with progress ticks)
            yield self._event("stage", "Running comprehensive AI analysis...", 50, analysis_id, correlation_id)

            try:
                analysis_task = asyncio.create_task(
                    asyncio.to_thread(
                        self.analysis_engine.run_all_analyses,
                        text=content.text_content,
                        tables=tables,
                        metrics=metrics,
                    )
                )
                _pct = 50
                while not analysis_task.done():
                    await asyncio.sleep(2.0)
                    if analysis_task.done():
                        break
                    if _pct < 63:
                        _pct += 2
                        yield self._event("stage", "Analyzing document...", _pct, analysis_id, correlation_id)
                analysis_results = await analysis_task
            except Exception as exc:
                logger.error("Analysis stage failed: %s", exc, exc_info=True)
                yield self._event("stage", f"AI analysis failed: {exc}", 55, analysis_id, correlation_id)
                warnings_list.append(f"AI analysis failed: {exc}")
                # Provide empty defaults so we can still return partial results
                from backend.app.schemas import (
                    SentimentLevel, SentimentAnalysis as SA, TextAnalytics as TA,
                    StatisticalAnalysis as StA, FinancialAnalysis as FA,
                )
                analysis_results = {
                    "summaries": {},
                    "sentiment": SA(overall_sentiment=SentimentLevel.NEUTRAL, overall_score=0.0, confidence=0.0),
                    "text_analytics": TA(),
                    "statistical_analysis": StA(),
                    "financial_analysis": FA(metrics_found=len(metrics)),
                    "insights": [], "risks": [], "opportunities": [], "action_items": [],
                }

            summaries = analysis_results["summaries"]
            sentiment = analysis_results["sentiment"]
            text_analytics = analysis_results["text_analytics"]
            statistical_analysis = analysis_results["statistical_analysis"]
            financial_analysis = analysis_results["financial_analysis"]
            insights = analysis_results["insights"]
            risks = analysis_results["risks"]
            opportunities = analysis_results["opportunities"]
            action_items = analysis_results["action_items"]

            yield self._event("stage", f"Generated {len(insights)} insights, {len(risks)} risks", 65, analysis_id, correlation_id)

            # Stage 5: Visualization (no LLM — pattern detection + chart specs)
            yield self._event("stage", "Generating visualizations...", 70, analysis_id, correlation_id)

            try:
                viz_results = self.visualization_engine.generate_all_visualizations(
                    tables=tables,
                    metrics=metrics,
                    max_charts=preferences.max_charts,
                )
                charts = viz_results["charts"]
                viz_suggestions = viz_results["suggestions"]
            except Exception as exc:
                logger.warning("Visualization failed: %s", exc)
                warnings_list.append(f"Chart generation failed: {exc}")
                charts = []
                viz_suggestions = []

            yield self._event("stage", f"Created {len(charts)} charts", 75, analysis_id, correlation_id)

            # Stage 6: Data Quality (no LLM)
            yield self._event("stage", "Assessing data quality...", 80, analysis_id, correlation_id)

            try:
                data_quality = self.export_service.assess_quality(tables)
            except Exception as exc:
                logger.warning("Quality assessment failed: %s", exc)
                data_quality = {}

            # Stage 7: Build RAG index for Q&A (no LLM)
            yield self._event("stage", "Building knowledge index for Q&A...", 88, analysis_id, correlation_id)

            rag = RAGRetriever(use_embeddings=False)
            rag.add_document(
                content=content.text_content,
                doc_id=analysis_id,
                metadata={"file_name": file_name, "analysis_id": analysis_id},
            )
            self._rag_retrievers[analysis_id] = rag

            # Stage 8: Suggested questions (no LLM)
            suggested_questions = self.ux_service.generate_suggested_questions(
                tables=tables,
                metrics=metrics,
                entities=entities,
            )

            # Calculate processing time
            processing_time_ms = int((time.time() - started) * 1000)

            yield self._event("stage", "Finalizing results...", 95, analysis_id, correlation_id)

            # Build final result
            result = EnhancedAnalysisResult(
                analysis_id=analysis_id,
                document_name=file_name,
                document_type=document_type,
                created_at=datetime.now(timezone.utc),
                processing_time_ms=processing_time_ms,

                # Extraction
                tables=tables,
                entities=entities,
                metrics=metrics,
                forms=forms,
                invoices=invoices,
                contracts=contracts,
                table_relationships=table_relationships,

                # Analysis
                summaries=summaries,
                sentiment=sentiment,
                text_analytics=text_analytics,
                financial_analysis=financial_analysis,
                statistical_analysis=statistical_analysis,

                # Visualizations
                chart_suggestions=charts,
                visualization_suggestions=viz_suggestions,

                # Insights
                insights=insights,
                risks=risks,
                opportunities=opportunities,
                action_items=action_items,

                # Quality
                data_quality=data_quality,

                # Metadata
                page_count=content.page_count,
                total_tables=len(tables),
                total_entities=len(entities),
                total_metrics=len(metrics),
                confidence_score=0.85 if not warnings_list else 0.6,

                # Settings
                preferences=preferences,

                # Warnings
                warnings=warnings_list,
            )

            # Cache the result
            _cache_put(analysis_id, result)
            try:
                self._store.save_result(result)
                self._store.save_context(analysis_id, content.text_content)
            except Exception as exc:
                logger.warning(f"Failed to persist analysis result: {exc}")

            yield self._event("stage", "Complete", 100, analysis_id, correlation_id)

            # Final result event (use mode="json" for datetime serialization)
            result_dict = result.model_dump(mode="json")
            result_dict["event"] = "result"
            result_dict["suggested_questions"] = suggested_questions

            if correlation_id:
                result_dict["correlation_id"] = correlation_id

            yield result_dict

        except asyncio.CancelledError:
            yield self._event("cancelled", "Analysis cancelled", 0, analysis_id, correlation_id)
            raise
        except Exception as e:
            logger.error(f"Analysis failed: {e}", exc_info=True)
            error_detail = str(e)
            if len(error_detail) > 200:
                error_detail = error_detail[:200] + "..."
            yield self._event("error", f"Analysis failed: {error_detail}", 0, analysis_id, correlation_id)

    def _event(
        self,
        event_type: str,
        detail: str,
        progress: int,
        analysis_id: str,
        correlation_id: Optional[str],
    ) -> Dict[str, Any]:
        """Build an event dict."""
        event = {
            "event": event_type if event_type != "stage" else "stage",
            "stage": event_type,
            "detail": detail,
            "progress": progress,
            "analysis_id": analysis_id,
        }
        if correlation_id:
            event["correlation_id"] = correlation_id
        return event

    def get_analysis(self, analysis_id: str) -> Optional[EnhancedAnalysisResult]:
        """Get a cached analysis result."""
        result = _ANALYSIS_CACHE.get(analysis_id)
        if result:
            return result
        stored = self._store.load_result(analysis_id)
        if stored:
            _cache_put(analysis_id, stored)
            if analysis_id not in self._rag_retrievers:
                text_content = self._store.load_context(analysis_id)
                if text_content:
                    rag = RAGRetriever(use_embeddings=False)
                    rag.add_document(
                        content=text_content,
                        doc_id=analysis_id,
                        metadata={"file_name": stored.document_name, "analysis_id": analysis_id},
                    )
                    self._rag_retrievers[analysis_id] = rag
        return stored

    def new_analysis_id(self) -> str:
        """Generate a new analysis ID."""
        return _generate_analysis_id()

    async def ask_question(
        self,
        analysis_id: str,
        question: str,
        include_sources: bool = True,
        max_context_chunks: int = 5,
    ) -> QuestionResponse:
        """Ask a question about the analyzed document."""
        # Get the analysis result
        result = self.get_analysis(analysis_id)
        if not result:
            return QuestionResponse(
                answer="Analysis not found. Please upload and analyze the document first.",
                confidence=0.0,
                sources=[],
                suggested_followups=[],
            )

        # Get RAG retriever
        rag = self._rag_retrievers.get(analysis_id)
        if not rag:
            return QuestionResponse(
                answer="Knowledge index not available. Please re-analyze the document.",
                confidence=0.0,
                sources=[],
                suggested_followups=[],
            )

        # Query with context (runs Claude CLI — must not block event loop)
        rag_result = await asyncio.to_thread(
            rag.query_with_context,
            question=question,
            top_k=max_context_chunks,
            include_sources=include_sources,
        )

        # Generate follow-up questions
        suggested_followups = await self._generate_followup_questions(question, rag_result["answer"])

        return QuestionResponse(
            answer=rag_result["answer"],
            confidence=0.8 if rag_result.get("context_used") else 0.5,
            sources=rag_result.get("sources", []) if include_sources else [],
            suggested_followups=suggested_followups,
        )

    async def _generate_followup_questions(self, question: str, answer: str) -> List[str]:
        """Generate follow-up questions based on Q&A."""
        try:
            client = get_llm_client()
            prompt = f"""Based on this Q&A, suggest 3 relevant follow-up questions.

Question: {question}
Answer: {answer[:500]}

Return JSON array of questions:
["Question 1", "Question 2", "Question 3"]"""

            response = await call_chat_completion_async(
                client,
                model=None,
                messages=[{"role": "user", "content": prompt}],
                description="followup_questions",
                temperature=0.5,
            )

            raw = response.choices[0].message.content or "[]"
            from backend.app.services.infra_services import extract_json_array_from_llm_response
            result = extract_json_array_from_llm_response(raw, default=[])
            if result:
                return result
        except Exception as e:
            logger.warning(f"Follow-up generation failed: {e}")

        return []

    async def generate_charts_from_query(
        self,
        analysis_id: str,
        query: str,
        include_trends: bool = True,
        include_forecasts: bool = False,
    ) -> Dict[str, Any]:
        """Generate charts from natural language query.

        Returns dict with 'charts' list and optional 'message' string.
        """
        result = self.get_analysis(analysis_id)
        if not result:
            return {"charts": [], "message": "Analysis not found. Please re-analyze the document."}

        if not result.tables:
            return {"charts": [], "message": "No tables were extracted from this document. Chart generation requires tabular data."}

        # Chart generation + intelligence both call Claude CLI — run in thread
        def _generate_sync():
            charts = self.visualization_engine.generate_from_query(
                query=query,
                tables=result.tables,
                metrics=result.metrics,
            )
            if include_trends or include_forecasts:
                charts = [
                    self.visualization_engine.add_intelligence_to_chart(
                        chart,
                        include_forecast=include_forecasts,
                    )
                    for chart in charts
                ]
            return charts

        try:
            charts = await asyncio.to_thread(_generate_sync)
        except Exception as exc:
            logger.error("Chart generation failed: %s", exc, exc_info=True)
            return {"charts": [], "message": f"Chart generation failed: {exc}"}

        if not charts:
            return {"charts": [], "message": "AI could not generate charts for this query. Try rephrasing or be more specific about the data you want to visualize."}

        return {"charts": [c.model_dump() for c in charts]}

    async def export_analysis(
        self,
        analysis_id: str,
        format: ExportFormat,
        include_raw_data: bool = True,
        include_charts: bool = True,
    ) -> tuple[bytes, str]:
        """Export analysis in specified format."""
        result = self.get_analysis(analysis_id)
        if not result:
            raise ValueError(f"Analysis not found: {analysis_id}")

        config = ExportConfiguration(
            format=format,
            include_raw_data=include_raw_data,
            include_charts=include_charts,
            include_analysis=True,
            include_insights=True,
        )

        return await self.export_service.export(result, config)

    async def compare_documents(
        self,
        analysis_id_1: str,
        analysis_id_2: str,
    ) -> Dict[str, Any]:
        """Compare two analyzed documents."""
        result1 = self.get_analysis(analysis_id_1)
        result2 = self.get_analysis(analysis_id_2)

        if not result1 or not result2:
            return {"error": "One or both analyses not found"}

        # Get text content from summaries
        text1 = result1.summaries.get("comprehensive", result1.summaries.get("executive"))
        text2 = result2.summaries.get("comprehensive", result2.summaries.get("executive"))

        text1_content = text1.content if text1 else ""
        text2_content = text2.content if text2 else ""

        comparison = compare_documents(
            text1=text1_content,
            text2=text2_content,
            metrics1=result1.metrics,
            metrics2=result2.metrics,
        )

        return comparison.model_dump()

    def get_industry_options(self) -> List[Dict[str, Any]]:
        """Get available industry options."""
        return self.ux_service.get_industry_options()

    def get_export_formats(self) -> List[Dict[str, str]]:
        """Get available export formats."""
        return [
            {"value": f.value, "label": f.value.upper()}
            for f in ExportFormat
        ]


# Singleton instance
_orchestrator: Optional[EnhancedAnalysisOrchestrator] = None


def get_orchestrator() -> EnhancedAnalysisOrchestrator:
    """Get the singleton orchestrator instance."""
    global _orchestrator
    if _orchestrator is None:
        _orchestrator = EnhancedAnalysisOrchestrator()
    return _orchestrator


# DATA TRANSFORM & EXPORT (merged from data_transform_export.py)

logger = logging.getLogger("neura.analyze.export")


# DATA QUALITY ASSESSMENT

def assess_data_quality(tables: List[EnhancedExtractedTable]) -> DataQualityReport:
    """Assess data quality across all tables."""
    total_rows = sum(t.row_count for t in tables)
    total_columns = sum(len(t.headers) for t in tables)

    missing_values: Dict[str, int] = {}
    missing_percentage: Dict[str, float] = {}
    unique_values_per_column: Dict[str, int] = {}
    invalid_values: Dict[str, List[Any]] = {}
    type_mismatches: Dict[str, List[int]] = {}
    format_inconsistencies: Dict[str, List[str]] = {}
    outliers_detected: Dict[str, List[int]] = {}

    duplicate_rows = 0

    for table in tables:
        # Check for duplicate rows
        seen_rows = set()
        for row in table.rows:
            row_key = tuple(str(v) for v in row)
            if row_key in seen_rows:
                duplicate_rows += 1
            seen_rows.add(row_key)

        # Analyze each column
        for col_idx, (header, dtype) in enumerate(zip(table.headers, table.data_types)):
            col_key = f"{table.id}.{header}"

            # Collect column values
            values = []
            missing_count = 0
            unique_set = set()
            numeric_values = []

            for row_idx, row in enumerate(table.rows):
                if col_idx < len(row):
                    val = row[col_idx]
                    values.append((row_idx, val))

                    # Check missing
                    if val is None or str(val).strip() == "" or str(val).lower() in ("null", "n/a", "na", "-"):
                        missing_count += 1
                    else:
                        unique_set.add(str(val))

                        # Type validation
                        if dtype == "numeric":
                            try:
                                cleaned = re.sub(r'[$,% ]', '', str(val))
                                num_val = float(cleaned)
                                numeric_values.append((row_idx, num_val))
                            except (ValueError, TypeError):
                                if col_key not in type_mismatches:
                                    type_mismatches[col_key] = []
                                type_mismatches[col_key].append(row_idx)

            # Record missing values
            if missing_count > 0:
                missing_values[col_key] = missing_count
                if len(table.rows) > 0:
                    missing_percentage[col_key] = round(missing_count / len(table.rows) * 100, 2)
                else:
                    missing_percentage[col_key] = 0.0

            unique_values_per_column[col_key] = len(unique_set)

            # Detect outliers for numeric columns
            if numeric_values and len(numeric_values) >= 5:
                vals = [v for _, v in numeric_values]
                mean = sum(vals) / len(vals)
                std = math.sqrt(sum((v - mean) ** 2 for v in vals) / len(vals))

                if std > 0:
                    outlier_indices = []
                    for row_idx, val in numeric_values:
                        if abs((val - mean) / std) > 3:
                            outlier_indices.append(row_idx)

                    if outlier_indices:
                        outliers_detected[col_key] = outlier_indices

            # Check format inconsistencies for text columns
            if dtype == "text" and unique_set:
                # Check for mixed case patterns
                patterns_found = set()
                for val in list(unique_set)[:50]:
                    if val.isupper():
                        patterns_found.add("UPPERCASE")
                    elif val.islower():
                        patterns_found.add("lowercase")
                    elif val.istitle():
                        patterns_found.add("Title Case")
                    else:
                        patterns_found.add("Mixed case")

                if len(patterns_found) > 1:
                    format_inconsistencies[col_key] = list(patterns_found)

    # Calculate quality score
    total_issues = (
        sum(missing_values.values()) +
        duplicate_rows * 2 +
        sum(len(v) for v in type_mismatches.values()) +
        sum(len(v) for v in outliers_detected.values())
    )

    max_possible_issues = total_rows * total_columns if total_rows and total_columns else 1
    quality_score = max(0, 1 - (total_issues / max_possible_issues))

    # Generate recommendations
    recommendations = []
    if missing_values:
        high_missing = [k for k, v in missing_percentage.items() if v > 20]
        if high_missing:
            recommendations.append(f"High missing data in columns: {', '.join(high_missing[:5])}")
    if duplicate_rows > 0:
        recommendations.append(f"Found {duplicate_rows} duplicate rows - consider deduplication")
    if type_mismatches:
        recommendations.append(f"Type mismatches detected in {len(type_mismatches)} columns")
    if outliers_detected:
        recommendations.append(f"Outliers detected in {len(outliers_detected)} numeric columns")

    return DataQualityReport(
        total_rows=total_rows,
        total_columns=total_columns,
        missing_values=missing_values,
        missing_percentage=missing_percentage,
        duplicate_rows=duplicate_rows,
        unique_values_per_column=unique_values_per_column,
        invalid_values=invalid_values,
        type_mismatches=type_mismatches,
        format_inconsistencies=format_inconsistencies,
        outliers_detected=outliers_detected,
        quality_score=round(quality_score, 3),
        recommendations=recommendations,
    )


# DATA CLEANING & TRANSFORMATION

def clean_table(
    table: EnhancedExtractedTable,
    operations: List[str] = None,
) -> EnhancedExtractedTable:
    """Apply cleaning operations to a table."""
    if operations is None:
        operations = ["trim", "normalize_case", "fill_missing"]

    cleaned_rows = [list(row) for row in table.rows]  # Deep copy

    for col_idx, (header, dtype) in enumerate(zip(table.headers, table.data_types)):
        for row_idx, row in enumerate(cleaned_rows):
            if col_idx >= len(row):
                continue

            val = row[col_idx]

            if "trim" in operations:
                if isinstance(val, str):
                    val = val.strip()

            if "normalize_case" in operations:
                if isinstance(val, str) and dtype == "text":
                    val = val.title()

            if "fill_missing" in operations:
                if val is None or str(val).strip() == "" or str(val).lower() in ("null", "n/a", "na"):
                    if dtype == "numeric":
                        # Fill with column median
                        col_values = []
                        for r in table.rows:
                            if col_idx < len(r):
                                try:
                                    col_values.append(float(str(r[col_idx]).replace(",", "").replace("$", "")))
                                except (ValueError, TypeError):
                                    pass
                        if col_values:
                            sorted_vals = sorted(col_values)
                            val = sorted_vals[len(sorted_vals) // 2]
                        else:
                            val = 0
                    else:
                        val = ""

            if "normalize_numbers" in operations:
                if dtype == "numeric":
                    try:
                        cleaned = re.sub(r'[$,% ]', '', str(val))
                        val = float(cleaned)
                    except (ValueError, TypeError):
                        pass

            cleaned_rows[row_idx][col_idx] = val

    return EnhancedExtractedTable(
        id=table.id,
        title=table.title,
        headers=table.headers,
        rows=cleaned_rows,
        data_types=table.data_types,
        source_page=table.source_page,
        source_sheet=table.source_sheet,
        confidence=table.confidence,
        row_count=len(cleaned_rows),
        column_count=len(table.headers),
        has_totals_row=table.has_totals_row,
        has_header_row=table.has_header_row,
        statistics=table.statistics,
    )


def apply_transformation(
    table: EnhancedExtractedTable,
    transformation: DataTransformation,
) -> EnhancedExtractedTable:
    """Apply a single transformation to a table."""
    if transformation.operation == "clean":
        return clean_table(table, transformation.parameters.get("operations"))

    elif transformation.operation == "normalize":
        # Normalize numeric columns to 0-1 range
        for col_name in transformation.source_columns:
            if col_name not in table.headers:
                continue
            col_idx = table.headers.index(col_name)

            values = []
            for row in table.rows:
                if col_idx < len(row):
                    try:
                        values.append(float(str(row[col_idx]).replace(",", "")))
                    except (ValueError, TypeError):
                        values.append(None)

            valid_values = [v for v in values if v is not None]
            if not valid_values:
                continue

            min_val = min(valid_values)
            max_val = max(valid_values)
            range_val = max_val - min_val if max_val != min_val else 1

            for row_idx, row in enumerate(table.rows):
                if col_idx < len(row) and values[row_idx] is not None:
                    row[col_idx] = round((values[row_idx] - min_val) / range_val, 4)

    elif transformation.operation == "aggregate":
        # Aggregate by group column
        group_col = transformation.parameters.get("group_by")
        agg_func = transformation.parameters.get("function", "sum")

        if group_col not in table.headers:
            return table

        group_idx = table.headers.index(group_col)
        value_cols = transformation.source_columns

        groups: Dict[str, Dict[str, List[float]]] = {}
        for row in table.rows:
            if group_idx >= len(row):
                continue
            group_key = str(row[group_idx])

            if group_key not in groups:
                groups[group_key] = {col: [] for col in value_cols}

            for col in value_cols:
                if col in table.headers:
                    col_idx = table.headers.index(col)
                    if col_idx < len(row):
                        try:
                            groups[group_key][col].append(float(str(row[col_idx]).replace(",", "")))
                        except (ValueError, TypeError):
                            pass

        # Build aggregated rows
        new_headers = [group_col] + value_cols
        new_rows = []

        for group_key, values_dict in groups.items():
            new_row = [group_key]
            for col in value_cols:
                vals = values_dict.get(col, [])
                if not vals:
                    new_row.append(0)
                elif agg_func == "sum":
                    new_row.append(sum(vals))
                elif agg_func == "mean":
                    new_row.append(sum(vals) / len(vals))
                elif agg_func == "count":
                    new_row.append(len(vals))
                elif agg_func == "min":
                    new_row.append(min(vals))
                elif agg_func == "max":
                    new_row.append(max(vals))
                else:
                    new_row.append(sum(vals))

            new_rows.append(new_row)

        return EnhancedExtractedTable(
            id=f"{table.id}_agg",
            title=f"{table.title or table.id} (Aggregated)",
            headers=new_headers,
            rows=new_rows,
            data_types=["text"] + ["numeric"] * len(value_cols),
            row_count=len(new_rows),
            column_count=len(new_headers),
        )

    return table


# EXPORT FUNCTIONS

def export_to_csv(
    tables: List[EnhancedExtractedTable],
    include_headers: bool = True,
) -> str:
    """Export tables to CSV format."""
    output = io.StringIO()
    writer = csv.writer(output)

    for table in tables:
        if include_headers:
            writer.writerow([f"# Table: {table.title or table.id}"])
            writer.writerow(table.headers)

        for row in table.rows:
            writer.writerow(row)

        writer.writerow([])  # Empty row between tables

    return output.getvalue()


def export_to_json(
    result: EnhancedAnalysisResult,
    include_raw_data: bool = True,
) -> str:
    """Export analysis result to JSON."""
    data = result.model_dump()

    if not include_raw_data:
        # Remove large data arrays
        for table in data.get("tables", []):
            table["rows"] = table["rows"][:10]  # Only first 10 rows

    return json.dumps(data, indent=2, default=str)


def export_to_markdown(result: EnhancedAnalysisResult) -> str:
    """Export analysis result to Markdown format."""
    lines = []

    # Title
    lines.append(f"# Analysis Report: {result.document_name}")
    lines.append(f"\n*Generated: {result.created_at.strftime('%Y-%m-%d %H:%M')}*\n")

    # Executive Summary
    if "executive" in result.summaries:
        lines.append("## Executive Summary\n")
        lines.append(result.summaries["executive"].content)
        lines.append("")

    # Key Metrics
    if result.metrics:
        lines.append("## Key Metrics\n")
        for metric in result.metrics[:10]:
            change_str = f" ({metric.change:+.1f}%)" if metric.change else ""
            lines.append(f"- **{metric.name}**: {metric.raw_value}{change_str}")
        lines.append("")

    # Tables
    if result.tables:
        lines.append("## Data Tables\n")
        for table in result.tables[:5]:
            lines.append(f"### {table.title or table.id}\n")
            # Header
            lines.append("| " + " | ".join(table.headers) + " |")
            lines.append("| " + " | ".join(["---"] * len(table.headers)) + " |")
            # Rows (limit to 10)
            for row in table.rows[:10]:
                lines.append("| " + " | ".join(str(v) for v in row) + " |")
            if len(table.rows) > 10:
                lines.append(f"\n*...and {len(table.rows) - 10} more rows*\n")
            lines.append("")

    # Insights
    if result.insights:
        lines.append("## Key Insights\n")
        for insight in result.insights:
            lines.append(f"### {insight.title}")
            lines.append(f"\n{insight.description}\n")
            if insight.suggested_actions:
                lines.append("**Suggested Actions:**")
                for action in insight.suggested_actions:
                    lines.append(f"- {action}")
            lines.append("")

    # Risks
    if result.risks:
        lines.append("## Risks Identified\n")
        for risk in result.risks:
            lines.append(f"- **{risk.title}** ({risk.risk_level.value}): {risk.description}")
        lines.append("")

    # Opportunities
    if result.opportunities:
        lines.append("## Opportunities\n")
        for opp in result.opportunities:
            lines.append(f"- **{opp.title}**: {opp.description}")
        lines.append("")

    return "\n".join(lines)


def export_to_html(result: EnhancedAnalysisResult) -> str:
    """Export analysis result to HTML format."""
    html_parts = ["""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>Analysis Report</title>
    <style>
        body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; margin: 40px; background: #f5f5f5; }
        .container { max-width: 1200px; margin: 0 auto; background: white; padding: 40px; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }
        h1 { color: #1a1a2e; border-bottom: 3px solid #4f46e5; padding-bottom: 10px; }
        h2 { color: #4f46e5; margin-top: 30px; }
        h3 { color: #374151; }
        .metric { display: inline-block; background: #f0f9ff; padding: 15px 25px; margin: 5px; border-radius: 8px; border-left: 4px solid #4f46e5; }
        .metric-value { font-size: 24px; font-weight: bold; color: #1a1a2e; }
        .metric-name { color: #6b7280; font-size: 14px; }
        table { width: 100%; border-collapse: collapse; margin: 20px 0; }
        th, td { padding: 12px; text-align: left; border-bottom: 1px solid #e5e7eb; }
        th { background: #f9fafb; font-weight: 600; }
        tr:hover { background: #f9fafb; }
        .insight { background: #fef3c7; padding: 15px; border-radius: 8px; margin: 10px 0; border-left: 4px solid #f59e0b; }
        .risk { background: #fee2e2; padding: 15px; border-radius: 8px; margin: 10px 0; border-left: 4px solid #ef4444; }
        .opportunity { background: #d1fae5; padding: 15px; border-radius: 8px; margin: 10px 0; border-left: 4px solid #10b981; }
        .badge { display: inline-block; padding: 4px 8px; border-radius: 4px; font-size: 12px; font-weight: 600; }
        .badge-high { background: #fee2e2; color: #dc2626; }
        .badge-medium { background: #fef3c7; color: #d97706; }
        .badge-low { background: #d1fae5; color: #059669; }
    </style>
</head>
<body>
<div class="container">
"""]

    # Title
    html_parts.append(f"<h1>Analysis Report: {html_mod.escape(str(result.document_name))}</h1>")
    html_parts.append(f"<p style='color:#6b7280;'>Generated: {html_mod.escape(str(result.created_at.strftime('%Y-%m-%d %H:%M')))}</p>")

    # Key Metrics
    if result.metrics:
        html_parts.append("<h2>Key Metrics</h2><div>")
        for metric in result.metrics[:8]:
            html_parts.append(f"""<div class="metric">
                <div class="metric-value">{html_mod.escape(str(metric.raw_value))}</div>
                <div class="metric-name">{html_mod.escape(str(metric.name))}</div>
            </div>""")
        html_parts.append("</div>")

    # Executive Summary
    if "executive" in result.summaries:
        html_parts.append("<h2>Executive Summary</h2>")
        html_parts.append(f"<p>{html_mod.escape(str(result.summaries['executive'].content))}</p>")

    # Tables
    if result.tables:
        html_parts.append("<h2>Data Tables</h2>")
        for table in result.tables[:3]:
            html_parts.append(f"<h3>{html_mod.escape(str(table.title or table.id))}</h3>")
            html_parts.append("<table><thead><tr>")
            for header in table.headers:
                html_parts.append(f"<th>{html_mod.escape(str(header))}</th>")
            html_parts.append("</tr></thead><tbody>")
            for row in table.rows[:15]:
                html_parts.append("<tr>")
                for val in row:
                    html_parts.append(f"<td>{html_mod.escape(str(val))}</td>")
                html_parts.append("</tr>")
            html_parts.append("</tbody></table>")
            if len(table.rows) > 15:
                html_parts.append(f"<p style='color:#6b7280;'>...and {len(table.rows) - 15} more rows</p>")

    # Insights
    if result.insights:
        html_parts.append("<h2>Key Insights</h2>")
        for insight in result.insights:
            html_parts.append(f"""<div class="insight">
                <strong>{html_mod.escape(str(insight.title))}</strong>
                <span class="badge badge-{html_mod.escape(str(insight.priority.value))}">{html_mod.escape(str(insight.priority.value))}</span>
                <p>{html_mod.escape(str(insight.description))}</p>
            </div>""")

    # Risks
    if result.risks:
        html_parts.append("<h2>Risks Identified</h2>")
        for risk in result.risks:
            html_parts.append(f"""<div class="risk">
                <strong>{html_mod.escape(str(risk.title))}</strong>
                <span class="badge badge-{html_mod.escape(str(risk.risk_level.value))}">{html_mod.escape(str(risk.risk_level.value))}</span>
                <p>{html_mod.escape(str(risk.description))}</p>
            </div>""")

    # Opportunities
    if result.opportunities:
        html_parts.append("<h2>Opportunities</h2>")
        for opp in result.opportunities:
            html_parts.append(f"""<div class="opportunity">
                <strong>{html_mod.escape(str(opp.title))}</strong>
                <p>{html_mod.escape(str(opp.description))}</p>
            </div>""")

    html_parts.append("</div></body></html>")

    return "\n".join(html_parts)


async def export_to_excel(
    result: EnhancedAnalysisResult,
    include_charts: bool = True,
) -> bytes:
    """Export analysis result to Excel format."""
    try:
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise RuntimeError("openpyxl is required for Excel export")

    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary["A1"] = "Analysis Report"
    ws_summary["A1"].font = Font(bold=True, size=16)
    ws_summary["A3"] = "Document:"
    ws_summary["B3"] = result.document_name
    ws_summary["A4"] = "Generated:"
    ws_summary["B4"] = result.created_at.strftime("%Y-%m-%d %H:%M")
    ws_summary["A5"] = "Tables Found:"
    ws_summary["B5"] = result.total_tables
    ws_summary["A6"] = "Metrics Extracted:"
    ws_summary["B6"] = result.total_metrics

    # Metrics Sheet
    if result.metrics:
        ws_metrics = wb.create_sheet("Metrics")
        headers = ["Name", "Value", "Type", "Period", "Change (%)", "Context"]
        for col, header in enumerate(headers, 1):
            cell = ws_metrics.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row, metric in enumerate(result.metrics, 2):
            ws_metrics.cell(row=row, column=1, value=metric.name)
            ws_metrics.cell(row=row, column=2, value=metric.raw_value)
            ws_metrics.cell(row=row, column=3, value=metric.metric_type.value)
            ws_metrics.cell(row=row, column=4, value=metric.period or "")
            ws_metrics.cell(row=row, column=5, value=metric.change or "")
            ws_metrics.cell(row=row, column=6, value=metric.context or "")

    # Data Tables
    for table in result.tables:
        # Sanitize sheet name
        sheet_name = (table.title or table.id)[:30].replace("/", "-").replace("\\", "-")
        try:
            ws_data = wb.create_sheet(sheet_name)
        except Exception:
            ws_data = wb.create_sheet(f"Table_{table.id[:20]}")

        # Headers
        for col, header in enumerate(table.headers, 1):
            cell = ws_data.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            ws_data.column_dimensions[get_column_letter(col)].width = max(12, len(header) + 2)

        # Data
        for row_idx, row in enumerate(table.rows, 2):
            for col_idx, val in enumerate(row, 1):
                ws_data.cell(row=row_idx, column=col_idx, value=val)

    # Insights Sheet
    if result.insights:
        ws_insights = wb.create_sheet("Insights")
        headers = ["Type", "Priority", "Title", "Description", "Actions"]
        for col, header in enumerate(headers, 1):
            cell = ws_insights.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for row, insight in enumerate(result.insights, 2):
            ws_insights.cell(row=row, column=1, value=insight.type)
            ws_insights.cell(row=row, column=2, value=insight.priority.value)
            ws_insights.cell(row=row, column=3, value=insight.title)
            ws_insights.cell(row=row, column=4, value=insight.description)
            ws_insights.cell(row=row, column=5, value="; ".join(insight.suggested_actions))

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


async def export_to_pdf(result: EnhancedAnalysisResult) -> bytes:
    """Export analysis result to PDF format."""
    html_content = export_to_html(result)

    # Try WeasyPrint first (HTML -> PDF)
    try:
        from weasyprint import HTML
        return HTML(string=html_content).write_pdf()
    except ImportError:
        pass
    except Exception as exc:
        logger.warning(f"WeasyPrint PDF export failed: {exc}")

    # Fallback to ReportLab if available
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas

        # Strip HTML tags for a basic text-only PDF
        text = re.sub(r"<[^>]+>", "", html_content)
        buffer = BytesIO()
        pdf = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter
        y = height - 72
        for line in text.splitlines():
            if y < 72:
                pdf.showPage()
                y = height - 72
            pdf.drawString(72, y, line[:120])
            y -= 14
        pdf.save()
        buffer.seek(0)
        return buffer.read()
    except ImportError:
        raise RuntimeError("PDF export requires weasyprint or reportlab to be installed.")
    except Exception as exc:
        raise RuntimeError(f"PDF export failed: {exc}") from exc


# EXPORT ORCHESTRATOR

class DataExportService:
    """Orchestrates data transformation and export operations."""

    def assess_quality(self, tables: List[EnhancedExtractedTable]) -> DataQualityReport:
        """Assess data quality."""
        return assess_data_quality(tables)

    def clean_data(
        self,
        tables: List[EnhancedExtractedTable],
        operations: List[str] = None,
    ) -> List[EnhancedExtractedTable]:
        """Clean all tables."""
        return [clean_table(t, operations) for t in tables]

    def apply_transformations(
        self,
        table: EnhancedExtractedTable,
        transformations: List[DataTransformation],
    ) -> EnhancedExtractedTable:
        """Apply multiple transformations to a table."""
        result = table
        for transform in transformations:
            result = apply_transformation(result, transform)
        return result

    async def export(
        self,
        result: EnhancedAnalysisResult,
        config: ExportConfiguration,
    ) -> Tuple[bytes, str]:
        """Export analysis result in specified format."""
        filename = config.filename or f"analysis_{result.analysis_id}"

        if config.format == ExportFormat.CSV:
            content = export_to_csv(result.tables)
            return content.encode('utf-8'), f"{filename}.csv"

        elif config.format == ExportFormat.JSON:
            content = export_to_json(result, config.include_raw_data)
            return content.encode('utf-8'), f"{filename}.json"

        elif config.format == ExportFormat.MARKDOWN:
            content = export_to_markdown(result)
            return content.encode('utf-8'), f"{filename}.md"

        elif config.format == ExportFormat.HTML:
            content = export_to_html(result)
            return content.encode('utf-8'), f"{filename}.html"

        elif config.format == ExportFormat.EXCEL:
            content = await export_to_excel(result, config.include_charts)
            return content, f"{filename}.xlsx"

        elif config.format == ExportFormat.PDF:
            content = await export_to_pdf(result)
            return content, f"{filename}.pdf"

        else:
            # Default to JSON
            content = export_to_json(result)
            return content.encode('utf-8'), f"{filename}.json"


# USER EXPERIENCE (merged from user_experience.py)


# 9.1 ANALYSIS CUSTOMIZATION

# Industry-specific configurations
INDUSTRY_CONFIGS = {
    "finance": {
        "focus_areas": ["financial", "risk", "compliance"],
        "key_metrics": ["revenue", "margin", "roi", "debt", "equity"],
        "terminology": ["EBITDA", "P/E ratio", "liquidity", "leverage"],
        "analysis_prompts": {
            "summary": "Focus on financial performance, risk indicators, and compliance matters.",
            "insights": "Identify financial risks, opportunities, and key performance drivers.",
        },
    },
    "healthcare": {
        "focus_areas": ["operational", "compliance", "patient"],
        "key_metrics": ["patient_volume", "readmission_rate", "cost_per_case"],
        "terminology": ["HIPAA", "CMS", "quality metrics", "patient outcomes"],
        "analysis_prompts": {
            "summary": "Focus on patient care metrics, compliance, and operational efficiency.",
            "insights": "Identify quality improvement opportunities and compliance risks.",
        },
    },
    "technology": {
        "focus_areas": ["growth", "innovation", "technical"],
        "key_metrics": ["mrr", "arr", "churn", "cac", "ltv"],
        "terminology": ["SaaS", "API", "scalability", "uptime"],
        "analysis_prompts": {
            "summary": "Focus on growth metrics, product performance, and technical capabilities.",
            "insights": "Identify growth opportunities, technical risks, and market trends.",
        },
    },
    "retail": {
        "focus_areas": ["sales", "inventory", "customer"],
        "key_metrics": ["same_store_sales", "inventory_turnover", "customer_acquisition"],
        "terminology": ["SKU", "foot traffic", "conversion rate", "basket size"],
        "analysis_prompts": {
            "summary": "Focus on sales performance, inventory management, and customer behavior.",
            "insights": "Identify sales trends, inventory optimization, and customer insights.",
        },
    },
    "manufacturing": {
        "focus_areas": ["operational", "quality", "supply_chain"],
        "key_metrics": ["oee", "defect_rate", "cycle_time", "inventory_days"],
        "terminology": ["lean", "six sigma", "yield", "throughput"],
        "analysis_prompts": {
            "summary": "Focus on production efficiency, quality metrics, and supply chain.",
            "insights": "Identify operational improvements, quality issues, and supply risks.",
        },
    },
}

# Output format configurations
OUTPUT_FORMATS = {
    "executive": {
        "max_summary_words": 150,
        "max_insights": 5,
        "include_technical_details": False,
        "visualization_style": "simple",
        "language_style": "concise",
    },
    "technical": {
        "max_summary_words": 500,
        "max_insights": 15,
        "include_technical_details": True,
        "visualization_style": "detailed",
        "language_style": "technical",
    },
    "visual": {
        "max_summary_words": 100,
        "max_insights": 8,
        "include_technical_details": False,
        "visualization_style": "rich",
        "chart_priority": True,
        "language_style": "brief",
    },
}


@dataclass
class AnalysisConfiguration:
    """Complete analysis configuration based on preferences."""
    preferences: AnalysisPreferences
    industry_config: Dict[str, Any] = field(default_factory=dict)
    output_config: Dict[str, Any] = field(default_factory=dict)
    custom_prompts: Dict[str, str] = field(default_factory=dict)


def build_analysis_configuration(
    preferences: AnalysisPreferences,
) -> AnalysisConfiguration:
    """Build complete analysis configuration from preferences."""
    # Get industry config
    industry_config = INDUSTRY_CONFIGS.get(preferences.industry, {})

    # Get output format config
    output_config = OUTPUT_FORMATS.get(preferences.output_format, OUTPUT_FORMATS["executive"])

    # Build custom prompts based on configuration
    custom_prompts = {}

    focus_str = ", ".join(preferences.focus_areas) if preferences.focus_areas else "general"
    depth_modifier = {
        AnalysisDepth.QUICK: "Be brief and highlight only the most critical points.",
        AnalysisDepth.STANDARD: "Provide a balanced analysis with key details.",
        AnalysisDepth.COMPREHENSIVE: "Provide thorough analysis with supporting details.",
        AnalysisDepth.DEEP: "Provide exhaustive analysis with all available details and nuances.",
    }

    base_prompt = f"Focus areas: {focus_str}. {depth_modifier.get(preferences.analysis_depth, '')}"

    if industry_config:
        base_prompt += f" Industry context: {preferences.industry}. "
        base_prompt += industry_config.get("analysis_prompts", {}).get("summary", "")

    custom_prompts["base"] = base_prompt
    custom_prompts["summary"] = industry_config.get("analysis_prompts", {}).get("summary", "")
    custom_prompts["insights"] = industry_config.get("analysis_prompts", {}).get("insights", "")

    return AnalysisConfiguration(
        preferences=preferences,
        industry_config=industry_config,
        output_config=output_config,
        custom_prompts=custom_prompts,
    )


def get_default_preferences() -> AnalysisPreferences:
    """Get default analysis preferences."""
    return AnalysisPreferences(
        analysis_depth=AnalysisDepth.STANDARD,
        focus_areas=["financial", "operational"],
        output_format="executive",
        language="en",
        currency_preference="USD",
        enable_predictions=True,
        enable_recommendations=True,
        auto_chart_generation=True,
        max_charts=10,
        summary_mode=SummaryMode.EXECUTIVE,
    )


# 9.2 COLLABORATION FEATURES

@dataclass
class AnalysisComment:
    """A comment on an analysis or specific element."""
    id: str
    analysis_id: str
    user_id: str
    user_name: str
    content: str
    element_type: Optional[str] = None  # table, chart, insight, metric
    element_id: Optional[str] = None
    created_at: datetime = field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = field(default_factory=lambda: datetime.now(timezone.utc))
    replies: List['AnalysisComment'] = field(default_factory=list)
    resolved: bool = False


@dataclass
class AnalysisShare:
    """Sharing configuration for an analysis."""
    id: str
    analysis_id: str
    share_type: str  # link, email, embed
    access_level: str  # view, comment, edit
    created_by: str
    created_at: datetime = field(default_factory=lambda: datetime.now(timezone.utc))
    expires_at: Optional[datetime] = None
    password_protected: bool = False
    access_count: int = 0
    allowed_emails: List[str] = field(default_factory=list)


@dataclass
class AnalysisVersion:
    """A version of an analysis."""
    version_id: str
    analysis_id: str
    version_number: int
    created_at: datetime
    created_by: str
    description: str
    changes: List[str]
    snapshot: Dict[str, Any]  # Serialized analysis state


class CollaborationService:
    """Manages collaboration features."""

    def __init__(self):
        self._comments: Dict[str, List[AnalysisComment]] = {}
        self._shares: Dict[str, List[AnalysisShare]] = {}
        self._versions: Dict[str, List[AnalysisVersion]] = {}
        self._store = get_analysis_store()

    def _parse_dt(self, value: Any) -> Optional[datetime]:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value
        try:
            return datetime.fromisoformat(str(value))
        except Exception:
            return None

    def _comment_to_dict(self, comment: AnalysisComment) -> Dict[str, Any]:
        return {
            "id": comment.id,
            "analysis_id": comment.analysis_id,
            "user_id": comment.user_id,
            "user_name": comment.user_name,
            "content": comment.content,
            "element_type": comment.element_type,
            "element_id": comment.element_id,
            "created_at": comment.created_at.isoformat() if comment.created_at else None,
            "updated_at": comment.updated_at.isoformat() if comment.updated_at else None,
            "replies": [self._comment_to_dict(r) for r in comment.replies],
            "resolved": comment.resolved,
        }

    def _comment_from_dict(self, data: Dict[str, Any]) -> AnalysisComment:
        replies = [self._comment_from_dict(r) for r in data.get("replies", [])]
        return AnalysisComment(
            id=data.get("id", f"comment_{uuid.uuid4().hex[:12]}"),
            analysis_id=data.get("analysis_id", ""),
            user_id=data.get("user_id", "anonymous"),
            user_name=data.get("user_name", "Anonymous"),
            content=data.get("content", ""),
            element_type=data.get("element_type"),
            element_id=data.get("element_id"),
            created_at=self._parse_dt(data.get("created_at")) or datetime.now(timezone.utc),
            updated_at=self._parse_dt(data.get("updated_at")) or datetime.now(timezone.utc),
            replies=replies,
            resolved=bool(data.get("resolved", False)),
        )

    def _share_to_dict(self, share: AnalysisShare) -> Dict[str, Any]:
        return {
            "id": share.id,
            "analysis_id": share.analysis_id,
            "share_type": share.share_type,
            "access_level": share.access_level,
            "created_by": share.created_by,
            "created_at": share.created_at.isoformat() if share.created_at else None,
            "expires_at": share.expires_at.isoformat() if share.expires_at else None,
            "password_protected": share.password_protected,
            "access_count": share.access_count,
            "allowed_emails": list(share.allowed_emails or []),
        }

    def _share_from_dict(self, data: Dict[str, Any]) -> AnalysisShare:
        return AnalysisShare(
            id=data.get("id", f"share_{uuid.uuid4().hex[:12]}"),
            analysis_id=data.get("analysis_id", ""),
            share_type=data.get("share_type", "link"),
            access_level=data.get("access_level", "view"),
            created_by=data.get("created_by", "api"),
            created_at=self._parse_dt(data.get("created_at")) or datetime.now(timezone.utc),
            expires_at=self._parse_dt(data.get("expires_at")),
            password_protected=bool(data.get("password_protected", False)),
            access_count=int(data.get("access_count", 0) or 0),
            allowed_emails=list(data.get("allowed_emails") or []),
        )

    def _version_to_dict(self, version: AnalysisVersion) -> Dict[str, Any]:
        return {
            "version_id": version.version_id,
            "analysis_id": version.analysis_id,
            "version_number": version.version_number,
            "created_at": version.created_at.isoformat() if version.created_at else None,
            "created_by": version.created_by,
            "description": version.description,
            "changes": list(version.changes or []),
            "snapshot": version.snapshot,
        }

    def _version_from_dict(self, data: Dict[str, Any]) -> AnalysisVersion:
        return AnalysisVersion(
            version_id=data.get("version_id", f"v_{uuid.uuid4().hex[:12]}"),
            analysis_id=data.get("analysis_id", ""),
            version_number=int(data.get("version_number", 1) or 1),
            created_at=self._parse_dt(data.get("created_at")) or datetime.now(timezone.utc),
            created_by=data.get("created_by", "api"),
            description=data.get("description", ""),
            changes=list(data.get("changes") or []),
            snapshot=data.get("snapshot") or {},
        )

    def _ensure_comments_loaded(self, analysis_id: str) -> None:
        if analysis_id not in self._comments:
            payload = self._store.load_comments(analysis_id)
            self._comments[analysis_id] = [self._comment_from_dict(p) for p in payload]

    def _ensure_shares_loaded(self, analysis_id: str) -> None:
        if analysis_id not in self._shares:
            payload = self._store.list_shares_for_analysis(analysis_id)
            self._shares[analysis_id] = [self._share_from_dict(p) for p in payload]

    def _ensure_versions_loaded(self, analysis_id: str) -> None:
        if analysis_id not in self._versions:
            payload = self._store.load_versions(analysis_id)
            self._versions[analysis_id] = [self._version_from_dict(p) for p in payload]

    def _persist_comments(self, analysis_id: str) -> None:
        payload = [self._comment_to_dict(c) for c in self._comments.get(analysis_id, [])]
        self._store.save_comments(analysis_id, payload)

    def _persist_versions(self, analysis_id: str) -> None:
        payload = [self._version_to_dict(v) for v in self._versions.get(analysis_id, [])]
        self._store.save_versions(analysis_id, payload)

    def add_comment(
        self,
        analysis_id: str,
        user_id: str,
        user_name: str,
        content: str,
        element_type: Optional[str] = None,
        element_id: Optional[str] = None,
        parent_comment_id: Optional[str] = None,
    ) -> AnalysisComment:
        """Add a comment to an analysis."""
        comment = AnalysisComment(
            id=f"comment_{uuid.uuid4().hex[:12]}",
            analysis_id=analysis_id,
            user_id=user_id,
            user_name=user_name,
            content=content,
            element_type=element_type,
            element_id=element_id,
        )

        self._ensure_comments_loaded(analysis_id)

        if parent_comment_id:
            # Find parent and add as reply
            for existing in self._comments[analysis_id]:
                if existing.id == parent_comment_id:
                    existing.replies.append(comment)
                    break
        else:
            self._comments[analysis_id].append(comment)

        self._persist_comments(analysis_id)
        return comment

    def get_comments(self, analysis_id: str) -> List[AnalysisComment]:
        """Get all comments for an analysis."""
        self._ensure_comments_loaded(analysis_id)
        return self._comments.get(analysis_id, [])

    def create_share_link(
        self,
        analysis_id: str,
        created_by: str,
        access_level: str = "view",
        expires_hours: Optional[int] = None,
        password_protected: bool = False,
        allowed_emails: List[str] = None,
    ) -> AnalysisShare:
        """Create a shareable link for an analysis."""
        share = AnalysisShare(
            id=f"share_{uuid.uuid4().hex[:12]}",
            analysis_id=analysis_id,
            share_type="link",
            access_level=access_level,
            created_by=created_by,
            expires_at=datetime.now(timezone.utc) + timedelta(hours=expires_hours) if expires_hours else None,
            password_protected=password_protected,
            allowed_emails=allowed_emails or [],
        )

        self._ensure_shares_loaded(analysis_id)
        self._shares[analysis_id].append(share)
        self._store.save_share(self._share_to_dict(share))

        return share

    def get_shares(self, analysis_id: str) -> List[AnalysisShare]:
        """List shares for an analysis."""
        self._ensure_shares_loaded(analysis_id)
        return self._shares.get(analysis_id, [])

    def get_share(self, share_id: str) -> Optional[AnalysisShare]:
        """Get a share by ID."""
        payload = self._store.load_share(share_id)
        if not payload:
            return None
        return self._share_from_dict(payload)

    def record_share_access(self, share_id: str) -> None:
        """Increment access count for a share."""
        share = self.get_share(share_id)
        if not share:
            return
        share.access_count += 1
        # Update in-memory list if loaded
        if share.analysis_id in self._shares:
            for idx, existing in enumerate(self._shares[share.analysis_id]):
                if existing.id == share.id:
                    self._shares[share.analysis_id][idx] = share
                    break
        self._store.save_share(self._share_to_dict(share))

    def save_version(
        self,
        analysis_id: str,
        created_by: str,
        description: str,
        analysis_snapshot: Dict[str, Any],
    ) -> AnalysisVersion:
        """Save a version of the analysis."""
        if analysis_id not in self._versions:
            self._ensure_versions_loaded(analysis_id)

        version_number = len(self._versions[analysis_id]) + 1

        # Detect changes from previous version
        changes = []
        if self._versions[analysis_id]:
            prev = self._versions[analysis_id][-1].snapshot
            # Compare key metrics
            prev_metrics = len(prev.get("metrics", []))
            curr_metrics = len(analysis_snapshot.get("metrics", []))
            if curr_metrics != prev_metrics:
                changes.append(f"Metrics: {prev_metrics} -> {curr_metrics}")

        version = AnalysisVersion(
            version_id=f"v_{uuid.uuid4().hex[:12]}",
            analysis_id=analysis_id,
            version_number=version_number,
            created_at=datetime.now(timezone.utc),
            created_by=created_by,
            description=description,
            changes=changes,
            snapshot=analysis_snapshot,
        )

        self._versions[analysis_id].append(version)
        self._persist_versions(analysis_id)
        return version

    def get_version_history(self, analysis_id: str) -> List[AnalysisVersion]:
        """Get version history for an analysis."""
        self._ensure_versions_loaded(analysis_id)
        return self._versions.get(analysis_id, [])


# Need to import timedelta
from datetime import timedelta


# 9.3 REAL-TIME FEATURES

@dataclass
class ProgressUpdate:
    """A progress update for streaming."""
    stage: str
    progress: float  # 0-100
    detail: str
    timestamp: datetime = field(default_factory=lambda: datetime.now(timezone.utc))
    data: Optional[Dict[str, Any]] = None


@dataclass
class IncrementalResult:
    """An incremental result during analysis."""
    result_type: str  # table, entity, metric, chart, insight
    data: Any
    is_final: bool = False


class StreamingAnalysisSession:
    """Manages a streaming analysis session."""

    def __init__(self, session_id: str):
        self.session_id = session_id
        self.started_at = datetime.now(timezone.utc)
        self.progress = 0.0
        self.current_stage = "initializing"
        self.is_cancelled = False
        self.is_complete = False
        self._progress_callbacks: List[Callable[[ProgressUpdate], None]] = []
        self._result_callbacks: List[Callable[[IncrementalResult], None]] = []
        self._incremental_results: List[IncrementalResult] = []

    def add_progress_callback(self, callback: Callable[[ProgressUpdate], None]) -> None:
        """Add a callback for progress updates."""
        self._progress_callbacks.append(callback)

    def add_result_callback(self, callback: Callable[[IncrementalResult], None]) -> None:
        """Add a callback for incremental results."""
        self._result_callbacks.append(callback)

    def update_progress(self, stage: str, progress: float, detail: str, data: Optional[Dict] = None) -> None:
        """Update progress and notify callbacks."""
        if self.is_cancelled:
            raise asyncio.CancelledError("Analysis was cancelled")

        self.current_stage = stage
        self.progress = progress

        update = ProgressUpdate(
            stage=stage,
            progress=progress,
            detail=detail,
            data=data,
        )

        for callback in self._progress_callbacks:
            try:
                callback(update)
            except Exception as e:
                logger.warning(f"Progress callback error: {e}")

    def emit_result(self, result_type: str, data: Any, is_final: bool = False) -> None:
        """Emit an incremental result."""
        result = IncrementalResult(
            result_type=result_type,
            data=data,
            is_final=is_final,
        )
        self._incremental_results.append(result)

        for callback in self._result_callbacks:
            try:
                callback(result)
            except Exception as e:
                logger.warning(f"Result callback error: {e}")

    def cancel(self) -> None:
        """Cancel the analysis session."""
        self.is_cancelled = True

    def complete(self) -> None:
        """Mark the session as complete."""
        self.is_complete = True
        self.progress = 100.0


async def stream_analysis_progress(
    session: StreamingAnalysisSession,
) -> AsyncGenerator[Dict[str, Any], None]:
    """Stream analysis progress as server-sent events."""
    last_progress = -1

    while not session.is_complete and not session.is_cancelled:
        if session.progress != last_progress:
            last_progress = session.progress
            yield {
                "event": "progress",
                "stage": session.current_stage,
                "progress": session.progress,
                "timestamp": datetime.now(timezone.utc).isoformat(),
            }

        # Check for new incremental results
        pending = list(session._incremental_results)
        session._incremental_results.clear()
        for result in pending:
            yield {
                "event": "result",
                "type": result.result_type,
                "data": result.data if isinstance(result.data, dict) else str(result.data),
                "is_final": result.is_final,
            }

        await asyncio.sleep(0.1)

    if session.is_cancelled:
        yield {"event": "cancelled", "timestamp": datetime.now(timezone.utc).isoformat()}
    else:
        yield {"event": "complete", "progress": 100, "timestamp": datetime.now(timezone.utc).isoformat()}


# SUGGESTED QUESTIONS

def generate_suggested_questions(
    tables: List[Any],
    metrics: List[Any],
    entities: List[Any],
) -> List[str]:
    """Generate suggested questions based on extracted data."""
    questions = []

    # Questions based on metrics
    if metrics:
        metric_names = [m.name for m in metrics[:5]]
        for name in metric_names:
            questions.append(f"What factors contributed to the {name}?")
            questions.append(f"How does the {name} compare to previous periods?")

        if len(metrics) >= 2:
            questions.append(f"What's the relationship between {metric_names[0]} and {metric_names[1]}?")

    # Questions based on tables
    if tables:
        for table in tables[:3]:
            if hasattr(table, 'title') and table.title:
                questions.append(f"What are the key insights from the {table.title} data?")
            if hasattr(table, 'headers'):
                numeric_cols = [h for h, d in zip(table.headers, getattr(table, 'data_types', [])) if d == 'numeric']
                if numeric_cols:
                    questions.append(f"What's the trend for {numeric_cols[0]}?")

    # Questions based on entities
    if entities:
        org_entities = [e for e in entities if hasattr(e, 'type') and e.type.value == 'organization']
        if org_entities:
            questions.append(f"What is the role of {org_entities[0].value} in this document?")

    # Generic insightful questions
    questions.extend([
        "What are the main risks identified in this document?",
        "What opportunities are suggested by the data?",
        "What actions should be taken based on these findings?",
        "Are there any anomalies or unusual patterns?",
        "How does this compare to industry benchmarks?",
    ])

    return questions[:10]  # Return top 10


# USER EXPERIENCE SERVICE ORCHESTRATOR

class UserExperienceService:
    """Orchestrates user experience features."""

    def __init__(self):
        self.collaboration = CollaborationService()
        self._active_sessions: Dict[str, StreamingAnalysisSession] = {}

    def build_configuration(self, preferences: AnalysisPreferences) -> AnalysisConfiguration:
        """Build analysis configuration from preferences."""
        return build_analysis_configuration(preferences)

    def get_industry_options(self) -> List[Dict[str, str]]:
        """Get available industry options."""
        return [
            {"value": key, "label": key.replace("_", " ").title(), "description": config.get("focus_areas", [])}
            for key, config in INDUSTRY_CONFIGS.items()
        ]

    def get_output_format_options(self) -> List[Dict[str, Any]]:
        """Get available output format options."""
        return [
            {"value": key, "label": key.title(), "config": config}
            for key, config in OUTPUT_FORMATS.items()
        ]

    def create_streaming_session(self) -> StreamingAnalysisSession:
        """Create a new streaming analysis session."""
        session_id = f"session_{uuid.uuid4().hex[:12]}"
        session = StreamingAnalysisSession(session_id)
        self._active_sessions[session_id] = session
        return session

    def get_session(self, session_id: str) -> Optional[StreamingAnalysisSession]:
        """Get an active session."""
        return self._active_sessions.get(session_id)

    def cancel_session(self, session_id: str) -> bool:
        """Cancel an active session."""
        session = self._active_sessions.get(session_id)
        if session:
            session.cancel()
            return True
        return False

    def generate_suggested_questions(
        self,
        tables: List[Any],
        metrics: List[Any],
        entities: List[Any],
    ) -> List[str]:
        """Generate suggested questions."""
        return generate_suggested_questions(tables, metrics, entities)

    # Collaboration methods
    def add_comment(self, *args, **kwargs) -> AnalysisComment:
        return self.collaboration.add_comment(*args, **kwargs)

    def get_comments(self, analysis_id: str) -> List[AnalysisComment]:
        return self.collaboration.get_comments(analysis_id)

    def create_share_link(self, *args, **kwargs) -> AnalysisShare:
        return self.collaboration.create_share_link(*args, **kwargs)

    def get_share(self, share_id: str) -> Optional[AnalysisShare]:
        return self.collaboration.get_share(share_id)

    def record_share_access(self, share_id: str) -> None:
        self.collaboration.record_share_access(share_id)

    def save_version(self, *args, **kwargs) -> AnalysisVersion:
        return self.collaboration.save_version(*args, **kwargs)

    def get_version_history(self, analysis_id: str) -> List[AnalysisVersion]:
        return self.collaboration.get_version_history(analysis_id)


# Convenience class reference
DocumentAnalysisService = type("DocumentAnalysisService", (), {
    "analyze_streaming": staticmethod(analyze_document_streaming),
    "get_analysis": staticmethod(get_analysis),
    "get_data": staticmethod(get_analysis_data),
    "suggest_charts": staticmethod(suggest_charts_for_analysis),
})
