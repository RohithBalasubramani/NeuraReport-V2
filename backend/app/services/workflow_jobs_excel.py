"""Consolidated Workflow, Excel & Generator Services (Phase B6).

Merged from: workflow/service.py, excel/ExcelVerify.py,
generator/GeneratorAssetsV1.py.
"""
from __future__ import annotations

"""Workflow Service.

Core workflow management and execution service.
"""

import asyncio
import logging
import uuid
from datetime import datetime, timezone
from typing import Any, Optional

from backend.app.common import utc_now, utc_now_iso
from backend.app.schemas import (
    CreateWorkflowRequest,
    ExecutionStatus,
    NodeExecutionResult,
    NodeType,
    UpdateWorkflowRequest,
    WorkflowEdge,
    WorkflowExecutionResponse,
    WorkflowNode,
    WorkflowResponse,
    WorkflowTrigger,
)

logger = logging.getLogger(__name__)

class WorkflowService:
    """Service for managing workflows."""

    def __init__(self):
        self._workflows: dict[str, dict] = {}
        self._executions: dict[str, dict] = {}
        self._pending_approvals: dict[str, dict] = {}
        self._running_tasks: set = set()

    def _safe_eval_condition(self, condition: str, context: dict[str, Any]) -> bool:
        """Safely evaluate a condition expression without using eval().

        Supports simple comparisons like:
        - "true" / "false"
        - "input.value > 10"
        - "input.status == 'active'"
        - "input.count >= 5 and input.enabled"
        """
        import operator

        condition_raw = condition.strip()
        condition_lower = condition_raw.lower()

        # Handle simple boolean literals
        if condition_lower in ("true", "1", "yes"):
            return True
        if condition_lower in ("false", "0", "no", ""):
            return False

        condition = condition_raw

        # Safe operators mapping
        ops = {
            "==": operator.eq,
            "!=": operator.ne,
            ">=": operator.ge,
            "<=": operator.le,
            ">": operator.gt,
            "<": operator.lt,
        }

        def get_value(path: str, ctx: dict) -> Any:
            """Safely get a nested value from context."""
            parts = path.strip().split(".")
            value = ctx
            for part in parts:
                if isinstance(value, dict):
                    value = value.get(part)
                else:
                    return None
            return value

        def parse_literal(s: str) -> Any:
            """Parse a literal value (string, number, bool)."""
            s = s.strip()
            if s.startswith(("'", '"')) and s.endswith(("'", '"')):
                return s[1:-1]
            if s.lower() == "true":
                return True
            if s.lower() == "false":
                return False
            if s.lower() == "none":
                return None
            try:
                if "." in s:
                    return float(s)
                return int(s)
            except ValueError:
                # Treat as context path
                return get_value(s, context)

        # Handle compound conditions (and/or)
        if " and " in condition:
            parts = condition.split(" and ")
            return all(self._safe_eval_condition(p.strip(), context) for p in parts)
        if " or " in condition:
            parts = condition.split(" or ")
            return any(self._safe_eval_condition(p.strip(), context) for p in parts)

        # Handle negation
        if condition.startswith("not "):
            return not self._safe_eval_condition(condition[4:].strip(), context)

        # Handle comparisons
        for op_str, op_func in ops.items():
            if op_str in condition:
                left, right = condition.split(op_str, 1)
                left_val = parse_literal(left)
                right_val = parse_literal(right)
                try:
                    return op_func(left_val, right_val)
                except (TypeError, ValueError):
                    return False

        # Handle simple truthiness check (e.g., "input.enabled")
        value = parse_literal(condition)
        return bool(value)

    async def create_workflow(
        self,
        request: CreateWorkflowRequest,
    ) -> WorkflowResponse:
        """Create a new workflow."""
        workflow_id = str(uuid.uuid4())
        now = utc_now()

        workflow = {
            "id": workflow_id,
            "name": request.name,
            "description": request.description,
            "nodes": [n.model_dump() for n in request.nodes],
            "edges": [e.model_dump() for e in request.edges],
            "triggers": [t.model_dump() for t in request.triggers],
            "is_active": request.is_active,
            "created_at": now,
            "updated_at": now,
            "last_run_at": None,
            "run_count": 0,
        }

        self._workflows[workflow_id] = workflow

        # Persist to state store
        try:
            from backend.app.repositories import state_store
            with state_store.transaction() as state:
                state["workflows"][workflow_id] = workflow
        except Exception as e:
            logger.warning(f"Failed to persist workflow to state store: {e}")

        return self._to_response(workflow)

    async def get_workflow(self, workflow_id: str) -> Optional[WorkflowResponse]:
        """Get a workflow by ID."""
        workflow = self._workflows.get(workflow_id)
        if not workflow:
            # Try loading from state store
            try:
                from backend.app.repositories import state_store
                with state_store.transaction() as state:
                    workflow = state.get("workflows", {}).get(workflow_id)
                    if workflow:
                        self._workflows[workflow_id] = workflow
            except Exception as e:
                logger.debug("Failed to load workflow from state store: %s", e)

        if not workflow:
            return None
        return self._to_response(workflow)

    async def list_workflows(
        self,
        active_only: bool = False,
        limit: int = 50,
        offset: int = 0,
    ) -> tuple[list[WorkflowResponse], int]:
        """List all workflows."""
        # Load from state store
        try:
            from backend.app.repositories import state_store
            with state_store.transaction() as state:
                self._workflows.update(state.get("workflows", {}))
        except Exception as e:
            logger.debug("Failed to load workflows from state store: %s", e)

        workflows = list(self._workflows.values())

        if active_only:
            workflows = [w for w in workflows if w.get("is_active")]

        workflows.sort(key=lambda w: w.get("updated_at", ""), reverse=True)
        total = len(workflows)
        workflows = workflows[offset:offset + limit]

        return [self._to_response(w) for w in workflows], total

    async def update_workflow(
        self,
        workflow_id: str,
        request: UpdateWorkflowRequest,
    ) -> Optional[WorkflowResponse]:
        """Update a workflow."""
        workflow = self._workflows.get(workflow_id)
        if not workflow:
            return None

        if request.name is not None:
            workflow["name"] = request.name
        if request.description is not None:
            workflow["description"] = request.description
        if request.nodes is not None:
            workflow["nodes"] = [n.model_dump() for n in request.nodes]
        if request.edges is not None:
            workflow["edges"] = [e.model_dump() for e in request.edges]
        if request.triggers is not None:
            workflow["triggers"] = [t.model_dump() for t in request.triggers]
        if request.is_active is not None:
            workflow["is_active"] = request.is_active

        workflow["updated_at"] = utc_now()

        # Persist to state store
        try:
            from backend.app.repositories import state_store
            with state_store.transaction() as state:
                state["workflows"][workflow_id] = workflow
        except Exception as e:
            logger.warning(f"Failed to persist workflow update: {e}")

        return self._to_response(workflow)

    async def delete_workflow(self, workflow_id: str) -> bool:
        """Delete a workflow."""
        if workflow_id not in self._workflows:
            return False

        del self._workflows[workflow_id]

        # Remove from state store
        try:
            from backend.app.repositories import state_store
            with state_store.transaction() as state:
                state["workflows"].pop(workflow_id, None)
        except Exception as e:
            logger.warning(f"Failed to delete workflow from state store: {e}")

        return True

    async def execute_workflow(
        self,
        workflow_id: str,
        input_data: dict[str, Any],
        async_execution: bool = True,
    ) -> WorkflowExecutionResponse:
        """Execute a workflow."""
        workflow = self._workflows.get(workflow_id)
        if not workflow:
            raise ValueError(f"Workflow {workflow_id} not found")

        execution_id = str(uuid.uuid4())
        now = utc_now()

        execution = {
            "id": execution_id,
            "workflow_id": workflow_id,
            "status": ExecutionStatus.PENDING.value,
            "input_data": input_data,
            "output_data": None,
            "node_results": [],
            "error": None,
            "started_at": now,
            "finished_at": None,
        }

        self._executions[execution_id] = execution

        # Update workflow stats
        workflow["last_run_at"] = now
        workflow["run_count"] = workflow.get("run_count", 0) + 1

        if async_execution:
            # Schedule async execution
            task = asyncio.create_task(self._run_workflow(execution_id, workflow, input_data))
            self._running_tasks.add(task)
            task.add_done_callback(self._running_tasks.discard)
        else:
            # Run synchronously
            await self._run_workflow(execution_id, workflow, input_data)
            execution = self._executions[execution_id]

        return self._execution_to_response(execution)

    async def _run_workflow(
        self,
        execution_id: str,
        workflow: dict,
        input_data: dict[str, Any],
    ) -> None:
        """Run workflow execution."""
        execution = self._executions[execution_id]
        execution["status"] = ExecutionStatus.RUNNING.value

        try:
            nodes = workflow.get("nodes", [])
            edges = workflow.get("edges", [])

            # Build execution order (topological sort)
            node_map = {n["id"]: n for n in nodes}
            incoming = {n["id"]: [] for n in nodes}
            outgoing = {n["id"]: [] for n in nodes}

            for edge in edges:
                outgoing[edge["source"]].append(edge["target"])
                incoming[edge["target"]].append(edge["source"])

            # Find start nodes (no incoming edges)
            queue = [nid for nid, inc in incoming.items() if not inc]
            enqueued = set(queue)
            context = {"input": input_data, "outputs": {}}

            while queue:
                node_id = queue.pop(0)
                node = node_map.get(node_id)
                if not node:
                    continue

                # Execute node
                result = await self._execute_node(node, context)
                execution["node_results"].append(result)

                if result["status"] == ExecutionStatus.FAILED.value:
                    execution["status"] = ExecutionStatus.FAILED.value
                    execution["error"] = result.get("error")
                    break

                if result["status"] == ExecutionStatus.WAITING_APPROVAL.value:
                    execution["status"] = ExecutionStatus.WAITING_APPROVAL.value
                    self._pending_approvals[execution_id] = {
                        "node_id": node_id,
                        "workflow_id": workflow["id"],
                    }
                    return

                # Store output for downstream nodes
                context["outputs"][node_id] = result.get("output", {})

                # Add downstream nodes to queue
                for target in outgoing.get(node_id, []):
                    if target in enqueued:
                        continue
                    # Check if all dependencies are satisfied
                    deps = incoming.get(target, [])
                    if all(d in context["outputs"] for d in deps):
                        queue.append(target)
                        enqueued.add(target)

            if execution["status"] == ExecutionStatus.RUNNING.value:
                execution["status"] = ExecutionStatus.COMPLETED.value
                execution["output_data"] = context["outputs"]

        except Exception as e:
            logger.error(f"Workflow execution failed: {e}")
            execution["status"] = ExecutionStatus.FAILED.value
            execution["error"] = "Workflow execution failed. Check logs for details."
            logger.error("Workflow execution %s failed: %s", execution_id, e, exc_info=True)

        finally:
            execution["finished_at"] = utc_now()

            # Persist execution
            try:
                from backend.app.repositories import state_store
                with state_store.transaction() as state:
                    state["workflow_executions"][execution_id] = execution
            except Exception as e:
                logger.error("Failed to persist execution %s: %s", execution_id, e)

    async def _execute_node(
        self,
        node: dict,
        context: dict[str, Any],
    ) -> dict:
        """Execute a single workflow node."""
        node_id = node["id"]
        node_type = node["type"]
        config = node.get("config", {})
        now = utc_now()

        result = {
            "node_id": node_id,
            "status": ExecutionStatus.RUNNING.value,
            "output": None,
            "error": None,
            "started_at": now,
            "finished_at": None,
        }

        try:
            if node_type == NodeType.TRIGGER.value:
                # Trigger node - pass input through
                result["output"] = context.get("input", {})

            elif node_type == NodeType.CONDITION.value:
                # Evaluate condition safely (no eval)
                condition = config.get("condition", "true")
                passed = self._safe_eval_condition(condition, context)
                result["output"] = {"passed": passed}

            elif node_type == NodeType.ACTION.value:
                # Execute action
                action_type = config.get("action_type", "log")
                if action_type == "log":
                    logger.info(f"Workflow action: {config.get('message', '')}")
                result["output"] = {"action": action_type, "executed": True}

            elif node_type == NodeType.EMAIL.value:
                # Send email (placeholder)
                result["output"] = {
                    "sent": True,
                    "to": config.get("to"),
                    "subject": config.get("subject"),
                }

            elif node_type == NodeType.APPROVAL.value:
                # Require approval
                result["status"] = ExecutionStatus.WAITING_APPROVAL.value
                result["output"] = {"awaiting_approval": True}
                return result

            elif node_type == NodeType.DATA_TRANSFORM.value:
                # Transform data
                transform = config.get("transform", {})
                result["output"] = {"transformed": True, "data": transform}

            elif node_type == NodeType.DELAY.value:
                # Wait for specified duration
                delay_ms = config.get("delay_ms", 1000)
                await asyncio.sleep(delay_ms / 1000)
                result["output"] = {"delayed": True, "duration_ms": delay_ms}

            elif node_type == NodeType.HTTP_REQUEST.value:
                # Make HTTP request
                import aiohttp
                async with aiohttp.ClientSession() as session:
                    method = config.get("method", "GET")
                    url = config.get("url", "")
                    async with session.request(method, url) as resp:
                        result["output"] = {
                            "status_code": resp.status,
                            "body": await resp.text(),
                        }

            else:
                result["output"] = {"node_type": node_type, "executed": True}

            result["status"] = ExecutionStatus.COMPLETED.value

        except Exception as e:
            logger.error("Workflow node execution failed: %s", e, exc_info=True)
            result["status"] = ExecutionStatus.FAILED.value
            result["error"] = "Workflow node execution failed"

        finally:
            result["finished_at"] = utc_now()

        return result

    async def get_execution(
        self,
        execution_id: str,
    ) -> Optional[WorkflowExecutionResponse]:
        """Get execution status."""
        execution = self._executions.get(execution_id)
        if not execution:
            # Try loading from state store
            try:
                from backend.app.repositories import state_store
                with state_store.transaction() as state:
                    execution = state.get("workflow_executions", {}).get(execution_id)
            except Exception as e:
                logger.debug("Failed to load execution from state store: %s", e)

        if not execution:
            return None
        return self._execution_to_response(execution)

    async def list_executions(
        self,
        workflow_id: Optional[str] = None,
        status: Optional[ExecutionStatus] = None,
        limit: int = 50,
    ) -> list[WorkflowExecutionResponse]:
        """List workflow executions."""
        # Load from state store
        try:
            from backend.app.repositories import state_store
            with state_store.transaction() as state:
                self._executions.update(state.get("workflow_executions", {}))
        except Exception as e:
            logger.debug("Failed to load executions from state store: %s", e)

        executions = list(self._executions.values())

        if workflow_id:
            executions = [e for e in executions if e.get("workflow_id") == workflow_id]
        if status:
            executions = [e for e in executions if e.get("status") == status.value]

        executions.sort(key=lambda e: e.get("started_at", ""), reverse=True)
        executions = executions[:limit]

        return [self._execution_to_response(e) for e in executions]

    async def approve_execution(
        self,
        execution_id: str,
        node_id: str,
        approved: bool,
        comment: Optional[str] = None,
    ) -> Optional[WorkflowExecutionResponse]:
        """Approve or reject a pending approval."""
        approval = self._pending_approvals.get(execution_id)
        if not approval or approval.get("node_id") != node_id:
            return None

        execution = self._executions.get(execution_id)
        if not execution:
            return None

        del self._pending_approvals[execution_id]

        if approved:
            # Continue execution
            workflow = self._workflows.get(execution["workflow_id"])
            if workflow:
                task = asyncio.create_task(
                    self._run_workflow(execution_id, workflow, execution["input_data"])
                )
                self._running_tasks.add(task)
                task.add_done_callback(self._running_tasks.discard)
        else:
            execution["status"] = ExecutionStatus.CANCELLED.value
            execution["error"] = f"Rejected: {comment or 'No reason provided'}"
            execution["finished_at"] = utc_now()

        return self._execution_to_response(execution)

    async def get_pending_approvals(
        self,
        workflow_id: Optional[str] = None,
    ) -> list[dict]:
        """Get pending approvals."""
        approvals = []
        for exec_id, approval in self._pending_approvals.items():
            if workflow_id and approval.get("workflow_id") != workflow_id:
                continue
            execution = self._executions.get(exec_id)
            if execution:
                approvals.append({
                    "execution_id": exec_id,
                    "workflow_id": approval["workflow_id"],
                    "node_id": approval["node_id"],
                    "requested_at": execution.get("started_at"),
                })
        return approvals

    def _to_response(self, workflow: dict) -> WorkflowResponse:
        """Convert workflow dict to response model."""
        return WorkflowResponse(
            id=workflow["id"],
            name=workflow["name"],
            description=workflow.get("description"),
            nodes=[WorkflowNode(**n) for n in workflow.get("nodes", [])],
            edges=[WorkflowEdge(**e) for e in workflow.get("edges", [])],
            triggers=[WorkflowTrigger(**t) for t in workflow.get("triggers", [])],
            is_active=workflow.get("is_active", True),
            created_at=workflow["created_at"],
            updated_at=workflow["updated_at"],
            last_run_at=workflow.get("last_run_at"),
            run_count=workflow.get("run_count", 0),
        )

    def _execution_to_response(self, execution: dict) -> WorkflowExecutionResponse:
        """Convert execution dict to response model."""
        return WorkflowExecutionResponse(
            id=execution["id"],
            workflow_id=execution["workflow_id"],
            status=ExecutionStatus(execution["status"]),
            input_data=execution.get("input_data", {}),
            output_data=execution.get("output_data"),
            node_results=[
                NodeExecutionResult(**r) for r in execution.get("node_results", [])
            ],
            error=execution.get("error"),
            started_at=execution["started_at"],
            finished_at=execution.get("finished_at"),
        )

# Singleton instance
workflow_service = WorkflowService()

import html
import json
import logging
import os
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Optional

try:
    import openpyxl  # type: ignore
except ImportError:  # pragma: no cover
    openpyxl = None  # type: ignore

from backend.app.services.ai_services import build_excel_llm_call_1_prompt
from backend.app.services.templates import MODEL, get_openai_client
from backend.app.services.infra_services import call_chat_completion, extract_tokens, normalize_token_braces, strip_code_fences
from backend.app.services.infra_services import render_html_to_png

logger = logging.getLogger("neura.excel.verify")

@dataclass
class ExcelInitialResult:
    html_path: Path
    png_path: Optional[Path]

def _extract_marked_section(text: str, begin: str, end: str) -> Optional[str]:
    pattern = re.compile(re.escape(begin) + r"([\s\S]*?)" + re.escape(end))
    match = pattern.search(text)
    if not match:
        return None
    return match.group(1).strip()

def _normalize_token(name: str) -> str:
    if name is None:
        return ""
    text = str(name).strip().lower()
    if not text:
        return ""
    text = re.sub(r"[^0-9a-z]+", "_", text)
    text = re.sub(r"_+", "_", text)
    return text.strip("_")

def _row_has_values(values) -> bool:
    for value in values:
        if value is None:
            continue
        if isinstance(value, str):
            if value.strip():
                return True
            continue
        return True
    return False

def _is_numeric_cell(text: str) -> bool:
    try:
        float(text.replace(",", ""))
        return True
    except ValueError:
        return False

def _is_sequential_numbers(cells: list[str]) -> bool:
    if len(cells) < 3:
        return False
    try:
        nums = [int(float(c)) for c in cells]
        return nums == list(range(nums[0], nums[0] + len(nums)))
    except (ValueError, TypeError):
        return False

def _detect_header_row(rows: list[tuple], *, max_scan: int = 10) -> int:
    """Score rows 0..max_scan and return the index of the best header row.

    Signals: unique text count, all-numeric penalty, sequential number penalty,
    merged-cell penalty, row position bonus.
    """
    best_index = -1
    best_score = float("-inf")

    for idx, row in enumerate(rows[:max_scan]):
        if not _row_has_values(row):
            continue
        cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
        if not cells:
            continue
        score = 0.0

        # Signal 1: Unique text count — headers have many distinct labels
        unique_text = len(set(c.lower() for c in cells))
        score += min(unique_text, 15) * 2.0

        # Signal 2: All-numeric penalty
        numeric_count = sum(1 for c in cells if _is_numeric_cell(c))
        if len(cells) > 0 and numeric_count / len(cells) > 0.8:
            score -= 20.0

        # Signal 3: Sequential number penalty
        if _is_sequential_numbers(cells):
            score -= 15.0

        # Signal 4: Merged-cell penalty — title rows have few cells spanning many columns
        non_empty = len(cells)
        total_cols = len(row)
        if total_cols > 3 and non_empty <= 2:
            score -= 10.0

        # Signal 5: Row position bonus
        if 1 <= idx <= 5:
            score += 1.0

        if score > best_score:
            best_score = score
            best_index = idx

    return best_index if best_index >= 0 else 0

def _ensure_label(value: object, idx: int) -> str:
    if value not in (None, ""):
        text = str(value).strip()
        if text:
            return text
    return f"Column {idx + 1}"

def _stringify_cell(value: object) -> str:
    if value is None:
        return ""
    try:
        text = str(value)
    except Exception:
        return ""
    return text.strip()

def _build_placeholder_samples(tokens: list[str], data_row: list[str]) -> dict[str, str]:
    samples: dict[str, str] = {}
    if not tokens:
        return samples
    for idx, token in enumerate(tokens):
        placeholder = token
        value = ""
        if data_row is not None and idx < len(data_row):
            cell = data_row[idx]
            if cell is not None:
                value = str(cell).strip()
        samples[placeholder] = value or "NOT_VISIBLE"
    return samples

def _sheet_snapshot_for_llm(sheet, *, max_rows: int = 20, max_preface_rows: int = 6) -> tuple[dict[str, Any], list[dict[str, Any]], list[str], int]:
    rows = list(sheet.iter_rows(values_only=True))
    header_index = _detect_header_row(rows)
    header_row = rows[header_index] if header_index < len(rows) else []

    header_labels = [_ensure_label(value, idx) for idx, value in enumerate(header_row)]
    preface_rows = rows[: max(header_index, 0)]

    data_rows = []
    if header_index >= 0:
        for row in rows[header_index + 1 :]:
            if _row_has_values(row):
                data_rows.append(row)
    else:
        data_rows = [row for row in rows if _row_has_values(row)]

    data_row_count = len(data_rows)
    sample_rows: list[dict[str, Any]] = []
    for offset, row in enumerate(data_rows[: max_rows]):
        cells = [_stringify_cell(row[idx] if idx < len(row) else "") for idx in range(len(header_labels))]
        sample_rows.append(
            {
                "row_index": header_index + offset + 2 if header_index >= 0 else offset + 1,
                "cells": cells,
            }
        )

    token_plan: list[dict[str, Any]] = []
    seen_tokens: dict[str, int] = {}
    for idx, label in enumerate(header_labels):
        norm = _normalize_token(label) or f"col_{idx + 1}"
        if norm in seen_tokens:
            seen_tokens[norm] += 1
            norm = f"{norm}_{seen_tokens[norm]}"
        else:
            seen_tokens[norm] = 1
        token_name = f"row_{norm}"
        sample_value = ""
        if sample_rows and idx < len(sample_rows[0]["cells"]):
            sample_value = sample_rows[0]["cells"][idx]
        token_plan.append(
            {
                "token": token_name,
                "header": label,
                "column_index": idx,
                "sample": sample_value,
            }
        )

    first_data_row = sample_rows[0]["cells"] if sample_rows else []
    grid_preview: list[list[str]] = []
    preview_limit = header_index + 1 + max_rows if header_index >= 0 else max_rows
    for row in rows[:preview_limit]:
        cols = max(len(header_labels), 1)
        grid_preview.append([_stringify_cell(row[idx] if idx < len(row) else "") for idx in range(cols)])

    snapshot = {
        "sheet_title": str(sheet.title or "Sheet1"),
        "preface_rows": [[_stringify_cell(cell) for cell in row] for row in preface_rows[-max_preface_rows:]],
        "headers": header_labels,
        "token_plan": token_plan,
        "sample_rows": sample_rows,
        "grid_preview": grid_preview,
        "sheet_notes": {
            "header_row_index": header_index,
            "data_row_count": data_row_count,
            "column_count": len(header_labels),
            "non_empty_rows": sum(1 for row in rows if _row_has_values(row)),
        },
    }

    return snapshot, token_plan, first_data_row, data_row_count

def _sheet_to_placeholder_html(sheet) -> tuple[str, list[str], list[str]]:
    rows = list(sheet.iter_rows(values_only=True))
    header_index = _detect_header_row(rows)
    header_row = rows[header_index] if header_index < len(rows) else None

    placeholder_tokens: list[str] = []
    if header_row:
        header_labels = [_ensure_label(value, idx) for idx, value in enumerate(header_row)]
    else:
        header_labels = []

    placeholder_tokens = []
    placeholder_cells: list[str] = []
    data_labels: list[str] = []
    seen_tokens: dict[str, int] = {}
    for idx, label in enumerate(header_labels):
        norm = _normalize_token(label) or f"col_{idx + 1}"
        if norm in seen_tokens:
            seen_tokens[norm] += 1
            norm = f"{norm}_{seen_tokens[norm]}"
        else:
            seen_tokens[norm] = 1
        token = f"row_{norm}"
        placeholder_tokens.append(token)
        placeholder_cells.append("<td>{" + token + "}</td>")
        data_labels.append(norm)

    if placeholder_tokens:
        th_cells = [
            f'<th data-label="{html.escape(data_label)}">{html.escape(label)}</th>'
            for data_label, label in zip(data_labels, header_labels)
        ]
        thead_html = f"<thead><tr>{''.join(th_cells)}</tr></thead>"
        tbody_html = f"<tbody><tr>{''.join(placeholder_cells)}</tr></tbody>"
    else:
        thead_html = "<thead><tr></tr></thead>"
        tbody_html = "<tbody><tr></tr></tbody>"

    styles = """
    <style>
      @page { size: A4; margin: 24mm; }
      body { font-family: Arial, sans-serif; }
      table { border-collapse: collapse; width: 100%; }
      td, th { border: 1px solid #999; padding: 6px 8px; vertical-align: top; }
    </style>
    """
    head = f"<head><meta charset='utf-8'>{styles}</head>"
    title = html.escape(str(sheet.title or "Sheet1"))
    caption = f"<caption style='caption-side:top;font-weight:700;text-align:left;margin:6px 0'>{title}</caption>"
    table = f'<table id="data-table">{caption}{thead_html}{tbody_html}</table>'
    html_text = f"<html>{head}<body>{table}</body></html>"

    first_data_row: list[str] = []
    if header_index >= 0:
        for row in rows[header_index + 1 :]:
            if _row_has_values(row):
                first_data_row = [_stringify_cell(row[idx] if idx < len(row) else "") for idx in range(len(header_labels))]
                break

    return html_text, placeholder_tokens, first_data_row

def _request_excel_llm_template(snapshot: dict[str, Any], sheet_html: str, schema_payload: Optional[dict[str, Any]] = None) -> tuple[str, Optional[dict[str, Any]]]:
    snapshot_json = json.dumps(snapshot, ensure_ascii=False, indent=2)
    sheet_html_payload = sheet_html or "<html></html>"
    schema_json = json.dumps(schema_payload or {}, ensure_ascii=False, separators=(",", ":"))
    prompt = build_excel_llm_call_1_prompt(snapshot_json, sheet_html_payload, schema_json)
    content = [{"type": "text", "text": prompt}]
    client = get_openai_client()
    response = call_chat_completion(
        client,
        model=MODEL,
        messages=[{"role": "user", "content": content}],
        description="excel_template_initial_html",
    )
    raw_content = strip_code_fences(response.choices[0].message.content or "")

    html_section = _extract_marked_section(raw_content, "<!--BEGIN_HTML-->", "<!--END_HTML-->")
    if html_section is None:
        raise RuntimeError("Excel LLM response missing HTML markers")
    html_clean = normalize_token_braces(html_section.strip())

    schema_section = _extract_marked_section(raw_content, "<!--BEGIN_SCHEMA_JSON-->", "<!--END_SCHEMA_JSON-->")
    schema_doc = None
    if schema_section:
        try:
            schema_doc = json.loads(schema_section)
        except json.JSONDecodeError:
            logger.warning(
                "excel_llm_schema_parse_failed",
                extra={"event": "excel_llm_schema_parse_failed", "snippet": schema_section[:200]},
            )
    return html_clean, schema_doc

def _sheet_to_reference_html(sheet, *, max_rows: int = 5) -> str:
    """
    Build a data-only HTML snapshot of the original Excel sheet (no placeholders),
    using the first non-empty row as header and up to `max_rows` subsequent data rows.
    This serves as the reference image for fidelity preview and LLM context.
    """
    rows = list(sheet.iter_rows(values_only=True))

    header_index = _detect_header_row(rows)
    header_row = rows[header_index] if header_index < len(rows) else []

    header_labels = [_ensure_label(value, idx) for idx, value in enumerate(header_row)]

    styles = """
    <style>
      @page { size: A4; margin: 24mm; }
      body { font-family: Arial, sans-serif; }
      table { border-collapse: collapse; width: 100%; }
      td, th { border: 1px solid #999; padding: 6px 8px; vertical-align: top; }
    </style>
    """
    head = f"<head><meta charset='utf-8'>{styles}</head>"
    title = html.escape(str(sheet.title or "Sheet1"))
    caption = f"<caption style='caption-side:top;font-weight:700;text-align:left;margin:6px 0'>{title}</caption>"

    th_cells = "".join(f"<th>{html.escape(label)}</th>" for label in header_labels)
    thead_html = f"<thead><tr>{th_cells}</tr></thead>"

    # Collect up to max_rows data rows after header
    body_rows: list[str] = []
    if header_index >= 0:
        for row in rows[header_index + 1 : header_index + 1 + max_rows]:
            if not _row_has_values(row):
                continue
            tds = [html.escape("" if v is None else str(v)) for v in row]
            body_rows.append("<tr>" + "".join(f"<td>{v}</td>" for v in tds) + "</tr>")
    if not body_rows:
        body_rows.append("<tr></tr>")
    tbody_html = "<tbody>" + "".join(body_rows) + "</tbody>"

    table = f'<table id="data-table">{caption}{thead_html}{tbody_html}</table>'
    return f"<html>{head}<body>{table}</body></html>"

def xlsx_to_html_preview(
    excel_path: Path,
    out_dir: Path,
    *,
    page_size: str = "A4",
    dpi: int = 144,
    db_path: Path | None = None,
) -> ExcelInitialResult:
    """
    Load the first worksheet of an Excel file and produce:
      - template_p1.html: LLM-generated HTML template (tokens preserved)
      - schema_ext.json: optional schema emitted by the LLM
      - sample_rows.json: literal samples for row_* tokens based on the first data row
      - report_final.png: screenshot of the generated template
      - reference_p1.html/png: literal snapshot of the worksheet data for fidelity preview
    """
    if openpyxl is None:  # pragma: no cover
        raise RuntimeError("openpyxl is required. Install with `pip install openpyxl`.")

    out_dir.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(filename=str(excel_path), data_only=True)
    sheet = wb.active
    # Enforce a simple safety/UX constraint for initial Excel uploads:
    # Limit the number of non-empty data rows to a maximum (default 30).
    # If exceeded, ask the user to delete extra rows and re-upload.
    try:
        max_rows_env = os.getenv("EXCEL_MAX_DATA_ROWS", "30").strip()
        MAX_DATA_ROWS = int(max_rows_env) if max_rows_env else 30
    except Exception:
        MAX_DATA_ROWS = 30

    snapshot, token_plan, first_data_row, data_row_count = _sheet_snapshot_for_llm(sheet, max_rows=MAX_DATA_ROWS)
    sheet_prototype_html, placeholder_tokens, placeholder_first_row = _sheet_to_placeholder_html(sheet)
    if data_row_count > MAX_DATA_ROWS:
        raise RuntimeError(
            f"Excel verification failed: found {data_row_count} data rows; maximum allowed is {MAX_DATA_ROWS}. "
            "Please delete extra rows and upload the file again."
        )

    placeholder_sample_map = _build_placeholder_samples(
        [token for token in placeholder_tokens],
        placeholder_first_row,
    )

    html_text, schema_payload = _request_excel_llm_template(snapshot, sheet_prototype_html)
    tokens_expected = set(placeholder_tokens)
    tokens_present = set(extract_tokens(normalize_token_braces(html_text)))
    missing_tokens = sorted(tokens_expected - tokens_present)
    if missing_tokens:
        logger.warning(
            "excel_llm_missing_tokens",
            extra={"event": "excel_llm_missing_tokens", "missing": missing_tokens},
        )
        html_text, schema_payload = sheet_prototype_html, None
    else:
        html_lower = html_text.lower()
        expected_labels = [
            token.replace("row_", "", 1)
            for token in placeholder_tokens
            if token.startswith("row_")
        ]
        missing_labels = [
            label for label in expected_labels if f'data-label="{label}"' not in html_lower
        ]
        if missing_labels:
            logger.warning(
                "excel_llm_missing_data_labels",
                extra={"event": "excel_llm_missing_data_labels", "missing": missing_labels},
            )
            html_text, schema_payload = sheet_prototype_html, None
        else:
            missing_exact = [
                label for label in expected_labels if f'<th data-label="{label}">' not in html_text
            ]
            if missing_exact:
                logger.warning(
                    "excel_llm_data_label_order",
                    extra={"event": "excel_llm_data_label_order", "missing": missing_exact},
                )
                html_text, schema_payload = sheet_prototype_html, None
    html_path = out_dir / "template_p1.html"
    html_path.write_text(html_text, encoding="utf-8")

    if schema_payload:
        schema_path = out_dir / "schema_ext.json"
        schema_path.write_text(json.dumps(schema_payload, ensure_ascii=False, indent=2), encoding="utf-8")

    tokens_in_template = set(extract_tokens(normalize_token_braces(html_text)))
    sample_row_map = {
        token: placeholder_sample_map.get(token, "NOT_VISIBLE") for token in tokens_in_template if token in placeholder_sample_map
    }
    sample_payload = {"sample_row": sample_row_map}
    sample_rows_path = out_dir / "sample_rows.json"
    sample_rows_path.write_text(json.dumps(sample_payload, ensure_ascii=False, indent=2), encoding="utf-8")

    # Build a data-only reference HTML to snapshot the original Excel content
    reference_html = _sheet_to_reference_html(sheet, max_rows=MAX_DATA_ROWS)
    reference_html_path = out_dir / "reference_p1.html"
    reference_html_path.write_text(reference_html, encoding="utf-8")

    reference_png_path: Optional[Path] = None
    try:
        reference_png_path = out_dir / "reference_p1.png"
        render_html_to_png(reference_html_path, reference_png_path, page_size=page_size, dpi=dpi)
    except Exception:
        logger.warning(
            "excel_reference_png_render_failed",
            extra={"event": "excel_reference_png_render_failed", "html": str(reference_html_path)},
            exc_info=True,
        )
        reference_png_path = None

    template_png_path: Optional[Path] = None
    try:
        template_png_path = out_dir / "report_final.png"
        render_html_to_png(html_path, template_png_path, page_size=page_size, dpi=dpi)
    except Exception:
        logger.warning(
            "excel_template_png_render_failed",
            extra={"event": "excel_template_png_render_failed", "html": str(html_path)},
            exc_info=True,
        )
        template_png_path = None

    thumbnail_path = template_png_path or reference_png_path
    return ExcelInitialResult(html_path=html_path, png_path=thumbnail_path)

# mypy: ignore-errors

import json
import logging
import re
from pathlib import Path
from typing import Any, Callable, Iterable, Mapping, Optional, Sequence

from backend.app.services.ai_services import (
    get_prompt_generator_assets as get_pdf_prompt_generator_assets,
)
from backend.app.services.templates import get_openai_client
from backend.app.services.infra_services import write_artifact_manifest, write_json_atomic, write_text_atomic
from backend.app.services.infra_services import call_chat_completion
from backend.app.services.infra_services import (
    validate_contract_v2,
    validate_generator_output_schemas,
    validate_generator_sql_pack,
)

logger = logging.getLogger(__name__)

DEFAULT_MODEL = "qwen"

class GeneratorAssetsError(RuntimeError):
    """Raised when generator asset creation fails."""

def _ensure_iter(values: Iterable[Any] | None) -> list[Any]:
    if not values:
        return []
    return list(values)

def _normalized_tokens(tokens: Iterable[str] | None) -> list[str]:
    cleaned: list[str] = []
    if not tokens:
        return cleaned
    seen: set[str] = set()
    for raw in tokens:
        text = str(raw or "").strip()
        if not text or text in seen:
            continue
        seen.add(text)
        cleaned.append(text)
    return cleaned

def _normalize_sql_dialect(value: str | None) -> str:
    """
    Normalize dialect names so the runtime consistently receives DuckDB-friendly SQL.
    Treat legacy "sqlite" declarations as DuckDB since the DB is now backed by pandas DataFrames.
    """
    text = str(value or "").strip().lower()
    if text in ("", "sqlite", "duckdb"):
        return "duckdb"
    if text in ("postgres", "postgresql"):
        return "postgres"
    return text or "duckdb"

_SQL_PLACEHOLDER_PATTERNS = (
    r"\.\.\.",
    r"\bTBD\b",
    r"\bTODO\b",
    r"ADD_SQL_HERE",
    r"<add_sql_here>",
)

def _extract_aliases(sql: str | None) -> list[str]:
    if not sql:
        return []
    pattern = re.compile(r"\bAS\s+([A-Za-z_][\w]*)", re.IGNORECASE)
    return pattern.findall(sql)

def _sql_contains_keyword(sql: str, keyword: str) -> bool:
    return bool(re.search(rf"\b{re.escape(keyword)}\b", sql, re.IGNORECASE))

def _sql_uses_table_reference(sql: str) -> bool:
    return bool(re.search(r"[A-Za-z_][\w]*\.[A-Za-z_][\w]*", sql))

def _validate_entrypoint_sql_shape(entrypoints: Mapping[str, str]) -> list[str]:
    issues: list[str] = []
    for section in ("header", "rows", "totals"):
        sql = (entrypoints.get(section) or "").strip()
        if not sql:
            continue
        if not _sql_contains_keyword(sql, "select"):
            issues.append(f"missing_select:{section}")
        needs_from = _sql_uses_table_reference(sql)
        if needs_from and not _sql_contains_keyword(sql, "from"):
            issues.append(f"missing_from:{section}")
        for pattern in _SQL_PLACEHOLDER_PATTERNS:
            if re.search(pattern, sql, re.IGNORECASE):
                issues.append(f"incomplete_sql:{section}")
                break
    return issues

def _derive_output_schemas(contract: Mapping[str, Any] | None) -> dict[str, list[str]]:
    """
    Build header/rows/totals token lists from the contract when the LLM response
    omits explicit output_schemas.
    """
    contract = contract or {}
    tokens_section = contract.get("tokens") if isinstance(contract, Mapping) else {}

    header_tokens = _normalized_tokens(contract.get("header_tokens") if isinstance(contract, Mapping) else None)
    row_tokens = _normalized_tokens(contract.get("row_tokens") if isinstance(contract, Mapping) else None)
    totals_tokens = _normalized_tokens(
        list((contract.get("totals") or {}).keys()) if isinstance(contract, Mapping) else None
    )

    if isinstance(tokens_section, Mapping):
        header_tokens = header_tokens or _normalized_tokens(tokens_section.get("scalars"))
        row_tokens = row_tokens or _normalized_tokens(tokens_section.get("row_tokens"))
        totals_tokens = totals_tokens or _normalized_tokens(tokens_section.get("totals"))

    return {
        "header": header_tokens,
        "rows": row_tokens,
        "totals": totals_tokens,
    }

def _validate_entrypoints_against_schema(
    entrypoints: Mapping[str, str],
    output_schemas: Mapping[str, Sequence[str]],
) -> list[str]:
    issues: list[str] = []
    for section, expected in output_schemas.items():
        expected_tokens = [str(token) for token in expected or []]
        sql = entrypoints.get(section, "") or ""
        if expected_tokens:
            aliases = _extract_aliases(sql)
            if aliases:
                alias_set = {alias.strip() for alias in aliases if alias}
                mismatch = [token for token in expected_tokens if token not in alias_set]
                if mismatch:
                    issues.append(f"schema_mismatch:{section}")
            else:
                # If we cannot infer aliases but schema expects values, record a warning
                issues.append(f"schema_ambiguous:{section}")
    return issues

def _default_entrypoints(existing: Mapping[str, str] | None) -> dict[str, str]:
    normalized = {}
    existing = existing or {}
    for name in ("header", "rows", "totals"):
        sql = existing.get(name)
        if sql:
            normalized[name] = str(sql)
        else:
            normalized[name] = "SELECT 1;"
    return normalized

def _render_sql_script(script: str | None, entrypoints: Mapping[str, str]) -> str:
    if script and script.strip():
        return script.strip() + ("\n" if not script.strip().endswith("\n") else "")
    sections = []
    for name in ("header", "rows", "totals"):
        sql = entrypoints.get(name, "").strip() or "SELECT 1;"
        section = f"-- {name.upper()} --\n{sql.strip()}"
        sections.append(section)
    return "\n\n".join(sections) + "\n"

def _json_safe(value: Any) -> Any:
    """Convert Paths and other non-serialisable types into JSON-friendly shapes."""
    if isinstance(value, Path):
        return value.as_posix()
    if isinstance(value, dict):
        return {str(k): _json_safe(v) for k, v in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_safe(v) for v in value]
    if isinstance(value, set):
        return [_json_safe(v) for v in value]
    return value

def _prepare_step4_for_prompt(step4_output: Mapping[str, Any]) -> dict[str, Any]:
    allowed_keys = (
        "contract",
        "overview_md",
        "step5_requirements",
        "assumptions",
        "warnings",
        "validation",
    )
    payload: dict[str, Any] = {}
    for key in allowed_keys:
        if key in step4_output and step4_output[key] is not None:
            payload[key] = step4_output[key]
    return _json_safe(payload)

_JSON_FENCE_RE = re.compile(r"```(?:json)?\s*(\{.*\})\s*```", re.DOTALL | re.IGNORECASE)
_JSON_ARRAY_CLOSURE_FIXES: tuple[tuple[str, str], ...] = (
    ('\n    },\n    "row_computed"', '\n    ],\n    "row_computed"'),
    ('\n    },\n    "header_tokens"', '\n    ],\n    "header_tokens"'),
    ('\n    },\n    "row_tokens"', '\n    ],\n    "row_tokens"'),
    ('\n    },\n    "row_order"', '\n    ],\n    "row_order"'),
)

def _repair_generator_json(text: str) -> Mapping[str, Any] | None:
    """
    Attempt to repair simple LLM mistakes where arrays are closed with `}` instead of `]`.
    """
    working = text
    max_attempts = len(_JSON_ARRAY_CLOSURE_FIXES) + 1
    for _ in range(max_attempts):
        try:
            return json.loads(working)
        except json.JSONDecodeError:
            replaced = False
            for needle, replacement in _JSON_ARRAY_CLOSURE_FIXES:
                if needle in working:
                    working = working.replace(needle, replacement, 1)
                    replaced = True
                    break
            if not replaced:
                return None
    try:
        return json.loads(working)
    except json.JSONDecodeError:
        return None

def _ensure_reshape_rule_purpose(contract: Mapping[str, Any]) -> None:
    """
    Backfill reshape rule purpose strings when the LLM omits them.
    """
    reshape_rules = contract.get("reshape_rules")
    if not isinstance(reshape_rules, list):
        return
    for idx, rule in enumerate(reshape_rules):
        if not isinstance(rule, dict):
            continue
        purpose = rule.get("purpose")
        if isinstance(purpose, str) and purpose.strip():
            continue
        alias = str(rule.get("alias") or "").strip()
        strategy = str(rule.get("strategy") or "").strip()
        if alias and strategy:
            summary = f"{alias} {strategy} rule"
        elif alias:
            summary = f"{alias} reshape rule"
        elif strategy:
            summary = f"{strategy} reshape rule"
        else:
            summary = f"Reshape rule {idx + 1}"
        rule["purpose"] = summary[:120]

def _ensure_row_order(contract: Mapping[str, Any]) -> None:
    """
    Normalise row_order to a non-empty list, deriving it from order_by when omitted.
    """
    row_order_raw = contract.get("row_order")
    cleaned: list[str] = []
    if isinstance(row_order_raw, str):
        text = row_order_raw.strip()
        if text:
            cleaned = [text]
    elif isinstance(row_order_raw, list):
        cleaned = [str(item).strip() for item in row_order_raw if str(item or "").strip()]

    order_block = contract.get("order_by")
    rows_spec: Any = None
    if isinstance(order_block, Mapping):
        rows_spec = order_block.get("rows")
    elif isinstance(order_block, list):
        rows_spec = list(order_block)
        contract["order_by"] = {"rows": list(rows_spec)}
    elif isinstance(order_block, str) and order_block.strip():
        rows_spec = [order_block.strip()]
        contract["order_by"] = {"rows": list(rows_spec)}
    else:
        contract["order_by"] = {"rows": []}
    if isinstance(rows_spec, list):
        rows_order = [str(item).strip() for item in rows_spec if str(item or "").strip()]
    elif isinstance(rows_spec, str) and rows_spec.strip():
        rows_order = [rows_spec.strip()]
        contract["order_by"]["rows"] = rows_order
    else:
        rows_order = []

    contract["row_order"] = cleaned or rows_order or ["ROWID"]

def _normalize_contract_join(contract: Mapping[str, Any]) -> None:
    """
    Drop or sanitise join blocks that the LLM emits with blank keys so schema validation
    is not tripped up by placeholder values.
    """
    join = contract.get("join")
    if not isinstance(join, Mapping):
        return

    parent_table = str(join.get("parent_table") or "").strip()
    parent_key = str(join.get("parent_key") or "").strip()
    child_table = str(join.get("child_table") or "").strip()
    child_key = str(join.get("child_key") or "").strip()

    if not parent_table or not parent_key:
        if any((parent_table, parent_key, child_table, child_key)):
            logger.info(
                "generator_contract_join_dropped",
                extra={
                    "event": "generator_contract_join_dropped",
                    "parent_table": parent_table,
                    "parent_key": parent_key,
                    "child_table": child_table,
                    "child_key": child_key,
                },
            )
        contract.pop("join", None)
        return

    normalized_join: dict[str, str] = {
        "parent_table": parent_table,
        "parent_key": parent_key,
    }
    if child_table and child_key:
        normalized_join["child_table"] = child_table
        normalized_join["child_key"] = child_key

    contract["join"] = normalized_join

def _parse_generator_response(raw_text: str) -> dict[str, Any]:
    text = (raw_text or "").strip()
    if not text:
        raise GeneratorAssetsError("Generator response was empty.")
    match = _JSON_FENCE_RE.search(text)
    if match:
        text = match.group(1).strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError as exc:
        repaired_payload = _repair_generator_json(text)
        if repaired_payload is not None:
            return repaired_payload
        start = text.find("{")
        end = text.rfind("}")
        if start != -1 and end != -1 and end > start:
            snippet = text[start : end + 1]
            try:
                return json.loads(snippet)
            except Exception:
                pass
        raise GeneratorAssetsError(f"Generator response was not valid JSON: {exc}") from exc

def _prepare_messages(
    payload: dict[str, Any],
    prompt_getter: Callable[[], dict[str, str]],
) -> list[dict[str, str]]:
    prompts = prompt_getter() or {}
    system_text = prompts.get("system") or "You generate SQL packs."
    user_template = prompts.get("user")
    user_payload = json.dumps(payload, ensure_ascii=False, indent=2)
    if user_template and "{payload}" in user_template:
        user_text = user_template.replace("{payload}", user_payload)
    else:
        user_text = f"{user_template or ''}\n{user_payload}"
    user_text = (
        f"{user_text.strip()}\n\nIMPORTANT: Output strictly valid JSON. Use double quotes for every key and string "
        f"value. Do not include trailing commas or comments."
    )
    return [
        {"role": "system", "content": system_text},
        {"role": "user", "content": user_text.strip()},
    ]

def _write_outputs(
    template_dir: Path,
    contract: Mapping[str, Any],
    entrypoints: Mapping[str, str],
    output_schemas: Mapping[str, Sequence[str]],
    params: dict[str, list[str]],
    dialect: str,
    needs_user_fix: list[str],
    invalid: bool,
    summary: Mapping[str, Any],
    key_tokens: Iterable[str] | None,
    script: str | None,
) -> dict[str, Path]:
    generator_dir = template_dir / "generator"
    generator_dir.mkdir(parents=True, exist_ok=True)
    contract_path = template_dir / "contract.json"
    write_json_atomic(contract_path, contract, ensure_ascii=False, indent=2, step="generator_contract")

    sql_path = generator_dir / "sql_pack.sql"
    script_text = _render_sql_script(script, entrypoints)
    write_text_atomic(sql_path, script_text, encoding="utf-8", step="generator_sql_pack")

    output_schemas_path = generator_dir / "output_schemas.json"
    write_json_atomic(
        output_schemas_path,
        output_schemas,
        ensure_ascii=False,
        indent=2,
        step="generator_output_schemas",
    )

    meta_payload = {
        "dialect": dialect,
        "entrypoints": entrypoints,
        "params": params,
        "needs_user_fix": needs_user_fix,
        "invalid": invalid,
        "summary": summary,
        "cached": False,
        "key_tokens": _normalized_tokens(key_tokens),
    }
    meta_path = generator_dir / "generator_assets.json"
    write_json_atomic(meta_path, meta_payload, ensure_ascii=False, indent=2, step="generator_assets_meta")

    write_artifact_manifest(
        template_dir,
        step="generator_assets_v1",
        files={
            "contract.json": contract_path,
            "sql_pack.sql": sql_path,
            "output_schemas.json": output_schemas_path,
            "generator_assets.json": meta_path,
        },
        inputs=["generator_assets_v1"],
        correlation_id=None,
    )

    return {
        "contract": contract_path,
        "sql_pack": sql_path,
        "output_schemas": output_schemas_path,
        "generator_assets": meta_path,
    }

def build_generator_assets_from_payload(
    *,
    template_dir: Path,
    step4_output: Mapping[str, Any],
    final_template_html: str,
    reference_pdf_image: Any = None,
    reference_worksheet_html: str | None = None,
    catalog_allowlist: Iterable[str] | None = None,
    params_spec: Sequence[str] | None = None,
    sample_params: Mapping[str, Any] | None = None,
    force_rebuild: bool = False,
    dialect: str | None = None,
    key_tokens: Iterable[str] | None = None,
    prompt_getter: Optional[Callable[[], dict[str, str]]] = None,
    require_contract_join: bool = True,
) -> dict[str, Any]:
    template_dir = Path(template_dir)
    template_dir.mkdir(parents=True, exist_ok=True)

    catalog_list = [str(item) for item in (catalog_allowlist or []) if str(item).strip()]
    params_list = list(params_spec or [])
    sample_params_dict = dict(sample_params or {})

    step4_prompt_payload = _prepare_step4_for_prompt(step4_output)

    request_payload = {
        "final_template_html": final_template_html,
        # For PDF-driven flows, this conveys the raster reference. For Excel flows,
        # callers may supply `reference_worksheet_html` instead (preferred).
        "reference_pdf_image": reference_pdf_image,
        "step4_output": step4_prompt_payload,
        "catalog_allowlist": catalog_list,
        "params_spec": params_list,
        "sample_params": _json_safe(sample_params_dict),
        "force_rebuild": bool(force_rebuild),
        "key_tokens": _normalized_tokens(key_tokens),
    }
    if isinstance(reference_worksheet_html, str) and reference_worksheet_html.strip():
        request_payload["reference_worksheet_html"] = reference_worksheet_html

    client = get_openai_client()
    if client is None:
        raise GeneratorAssetsError("OpenAI client is not configured.")

    prompt_factory = prompt_getter or get_pdf_prompt_generator_assets
    messages = _prepare_messages(request_payload, prompt_factory)
    try:
        raw_response = call_chat_completion(
            client,
            model=DEFAULT_MODEL,
            messages=messages,
            description="generator_assets_v1",
        )
    except Exception as exc:  # pragma: no cover - network failures bubble up
        raise GeneratorAssetsError(f"Generator LLM call failed: {exc}") from exc

    try:
        content = raw_response.choices[0].message.content or ""
        response_payload = _parse_generator_response(content)
    except GeneratorAssetsError:
        raise
    except Exception as exc:  # pragma: no cover - malformed response
        raise GeneratorAssetsError(f"Generator response was not valid JSON: {exc}") from exc

    sql_pack_raw = response_payload.get("sql_pack") or {}
    contract = response_payload.get("contract")
    if not isinstance(contract, Mapping) or not contract:
        raise GeneratorAssetsError("Generator LLM response did not include a contract payload.")
    try:
        _ensure_reshape_rule_purpose(contract)
        _ensure_row_order(contract)
        _normalize_contract_join(contract)
        validate_contract_v2(contract, require_join=require_contract_join)
    except Exception as exc:
        raise GeneratorAssetsError(f"Generator contract failed validation: {exc}") from exc

    output_schemas_payload = response_payload.get("output_schemas")
    if isinstance(output_schemas_payload, Mapping):
        output_schemas = {
            "header": _normalized_tokens(output_schemas_payload.get("header")),
            "rows": _normalized_tokens(output_schemas_payload.get("rows")),
            "totals": _normalized_tokens(output_schemas_payload.get("totals")),
        }
    else:
        output_schemas = _derive_output_schemas(contract)

    # Validate generator structures
    try:
        validate_generator_output_schemas(output_schemas)
    except Exception as exc:
        raise GeneratorAssetsError(f"Invalid output schemas: {exc}") from exc

    # Normalise entrypoints and params
    entrypoints_raw = sql_pack_raw.get("entrypoints")
    if isinstance(entrypoints_raw, Mapping) and entrypoints_raw:
        entrypoints = _default_entrypoints(entrypoints_raw)
    else:
        legacy_entrypoints = {
            key: value
            for key, value in {
                "header": sql_pack_raw.get("header"),
                "rows": sql_pack_raw.get("rows"),
                "totals": sql_pack_raw.get("totals"),
            }.items()
            if isinstance(value, str)
        }
        entrypoints = _default_entrypoints(legacy_entrypoints)

    script_text = sql_pack_raw.get("script")
    if isinstance(script_text, str) and script_text.strip():
        script_for_validation = script_text
    else:
        script_for_validation = _render_sql_script(script_text, entrypoints)

    params_section = sql_pack_raw.get("params")
    required_params: list[str] = []
    optional_params: list[str] = []
    if isinstance(params_section, Mapping):
        required_params.extend(params_section.get("required") or [])
        optional_params.extend(params_section.get("optional") or [])
    elif isinstance(params_section, Sequence):
        required_params.extend(params_section)

    base_params = _normalized_tokens(params_list)
    key_param_tokens = _normalized_tokens(key_tokens)
    for token in key_param_tokens:
        if token not in base_params:
            base_params.append(token)
    for item in base_params:
        if item not in required_params:
            required_params.append(item)
    optional_params = [p for p in optional_params if p not in required_params]
    params_normalized = {"required": required_params, "optional": optional_params}

    sql_pack_normalized = {
        "dialect": _normalize_sql_dialect(sql_pack_raw.get("dialect") or response_payload.get("dialect") or dialect),
        "script": script_for_validation,
        "entrypoints": entrypoints,
        "params": params_normalized,
    }

    try:
        validate_generator_sql_pack(sql_pack_normalized)
    except Exception as exc:
        raise GeneratorAssetsError(f"Invalid SQL pack: {exc}") from exc

    schema_issues = _validate_entrypoints_against_schema(entrypoints, output_schemas)
    shape_issues = _validate_entrypoint_sql_shape(entrypoints)
    if schema_issues:
        logger.warning(
            "generator_assets_schema_issues", extra={"event": "generator_assets_schema_issues", "issues": schema_issues}
        )
    if shape_issues:
        logger.warning(
            "generator_assets_sql_shape_issues",
            extra={"event": "generator_assets_sql_shape_issues", "issues": shape_issues},
        )

    needs_user_fix = _ensure_iter(response_payload.get("needs_user_fix")) + schema_issues + shape_issues
    invalid = bool(response_payload.get("invalid")) or bool(schema_issues) or bool(shape_issues)
    summary = response_payload.get("summary") or {}
    selected_dialect = _normalize_sql_dialect(response_payload.get("dialect") or dialect)

    artifacts = _write_outputs(
        template_dir=template_dir,
        contract=contract,
        entrypoints=entrypoints,
        output_schemas=output_schemas,
        params=params_normalized,
        dialect=selected_dialect,
        needs_user_fix=needs_user_fix,
        invalid=invalid,
        summary=summary,
        key_tokens=key_param_tokens,
        script=script_text,
    )

    result = {
        "artifacts": artifacts,
        "needs_user_fix": needs_user_fix,
        "invalid": invalid,
        "dialect": selected_dialect,
        "params": params_normalized,
        "dry_run": response_payload.get("dry_run"),
        "summary": summary,
        "cached": False,
    }
    return result

def load_generator_assets_bundle(template_dir: Path) -> dict[str, Any] | None:
    generator_dir = Path(template_dir) / "generator"
    meta_path = generator_dir / "generator_assets.json"
    if not meta_path.exists():
        return None

    try:
        meta = json.loads(meta_path.read_text(encoding="utf-8"))
    except Exception:
        return None

    sql_path = generator_dir / "sql_pack.sql"
    output_schemas_path = generator_dir / "output_schemas.json"
    contract_path = Path(template_dir) / "contract.json"

    artifacts: dict[str, Path] = {}
    if contract_path.exists():
        artifacts["contract"] = contract_path
    if sql_path.exists():
        artifacts["sql_pack"] = sql_path
    if output_schemas_path.exists():
        artifacts["output_schemas"] = output_schemas_path
    if meta_path.exists():
        artifacts["generator_assets"] = meta_path

    bundle = {
        "artifacts": artifacts,
        "meta": meta,
        "needs_user_fix": _ensure_iter(meta.get("needs_user_fix")),
        "invalid": bool(meta.get("invalid")),
        "dialect": meta.get("dialect"),
        "params": meta.get("params") or {"required": [], "optional": []},
        "summary": meta.get("summary") or {},
        "dry_run": None,
        "cached": True,
        "key_tokens": meta.get("key_tokens") or [],
    }
    return bundle

# Ensure dramatiq broker has a results backend so @dramatiq.actor(store_results=True)
# doesn't fail at import time.  In the worker process the real Redis backend is
# configured; here we add a stub only if no results middleware is present yet.
try:
    import dramatiq as _dm
    _broker = _dm.get_broker()
    if not any("Results" in type(m).__name__ for m in _broker.middleware):
        from dramatiq.results import Results
        from dramatiq.results.backends import StubBackend
        _broker.add_middleware(Results(backend=StubBackend()))
except Exception:
    pass

"""Prometheus metrics for Dramatiq worker monitoring."""
import time
from dramatiq.middleware import Middleware

try:
    from prometheus_client import Counter, Histogram, Gauge, start_http_server
    HAS_PROMETHEUS = True
except ImportError:
    HAS_PROMETHEUS = False

if HAS_PROMETHEUS:
    TASK_COMPLETED = Counter("dramatiq_task_completed_total", "Tasks completed", ["actor", "queue", "status"])
    TASK_DURATION = Histogram("dramatiq_task_duration_seconds", "Task duration", ["actor", "queue"], buckets=[0.1, 0.5, 1, 5, 10, 30, 60, 120, 300, 600])
    TASK_ACTIVE = Gauge("dramatiq_tasks_active", "Active tasks", ["actor", "queue"])
    TASK_ENQUEUED = Counter("dramatiq_task_enqueued_total", "Tasks enqueued", ["actor", "queue"])
    DLQ_SIZE = Gauge("dramatiq_dlq_size", "Dead-lettered tasks", ["queue"])

class WorkerMetricsMiddleware(Middleware):
    """Export Dramatiq metrics to Prometheus."""

    def before_process_message(self, broker, message):
        if not HAS_PROMETHEUS:
            return
        message.options["_prom_start"] = time.monotonic()
        TASK_ACTIVE.labels(actor=message.actor_name, queue=message.queue_name).inc()

    def after_enqueue(self, broker, message, delay):
        if not HAS_PROMETHEUS:
            return
        TASK_ENQUEUED.labels(actor=message.actor_name, queue=message.queue_name).inc()

    def after_process_message(self, broker, message, *, result=None, exception=None):
        if not HAS_PROMETHEUS:
            return
        status = "error" if exception else "success"
        elapsed = time.monotonic() - message.options.get("_prom_start", time.monotonic())
        TASK_COMPLETED.labels(actor=message.actor_name, queue=message.queue_name, status=status).inc()
        TASK_DURATION.labels(actor=message.actor_name, queue=message.queue_name).observe(elapsed)
        TASK_ACTIVE.labels(actor=message.actor_name, queue=message.queue_name).dec()

    def after_skip_message(self, broker, message):
        if not HAS_PROMETHEUS:
            return
        DLQ_SIZE.labels(queue=message.queue_name).inc()

def start_metrics_server(port: int = 9191) -> None:
    """Start a standalone Prometheus HTTP metrics server for the worker process."""
    if not HAS_PROMETHEUS:
        return
    start_http_server(port)

"""Agent execution tasks - durable via Dramatiq + Redis."""
import os
import dramatiq
import logging

from backend.app.repositories import agent_task_repository

logger = logging.getLogger("neura.worker.agents")

try:
    from dramatiq.rate_limits import ConcurrentRateLimiter
    from dramatiq.rate_limits.backends import RedisBackend
    _rate_backend = RedisBackend(url=os.getenv("NEURA_REDIS_URL", "redis://localhost:6379/0"))
    AGENT_MUTEX = ConcurrentRateLimiter(_rate_backend, key="agent-execution", limit=4)
except (ImportError, ConnectionError, OSError, ValueError):
    logger.warning("Rate limiter unavailable; agent concurrency will not be limited")
    AGENT_MUTEX = None

@dramatiq.actor(
    queue_name="agents",
    max_retries=3,
    min_backoff=5000,
    max_backoff=120000,
    time_limit=300_000,  # 5 min
    store_results=True,
)
def run_agent(task_id: str, agent_type: str, params: dict):
    """Execute an agent task. Survives worker crashes via Redis persistence."""
    from backend.app.services.agents import agent_service_v2

    # Idempotency: skip if already completed
    existing = agent_task_repository.get_task(task_id)
    if existing and existing.status in ("completed", "failed", "cancelled"):
        logger.info("agent_skipped_idempotent", extra={
            "event": "agent_skipped_idempotent", "task_id": task_id, "status": existing.status,
        })
        return

    if AGENT_MUTEX is not None:
        with AGENT_MUTEX.acquire():
            _run_agent(task_id, agent_type, params, agent_service_v2)
    else:
        _run_agent(task_id, agent_type, params, agent_service_v2)

def _run_agent(task_id: str, agent_type: str, params: dict, service):
    """Core agent execution logic, extracted for rate-limiter wrapping."""
    try:
        service.execute_task_sync(task_id, agent_type, params)
        logger.info("agent_task_completed", extra={"event": "agent_task_completed", "task_id": task_id})
    except Exception:
        logger.exception("agent_task_failed", extra={"event": "agent_task_failed", "task_id": task_id})
        raise

"""Export tasks - durable via Dramatiq + Redis."""

import asyncio
import logging

import dramatiq

logger = logging.getLogger("neura.worker.exports")

@dramatiq.actor(
    queue_name="exports",
    max_retries=2,
    min_backoff=3000,
    max_backoff=60000,
    time_limit=300_000,  # 5 min
    store_results=True,
)
def export_document(document_id: str, output_format: str, options: dict | None = None) -> dict:
    """Create an export job for a document.

    Note: The export service currently persists jobs and returns metadata; actual
    export execution can be implemented as a follow-up step without changing the
    task contract.
    """
    from backend.app.services.infra_services import export_service

    opts = options or {}
    try:
        return asyncio.run(export_service.create_export_job(document_id=document_id, format=output_format, options=opts))
    except RuntimeError:
        # If already in an event loop (rare in worker context), run synchronously.
        return _run_export_in_loop(export_service, document_id, output_format, opts)

def _run_export_in_loop(export_service, document_id: str, output_format: str, opts: dict) -> dict:
    loop = asyncio.new_event_loop()
    try:
        return loop.run_until_complete(
            export_service.create_export_job(document_id=document_id, format=output_format, options=opts)
        )
    finally:
        loop.close()

"""Document ingestion tasks - durable via Dramatiq + Redis."""
import os
import dramatiq
import logging

logger = logging.getLogger("neura.worker.ingestion")

try:
    from dramatiq.rate_limits import ConcurrentRateLimiter
    from dramatiq.rate_limits.backends import RedisBackend
    _rate_backend = RedisBackend(url=os.getenv("NEURA_REDIS_URL", "redis://localhost:6379/0"))
    INGESTION_MUTEX = ConcurrentRateLimiter(_rate_backend, key="ingestion-pipeline", limit=5)
except Exception:
    INGESTION_MUTEX = None

@dramatiq.actor(
    queue_name="ingestion",
    max_retries=2,
    min_backoff=3000,
    max_backoff=60000,
    time_limit=120_000,
    store_results=True,
)
def ingest_document(doc_id: str, source_type: str, source_url: str, **kwargs):
    """Ingest a document from an external source. Survives worker crashes via Redis persistence."""
    logger.info("ingestion_started", extra={"event": "ingestion_started", "doc_id": doc_id, "source_type": source_type})

    # Acquire concurrent rate limiter (max 5 simultaneous ingestions).
    if INGESTION_MUTEX is not None:
        with INGESTION_MUTEX.acquire():
            return _run_ingestion(doc_id, source_type, source_url, **kwargs)
    return _run_ingestion(doc_id, source_type, source_url, **kwargs)

def _run_ingestion(doc_id: str, source_type: str, source_url: str, **kwargs):
    """Core ingestion logic, extracted for rate-limiter wrapping."""
    from backend.app.services.ingestion_service import IngestService
    service = IngestService()
    result = service.ingest(doc_id=doc_id, source_type=source_type, source_url=source_url, **kwargs)
    logger.info("ingestion_completed", extra={"event": "ingestion_completed", "doc_id": doc_id})
    return result

"""Report generation tasks - durable via Dramatiq + Redis."""
import os

import dramatiq
import logging

logger = logging.getLogger("neura.worker.reports")

try:
    from dramatiq.rate_limits import ConcurrentRateLimiter
    from dramatiq.rate_limits.backends import RedisBackend
    _rate_backend = RedisBackend(url=os.getenv("NEURA_REDIS_URL", "redis://localhost:6379/0"))
    REPORT_MUTEX = ConcurrentRateLimiter(_rate_backend, key="report-generation", limit=3)
except Exception:
    REPORT_MUTEX = None

@dramatiq.actor(
    queue_name="reports",
    max_retries=3,
    min_backoff=5000,
    max_backoff=300000,
    time_limit=600_000,  # 10 min hard limit
    store_results=True,
)
def generate_report(job_id: str, template_id: str, connection_id: str, output_format: str = "pdf", **kwargs):
    """Generate a report. Survives worker crashes via Redis persistence."""
    from backend.app.repositories import state_store

    # Idempotency: skip if already completed
    existing = state_store.get_job(job_id)
    if existing and existing.get("status") in ("succeeded", "failed"):
        logger.info("report_skipped_idempotent", extra={"event": "report_skipped_idempotent", "job_id": job_id})
        return existing.get("result", {})

    # Acquire concurrent rate limiter (max 3 simultaneous report generations).
    # If the limiter is unavailable (Redis down, import failure), proceed without it.
    if REPORT_MUTEX is not None:
        with REPORT_MUTEX.acquire():
            return _run_report(job_id, template_id, connection_id, output_format, state_store)
    return _run_report(job_id, template_id, connection_id, output_format, state_store)

def _run_report(job_id: str, template_id: str, connection_id: str, output_format: str, state_store):
    """Core report generation logic, extracted for rate-limiter wrapping."""
    try:
        state_store.record_job_start(job_id)
        # record_job_step enforces keyword-only args after `name`
        state_store.record_job_step(
            job_id,
            "generate",
            status="running",
            label="Starting report generation",
        )

        from backend.engine_all import ReportPipeline
        pipeline = ReportPipeline()
        result = pipeline.run(
            template_id=template_id,
            connection_id=connection_id,
            output_format=output_format,
        )

        state_store.record_job_completion(job_id, status="succeeded", result=result)
        logger.info("report_generated", extra={"event": "report_generated", "job_id": job_id})
        return result
    except Exception as exc:
        state_store.record_job_completion(job_id, status="failed", error=str(exc))
        logger.exception("report_generation_failed", extra={"event": "report_generation_failed", "job_id": job_id})
        raise

"""Webhook delivery tasks - durable via Dramatiq + Redis."""

import asyncio
import logging

import dramatiq

from backend.app.utils import validate_url

logger = logging.getLogger("neura.worker.webhooks")

@dramatiq.actor(
    queue_name="webhooks",
    max_retries=5,
    min_backoff=1000,
    max_backoff=120_000,
    time_limit=30_000,  # 30s
    store_results=True,
)
def send_webhook(url: str, payload: dict, headers: dict | None = None, method: str = "POST") -> dict:
    """Deliver a webhook notification with SSRF protection."""
    validate_url(url)
    from backend.app.services.infra_services import distribution_service

    safe_headers = headers or {}
    safe_payload = payload or {}
    try:
        return asyncio.run(
            distribution_service.send_webhook(
                document_id=str(safe_payload.get("document_id") or "unknown"),
                webhook_url=url,
                method=method,
                headers=safe_headers,
                payload=safe_payload,
            )
        )
    except RuntimeError:
        return _run_webhook_in_loop(distribution_service, url, method, safe_headers, safe_payload)

def _run_webhook_in_loop(distribution_service, url: str, method: str, headers: dict, payload: dict) -> dict:
    loop = asyncio.new_event_loop()
    try:
        return loop.run_until_complete(
            distribution_service.send_webhook(
                document_id=str(payload.get("document_id") or "unknown"),
                webhook_url=url,
                method=method,
                headers=headers,
                payload=payload,
            )
        )
    finally:
        loop.close()

